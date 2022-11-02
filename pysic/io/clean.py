#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
    pysic.io.update.py : function to update ice core spreadsheet
"""

from pysic.core.core import __CoreVersion__
import logging
import os
import openpyxl
import numpy as np

from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

from pysic.tools import version2int
from pysic.property import prop_associated_tab

import pysic
pysic_fp = pysic.__path__[0]

#ic_path = '/home/megavolts/git/pysic/pysic/ressources/AAA_BB-YYYYXXZZ-N_P-1.4.6.xlsx'


data_row_n = 42

# Cell style
# unit
unit_style = Font(name='Geneva', charset=1, family=2.0, b=False, i=True, strike=None, outline=None,
                  shadow=None, condense=None, color=None, extend=None, sz=9.0, u=None, vertAlign=None,
                  scheme=None)
unit_alignment = Alignment(horizontal='right')

# Border style
# bottom border
b_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# bottom_left_border
bl_border = Border(left=Side(border_style='hair', color='00000000'), right=Side(border_style=None),
                   top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# left_border
l_border = Border(left=Side(border_style='hair', color='00000000'), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style=None))
# right border
r_border = Border(left=Side(border_style=None), right=Side(border_style='hair', color='00000000'),
                  top=Side(border_style=None), bottom=Side(border_style=None))
# bottom right border
br_border = Border(left=Side(border_style=None), right=Side(border_style='hair', color='00000000'),
                  top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# no_border
no_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                   top=Side(border_style=None), bottom=Side(border_style=None))
noBorder = no_border

## METADATA TAB STYLE
m_header_style = NamedStyle(name="m_header_style")
m_header_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=True, italic=False, color='FF000000')  # black
m_header_style.alignment = Alignment(horizontal='center', wrapText=True, vertical='top')
m_header_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFE0E0E0", fill_type="solid")

m_subheader_style = NamedStyle(name="m_subheader_style")
m_subheader_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=False, color='FF000000')
m_subheader_style.alignment = Alignment(horizontal='center', wrapText=True, vertical='center')
m_subheader_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFE0E0E0", fill_type="solid")

m_unit_style = NamedStyle(name="m_unit_style")
m_unit_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=True, color='FF000000')
m_unit_style.alignment = Alignment(horizontal='center', wrapText=True, vertical='center')
m_unit_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFE0E0E0", fill_type="solid")

m_unit_l_style = NamedStyle(name="m_unit_l_style")
m_unit_l_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=True, color='FF000000')
m_unit_l_style.alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
m_unit_l_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFE0E0E0", fill_type="solid")

m_subheader_style_coring = NamedStyle(name="m_subheader_style_coring")
m_subheader_style_coring.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=True, italic=False, color='FF000000')  # black
m_subheader_style_coring.alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
m_subheader_style_coring.fill = PatternFill(start_color="FFCCCCCC", end_color="FFBFBFBF", fill_type="solid")

m_unit_style_coring = NamedStyle(name="m_unit_style_coring")
m_unit_style_coring.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=True, color='FF000000')
m_unit_style_coring.alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
m_unit_style_coring.fill = PatternFill(start_color="FFCCCCCC", end_color="FFBFBFBF", fill_type="solid")

m_comment_style_coring = NamedStyle(name="m_comment_style_coring")
m_comment_style_coring.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=True, color='FF000000')
m_comment_style_coring.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_comment_style_coring.fill = PatternFill(start_color="FFCCCCCC", end_color="FFBFBFBF", fill_type="solid")

## PROPERTY TAB
mid_grey = "FFCCCCCC"
light_grey = "FFDDDDDD"
dark_grey = "FFB2B2B2"
white = "ffFFFFFF"
noFill = PatternFill(fill_type=None)

# data right: black font, right aligned, white/no background
m_data_style = NamedStyle(name="m_data_style")
m_data_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=False, color='FF000000')  # black
m_data_style.alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
#m_data_style.fill = PatternFill(fgColor=white, bgColor=white, fill_type="solid")
m_data_style.fill = PatternFill(fill_type=None)

# data entry style, left aligned: black font, left alignment, white/no background
m_data_l_style = NamedStyle(name="m_data_l_style")
m_data_l_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=False, color='FF000000')  # black
m_data_l_style.alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
#m_data_l_style.fill = PatternFill(fgColor=white, bgColor=white, fill_type="solid")
m_data_l_style.fill = PatternFill(fill_type=None)

# background style: dark grey
m_bkg_style = NamedStyle(name="m_bkg_style")
m_bkg_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=False, color='FF000000')
m_bkg_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_bkg_style.fill = PatternFill(start_color=dark_grey, end_color=dark_grey, fill_type="solid")
m_bottom_style = m_bkg_style

# subheader right: black font, right aligned, light grey background
m_subheader_r_style = NamedStyle(name="m_subheader_r_style")
m_subheader_r_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=False, color='FF000000')
m_subheader_r_style.alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
m_subheader_r_style.fill = PatternFill(start_color=light_grey, end_color=light_grey, fill_type="solid")

# subheader left: black font, left aligned, light grey background
m_subheader_l_style = NamedStyle(name="m_subheader_l_style")
m_subheader_l_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=False, italic=False, color='FF000000')
m_subheader_l_style.alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
m_subheader_l_style.fill = PatternFill(start_color=light_grey, end_color=light_grey, fill_type="solid")

# header right: black bold font, right aligned, light grey background
m_header_r_style = NamedStyle(name="m_header_r_style")
m_header_r_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=True, italic=False, color='FF000000')
m_header_r_style.alignment = Alignment(horizontal='right', wrapText=False, vertical='center')
m_header_r_style.fill = PatternFill(start_color=light_grey, end_color=light_grey, fill_type="solid")

# header left: black bold font, left aligned, light grey background
m_header_l_style = NamedStyle(name="m_header_l_style")
m_header_l_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=True, italic=False, color='FF000000')
m_header_l_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_header_l_style.fill = PatternFill(start_color=light_grey, end_color=light_grey, fill_type="solid")

# header title: black bold font, left aligned, dark grey background
m_header_title_style = NamedStyle(name="m_header_title_style")
m_header_title_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=True, italic=False, color='FF000000')
m_header_title_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_header_title_style.fill = PatternFill(start_color=mid_grey, end_color=mid_grey, fill_type="solid")

# version: black bold font, left aligned, bkg background
m_version_style = NamedStyle(name="m_version_style")
m_version_style.font = Font(name='Geneva', charset=1, family=2.0, sz=9.0, bold=True, italic=False, color='FF000000')
m_version_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_version_style.fill = PatternFill(start_color=dark_grey, end_color=dark_grey, fill_type="solid")


# Data validation
vert_ref_list = DataValidation(type="list", formula1='"ice surface, ice/water interface, snow surface, water level"')
vert_dir_list = DataValidation(type="list", formula1='"up, down"')
texture_dv = DataValidation(type="list", formula1='{0}!$F$3:$F$27'.format(quote_sheetname('lists')), allow_blank=True)
inclusion_dv = DataValidation(type="list", formula1='{0}!$G$3:$G$27'.format(quote_sheetname('lists')), allow_blank=True)
description_dv = DataValidation(type="list", formula1='{0}!$H$3:$H$27'.format(quote_sheetname('lists')), allow_blank=True)


#TODO: find surname and given name form os
user = 'Marc Oggier'

def clean_spreadsheet(wb):
    metadata_core(wb)


def metadata_core(wb, user=user):
    sheetname = 'metadata-core'

    max_row = wb[sheetname].max_row
    # find INSTRUMENT section start row
    row_idx = 20
    while 'INSTRUMENTS:' not in wb[sheetname].cell(row_idx, 1).value and row_idx < max_row:
        row_idx += 1
    row_instrument = row_idx
    if row_idx == wb[sheetname].max_row:
        logging.error('\t%s - no INSTRUMENTS section found in %s' %(wb[sheetname]['C1'].value, sheetname))

    # find VERSION section  start row
    row_idx = row_instrument + 1
    while wb[sheetname].cell(row_idx, 1).value != 'VERSION' and row_idx < max_row:
        row_idx += 1
    if row_idx == wb[sheetname].max_row:
        logging.error('\t%s - no VERSION section found in %s' %(wb[sheetname]['C1'].value, sheetname))
    # insert blank line in between both section if not present
    row_insert_idx = row_idx
    if wb[sheetname].cell(row_idx - 1, 1).value != None:
        for merged_cell in wb[sheetname].merged_cells.ranges:
            if merged_cell.right[0][0] >= row_insert_idx:
                merged_cell.shift(0, 1)
        wb[sheetname].insert_rows(row_insert_idx, 1)

    # formatting
    for row_idx in range(row_instrument + 1, row_insert_idx):
        for col_idx in range(1, 3):
            wb[sheetname].cell(row_idx, col_idx).style = 'm_subheader_l_style'
            if row_idx < row_insert_idx - 2:
                wb[sheetname].cell(row_idx, col_idx).border = b_border
            else:
                wb[sheetname].cell(row_idx, col_idx).border = noBorder
        new_range = 'C' + str(row_idx) + ':H' + str(row_idx)
        wb[sheetname].merge_cells(new_range)
        wb[sheetname].cell(row_idx, col_idx + 1).style = 'm_data_l_style'
        wb[sheetname].cell(row_idx, col_idx + 1).border = noBorder
        if row_idx < row_insert_idx - 2:
            wb[sheetname].cell(row_idx, col_idx + 1).border = b_border
        else:
            wb[sheetname].cell(row_idx, col_idx + 1).border = noBorder
    for col_idx in range(1, wb[sheetname].max_column):
        wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'
        wb[sheetname].cell(row_idx, col_idx).border = noBorder
    wb[sheetname].cell(row_insert_idx, 1).border = noBorder  # no bottom border on the last line
    
    # CLEANING VERSION NUMBER
    # find VERSION section  start row
    row_idx = row_instrument + 1
    max_row = wb[sheetname].max_row
    while wb[sheetname].cell(row_idx, 1).value != 'VERSION' and row_idx < max_row:
        row_idx += 1
    row_version = row_idx

    # Find line number with first version entries
    row_idx = row_version + 2
    first_entry = row_idx
    last_entry = 0
    while last_entry == 0 and row_idx <= max_row:
        if wb['metadata-core'].cell(row_idx, 1).value is None:
            last_entry = row_idx - 1
        row_idx += 1

    # check incremental number (col 0)
    incremental_number = []
    incremental_error = []
    date = []
    author = []
    modification = []
    for row_idx in range(first_entry, last_entry + 1):
        incremental_number += [int(wb[sheetname].cell(row_idx, 1).value)]
        date += [wb[sheetname].cell(row_idx, 2).value]
        _author = wb[sheetname].cell(row_idx, 3).value
        if _author is None and 'PS122' in wb['metadata-core']['C1'].value:
            _author = 'Marc Oggier'
        author += [_author]
        modification += [wb[sheetname].cell(row_idx, 5).value]
        if row_idx > int(first_entry) and incremental_number[-1] <= incremental_number[-2]:
            incremental_error += [row_idx]

    if len(incremental_error) > 0:
        if 'Initial data entry' in modification:
            if modification.index('Initial data entry') > 0 and len(incremental_number) < 3:
                # inverse row in incremental version
                wb[sheetname].cell(first_entry, 2).value = date[1]
                wb[sheetname].cell(first_entry, 3).value = author[1]
                wb[sheetname].cell(first_entry, 5).value = modification[1]
                wb[sheetname].cell(last_entry, 1).value = 2
                wb[sheetname].cell(last_entry, 2).value = date[0]
                wb[sheetname].cell(last_entry, 3).value = author[0]
                wb[sheetname].cell(last_entry, 5).value = modification[0]
            else:
                logging.error('%s - %s: error in incremental version' % (wb['metadata-core']['C1'].value, sheetname))