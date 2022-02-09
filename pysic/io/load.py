#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
    pysic.io.load.py : function to import ice core data from ice core data spreadsheet
"""

import logging
import openpyxl
import pandas as pd
import numpy as np

from pysic.io import subvariable_dict, property2sheet
from pysic.io import update
from pysic.tools import inverse_dict, parse_datetimetz, parse_coordinate
from pysic.core.core import __CoreVersion__

MAX_ROW = 1000
TOL = 1e-12

# DEGUG
logging.basicConfig(level=logging.DEBUG)

# TODO: modifiy reading to account for false bottom
def ic_from_path(ic_path, ic_property=None, drop_empty=False, fill_missing=True):
    """
    :param ic_path:
        string: path to the ice core data. Ice core data should be stored in defined spreadsheet
    :param ic_property:
        list of string: properties to import. If not specified all properties are imported.
    :param drop_empty:
        boolean: True drop empty profile
    :param fill_missing:
        boolean, default True: If true, missing section are filled with np.nan/none
    :return:
    """
    from pysic.core.core import Core

    logger = logging.getLogger(__name__)

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet

    ws_summary = wb['metadata-station']
    ws_metadata_core = wb['metadata-core']

    name = ws_metadata_core['C1'].value
    if ws_summary['C1'].value:
        version = ws_summary['C1'].value
        if version != __CoreVersion__:
            logger.warning("%s: ice core data is updated from %s to %s" % (name, version, __CoreVersion__))
            # TODO: update function pysic.core.update_data_spreadsheet
            update.ice_core_data(ic_path)
            ic_from_path(ic_path, ic_property, drop_empty=drop_empty, fill_missing=fill_missing)
    else:
        logger.error("(%s) ice core spreadsheet version not unavailable" % name)
        wb.close()

    comments = []

    logger.info("importing data for %s" % ic_path)

    # Import datetime with tz
    datetime = parse_datetimetz(ws_metadata_core['C2'].value, ws_metadata_core['C3'].value, ws_metadata_core['D3'].value)

    # project
    campaign = ws_summary['C3'].value
    site = ws_summary['C4'].value

    # Location
    lat_start_deg = parse_coordinate(ws_summary['C9'].value, ws_summary['D9'].value, ws_summary['E9'].value)
    lon_start_deg = parse_coordinate(ws_summary['C10'].value, ws_summary['D10'].value, ws_summary['E10'].value)
    lat_end_deg = parse_coordinate(ws_summary['F9'].value, ws_summary['G9'].value, ws_summary['H9'].value)
    lon_end_deg = parse_coordinate(ws_summary['F10'].value, ws_summary['G10'].value, ws_summary['H10'].value)

    # Station time
    station_date_start = parse_datetimetz(ws_summary['C14'].value, ws_summary['C15'].value, ws_summary['C16'].value)
    station_date_end = parse_datetimetz(ws_summary['D14'].value, ws_summary['D15'].value, ws_summary['D16'].value)

    # Snow Depth
    n_snow = 1
    row_snow = 19
    snow_depth = []
    while ws_summary.cell(row=row_snow, column=3+n_snow).value is not None:
        snow_depth.append(ws_summary.cell(row=row_snow, column=3+n_snow).value)
        n_snow += 1
    snow_depth = np.array(snow_depth)
    if any(snow_depth) > 1:  # check for metric unit
        logger.warning('%s: check if snow depth are reported in cm rather than m' % name)

    # Snow depth measurement
    if len(snow_depth) > 0:
        if isinstance(snow_depth[-1], str):
            comments.append(snow_depth[-1])
            snow_depth = pd.to_numeric(snow_depth[:-1], errors='coerce')
        else:
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')

    # Snow average
    if isinstance(ws_summary.cell(row=row_snow, column=3).value, (float, int)):
        snow_depth_avg = ws_summary.cell(row=row_snow, column=3).value
    elif not np.isnan(snow_depth).all():
        snow_depth_avg = np.nanmean(snow_depth)
    else:
        snow_depth_avg = [np.nan]
    if len(snow_depth) == 0:
        snow_depth = snow_depth_avg

    # Ice Thickness
    try:
        h_i = ws_metadata_core['D7'].value
    except ValueError:
        logger.info('(%s) no ice thickness information ' % name)
        h_i = np.nan
    else:
        if h_i == 'n/a':
            comments.append('ice thickness not available')
            logger.info('(%s) ice thickness is not available (n/a)' % name)
            h_i = np.nan
        elif not isinstance(h_i, (int, float)):
            logger.info('%s ice thickness is not a number' % name)
            comments.append('ice thickness not available')
            h_i = np.nan
    if h_i > 10:  # check for metric unit
        logger.warning('%s: check if ice thickness is reported in cm rather than m' % name)

    # Ice Draft
    try:
        h_d = ws_metadata_core['D8'].value
    except ValueError:
        logger.info('(%s) no ice draft information ' % name)
        h_d = np.nan
    else:
        if h_d == 'n/a':
            comments.append('ice draft not available')
            logger.info('(%s) ice draft is not available (n/a)' % name)
            h_d = np.nan
        elif not isinstance(h_d, (int, float)):
            logger.info('%s ice draft is not a number' % name)
            comments.append('ice draft not available')
            h_d = np.nan
    if h_d > 10:  # check for metric unit
        logger.warning('%s: check if ice draft is reported in cm rather than m' % name)

    # Ice freeboard
    if not (np.isnan(h_d) and np.isnan(h_i)):
        h_f = h_i - h_d
    else:
        try:
            h_f = ws_metadata_core['D9'].value
        except ValueError:
            logger.info('(%s) no ice draft information ' % name)
            h_f = np.nan
        else:
            if h_f == 'n/a':
                comments.append('ice draft not available')
                logger.info('(%s) ice draft is not available (n/a)' % name)
                h_f = np.nan
            elif not isinstance(h_f, (int, float)):
                logger.info('%s ice draft is not a number' % name)
                comments.append('ice draft not available')
                h_f = np.nan
    if h_f > 10:  # check for metric unit
        logger.warning('%s: check if ice freeboard is reported in cm rather than m' % name)

    # Core Length l_c (measured in with ruler)
    try:
        l_c = ws_metadata_core['D10'].value
    except ValueError:
        logger.info('(%s) no ice core length' % name)
        l_c = np.nan
    else:
        if l_c == 'n/a':
            comments.append('ice core length not available')
            logger.info('(%s) ice core length is not available (n/a)' % name)
            l_c = np.nan
        elif not isinstance(l_c, (int, float)):
            logger.info('%s ice core length is not a number' % name)
            comments.append('ice core length not available')
            l_c = np.nan
    if l_c > 10:  # check for metric unit
        logger.warning('%s: check if ice freeboard is reported in cm rather than m' % name)

    core = Core(name, datetime, campaign, lat_start_deg, lon_start_deg, l_c, h_i, h_f, snow_depth)

    # Temperature values
    if ws_summary['C29'].value:
        core.t_air = ws_summary['C29'].value
    if ws_summary['C30'].value:
        core.t_snow_surface = ws_summary['C30'].value
    if ws_summary['C31'].value:
        core.t_ice_surface = ws_summary['C31'].value
    if ws_summary['C32'].value:
        core.t_water = ws_summary['C32'].value
    if ws_summary['C33'].value:
        core.s_water = ws_summary['C33'].value

    # Sampling event
    if ws_summary['C36'].value:
        core.station = ws_summary['C36'].value

    # Sampling protocol
    if ws_summary.cell(3, 39).value:
        core.protocol = ws_summary.cell(3, 39).value
    else:
        core.protocol = 'N/A'

    # Sampling instrument
    instrument_d = {}
    row_instrument = 21
    while ws_metadata_core.cell(row_instrument, 1).value is not None:
        instrument_d[ws_metadata_core.cell(row_instrument, 1).value] = ws_metadata_core.cell(row_instrument, 3).value
        row_instrument += 1
    core.instrument = instrument_d

    # Core collection
    m_col = 3
    row_collection = 37
    while ws_summary.cell(row_collection, m_col).value:
        core.add_to_collection(ws_summary.cell(row_collection, m_col).value)
        m_col += 1

    # comment
    if ws_summary['A49'].value is not None:
        comments.append(ws_summary['A42'].value)
    core.add_comment('; '.join(comments))

    # weather
    # TODO: read weather information

    # References
    reference_d = {}
    if ws_metadata_core['D13'].value is not None:
        if ws_metadata_core['D14'].value is not None:
            reference_d['ice'] = [ws_metadata_core['D13'].value, ws_metadata_core['D14'].value]
        else:
            reference_d['ice'] = [ws_metadata_core['D13'].value, 'up']
    if ws_metadata_core['D15'].value is not None:
        if ws_metadata_core['D16'].value is not None:
            reference_d['snow'] = [ws_metadata_core['D15'].value, ws_metadata_core['D16'].value]
        else:
            reference_d['snow'] = [ws_metadata_core['D15'].value, 'up']
    if ws_metadata_core['D17'].value is not None:
        if ws_metadata_core['D18'].value is not None:
            reference_d['seawater'] = [ws_metadata_core['D17'].value, ws_metadata_core['D18'].value]
        else:
            reference_d['seawater'] = [ws_metadata_core['D18'].value, 'up']
    core.reference.update(reference_d)

    # PAR
    if isinstance(ws_summary['C45'].value, (float, int)):
        core.par_incoming = ws_summary['C45'].value
        # read units
        par_unit = ws_summary['B45'].value
        core.unit.update({'par': par_unit})
    if isinstance(ws_summary['C46'].value, (float, int)):
        core.par_transmitted = ws_summary['C46'].value
        par_unit = ws_summary['B45'].value
        if 'par' in core.unit and par_unit != core.unit['par']:
            logger.error('%s\t\tPar unit do not match between incoming and transmitted. Par conversion not implemented')
        else:
            core.unit.update({'par': par_unit})

    # import property profile
    sheets = wb.sheetnames
    if ic_property is None:
        # TODO: special import for TEX, snow
        sheets = [sheet for sheet in sheets if (sheet not in ['tex', 'TM',  'summary', 'abreviation', 'locations',
                                                              'lists', 'Vf_oil_calculation', 'metadata-core',
                                                              'metadata-station', 'density-volume', 'sediment',
                                                              'ct']) and
                  (sheet.lower().find('fig') == -1)]
        # Always import 'salo18' then 'temperature' before the other
        if 'temp' in sheets:
            sheets.remove('temp')
            sheets = ['temp'] + sheets
        if 'salo18' in sheets:
            sheets.remove('salo18')
            sheets = ['salo18'] + sheets

        for sheet in sheets:
            ws_property = wb[sheet]
            if sheet == 'snow':
                profile = read_snow_profile(ws_property, ic_property=None, reference_d=reference_d)
                matter = 'snow'
            elif sheet == 'seawater':
                profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
                matter = 'seawater'
            elif sheet == 'sackhole' or sheet == 'brine':
                profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
                matter = 'brine'
            else:
                profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
                matter = 'ice'

            profile['matter'] = matter

#           def add_sw_salinity(profile):
            if matter == 'ice' and 'salinity' in profile.get_property() and core.s_water:
                # add sea water salinity at the correct
                #def profile.get_vref():
                v_ref_loc = profile.v_ref_loc.unique()
                if len(v_ref_loc) != 1:
                    logger.error('%s\t\tvertical reference location not unique')
                else:
                    v_ref_loc = v_ref_loc[0]
                    v_ref_dir = profile.v_ref_dir.unique()
                    if len(v_ref_dir) != 1:
                        logger.error('%s\t\tvertical reference direction not unique')
                    else:
                        v_ref_dir = v_ref_dir[0]
                        v_ref_h = profile.v_ref_h.unique()
                        if len(v_ref_h) != 1:
                            logger.error('%s\t\tvertical reference height not unique')
                        else:
                            v_ref_h = v_ref_h[0]
                headers = ['y_low', 'y_sup', 'salinity_ID', 'salinity_value', 'salinity_quality', 'v_ref_loc', 'v_ref_dir', 'v_ref_h', 'matter']
                # if v_ref_loc == 'ice bottom':
                #     # y_low =
                #     # y_sup =
                #     data = [y_low, y_sup, 's_sw', core.s_water, 0, v_ref_loc, v_ref_dir, v_ref_h, 'seawater']
                # elif v_ref_loc == 'ice surface':
                #     # y_low =
                #     # y_sup =
                #     data = [y_low, y_sup, 's_sw', core.s_water, 0, v_ref_loc, v_ref_dir, v_ref_h, 'seawater']
                # else:
                #     logger.error('%s\t\tTODO: implement v_ref_loc %s for profile%' %(core.name, v_ref_loc))


        #             if matter == 'ice' and profile.get_property() is not None and 'temperature' in profile.get_property():
#                 headers = ['y_mid', 'temperature_value', 'comment', 'variable', 'v_ref', 'matter']
#                 v_ref = profile.v_ref.unique()
#                 if len(v_ref) == 1:
#                     v_ref = v_ref[0]
#                     if v_ref == 'top':
#                         if not -1 in profile.y_mid.values:
#                             # air temperature 1 m above snow surface
#                             if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
#                                 data_surface = [-1, core.t_air, 'Air temperature', 'temperature', 'top', 'air']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                         if not 0 in profile.y_mid.values:
#                             # ice surface
#                             if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
#                                 data_surface = [0, core.t_ice_surface, 'Ice surface temperature', 'temperature', 'top', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                         if core.length - profile.y_mid.max() > TOL:
#                             # ice bottom / seawtaer
#                             if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
#                                 data_bottom = [core.length, core.t_water, 'Ice bottom temperature', 'temperature', 'top', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                                 data_bottom = [core.length, core.t_water, 'Seawater temperature', 'temperature', 'top', 'seawater']
#                                 profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                     elif v_ref == 'bottom':
#                         # air temperature
#                         if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
#                             data_surface = [1, core.t_air, 'Air temperature', 'temperature', 'bottom', 'air']
#                             profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                         if not 0 in profile.y_mid.values:
#                             # ice bottom
#                             if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
#                                 data_bottom = [0, core.t_water, 'Seawater temperature', 'temperature', 'bottom', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                         if np.abs(core.length - profile.y_mid.max()) < TOL:
#                             # ice surface
#                             if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
#                                 data_surface = [core.length, core.t_ice_surface, 'Ice surface temperature', 'temperature', 'bottom', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                 else:
#                     logger.error('%s - %s: vertical references mixed up ' %(core.name, sheet))
#                 profile = profile.sort_values(by='y_mid')
#             profile['name'] = name
#             if drop_empty:
#                 profile.drop_empty_property()
#
#             if not profile.empty:
#                 core.add_profile(profile)
#                 logger.info('(%s) data imported with success: %s' % (core.name, ", ".join(profile.get_property())))
#             else:
#                 logger.info('(%s) no data to import from %s ' % (core.name, sheet))
    else:
        if not isinstance(ic_property, list):
            if ic_property.lower().find('state variable')+1:
                ic_property = ['temperature', 'salinity']
            else:
                ic_property = [ic_property]

        _imported_variables = []
        for ic_prop in ic_property:
            if property2sheet[ic_prop] in sheets and ic_prop not in _imported_variables:
                sheet = property2sheet[ic_prop]
                ws_property = wb[sheet]
                property2import = [p for p in ic_property if p in inverse_dict(property2sheet)[sheet]]
                if sheet == 'snow':
                    profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d)
                    matter = 'snow'
                elif sheet == 'seawater':
                    matter = 'seawater'
                    profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, fill_missing=True)
                elif sheet == 'sackhole' or sheet == 'brine':
                    matter = 'brine'
                    profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, fill_missing=True)
                else:
                    matter = 'ice'
                    profile = read_generic_profile(ws_property, ic_property=None, reference_d=reference_d, fill_missing=True)
                profile['matter'] = matter

    # Add air, snow surface, ice surface, seawater temperature to temperature profile if exist


    # Add seawater salinity to salinity profile

        #
        #                 # Add temperature at ice surface for temperautre profile
        #                 if matter == 'ice' and profile.get_property() is not None and 'temperature' in profile.get_property():
        #                     headers = ['y_mid', 'temperature_value', 'comment', 'variable', 'v_ref', 'matter']
        #                     v_ref = profile.v_ref.unique()
        #                     if len(v_ref) == 1:
        #                         v_ref = v_ref[0]
        #                         if v_ref == 'top':
        #                             if not -1 in profile.y_mid.values:
        #                                 # air temperature 1 m above snow surface
        #                                 if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
        #                                     data_surface = [-1, core.t_air, 'Air temperature', 'temperature', 'top', 'air']
        #                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                             if not 0 in profile.y_mid.values:
        #                                 # ice surface
        #                                 if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
        #                                     data_surface = [0, core.t_ice_surface, 'Ice surface temperature', 'temperature',
        #                                                     'top', 'ice']
        #                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                             if core.length - profile.y_mid.max() > TOL:
        #                                 # ice bottom / seawtaer
        #                                 if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
        #                                     data_bottom = [core.length, core.t_water, 'Ice bottom temperature', 'temperature',
        #                                                    'top', 'ice']
        #                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
        #                                     data_bottom = [core.length, core.t_water, 'Seawater temperature', 'temperature',
        #                                                    'top', 'seawater']
        #                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
        #                         elif v_ref == 'bottom':
        #                             # air temperature
        #                             if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
        #                                 data_surface = [1, core.t_air, 'Air temperature', 'temperature', 'bottom', 'air']
        #                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                             if not 0 in profile.y_mid.values:
        #                                 # ice bottom
        #                                 if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
        #                                     data_bottom = [0, core.t_water, 'Seawater temperature', 'temperature', 'bottom',
        #                                                    'ice']
        #                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
        #                             if np.abs(core.length - profile.y_mid.max()) < TOL:
        #                                 # ice surface
        #                                 if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
        #                                     data_surface = [core.length, core.t_ice_surface, 'Ice surface temperature',
        #                                                     'temperature', 'bottom', 'ice']
        #                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
        #                     else:
        #                         logger.error('%s - %s: vertical references mixed up ' % (core.name, sheet))
        #                     profile = profile.sort_values(by='y_mid')

    return core


def read_generic_profile(ws_property, ic_property=None, reference_d={'ice': ['ice surface', 'down']}, core_length=np.nan, fill_missing=True):
    """
    :param ws_property:
        openpyxl.worksheet, worksheet property to import
    :param ic_property:
        str array, containing property to import. Default None import all properties available
    :param reference_d:
        dict, contain information of vertical reference system
    :param core_length:
        float, default np.nan (optional). Ice core length
    """
    from pysic.core.profile import Profile

    logger = logging.getLogger(__name__)

    # find last column number with column header and/or subheaders
    n_row = 1  # header row
    n_col_min = 1  # start column
    n_col = n_col_min
    empty_header = 0
    headers = []
    subheaders = []
    units = []
    header_dict = {}
    property_dict = {}
    while not empty_header and n_col < ws_property.max_column:
        # Read depth
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value  # header
            hs_ = ws_property.cell(n_row + 1, n_col).value  # subheader
            hu_ = ws_property.cell(n_row + 2, n_col).value

            headers.append(h_)
            subheaders.append(hs_)
            units.append(hu_)
            n_col += 1
        elif ws_property.cell(n_row + 1, n_col).value == 'ID':
            # Look for next property column number
            property_col = n_col
            while ws_property.cell(n_row + 1, property_col).value != 'value':
                property_col += 1




            hs_ = ws_property.cell(n_row + 1, n_col).value  # subheader
            headers.append(h_)
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row+2, n_col).value)
            n_col += 1
        elif isinstance(ws_property.cell(n_row+1, n_col).value, str):
            hs_ = ws_property.cell(n_row + 1, n_col).value   # subheader
            headers.append(h_)
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row+2, n_col).value)
            n_col += 1
        else:
            empty_header = 1

#            subheaders.append(hs_)
#            n_col += 1

    #
    # # Old method
    # # stop when header is empty or max column number is reached
    # while not empty_header and n_col < ws_property.max_column:
    #     if isinstance(ws_property.cell(n_row, n_col).value, str):
    #         h_ = ws_property.cell(n_row, n_col).value  # header
    #         hs_ = ws_property.cell(n_row + 1, n_col).value  # subheader
    #         headers.append(h_)
    #         subheaders.append(hs_)
    #         units.append(ws_property.cell(n_row + 2, n_col).value)
    #         n_col += 1
    #     elif isinstance(ws_property.cell(n_row+1, n_col).value, str):
    #         hs_ = ws_property.cell(n_row + 1, n_col).value   # subheader
    #         headers.append(h_)
    #         subheaders.append(hs_)
    #         units.append(ws_property.cell(n_row+2, n_col).value)
    #         n_col += 1
    #     else:
    #         empty_header = 1
    # n_col_max = n_col - 1

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Read center depth for continuous profile and step profile (if available)
    if 'depth center' in headers:
        loc3 = headers.index('depth center')
        headers[loc3] = 'y_mid'

        # look for 1st numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc3 + 1).value, (float, int)):
                min_row_3 = ii_row
                break

        y_mid = np.array(
            [ws_property.cell(row, loc3+1).value for row in range(min_row_3, max_row)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_mid))[::-1] if np.isnan(y_mid[ii])]
        y_mid = np.delete(y_mid, y_nan_loc)
    else:
        y_mid = []
        min_row_3 = np.nan

    # Read lower and upper section depth for step profile or set to nan for continuous profile
    if 'depth 1' in headers and 'depth 2' in headers:
        # find column with depth 1
        loc1 = headers.index('depth 1')
        headers[loc1] = 'y_low'
        loc2 = headers.index('depth 2')
        headers[loc2] = 'y_sup'

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc1 + 1).value, (float, int)):
                min_row_1 = ii_row
                break
            else:
                min_row_2 = np.nan
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc2 + 1).value, (float, int)):
                min_row_2 = ii_row
                break
            else:
                min_row_2 = np.nan

        min_row_12 = min(min_row_1, min_row_2)
        y_low = np.array([ws_property.cell(ii_row, loc1+1).value for ii_row in range(min_row_12, max_row + 1)]).astype(float)
        y_sup = np.array([ws_property.cell(ii_row, loc2+1).value for ii_row in range(min_row_12, max_row + 1)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_low))[::-1] if np.isnan(y_low[ii]) and np.isnan(y_sup[ii])]
        y_low = np.delete(y_low, y_nan_loc)
        y_sup = np.delete(y_sup, y_nan_loc)

        if len(y_mid) == len(y_low) and (min_row_3 == min_row_12 or np.isnan(min_row_3)):
            if np.any(np.abs(((y_low + y_sup) / 2) - y_mid > TOL)):
                logger.error('\t\t%s: y_mid are not mid points of y_low and y_sup. Computing y_mid = (y_low+y_sup)/2'
                             % ws_property.title)
                y_mid = (y_low + y_sup) / 2
            else:
                logger.info('\t\t%s: y_mid are mid points of y_low and y_sup. Do nothing' % ws_property.title)
        elif np.isnan(y_mid).any() or min_row_3 != min_row_12:
            y_mid = (y_low + y_sup) / 2
            logger.warning('\t\t%s: not all y_mid exit. Computing y_mid = (y_low+y_sup)/2'
                        % ws_property.title)
        else:
            y_mid = (y_low + y_sup) / 2
            logger.info('\t\t%s: y_mid does not exist. Computing y_mid = (y_low+y_sup)/2'
                        % ws_property.title)
    elif 'depth 1' in headers:
        loc1 = headers.index('depth 1')
        headers[loc1] = 'y_low'

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc1 + 1).value, (float, int)):
                min_row_1 = ii_row
                break
        y_low = np.array([ws_property.cell(ii_row, loc1+1).value for ii_row in range(min_row_1, max_row + 1)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_low))[::-1] if np.isnan(y_low[ii])]
        y_low = np.delete(y_low, y_nan_loc)

        # Fill y_sup
        logger.warning('\t\t%s: y_sup do not exist. Attempting to infer y_sup from y_low'
                       % ws_property.title)
        if not np.isnan(core_length):
            logger.info('\t\t%s: using core length, y_sup[-1] = l_c' % ws_property.title)
            y_sup = np.concatenate([y_low[1:], [core_length]])
        elif min_row_1 == min_row_3 and len(y_low) == len(y_mid):
            logger.warning('\t\t%s: using y_mid and y_low,  y_sup[-1] = y_low[-1] + 2 * (y_mid[-1] - y_low[-1])'
                           % ws_property.title)
            dy = 2 * (y_mid[-1] - y_low[-1])
            y_sup = np.concatenate([y_low[1:], [y_low[-1]+dy]])
        else:
            logger.warning('\t\t%s: core length not available, y_sup[-1] = y_low[-1] + (y_low[-1]-y_low[-2])'
                           % ws_property.title)
            dy = np.diff(y_low[-2:])
            y_sup = np.concatenate([y_low[1:], [y_low[-1]+dy]])
    elif 'depth 2' in headers:
        loc2 = headers.index('depth 2')
        headers[loc2] = 'y_sup'
        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc2 + 1).value, (float, int)):
                min_row_2 = ii_row
                break
        y_sup = np.array([ws_property.cell(ii_row, loc2+1).value for ii_row in range(min_row_2, max_row + 1)]).astype(float)

        logger.warning('\t\t%s: y_low do not exist. Attempting to infer y_low from y_sup'% ws_property.title)
        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_sup))[::-1] if np.isnan(y_sup[ii])]
        y_sup = np.delete(y_sup, y_nan_loc)

        # Fill y_low
        if min_row_2 == min_row_3 and len(y_sup) == len(y_mid):
            dy = 2 * (y_sup[0] - y_mid[0])
            y_sup0 = y_sup[0] - dy
            if y_sup0 >= 0:
                logger.warning('\t\t%s: using y_mid and y_low,  y_low[0] = y_sup[0] - 2 * (y_sup[0]-y_mid[0])'
                               % ws_property.title)
                y_low = np.concatenate([[y_sup0], y_sup[1:]])
            else:
                logger.info('\t\t%s: For lower y_low, using y_low[0] = 0' % ws_property.title)
                y_low = np.concatenate([[0], y_sup[:-1]])
        else:
            logger.info('\t\t%s: For lower y_low, using y_low[0] = 0' % ws_property.title)
            y_low = np.concatenate([[0], y_sup[:-1]])
    else:
        logger.info('\t\t%s: y_low and y_sup do not exist. Creating nan array of same size as y_mid'% (ws_property.title))
        y_low = np.nan * np.ones(y_mid.__len__())
        y_sup = np.nan * np.ones(y_mid.__len__())

    # Check length consistency
    if len(y_low) != len(y_mid):
        logger.error('\t\t%s: y_low/y_sup and y_mid of different size'% (ws_property.title))

    # Read data, according to depth value
    # look up for first numeric or standard entry value
    row_min = min(min_row_12, min_row_3)
    row_max = row_min + len(y_mid) - 1

    # Drop column with depth:
    # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
    #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
    _data = [[cell.value for cell in row] for row in ws_property.iter_rows(row_min, row_max, n_col_min, n_col_max)]

    # concatenate header and subheader
    if ic_property is None:
        variable_prefix = ''
        matter = 'sea_ice'
    elif any(map(ic_property.__contains__, ['brine', 'sackhole'])):
        variable_prefix = 'brine_'
        matter = 'brine'
    elif any(map(ic_property.__contains__, ['seawater'])):
        variable_prefix = 'seawater_'
        matter = 'seawater'
    elif any(map(ic_property.__contains__, ['snow'])):
        variable_prefix = 'snow_'
        matter = 'snow'
    else:
        variable_prefix = ''
        matter = 'seaice'

    # concatenate header and subheader to generate header
    subheaders = [sh if sh is not None else '' for sh in subheaders]
    profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid', 'comment']) else h
                       for ii, h in enumerate(headers)]

    profile = pd.DataFrame(_data, columns=profile_headers)

    # Add 'y_mid' column if does not exist, and fill it
    if 'y_mid' not in profile.keys():
        profile['y_mid'] = profile['y_mid'] = (profile.y_low + profile.y_sup)/2
    elif any(profile.y_mid.isna()):
        profile.loc[profile.y_mid.isna(), 'y_mid'] = profile.loc[profile.y_mid.isna(), ['y_low', 'y_sup']].sum(axis=1)/2

    # Add 'comment' column if does not exist and fill it with none value
    if 'comment' not in profile.columns:
        profile['comment'] = None

    # Fill row with np.nan/None for missing section
    if fill_missing:
        for ii_row in np.where(profile.y_sup.values[:-1] - profile.y_low.values[1:] < -TOL)[0]:
            y_low = profile.loc[profile.index == ii_row, 'y_sup'].values
            y_sup = profile.loc[profile.index == ii_row + 1, 'y_low'].values
            empty_row = pd.DataFrame([[np.nan] * len(profile_headers)], columns=profile_headers)
            empty_row['y_low'] = y_low
            empty_row['y_sup'] = y_sup
            empty_row['y_mid'] = (y_low + y_sup) / 2
            profile = pd.concat([profile, empty_row]).reset_index(drop=True)
            logger.info('\t\t%s: filling missing section (%.3f - %.3f) with nan/none values'
                        % (ws_property.title, y_low, y_sup))
        if any(np.isnan(profile.y_mid)):
            profile = profile.drop('y_mid', axis=1)
            profile = profile.sort_values('y_low').reset_index(drop=True)
        else:
            profile = profile.sort_values('y_mid').reset_index(drop=True)

    # Clean profile:
    # Drop empty headers
    if None in profile.columns:
        profile = profile.drop(labels=[None], axis=1)

    # get profile property from headers (e.g. salinity, temperature, ...)
    ic_property = [h.split('_ID')[0] for h in profile.columns if 'ID' in h]

    # old method
    # ic_property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
    # ic_property = [prop.split('_')[0] for prop in ic_property]
    # ic_property = list(set(ic_property))
    #
    # # remove property parameters (e.g. conductivity temperature measurement for conductivity
    # ic_property = [prop for prop in ic_property if prop not in inverse_dict(prop_parameter_dict)]

    # generate list of headers, with float type, and set column type to float
    string_header = ['comment']
    float_header = [h for h in profile.columns if h not in string_header]
    profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

    # drop property if there is neither an ID and a value for each sections
    for ic_prop in ic_property:
        ic_prop_headers = [h for h in profile.columns if ic_prop+'_' in h]
        if all(profile[ic_prop+'_ID'].isna()) and all(profile[ic_prop+'_value'].notna()):
            profile.drop(labels=ic_prop_headers)

    # remove empty line if all element of depth are nan:
    subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
    profile = profile.dropna(axis=0, subset=subset, how='all')

    # set property by row
    def add_property(x, ic_prop):
        if x is None:
            x = ic_prop
        else:
            x = ', '.join(list(set(filter(None, x.split(', ')))) + [ic_prop])
        return x
    profile['property'] = None
    for ic_prop in ic_property:
        profile_index = profile[(profile[ic_prop + '_ID'].notna() | profile[ic_prop + '_value'].notna())].index
        profile.loc[profile_index, 'property'] = profile.loc[profile_index, 'property'].apply(lambda x: add_property(x, ic_prop))

    # TODO: add property to profile even if na (override)

    # set location, direction and depth of the vertical referential system
    # TODO: improve vertical reference
    if reference_d['ice'][0] == 'ice surface':
        v_ref_loc = 'ice surface'
    elif reference_d['ice'][0] == 'ice/water interface':
        v_ref_loc = 'ice bottom'
    else:
        logger.error('%s\t\tVertical reference %s not defined.' % (ws_property, reference_d['ice'][0]))
    profile['v_ref_loc'] = [v_ref_loc] * len(profile.index)

    if reference_d['ice'][1] == 'up':
        v_ref_dir = 'positive'
    elif reference_d['ice'][1] == 'down':
        v_ref_dir = 'negative'
    else:
        logger.error('%s\t\tVertical reference %s not defined.' % (ws_property, reference_d['ice'][0]))

    profile['v_ref_dir'] = [v_ref_dir] * len(profile.index)
    profile['v_ref_h'] = [0] * len(profile.index)

    # set columns type
    string_header.extend(['property', 'v_ref_dir', 'v_ref_h'])
    float_header.extend(['v_ref_h'])
    c_string = [h for h in string_header if h in profile.columns]
    profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

    profile = Profile(profile)

    # import pickle
    # with open('/home/megavolts/temp.pkl', 'w') as f:
    #     pickle.dump(profile, f)
    return profile


def read_snow_profile(ws_property, ic_property=None, reference_d={'ice': ['ice surface', 'down']}):
    """
    :param ws_property:
        openpyxl.worksheet
    :param ic_property:
    :param reference_d:
        top, or bottom
    """
    from pysic.core.profile import Profile
    logger = logging.getLogger(__name__)

    # Read Salinity block
    # read headers:
    n_row = 1
    n_col = 1
    n_col_min = n_col
    cell_flag = 2
    headers = []
    subheaders = []
    units = []
    while cell_flag >= 1:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value
            headers.append(ws_property.cell(n_row, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(ws_property.cell(n_row + 1, n_col).value)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            cell_flag = 2
            n_col += 1
        elif isinstance(ws_property.cell(n_row+1, n_col).value, str):
            headers.append(h_)
            subheaders.append(ws_property.cell(n_row+1, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            units.append(ws_property.cell(n_row+2, n_col).value)
            n_col += 1
        else:
            cell_flag = 0
            cell_mark = n_col + 1
    n_col_max = n_col - n_col_min

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Check for step or continuous profiles:
    if 'depth center' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth center'][0] + 1
        headers[loc1-1] = 'y_mid'
        y_mid = np.array(
            [ws_property.cell(row, loc1).value for row in range(min_row, max_row)]).astype(float)

        # discard trailing nan value from the end up
        # find nan value in y_low and y_sup
        y_nan_loc = np.where(np.isnan(y_mid))[0]
        # discard trailing nan value starting at the end
        if len(y_nan_loc) > 0 and len(y_mid) > 1 and y_nan_loc[-1] == len(y_mid)-1:
            y_nan_loc = [len(y_mid)-1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if val == y_nan_loc[::-1][ii]-1]
            y_nan_loc = y_nan_loc[::-1]
            y_mid = [y for ii, y in enumerate(y_mid) if ii not in y_nan_loc]

    if 'depth 1' in headers and 'depth 2' in headers:
        step_flag = 1

        # find column with depth 1
        # TODO find a better way to find the location
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 1'][0]+1
        headers[loc1 - 1] = 'y_low'
        loc2 = [ii for ii, h in enumerate(headers) if h == 'depth 2'][0]+1
        headers[loc2 - 1] = 'y_sup'
        # TODO: remove 'depth center'
        y_low = np.array([ws_property.cell(row, loc1).value for row in range(min_row, max_row + 1)]).astype(float)
        y_sup = np.array([ws_property.cell(row, loc2).value for row in range(min_row, max_row + 1)]).astype(float)

        # discard trailing nan value from the end up
        # find nan value in y_low and y_sup
        y_nan_loc = [ii for ii in np.where(np.isnan(y_sup))[0] if ii in np.where(np.isnan(y_low))[0]]

        # discard trailing nan value starting at the end
        if len(y_sup) > 0 and len(y_nan_loc) > 0 and y_nan_loc[-1] == len(y_sup)-1:
            y_nan_loc = [len(y_sup)-1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if val == y_nan_loc[::-1][ii]-1]
            y_nan_loc = y_nan_loc[::-1]
            y_low = np.array([y for ii, y in enumerate(y_low) if ii not in y_nan_loc])
            y_sup = np.array([y for ii, y in enumerate(y_sup) if ii not in y_nan_loc])

            # TODO: replace missing y_low and y_sup with y_mid if possible

        if len(y_low) == 0:
            profile = Profile()
            y_mid = []
        elif 'y_mid' in headers:
            if np.isnan(y_mid).any() or len(y_mid) == 0:
                y_mid = (y_low + y_sup) / 2
                logger.info('(%s ) not all y_mid exits, calculating y_mid = (y_low+y_sup)/2'
                            % (ws_property.title))
            elif np.any(np.abs(((y_low + y_sup) / 2) - y_mid > 1e-12)):
                logger.error('(%s ) y_mid are not mid point between y_low and y_sup. \\'
                             'Replacing with y_mid = (y_low+y_sup)/2'
                             % (ws_property.title))
                y_mid = (y_low + y_sup) / 2
            else:
                logger.info('(%s ) y_low, y_mid and y_sup read with success'
                            % (ws_property.title))
        else:
            y_mid = (y_low + y_sup) / 2
    elif 'depth 1' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 1'][0]+1
        headers[loc1 - 1] = 'y_low'
        # TODO : fill y_sup
    elif 'depth 2' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 2'][0]+1
        headers[loc1 - 1] = 'y_sup'
        # TODO : fill y_low
    # Continuous profile
    else:
        y_low = np.nan * np.ones(y_mid.__len__())
        y_sup = np.nan * np.ones(y_mid.__len__())

    # Read data:
    # look up for first numeric or standard entry value
    # n_row_min = 4

    if len(y_mid) > 0:
        n_row_min = 1
        n_col_min = 1
        while not isinstance(ws_property.cell(n_row_min, n_col_min).value, (float, int)):
            n_row_min += 1
            if n_row_min > 1000:
                break
        n_row_max = n_row_min + len(y_mid) - 1

        # Drop column with depth:
        # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
        #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
        _data = [[cell.value for cell in row] for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]

        # TODO:  fill missing section with np.nan
        # if fill_missing:
        #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
        #     for ii_idx in idx:
        #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
        #         empty += [np.nan] * (variable_headers.__len__()+1)
        #     data = np.vstack([data, empty])

        # concatenate header and subheader
        if ic_property is None:
            variable_prefix = ''
            phase = 'N/A'
        elif any(map(ic_property.__contains__, ['brine', 'sackhole'])):
            variable_prefix = 'brine_'
            phase = 'brine'
        elif any(map(ic_property.__contains__, ['seawater'])):
            variable_prefix = 'seawater_'
            phase = 'seawater'
        elif any(map(ic_property.__contains__, ['snow'])):
            variable_prefix = 'snow_'
            phase = 'snow'
        else:
            variable_prefix = ''
            phase = 'seaice'
        subheaders = [sh if sh is not None else '' for sh in subheaders]
        profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid']) else h
                           for ii, h in enumerate(headers)]
        # TODO: double header for dataframe with header and subheader

        profile = pd.DataFrame(_data, columns=profile_headers)
        if 'y_mid' not in profile.keys():
            profile['y_mid'] = y_mid

        # drop empty variable header
        if None in profile.columns:
            profile = profile.drop(labels=[None], axis=1)

        # sample ID columns is string:
        string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

        # convert string to float:
        float_header = [h for h in profile.columns if h not in string_header]
        profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

        # drop property with all nan value
        profile = profile.dropna(axis=1, how='all')

        # remove empty line if all element of depth are nan:
        subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
        profile = profile.dropna(axis=0, subset=subset, how='all')

        # singularize comments
        if 'comments' in profile.columns:
            profile.rename(columns={'comments': "comment"}, inplace=True)
        # add comment column if it does not exist
        if 'comment' not in profile.columns:
            profile['comment'] = None
        else:
            profile['comment'] = profile['comment'].astype(str).replace({'nan': None})

        # get all property variable (e.g. salinity, temperature, ...)
        property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
        property = [prop.split('_')[0] for prop in property]
        property = list(set(property))

        # remove subvariable (e.g. conductivity temperature measurement for conductivity
        property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

        # set variable to string of property
        profile['variable'] = [', '.join(property)] * len(profile.index)

        # set vertical references
        # TODO: improve vertical reference
        if reference_d['snow'][0] == 'ice surface':
            v_ref = 'bottom'
        elif reference_d['snow'][0] == 'snow interface':
            v_ref = 'bottom'
        else:
            logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
        profile['v_ref'] = [v_ref] * len(profile.index)

        # set columns type
        col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
        col_date = ['date']
        col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
        col_string = col_string + [h for h in profile.columns if 'ID' in h]
        profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
        c_string = [h for h in col_string if h in profile.columns]
        profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

        profile = Profile(profile)
        # remove variable not in variables
        if ic_property is not None:
            for property in profile.properties():
                if property not in ic_property:
                    profile.delete_property(property)

    if ic_property is not None and 'salinity' not in ic_property:
        profile_S = Profile()
    else:
        profile_S = profile

    cell_mark = 26

    del profile
    # Read Temperature block
    # read headers:
    n_row = 1
    n_col = cell_mark
    n_col_min = n_col
    cell_flag = 2
    headers = []
    subheaders = []
    units = []
    while cell_flag >= 1:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value
            headers.append(ws_property.cell(n_row, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(ws_property.cell(n_row + 1, n_col).value)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            cell_flag = 2
            n_col += 1
        elif isinstance(ws_property.cell(n_row + 1, n_col).value, str):
            headers.append(h_)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            n_col += 1
        else:
            cell_flag = 0
            cell_mark = n_col + 1
    n_col_max = n_col

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Check for step or continuous profiles:
    #loc1 = [ii for ii, h in enumerate(headers) if h == 'depth'][0] + 1
    headers[loc1 - 1] = 'depth'
    subheaders[loc1 - 1] = 'y_mid'
    headers = ['y_mid', 'temperature', 'temperature']
    subheaders = ['', 'value', 'quality']
    y_mid = np.array(
        [ws_property.cell(row, n_col_min).value for row in range(min_row, max_row)]).astype(float)

    # discard trailing nan value from the end up
    # find nan value in y_low and y_sup
    y_nan_loc = np.where(np.isnan(y_mid))[0]
    # discard trailing nan value starting at the end
    if len(y_nan_loc) > 0 and len(y_mid) > 1 and y_nan_loc[-1] == len(y_mid) - 1:
        y_nan_loc = [len(y_mid) - 1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if
                                        val == y_nan_loc[::-1][ii] - 1]
        y_nan_loc = y_nan_loc[::-1]
        y_mid = [y for ii, y in enumerate(y_mid) if ii not in y_nan_loc]

    y_low = np.nan * np.ones(y_mid.__len__())
    y_sup = np.nan * np.ones(y_mid.__len__())

    # Read data:
    # look up for first numeric or standard entry value
    # n_row_min = 4

    if len(y_mid) > 0:
        n_row_min = 1
        while not isinstance(ws_property.cell(n_row_min, n_col_min).value, (float, int)):
            n_row_min += 1
            if n_row_min > 1000:
                break
        n_row_max = n_row_min + len(y_mid) - 1

        # Drop column with depth:
        # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
        #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
        _data = [[cell.value for cell in row] for row in
                 ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max-1)]

        # TODO:  fill missing section with np.nan
        # if fill_missing:
        #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
        #     for ii_idx in idx:
        #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
        #         empty += [np.nan] * (variable_headers.__len__()+1)
        #     data = np.vstack([data, empty])

        # concatenate header and subheader
        if ic_property is None:
            variable_prefix = ''
            phase = 'N/A'
        elif any(map(ic_property.__contains__, ['brine', 'sackhole'])):
            variable_prefix = 'brine_'
            phase = 'brine'
        elif any(map(ic_property.__contains__, ['seawater'])):
            variable_prefix = 'seawater_'
            phase = 'seawater'
        elif any(map(ic_property.__contains__, ['snow'])):
            variable_prefix = 'snow_'
            phase = 'snow'
        else:
            variable_prefix = ''
            phase = 'seaice'
        subheaders = [sh if sh is not None else '' for sh in subheaders]
        profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (
                    len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid']) else h
                           for ii, h in enumerate(headers)]
        # TODO: double header for dataframe with header and subheader

        profile = pd.DataFrame(_data, columns=profile_headers)
        if 'y_mid' not in profile.keys():
            profile['y_mid'] = y_mid

        # drop empty variable header
        if None in profile.columns:
            profile = profile.drop(labels=[None], axis=1)

        # sample ID columns is string:
        string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

        # convert string to float:
        float_header = [h for h in profile.columns if h not in string_header]
        profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

        # drop property with all nan value
        profile = profile.dropna(axis=1, how='all')

        # remove empty line if all element of depth are nan:
        subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
        profile = profile.dropna(axis=0, subset=subset, how='all')

        # singularize comments
        if 'comments' in profile.columns:
            profile.rename(columns={'comments': "comment"}, inplace=True)
        # add comment column if it does not exist
        if 'comment' not in profile.columns:
            profile['comment'] = None
        else:
            profile['comment'] = profile['comment'].astype(str).replace({'nan': None})

        # get all property variable (e.g. salinity, temperature, ...)
        property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
        property = [prop.split('_')[0] for prop in property]
        property = list(set(property))

        # remove subvariable (e.g. conductivity temperature measurement for conductivity
        property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

        # set variable to string of property
        profile['variable'] = [', '.join(property)] * len(profile.index)

        # set vertical references
        # TODO: improve vertical reference
        if reference_d['snow'][0] == 'ice surface':
            v_ref = 'bottom'
        elif reference_d['snow'][0] == 'snow interface':
            v_ref = 'bottom'
        else:
            logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
        profile['v_ref'] = [v_ref] * len(profile.index)

        # set columns type
        col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
        col_date = ['date']
        col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
        col_string = col_string + [h for h in profile.columns if 'ID' in h]
        profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
        c_string = [h for h in col_string if h in profile.columns]
        profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

        profile = Profile(profile)
        # remove variable not in variables
        if ic_property is not None:
            for _property in profile.properties():
                if _property not in ic_property:
                    profile.delete_property(_property)

        if ic_property is not None and 'temperature' not in ic_property:
            profile = profile_S
        else:
            profile = profile.append(profile_S)
        profile.reset_index(drop=True, inplace=True)
    else:
        profile = Profile()
    return profile
