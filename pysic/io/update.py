#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
    pysic.io.update.py : function to update ice core spreadsheet
"""

from pysic.core.core import __CoreVersion__
import logging
import os
import openpyxl
import numpy as np
from copy import copy

from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

from pysic.tools import version2int
from pysic.property import prop_associated_tab

import pysic

__all__ = ['ic_data', 'ic_data_from_path']

pysic_fp = pysic.__path__[0]

data_row_n = 42
user = 'Marc Oggier'

# Border style
# top and bottom border
tb_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                   top=Side(border_style='hair', color='00000000'), bottom=Side(border_style='hair', color='00000000'))
# bottom border
b_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# bottom_left_border
bl_border = Border(left=Side(border_style='hair', color='00000000'), right=Side(border_style=None),
                   top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# left_border
l_border = Border(left=Side(border_style='hair', color='00000000'), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style=None))
# right border
r_border = Border(left=Side(border_style=None), right=Side(border_style='hair', color='00000000'),
                  top=Side(border_style=None), bottom=Side(border_style=None))
# bottom right border
br_border = Border(left=Side(border_style=None), right=Side(border_style='hair', color='00000000'),
                   top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# noBorder
noBorder = Border(left=Side(border_style=None), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style=None))

# Color scheme
black = "FF000000"
dark_grey = "FFB2B2B2"  # background
mid_grey = "FFDDDDDD"  # header
light_grey = "FFEEEEEE"  # data computed
white = "FFFFFFFF"
noFill = PatternFill(fill_type=None)  # data entry

# METADATA TAB STYLE
# - suptitle: black bold font, left aligned, dark grey background
m_title_style = NamedStyle(name="m_title_style")
m_title_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True, italic=False, color='FF000000')
m_title_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_title_style.fill = PatternFill(start_color=dark_grey, end_color="FFDDDDDD", fill_type="solid")

# - suptitle: black bold font, left aligned, dark grey background
m_title_c_style = NamedStyle(name="m_title_c_style")
m_title_c_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True, italic=False, color='FF000000')
m_title_c_style.alignment = Alignment(horizontal='center', wrapText=False, vertical='center')
m_title_c_style.fill = PatternFill(start_color=dark_grey, end_color="FFDDDDDD", fill_type="solid")

# - header: black, bold font; dark grey background
m_header_style = NamedStyle(name="m_header_style")
m_header_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True, italic=False, color='FF000000')
m_header_style.alignment = Alignment(horizontal='center', wrapText=False, vertical='top')
m_header_style.fill = PatternFill(start_color=mid_grey, end_color="FFE0E0E0", fill_type="solid")

# header right: black bold font, right aligned, light grey background
m_header_r_style = NamedStyle(name="m_header_r_style")
m_header_r_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True, italic=False, color='FF000000')
m_header_r_style.alignment = Alignment(horizontal='right', wrapText=False, vertical='center')
m_header_r_style.fill = PatternFill(start_color=mid_grey, end_color="FFDDDDDD", fill_type="solid")

# header left: black bold font, left aligned, light grey background
m_header_l_style = NamedStyle(name="m_header_l_style")
m_header_l_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True, italic=False, color='FF000000')
m_header_l_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_header_l_style.fill = PatternFill(start_color=mid_grey, end_color="FFDDDDDD", fill_type="solid")

# - subheader right: black, bold font; light grey background
m_subheader_r_style = NamedStyle(name="m_subheader_r_style")
m_subheader_r_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
m_subheader_r_style.alignment = Alignment(horizontal='right', wrapText=True, vertical='center')
m_subheader_r_style.fill = PatternFill(start_color=mid_grey, end_color="FFDDDDDD", fill_type="solid")

# - subheader left: black, bold font; light grey background
m_subheader_l_style = NamedStyle(name="m_subheader_l_style")
m_subheader_l_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
m_subheader_l_style.alignment = Alignment(horizontal='left', wrapText=True, vertical='center')
m_subheader_l_style.fill = PatternFill(start_color=mid_grey, end_color="FFDDDDDD", fill_type="solid")

# - unit: right, black, italic font; light grey background
m_unit_r_style = NamedStyle(name="m_unit_r_style")
m_unit_r_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=True, color='FF000000')
m_unit_r_style.alignment = Alignment(horizontal='right', wrapText=False, vertical='center')
m_unit_r_style.fill = PatternFill(start_color=mid_grey, end_color="FFDDDDDD", fill_type="solid")

# - comment: left, black, italic font; light grey background
m_comment_style = NamedStyle(name="m_comment_style")
m_comment_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=True, color='FF000000')
m_comment_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_comment_style.fill = PatternFill(start_color=dark_grey, end_color="FFDDDDDD", fill_type="solid")

# data, right: black; no background
m_data_r_style = NamedStyle(name="m_data_r_style")
m_data_r_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
m_data_r_style.alignment = Alignment(horizontal='right', wrapText=False, vertical='center')
m_data_r_style.fill = PatternFill(fill_type=None)

# data, left: black; no background
m_data_l_style = NamedStyle(name="m_data_l_style")
m_data_l_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
m_data_l_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_data_l_style.fill = PatternFill(fill_type=None)

# computed data, right: black; no background
m_cdata_r_style = NamedStyle(name="m_cdata_r_style")
m_cdata_r_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
m_cdata_r_style.alignment = Alignment(horizontal='right', wrapText=False, vertical='center')
m_cdata_r_style.fill = PatternFill(start_color=light_grey, end_color="FFDDDDDD", fill_type="solid")

# PROPERTY TAB
# - header
p_header_style = NamedStyle(name="p_header_style")
p_header_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True, italic=False, color='FF000000')
p_header_style.alignment = Alignment(horizontal='center', wrapText=False, vertical='center')
p_header_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")

# - subheader
p_subheader_style = NamedStyle(name="p_subheader_style")
p_subheader_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
p_subheader_style.alignment = Alignment(horizontal='center', wrapText=False, vertical='center')
p_subheader_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")

# - unit
p_unit_style = NamedStyle(name="p_unit_style")
p_unit_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=True, color='FF000000')
p_unit_style.alignment = Alignment(horizontal='center', wrapText=False, vertical='center')
p_unit_style.fill = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")

# data, right: black, no background (for number)
p_data_r_style = NamedStyle(name="p_data_r_style")
p_data_r_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
p_data_r_style.alignment = Alignment(horizontal='right', wrapText=False, vertical='top')
p_data_r_style.fill = PatternFill(fill_type=None)

# data, left:  black, no background (for string)
p_data_l_style = NamedStyle(name="p_data_l_style")
p_data_l_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
p_data_l_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='top')
p_data_l_style.fill = PatternFill(fill_type=None)

# BACKGROUND
m_bkg_style = NamedStyle(name="m_bkg_style")
m_bkg_style.font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False, color='FF000000')
m_bkg_style.alignment = Alignment(horizontal='left', wrapText=False, vertical='center')
m_bkg_style.fill = PatternFill(start_color=dark_grey, end_color=dark_grey, fill_type="solid")

# Data validation
vert_ref_list = DataValidation(type="list", formula1='"ice surface, ice/water interface, snow surface, water level, N/A"')
vert_dir_list = DataValidation(type="list", formula1='"up, down, N/A"')
dv_lists_dict = {'tz_dv': 'A', 'surface_dv': 'B', 'age_dv': 'C', 'topography_dv': 'D', 'environement_dv': 'E',
            'texture_dv': 'F', 'inclusion_dv': 'G', 'description_dv': 'H', 'ice_type_dv': 'I', 'hardness_dv': 'J',
            'snow_type_dv': 'K', 'cloud': 'P'
                                          ''}
wb_source = openpyxl.load_workbook(
    os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYMMDD-P-1.4.1-lists.xlsx'), data_only=True)
for k in dv_lists_dict.keys():
    col = dv_lists_dict[k]
    col_idx = openpyxl.utils.column_index_from_string(col)
    row_idx = 1
    while wb_source['lists'].cell(row_idx, col_idx).value is not None:
        row_idx += 1
    cr = "$" + col + "$3:$" + col + "$" + str(row_idx)
    dv_lists_dict[k] = eval("DataValidation(type='list', formula1='{0}!" + cr +
                            "'.format(quote_sheetname('lists')), allow_blank=True)")

dv_units_dict = {'length': ['m', 'mm'],
                 'mass': ['kg', 'g']}

m_styles_list = ['m_title_style', 'm_title_c_style', 'm_header_style', 'm_header_r_style', 'm_header_l_style',
                 'm_subheader_r_style', 'm_subheader_l_style', 'm_unit_r_style', 'm_comment_style', 'm_data_r_style',
                 'm_data_l_style', 'm_cdata_r_style', 'p_header_style', 'p_subheader_style', 'p_unit_style',
                 'p_data_r_style', 'p_data_l_style', 'm_bkg_style']


def ic_data_from_path(ic_path, backup=True, user=user):
    """
    :param ic_path:
        path; Filepath to the data spreadsheet to update
    :param backup:
        boolean, default True; Make a backup copy in a subfolder
    """
    import shutil
    logger = logging.getLogger(__name__)

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet

    # Read ice core data version
    if 'metadata-coring' in wb.sheetnames:
        ws_summary = wb['metadata-coring']  # load the data from the summary sheet
    else:
        ws_summary = wb['metadata-station']  # load the data from the summary sheet
    version = ws_summary['C1'].value

    if version == __CoreVersion__:
        logger.debug('%s\t\talready at latest (%s) version' % (wb['metadata-core']['A1'].value, __CoreVersion__))
    else:
        # Backup old version
        if backup:
            backup_dir = os.path.join(os.path.dirname(ic_path), 'bkp-ic_version_' + str(version))
            ic_bkp = os.path.join(backup_dir, os.path.basename(ic_path))
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
            if not os.path.exists(ic_bkp):
                shutil.copy(ic_path, ic_bkp)
                logger.info('\t%s\t\tbackup file (version %s) already exist at %s' \
                            % (wb['metadata-core']['C1'].value, version, backup_dir))
            else:
                logger.info('\t%s\t\tsaving backup file (version %s) to %s' \
                            % (wb['metadata-core']['C1'].value, version, backup_dir))

        # udpate ice core data
        wb = ic_data(wb, user=user)

        logger.info('%s\t\tsaving updated file (version %s) to %s' % (
            wb['metadata-core']['C1'].value, __CoreVersion__, ic_path))

        wb.save(ic_path)
        wb.close()


def ic_data(wb, user=user):
    """
    """
    from datetime import date
    logger = logging.getLogger(__name__)

    if 'metadata-coring' in wb.sheetnames:
        ws_summary = wb['metadata-coring']  # load the data from the summary sheet
    else:
        ws_summary = wb['metadata-station']  # load the data from the summary sheet
    version = ws_summary['C1'].value
    version_int = version2int(version)

    if version_int[1] < 2 and version_int[2] < 1:
        logger.error('%s\t\tversion update from %s not supported' % (wb['metadata-core']['C1'].value, str(version)))
    elif version == __CoreVersion__:
        logger.debug('%s\t\talready at latest (%s) version' % (wb['metadata-core']['A1'].value, __CoreVersion__))
    else:
        # Add defined style
        wb = add_style(wb)

        # Update from 1.2 to 1.3:
        wb = version_1_2_x_to_1_3_0(wb)
        wb = version_1_3_0_to_1_4_1(wb)
        wb = version_1_4_1_to_1_4_2(wb)
        wb = version_1_4_2_to_1_4_3(wb)
        wb = version_1_4_3_to_1_4_4(wb)
        wb = version_1_4_4_to_1_4_5(wb)
        wb = version_1_4_5_to_1_4_6(wb)
        wb = version_1_4_6_to_1_4_7(wb)
        wb = version_1_4_7_to_1_4_8(wb)
        wb = version_1_4_8_to_1_4_9(wb)

        # Add an update is the 'metadata-core'
        sheetname = 'metadata-core'
        version_update = wb['metadata-station']['C1'].value

        # Find where 'VERSION' entries are located
        row = 1
        version_row = 0
        while version_row == 0 and row <= wb[sheetname].max_row:
            if wb[sheetname].cell(row, 1).value == 'VERSION' and wb[sheetname].cell(row + 1, 1).value == 'number':
                version_row = row
                break
            row += 1

        # Find line number with the latest version entry
        row_idx = version_row + 2
        while wb[sheetname].cell(row_idx, 1).value is not None or row_idx > wb[sheetname].max_row:
            row_idx += 1

        # Add new line afterwards
        insert_row_with_merge(wb[sheetname], row_idx, 1)
        wb[sheetname].cell(row_idx, 1).value = int(wb[sheetname].cell(row_idx - 1, 1).value) + 1
        wb[sheetname].cell(row_idx, 2).value = date.today().isoformat()
        wb[sheetname].cell(row_idx, 3).value = user
        wb[sheetname].cell(row_idx, 5).value = 'Updated ice core data from ' + str(version) + ' to ' + version_update

        # apply style and format
        wb = spreadsheet_style(wb)

        # remove all external links
        wb = remove_external_link(wb)

        logger.info('%s\t\tupdated from from %s to %s' % (wb[sheetname]['C1'].value, version, version_update))
    return wb


def remove_external_link(wb):
    external_links = wb._external_links
    wb._external_links = {}
    return wb
def add_style(wb):
    for style in m_styles_list:
        if style not in wb.style_names:
            wb.add_named_style(eval(style))
    return wb


# Update scripts:
def version_1_2_x_to_1_3_0(wb):
    logger = logging.getLogger(__name__)

    try:
        version = version2int(wb['metadata-station']['C1'].value)
    except KeyError:
        version = version2int(wb['metadata-coring']['C1'].value)

    if version[0] == 1 and version[1] == 2 and version[2] == 22:
        wb = version_1_2m_to_1_3_0(wb)
    elif version[0] == 1 and version[1] < 3:
        logger.error('\t%s - not implemented yet' % (wb['metadata-core']['C1'].value))
    # # TODO: update to version 1.3.1
    #
    # if version == '2.3':
    #     new_version = '1.3.1'
    #     print('Update to future version ' + new_version)
    #     ws_summary['C1'] = new_version
    #

    #     if 'ECO' in wb.sheetnames:
    #         ws = wb["ECO"]
    #         if ws['F2'].value == 'Nutrient':
    #             ws['F2'] = 'nutrient'
    #
    #         # TODO: update nutrient with 5 nutrient
    #
    #     if 'snow' in wb.sheetnames:
    #         ws = wb["snow"]
    #         ws['L2'].value = 'snow'
    #         ws['L3'].value = 'weight'
    #         ws['M3'].value = 'density'
    #         ws['M3'].value = 'snow'
    #         ws['M3'].value = 'density'
    #         ws['N3'].value = 'comment'
    #         ws['N3'].value = '-'
    #         ws['N4'].value = '-'
    #         ws['O3'].value = 'temperature'
    #         ws['P3'].value = 'temperature'
    #         ws['P3'].value = 'value'
    else:
        logger.info("\t%s: already update to version %s " % (
            wb['metadata-core']['C1'].value, wb['metadata-coring']['C1'].value))
    wb = fix_merged_cells(wb)
    return wb


def version_1_2m_to_1_3_0(wb):
    logger = logging.getLogger(__name__)

    try:
        version = wb['metadata-station']['C1'].value
    except KeyError:
        version = wb['metadata-coring']['C1'].value

    if version == '1.2M':
        # METADATA-CORE
        sheetname = 'metadata-core'
        # find last non-empty cell in column A
        for row_ii in range(1, wb[sheetname].max_row + 1):
            if wb[sheetname].cell(row_ii, 1).value == 'DATA VERSION':
                row_idx = 0
                break
            elif wb[sheetname].cell(row_ii, 1).value is not None:
                row_idx = row_ii
        if row_idx != 0:
            row_idx = row_idx + 2
            wb[sheetname].cell(row_idx, 1).value = 'DATA VERSION'
            wb[sheetname].cell(row_idx + 1, 1).value = 1
            wb[sheetname].cell(row_idx + 1, 2).value = None
            wb[sheetname].cell(row_idx + 1, 3).value = 'initial data entry'

        # METADATA-CORING
        sheetname = 'metadata-coring'
        wb[sheetname]['D28'].value = None

        # SALINITY
        sheetname = 'SALO18'
        wb[sheetname]['D2'].value = 'Field Sample ID'
        wb[sheetname]['D3'].value = None
        wb[sheetname].merge_cells('I2:J2')
        wb[sheetname].merge_cells('K2:L2')

        # TEMPERATURE
        sheetname = 'TEMP'
        wb[sheetname].insert_cols(3, 1)
        wb[sheetname]['A3'].value = 'value'
        wb[sheetname]['B3'].value = 'value'
        wb[sheetname]['C3'].value = '-'

        # TEX
        sheetname = 'TEX'
        wb[sheetname].insert_rows(3, 1)
        wb[sheetname]['A3'] = 'value'
        wb[sheetname]['B3'] = 'value'
        wb[sheetname]['C2'] = 'depth center'
        wb[sheetname]['C3'] = 'value'
        wb[sheetname]['D3'] = 'value'
        wb[sheetname]['F3'] = '-'

        # DENSITY-VOLUME
        sheetname = 'Density-volume'
        wb[sheetname].insert_rows(3)
        wb[sheetname]['A3'] = 'value'
        wb[sheetname]['B3'] = 'value'
        wb[sheetname]['C2'] = 'field'
        wb[sheetname]['C3'] = 'sample ID'
        wb[sheetname]['D3'] = 'value'
        wb[sheetname]['E2'] = 'thickness'
        wb[sheetname]['E3'] = 'value1'
        wb[sheetname]['F2'] = 'thickness'
        wb[sheetname]['F3'] = 'value2'
        wb[sheetname]['G2'] = 'thickness'
        wb[sheetname]['G3'] = 'value3'
        wb[sheetname]['H2'] = 'thickness'
        wb[sheetname]['H3'] = 'value4'
        wb[sheetname]['I2'] = 'thickness'
        wb[sheetname]['I3'] = 'average'
        wb[sheetname]['J3'] = 'value'
        wb[sheetname]['K2'] = 'density'
        wb[sheetname]['K3'] = 'value'
        wb[sheetname]['L3'] = '-'

        # DENSITY-DENSIMETRY
        sheetname = 'Density-densimetry'
        wb[sheetname].insert_rows(3)
        wb[sheetname]['A3'] = 'value'
        wb[sheetname]['B3'] = 'value'
        wb[sheetname]['C2'] = 'field'
        wb[sheetname]['C3'] = 'sample ID'
        wb[sheetname]['D3'] = 'value'
        wb[sheetname]['E3'] = 'value'
        wb[sheetname]['F3'] = 'value'
        wb[sheetname]['G3'] = 'value'
        wb[sheetname]['H3'] = '-'

        # SNOW
        sheetname = 'snow'
        wb[sheetname].insert_rows(1, 1)
        wb[sheetname]['A1'] = 'Reference'
        wb[sheetname]['B1'] = 'zero vertical'
        wb[sheetname]['C1'] = 'ice surface'
        wb[sheetname]['C2'] = 'field'
        wb[sheetname]['C3'] = 'sample ID'
        wb[sheetname]['D1'] = 'direction'
        wb[sheetname]['E1'] = 'up'

        wb['metadata-coring']['C1'].value = '1.3.0'
        wb = fix_merged_cells(wb)
        logger.debug(
            '\t %s updated from 1.2M to %s' % (wb['metadata-core']['C1'].value, wb['metadata-coring']['C1'].value))
    return wb


def version_1_3_0_to_1_4_1(wb):
    logger = logging.getLogger(__name__)
    try:
        version = version2int(wb['metadata-station']['C1'].value)
    except KeyError:
        version = version2int(wb['metadata-coring']['C1'].value)

    if version[0] < 2 and version[1] < 4 and version[2] < 1:
        # Update from 1.3 to 1.4.1
        # - METADATA-CORE
        sheetname = 'metadata-core'
        wb[sheetname]['A1'].value = 'core name'
        wb[sheetname]['A2'].value = 'date'
        wb[sheetname]['A3'].value = 'time'
        wb[sheetname]['A4'].value = 'note'
        wb[sheetname]['A7'].value = 'ice thickness'
        wb[sheetname]['A8'].value = 'draft'
        wb[sheetname]['A9'].value = 'freeboard'
        wb[sheetname]['A10'].value = 'core length'

        # formatting
        wb[sheetname]['E1'].style = 'm_bkg_style'
        for row in range(1, 4):
            for col in range(1, openpyxl.utils.column_index_from_string('D') + 1):
                wb[sheetname].cell(row, col).border = b_border
        for col in range(1, openpyxl.utils.column_index_from_string('C') + 1):
            wb[sheetname].cell(7, col).border = b_border
            wb[sheetname].cell(10, col).border = noBorder
            wb[sheetname].cell(9, col).border = b_border

        # - Add REFERENCE BLOCK
        row_n = 8
        row_insert_idx = 12
        # move merged cells downwards below row_insert_idx row
        for cr in wb[sheetname].merged_cells.ranges:
            if cr.right[0][0] > row_insert_idx:
                cr.shift(0, row_n)
        # insert row_n rows starting at row_insert_idx row
        wb[sheetname].insert_rows(row_insert_idx, row_n)

        # insert text
        wb[sheetname]['A12'].value = 'REFERENCE'
        wb[sheetname]['A13'].value = 'ICE'
        wb[sheetname]['A15'].value = 'SNOW'
        wb[sheetname]['A17'].value = 'SEAWATER'
        wb[sheetname]['B13'].value = 'zero vertical'
        wb[sheetname]['B14'].value = 'direction'
        wb[sheetname]['C13'].value = '(0m @)'
        wb[sheetname]['E14'].value = 'down: positive away from the sky; up: positive toward the sky'
        wb[sheetname]['B15'].value = 'zero vertical'
        wb[sheetname]['B16'].value = 'direction'
        wb[sheetname]['C15'].value = '(0m @)'
        wb[sheetname]['E16'].value = 'down: positive away from the sky; up: positive toward the sky'
        wb[sheetname]['B17'].value = 'zero vertical'
        wb[sheetname]['B18'].value = 'direction'
        wb[sheetname]['C17'].value = '(0m @)'
        wb[sheetname]['E18'].value = 'down: positive away from the sky; up: positive toward the sky'

        # add data validation to worksheet
        wb[sheetname].add_data_validation(vert_ref_list)
        wb[sheetname].add_data_validation(vert_dir_list)
        vert_ref_list.add('D13')
        vert_ref_list.add('D15')
        vert_ref_list.add('D17')
        vert_dir_list.add('D14')
        vert_dir_list.add('D16')
        vert_dir_list.add('D18')

        # formatting
        max_col = 26
        for row in range(row_insert_idx, row_insert_idx + row_n + 1):
            for col in range(1, max_col + 1):
                wb[sheetname].cell(row, col).style = 'm_bkg_style'
        # paint upper right cell into bkg
        for row in range(1, row_insert_idx):
            for col in range(7, max_col + 1):
                wb[sheetname].cell(row, col).style = 'm_bkg_style'

        wb[sheetname]['A12'].style = 'm_title_style'
        for row in range(row_insert_idx + 1, row_insert_idx + 6 + 1):
            # A header
            wb[sheetname].cell(row, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row, 2).style = 'm_subheader_l_style'
            wb[sheetname].cell(row, 3).style = 'm_subheader_r_style'
            wb[sheetname].cell(row, 4).style = 'm_data_l_style'

        # cell line
        for row in np.arange(row_insert_idx + 2, row_insert_idx + row_n - 2, 2):
            for col in range(1, openpyxl.utils.column_index_from_string('H') + 1):
                wb[sheetname].cell(row, col).border = b_border

        # INSTRUMENTS & VERSION blocks
        # look for INSTRUMENTS and VERSION header row
        row = row_insert_idx + row_n
        flag_instrument = True
        flag_version = True
        while flag_instrument or flag_version:
            if isinstance(wb[sheetname].cell(row, 1).value, str) and \
                    'INSTRUMENT' in wb[sheetname].cell(row, 1).value:
                flag_instrument = False
                row_instrument_idx = row
            elif isinstance(wb[sheetname].cell(row, 1).value, str) and \
                    'DATA VERSION' in wb[sheetname].cell(row, 1).value:
                flag_version = False
                row_version_idx = row
            row += 1

        # - apply background format
        for row in range(row_instrument_idx - 1, wb[sheetname].max_row + 1 + 1):
            for col in range(1, max_col + 1):
                wb[sheetname].cell(row, col).style = 'm_bkg_style'
            for col in range(max_col + 1, wb[sheetname].max_column + 1):
                wb[sheetname].cell(row, col).fill = noFill
                wb[sheetname].cell(row, col).border = noBorder

        # Format INSTRUMENTS block
        # - title formatting
        wb[sheetname].cell(row_instrument_idx, 1).style = 'm_title_style'
        wb[sheetname].cell(row_instrument_idx, 4).value = 'one by line'

        # - entries formatting
        for row in range(row_instrument_idx + 1, row_version_idx - 1):
            wb[sheetname].cell(row, 1).style = 'm_subheader_r_style'
            wb[sheetname].cell(row, 2).style = 'm_subheader_r_style'
            # merge cell from col C (3) to col H (8)
            merge_cells_list = sorted(list(wb[sheetname].merged_cells.ranges))
            for merge_cells in merge_cells_list:
                if merge_cells.left[0][0] == row and merge_cells.left[0][1] == 3:
                    if merge_cells in wb[sheetname].merged_cells.ranges:
                        wb[sheetname].unmerge_cells(merge_cells.coord)
                    # # merge instrument cells
                    wb[sheetname].cell(row, 3).style = 'm_data_l_style'
                    new_range = 'C' + str(row) + ':H' + str(row)
                    wb[sheetname].merge_cells(new_range)
                    for col in range(9, max_col + 1):
                        wb[sheetname].cell(row, col).style = 'm_bkg_style'
                    if row < row_version_idx - 2:
                        for col in range(1, 8 + 1):
                            wb[sheetname].cell(row, col).border = b_border

        # Format VERSION block
        # - title formatting
        wb[sheetname].cell(row_version_idx, 1).style = 'm_title_style'
        # - header formatting
        row_n = 1
        row_insert_idx = row_version_idx + 1
        # move merged cells downwards below row_insert_idx row
        for cr in wb[sheetname].merged_cells.ranges:
            if cr.right[0][0] > row_insert_idx:
                cr.shift(0, row_n)
        # insert row_n rows starting at row_insert_idx row
        wb[sheetname].insert_rows(row_insert_idx, row_n)
        wb[sheetname].cell(row_version_idx, 1).value = 'VERSION'

        # insert text
        wb[sheetname].cell(row_version_idx + 1, 1).value = 'number'
        wb[sheetname].cell(row_version_idx + 1, 2).value = 'date'
        wb[sheetname].cell(row_version_idx + 1, 3).value = 'author'
        wb[sheetname].cell(row_version_idx + 1, 5).value = 'modification'

        # move modification data from col C to col E
        for row in range(row_version_idx + 2, wb[sheetname].max_row + 1):
            wb[sheetname].cell(row, 5).value = wb[sheetname].cell(row, 3).value
            wb[sheetname].cell(row, 3).value = None

        # version subheader
        row = row_version_idx + 1
        for col_idx in range(1, 3):
            wb[sheetname].cell(row, col_idx).style = 'm_subheader_r_style'
        for col_idx in range(3, 8 + 1):
            wb[sheetname].cell(row, col_idx).style = 'm_subheader_l_style'
        for col_idx in range(9, max_col + 1):
            wb[sheetname].cell(row, col_idx).style = 'm_bkg_style'

        # version entries
        for row in range(row_version_idx + 2, wb[sheetname].max_row):
            for col_idx in range(1, max_col + 1):
                if col_idx < 3:
                    wb[sheetname].cell(row, col_idx).style = 'm_data_r_style'
                elif col_idx < 9:
                    wb[sheetname].cell(row, col_idx).style = 'm_data_l_style'
                else:
                    wb[sheetname].cell(row, col_idx).style = 'm_bkg_style'
                if col_idx == 2:
                    wb[sheetname].cell(row, col_idx).number_format = 'yyyy-dd-mm'

        # remove extra columns
        wb[sheetname].delete_cols(27, wb[sheetname].max_column - 26)

        # METADATA-STATION
        sheetname = 'metadata-station'
        # rename metadata-coring sheetname to metadata-station
        if 'metadata-coring' in wb.sheetnames:
            wb['metadata-coring'].title = sheetname

        wb[sheetname]['A1'].value = 'ICE CORE DATA SHEET'
        wb[sheetname]['B1'].value = 'version'
        wb[sheetname]['D1'].value = 'to be used with PySIC python module'

        # Formatting
        for col in range(1, 5):
            wb[sheetname].cell(1, col).style = 'm_title_style'
            if col == 3:
                wb[sheetname].cell(1, col).font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False,
                                                       italic=True, color='FF000000')
            if col == 4:
                wb[sheetname].cell(1, col).font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False,
                                                       italic=False, color='FF000000')
        # - POSITION
        # add line below sampling name
        row_n = 1
        row_insert_idx = 5
        for cr in wb[sheetname].merged_cells.ranges:  # move merged cells downwards below row_insert_idx row
            if cr.right[0][0] >= row_insert_idx:
                cr.shift(0, row_n)
        wb[sheetname].insert_rows(row_insert_idx, row_n)  # insert row_n rows starting at row_insert_idx row
        for row in range(row_insert_idx, row_insert_idx + row_n):
            for col in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col).style = 'm_bkg_style'

        wb[sheetname]['A6'].value = 'position'
        wb[sheetname]['C7'].value = 'start'
        wb[sheetname]['F7'].value = 'end'
        wb[sheetname]['C8'].value = 'degree'
        wb[sheetname]['D8'].value = 'minute'
        wb[sheetname]['E8'].value = 'second'
        wb[sheetname]['F8'].value = 'degree'
        wb[sheetname]['G8'].value = 'minute'
        wb[sheetname]['H8'].value = 'second'
        wb[sheetname]['A9'].value = 'latitude'
        wb[sheetname]['A10'].value = 'longitude'

        # Formatting
        for col in range(1, 8 + 1):
            wb[sheetname].cell(9, col).border = b_border
            if col > 2:
                wb[sheetname].style = 'm_data_r_style'
        for row in range(7, 10 + 1):
            wb[sheetname].cell(row, 5).border = r_border
            wb[sheetname].cell(row, 6).border = l_border
            if row == 9:
                wb[sheetname].cell(row, 5).border = br_border
                wb[sheetname].cell(row, 6).border = bl_border
        for row in range(10, 11):
            for col in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col).border = noBorder

        # add line below sampling name
        row_n = 1
        row_insert_idx = 11
        for cr in wb[sheetname].merged_cells.ranges:  # move merged cells downwards below row_insert_idx row
            if cr.right[0][0] >= row_insert_idx:
                cr.shift(0, row_n)
        wb[sheetname].insert_rows(row_insert_idx, row_n)  # insert row_n rows starting at row_insert_idx row
        for row in range(row_insert_idx, row_insert_idx + row_n):
            for col in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col).style = 'm_bkg_style'

        wb[sheetname]['A12'].value = 'date and time'
        wb[sheetname]['C13'].value = 'start'
        wb[sheetname]['D13'].value = 'end'
        wb[sheetname]['A14'].value = 'date'
        wb[sheetname]['A15'].value = 'time'
        wb[sheetname]['A16'].value = 'timezone'

        # Formatting
        for row in [14, 15]:
            for col in range(1, 4 + 1):
                wb[sheetname].cell(row, col).border = b_border
        for row in range(13, 16 + 1):
            col = 4
            wb[sheetname].cell(row, col).border = l_border
            if row in [14, 15]:
                wb[sheetname].cell(row, col).border = bl_border

        # - ICE GEOMETRY
        # add line below sampling name
        row_n = 1
        row_insert_idx = 17
        for cr in wb[sheetname].merged_cells.ranges:  # move merged cells downwards below row_insert_idx row
            if cr.right[0][0] >= row_insert_idx:
                cr.shift(0, row_n)
        wb[sheetname].insert_rows(row_insert_idx, row_n)  # insert row_n rows starting at row_insert_idx row
        for row in range(row_insert_idx, row_insert_idx + row_n):
            for col_idx in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col_idx).style = 'm_bkg_style'

        value_entry = ['snow depth', 'average freeboard', 'average ice thickness', 'water depth', 'age', 'topography',
                       'environment', 'surface condition']
        for ii, row in enumerate(range(19, 27)):
            col_idx = 1
            wb[sheetname].cell(row, col_idx).value = value_entry[ii]

        # - ICE TEMPERATURE
        # add line below sampling name
        row_n = 1
        row_insert_idx = 27
        for cr in wb[sheetname].merged_cells.ranges:  # move merged cells downwards below row_insert_idx row
            if cr.right[0][0] >= row_insert_idx:
                cr.shift(0, row_n)
        wb[sheetname].insert_rows(row_insert_idx, row_n)  # insert row_n rows starting at row_insert_idx row
        for row in range(row_insert_idx, row_insert_idx + row_n):
            for col_idx in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col_idx).style = 'm_bkg_style'

        # - SAMPLING EVENT
        # add line below sampling name
        row_n = 1
        row_insert_idx = 33
        for cr in wb[sheetname].merged_cells.ranges:  # move merged cells downwards below row_insert_idx row
            if cr.right[0][0] >= row_insert_idx:
                cr.shift(0, row_n)
        wb[sheetname].insert_rows(row_insert_idx, row_n)  # insert row_n rows starting at row_insert_idx row
        for row in range(33, 35):
            for col_idx in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col_idx).border = noBorder
        for row in range(row_insert_idx, row_insert_idx + row_n):
            for col_idx in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col_idx).style = 'm_bkg_style'

        wb[sheetname]['A34'].value = 'sampling event'

        # - WEATHER INFORMATION
        wb[sheetname]['A40'].value = 'weather information'

        # - SAMPLING EVENT
        # add line below sampling name
        row_n = 1
        row_insert_idx = 46
        for cr in wb[sheetname].merged_cells.ranges:  # move merged cells downwards below row_insert_idx row
            if cr.right[0][0] >= row_insert_idx:
                cr.shift(0, row_n)
        wb[sheetname].insert_rows(row_insert_idx, row_n)  # insert row_n rows starting at row_insert_idx row
        for row in range(row_insert_idx, row_insert_idx + row_n):
            for col in range(1, wb[sheetname].max_column):
                wb[sheetname].cell(row, col).style = 'm_bkg_style'

        wb[sheetname]['A47'].value = 'general comments'

        # set cell geometry
        max_row = wb[sheetname].max_row
        for row in range(1, max_row + 1):
            wb[sheetname].row_dimensions[row].height = 12.75
            wb[sheetname].cell(row, wb['metadata-station'].max_column).style = 'm_bkg_style'
        wb[sheetname].column_dimensions['A'].width = 1.76 * 14  # 1.76 in, fudge_factor = 13.97

        # - PROFILE ORIENTATION
        # -- ICE
        # Look for orientation of ice data
        data_sheetnames = wb.sheetnames
        for element in ['metadata-core', 'metadata-station', 'lists', 'snow', 'seawater']:
            if element in data_sheetnames:
                data_sheetnames.remove(element)
        # find sheetname with data
        data_sheet = {}
        for sheetname in data_sheetnames:
            for row in range(1, 4):
                for col in range(1, wb[sheetname].max_column):
                    str_condition = isinstance(wb[sheetname].cell(row, col).value, str)
                    # For column with potential depth entry
                    if str_condition and 'depth' in wb[sheetname].cell(row, col).value:
                        for row_idx in range(row, wb[sheetname].max_row):
                            if isinstance(wb[sheetname].cell(row_idx, col).value, (int, float)):
                                start_row = row_idx
                                depth_data = []
                                for row_idx in range(start_row, wb[sheetname].max_row):
                                    depth_data.append(wb[sheetname].cell(row_idx, col).value)
                                depth_data = np.array(depth_data).astype(float)
                                depth_data = depth_data[np.where(~np.isnan(depth_data))]
                                if len(depth_data) > 0:
                                    # read vertical reference if existing
                                    if sheetname.lower() == 'salo18':
                                        data_sheet[sheetname] = [wb[sheetname]['C1'].value, wb[sheetname]['F1'].value]
                                    else:
                                        data_sheet[sheetname] = [wb[sheetname]['C1'].value, wb[sheetname]['E1'].value]
                                    break
                            if sheetname in data_sheet:
                                break
                    if sheetname in data_sheet:
                        break
        # check orientation consistency:
        for ii, sheet in enumerate(data_sheet.keys()):
            if data_sheet[sheet][0] is None:
                logger.warning('%s - %s vertical reference is None' % (wb['metadata-core']['C1'].value, sheet))
            if data_sheet[sheet][1] is None:
                logger.warning('%s - %s vertical direction is None' % (wb['metadata-core']['C1'].value, sheet))

            if ii == 0 or vert_ref is None:
                vert_ref = data_sheet[sheet][0]
            if ii == 0 or vert_dir is None:
                vert_dir = data_sheet[sheet][1]
            else:
                if data_sheet[sheet][0] is not None and data_sheet[sheet][0] != vert_ref:
                    logger.error('%s - vertical references are not consistant' % wb['metadata-core']['C1'].value)
                if data_sheet[sheet][1] is not None and data_sheet[sheet][1] != vert_dir:
                    logger.error('%s - vertical direction are not consistant' % wb['metadata-core']['C1'].value)
        # update ice direction in metadata-core
        if vert_ref in vert_ref_list.formula1:
            wb['metadata-core']['D13'] = vert_ref
        if vert_dir in vert_dir_list.formula1:
            wb['metadata-core']['D14'] = vert_dir

        # -- SNOW
        sheetname = 'snow'
        data_sheet = {}
        for row in range(1, 4):
            for col in range(1, wb[sheetname].max_column):
                str_condition = isinstance(wb[sheetname].cell(row, col).value, str)
                # For column with potential depth entry
                if str_condition and 'depth' in wb[sheetname].cell(row, col).value:
                    for row_idx in range(row, wb[sheetname].max_row):
                        if isinstance(wb[sheetname].cell(row_idx, col).value, (int, float)):
                            start_row = row_idx
                            depth_data = []
                            for row_idx in range(start_row, wb[sheetname].max_row):
                                depth_data.append(wb[sheetname].cell(row_idx, col).value)
                            depth_data = np.array(depth_data).astype(float)
                            depth_data = depth_data[np.where(~np.isnan(depth_data))]
                            if len(depth_data) > 0:
                                # read vertical reference if existing
                                data_sheet[sheetname] = [wb[sheetname]['C1'].value, wb[sheetname]['E1'].value]
                            break
                        if sheetname in data_sheet:
                            break
                if sheetname in data_sheet:
                    break
        if len(data_sheet.keys()) > 1:
            # update ice direction in metadata-core
            if vert_ref in vert_ref_list.formula1:
                wb['metadata-core']['D15'] = vert_ref
            if vert_dir in vert_dir_list.formula1:
                wb['metadata-core']['D16'] = vert_dir
        else:
            wb['metadata-core']['D15'] = 'N/A'
            wb['metadata-core']['D16'] = 'N/A'

        # -- SEAWATER
        sheetname = 'seawater'
        data_sheet[sheetname] = ['N/A', 'N/A']
        for row in range(1, 4):
            for col in range(1, wb[sheetname].max_column):
                str_condition = isinstance(wb[sheetname].cell(row, col).value, str)
                # For column with potential depth entry
                if str_condition and 'depth' in wb[sheetname].cell(row, col).value:
                    for row_idx in range(row, wb[sheetname].max_row):
                        if isinstance(wb[sheetname].cell(row_idx, col).value, (int, float)):
                            start_row = row_idx
                            depth_data = []
                            for row_idx in range(start_row, wb[sheetname].max_row):
                                depth_data.append(wb[sheetname].cell(row_idx, col).value)
                            depth_data = np.array(depth_data).astype(float)
                            depth_data = depth_data[np.where(~np.isnan(depth_data))]
                            if len(depth_data) > 0:
                                # read vertical reference if existing
                                data_sheet[sheetname] = [wb[sheetname]['C1'].value, wb[sheetname]['E1'].value]
                            break
                        if sheetname in data_sheet:
                            break
                if sheetname in data_sheet:
                    break
        if len(data_sheet.keys()) > 1:
            # update ice direction in metadata-core
            if vert_ref in vert_ref_list.formula1:
                wb['metadata-core']['D17'] = vert_ref
            if vert_dir in vert_dir_list.formula1:
                wb['metadata-core']['D18'] = vert_dir

        # DATA UPDATE
        # - LISTS
        # Override latest `list` sheet
        if 'lists' in wb.sheetnames:
            list_sheet_index = wb.worksheets.index(wb['lists'])
            del wb['lists']
        wb.create_sheet("lists", list_sheet_index)
        wb_source = openpyxl.load_workbook(
            os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYMMDD-P-1.4.1-lists.xlsx'), data_only=True)
        copy_cells(wb_source['lists'], wb['lists'])  # copy all the cel values and styles
        copy_sheet_attributes(wb_source['lists'], wb['lists'])

        # - SALO18
        sheetname = 'salo18'
        if 'SALO18' in wb.sheetnames:
            wb['SALO18'].title = sheetname
            wb['salo181'].title = sheetname
        if 'salo18' in wb.sheetnames:
            # unmerge all cells
            while wb[sheetname].merged_cells.ranges:
                wb[sheetname].merged_cells.ranges.pop()
            # remove all data validation
            row_idx = wb[sheetname].max_row
            for col in range(1, wb[sheetname].max_column):
                removeExistingCellDataValidation(wb[sheetname], wb[sheetname].cell(row_idx, col))

            # remove first line
            delete_row_with_merge(wb[sheetname], 1, 1)

            # remove depth center (C) if empty
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('C')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)
            wb[sheetname]['C1'].value = 'salinity'
            wb[sheetname]['D1'].value = None
            wb[sheetname]['C2'].value = 'ID'
            wb[sheetname]['G2'].value = 'temperature'

            # insert quality column for salinity,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('E')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('C1:E1')
            wb[sheetname].merge_cells('F1:G1')

            # insert quality column for d18O,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('K')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('I1:K1')
            wb[sheetname]['I2'].value = 'ID'

            # insert quality column for dD,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('N')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('L1:N1')
            wb[sheetname]['L2'].value = 'ID'

            # delete column O (density)
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('O')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['C', 'I', 'L', 'O']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - TEMP
        sheetname = 'temp'
        if 'TEMP' in wb.sheetnames:
            wb['TEMP'].title = sheetname
            if 'temp1' in wb.sheetnames:
                wb['temp1'].title = sheetname
        if sheetname in wb.sheetnames:
            # unmerge all cells
            # while len(wb[sheetname].merged_cells.ranges) > 0:
            #     crs = wb[sheetname].merged_cells.ranges[0].coord
            #     wb[sheetname].unmerge_cells(crs)
            while wb[sheetname].merged_cells.ranges:
                wb[sheetname].merged_cells.ranges.pop()
            # remove first line
            row_n = 1
            row_delete_idx = 1
            delete_row_with_merge(wb[sheetname], row_delete_idx, row_n)
            removeExistingCellDataValidation(wb[sheetname], 'C1')
            removeExistingCellDataValidation(wb[sheetname], 'E1')

            # insert 1 column to move comment column to the right
            col_n = 1
            col_insert_idx = 3
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)

            wb[sheetname]['D1'].value = 'comment'
            wb[sheetname]['C2'].value = 'quality'
            wb[sheetname]['D2'].value = 'value'
            wb[sheetname]['C3'].value = '[0-9]'
            wb[sheetname]['D3'].value = '-'
            wb[sheetname].merge_cells('B1:C1')

            # formating
            max_row = 3 + data_row_n + 1
            l_border_col = ['B', 'D']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - TEX
        sheetname = 'tex'
        if 'TEX' in wb.sheetnames:
            wb['TEX'].title = sheetname
            if 'tex1' in wb.sheetnames:
                wb['tex1'].title = sheetname
        if sheetname in wb.sheetnames:
            # unmerge all cell
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)

            # remove first line
            row_n = 1
            row_delete_idx = 1
            delete_row_with_merge(wb[sheetname], row_delete_idx, row_n)
            removeExistingCellDataValidation(wb[sheetname], 'C1')
            removeExistingCellDataValidation(wb[sheetname], 'E1')

            # remove data validation for the last line
            row_idx = wb[sheetname].max_row
            for col in range(1, wb[sheetname].max_column):
                removeExistingCellDataValidation(wb[sheetname], wb[sheetname].cell(row_idx, col))

            # remove column C to move texture column to the left
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('C')
            delete_col_with_merge(wb[sheetname], col_insert_idx, col_n, True, 4)

            # insert column to move comment column to the right
            col_n = 6
            col_insert_idx = openpyxl.utils.column_index_from_string('E')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)

            wb[sheetname]['C1'].value = 'texture'
            wb[sheetname]['E1'].value = 'inclusion'
            wb[sheetname]['F1'].value = 'description'
            wb[sheetname]['G1'].value = 'stratigraphy'
            wb[sheetname]['K1'].value = 'comment'
            subheader_list = ['value', 'quality', 'value', 'value', 'ID', 'type', 'quality', 'section', 'value']
            unit_list = ['-', '[0-9]', '-', '-', '-', '-', '[0-9]', '-', '-']

            for ii, col_idx in enumerate(range(openpyxl.utils.column_index_from_string('C'),
                                               openpyxl.utils.column_index_from_string('K') + 1)):
                wb[sheetname].cell(2, col_idx).value = subheader_list[ii]
                wb[sheetname].cell(3, col_idx).value = unit_list[ii]
            wb[sheetname].merge_cells('C1:D1')
            wb[sheetname].merge_cells('G1:J1')

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['C', 'G', 'K']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

            # remove data validation
            for row in range(3, max_row):
                removeExistingCellDataValidation(wb[sheetname], 'C' + str(row))
            dv_range = 'C4:C' + str(max_row - 1)
            wb[sheetname].add_data_validation(dv_lists_dict['texture_dv'])
            dv_lists_dict['texture_dv'].add(dv_range)
            dv_range = 'E4:E' + str(max_row - 1)
            wb[sheetname].add_data_validation(dv_lists_dict['inclusion_dv'])
            dv_lists_dict['inclusion_dv'].add(dv_range)
            dv_range = 'F4:F' + str(max_row - 1)
            wb[sheetname].add_data_validation(dv_lists_dict['description_dv'])
            dv_lists_dict['description_dv'].add(dv_range)

        # - ECO
        sheetname = 'eco'
        if 'ECO' in wb.sheetnames:
            wb['ECO'].title = sheetname
            if 'eco1' in wb.sheetnames:
                wb['eco1'].title = sheetname
        if sheetname in wb.sheetnames:
            # unmerge all cell
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)

            # remove first line
            row_n = 1
            row_delete_idx = 1
            delete_row_with_merge(wb[sheetname], row_delete_idx, row_n)
            removeExistingCellDataValidation(wb[sheetname], 'C1')
            removeExistingCellDataValidation(wb[sheetname], 'E1')

            # add column for field sample quality
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('D')
            wb[sheetname].insert_cols(col_insert_idx, col_n)

            wb[sheetname]['C1'].value = 'field sample'
            wb[sheetname]['C2'].value = 'ID'
            wb[sheetname]['D2'].value = 'quality'
            wb[sheetname]['D3'].value = '-'
            wb[sheetname]['D3'].value = '[0-9]'
            wb[sheetname].merge_cells('C1:D1')

            # -- melted volume
            # add 2 column to the left of total volume:
            col_n = 2
            col_insert_idx = openpyxl.utils.column_index_from_string('E')
            wb[sheetname].insert_cols(col_insert_idx, col_n)

            # move added seawater column to the left of total volume
            wb[sheetname].move_range('H1:H' + str(wb[sheetname].max_row), rows=0, cols=-2, translate=True)
            wb[sheetname]['E1'].value = 'melted volume'
            wb[sheetname]['E2'].value = 'ID'
            wb[sheetname]['F2'].value = 'added seawater volume'
            wb[sheetname]['G2'].value = 'total volume'
            wb[sheetname]['H2'].value = 'value'
            wb[sheetname]['E3'].value = '-'
            wb[sheetname]['F3'].value = 'L'
            wb[sheetname]['G3'].value = 'L'
            wb[sheetname]['H3'].value = 'L'
            wb[sheetname].merge_cells('E1:H1')

            # -- 'eco. data example'
            # insert 4 column
            col_n = 4
            col_insert_idx = openpyxl.utils.column_index_from_string('I')
            wb[sheetname].insert_cols(col_insert_idx, col_n)

            wb[sheetname]['I1'].value = 'eco. data example'
            wb[sheetname]['I2'].value = 'ID'
            wb[sheetname]['J2'].value = 'volume'
            wb[sheetname]['K2'].value = 'value'
            wb[sheetname]['L2'].value = 'quality'
            wb[sheetname]['I3'].value = '-'
            wb[sheetname]['J3'].value = 'L'
            wb[sheetname]['K3'].value = 'μg/l'
            wb[sheetname]['L3'].value = '[0-9]'
            wb[sheetname].merge_cells('I1:L1')

            # modified nutrient entry, add quality column
            col_n = 1
            col_idx = openpyxl.utils.column_index_from_string('O')
            insert_col_with_merge(wb[sheetname], col_idx, col_n)
            wb[sheetname]['M1'].value = 'nutrient'
            wb[sheetname]['M2'].value = 'ID'
            wb[sheetname]['N2'].value = 'value'
            wb[sheetname]['O2'].value = 'quality'
            wb[sheetname]['M3'].value = '-'
            wb[sheetname]['N3'].value = '-'
            wb[sheetname]['O3'].value = '[0-9]]'
            wb[sheetname].merge_cells('M1:O1')

            # remove extra column
            col_n = 24
            col_delete_idx = openpyxl.utils.column_index_from_string('P')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['E', 'M', 'I', 'P']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - SNOW
        sheetname = 'snow'
        if 'SNOW' in wb.sheetnames:
            wb['SNOW'].title = sheetname + '_temp'
            wb[sheetname + '_temp'].title = sheetname
        if sheetname in wb.sheetnames:
            # unmerge all cell
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)

            # remove first line
            delete_row_with_merge(wb[sheetname], 1, 1)
            removeExistingCellDataValidation(wb[sheetname], 'C1')
            removeExistingCellDataValidation(wb[sheetname], 'E1')

            wb[sheetname]['C1'].value = 'salinity'
            wb[sheetname]['C2'].value = 'ID'
            wb[sheetname]['C3'].value = '-'

            # insert quality column for salinity,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('E')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('C1:E1')
            wb[sheetname]['G1'].value = None
            wb[sheetname].merge_cells('F1:G1')
            wb[sheetname]['H2'].value = 'value'

            # insert quality column for d18O,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('K')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('I1:K1')
            wb[sheetname]['I2'].value = 'ID'

            # insert quality column for dD,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('N')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('L1:N1')
            wb[sheetname]['L2'].value = 'ID'

            # insert quality column for density,
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('Q')
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(1, col_insert_idx - 2).value = 'density'
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(2, col_insert_idx + 1).value = 'value'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('O1:P1')

            # insert quality column for comment
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('R')
            insert_col_with_merge(wb[sheetname], col_insert_idx + 1, col_n)
            wb[sheetname].cell(1, col_insert_idx).value = 'comment'
            wb[sheetname].cell(2, col_insert_idx).value = 'value'
            wb[sheetname].cell(3, col_insert_idx).value = '-'

            # insert quality column for temperature_quality
            col_n = 1
            col_insert_idx = openpyxl.utils.column_index_from_string('V')
            col_tmp_idx = col_insert_idx - 2
            insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
            wb[sheetname].cell(1, col_insert_idx - 2).value = 'temperature'
            wb[sheetname].cell(2, col_insert_idx).value = 'quality'
            wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
            wb[sheetname].merge_cells('T1:V1')

            # if nutrient
            delta_nuts = 0
            for col_idx in range(1, wb[sheetname].max_column):
                if wb[sheetname].cell(1, col_idx).value in ['Nutrient', 'nutrient']:
                    col_nuts = col_idx
                    delta_nuts = 3
                    # insert quality column for nutrient_quality
                    col_n = 1
                    col_insert_idx = openpyxl.utils.column_index_from_string('Z')
                    insert_col_with_merge(wb[sheetname], col_insert_idx, col_n)
                    wb[sheetname].cell(1, col_nuts).value = 'nutrient'
                    wb[sheetname].cell(2, col_nuts).value = 'ID'
                    wb[sheetname].cell(2, col_nuts + 1).value = 'value'
                    wb[sheetname].cell(3, col_nuts + 1).value = '-'
                    wb[sheetname].cell(2, col_insert_idx).value = 'quality'
                    wb[sheetname].cell(3, col_insert_idx).value = '[0-9]'
                    wb[sheetname].merge_cells('X1:Z1')
                    break
                else:
                    pass

            # - eco sample: insert 3 column
            col_n = 4
            col_eco_n = col_n
            col_idx = openpyxl.utils.column_index_from_string('X') + delta_nuts + 1
            col_eco_idx = col_idx
            insert_col_with_merge(wb[sheetname], col_idx, col_n)
            wb[sheetname].cell(1, col_idx).value = 'eco sample'
            wb[sheetname].cell(2, col_idx).value = 'ID'
            wb[sheetname].cell(2, col_idx + 1).value = 'value'
            wb[sheetname].cell(2, col_idx + 2).value = 'quality'
            wb[sheetname].cell(3, col_idx).value = '-'
            wb[sheetname].cell(3, col_idx + 1).value = '-'
            wb[sheetname].cell(3, col_idx + 2).value = '[0-9]'
            merge_range = openpyxl.utils.get_column_letter(col_idx) + '1:' + openpyxl.utils.get_column_letter(
                col_idx + 2) + '1'
            wb[sheetname].merge_cells(merge_range)

            # - extra sample
            col_idx = openpyxl.utils.column_index_from_string('X') + delta_nuts + 1 + col_eco_n
            col_exs_idx = col_idx
            wb[sheetname].cell(1, col_idx + 2).value = 'extra sample'
            wb[sheetname].cell(2, col_idx + 2).value = 'ID'

            # - snow micropen
            col_idx = openpyxl.utils.column_index_from_string('X') + delta_nuts + 1 + col_eco_n + 4
            col_smp_idx = col_idx
            wb[sheetname].cell(1, col_idx).value = 'Snow micro pen SMP'
            wb[sheetname].cell(2, col_idx + 2).value = 'ID'
            wb[sheetname].cell(3, col_idx).value = '-'
            wb[sheetname].cell(3, col_idx + 1).value = '-'
            wb[sheetname].cell(3, col_idx + 2).value = '-'
            merge_range = openpyxl.utils.get_column_letter(col_idx) + '1:' + openpyxl.utils.get_column_letter(
                col_idx + 2) + '1'
            wb[sheetname].merge_cells(merge_range)

            # - snow water equivalent
            col_n = 3
            col_idx = openpyxl.utils.column_index_from_string('X') + delta_nuts + 1 + col_eco_n + 4 + 4
            col_swe_idx = col_idx
            wb[sheetname].insert_cols(col_idx + 1, col_n)
            wb[sheetname].cell(1, col_idx).value = 'SWE'
            wb[sheetname].cell(2, col_idx).value = 'value'
            wb[sheetname].cell(2, col_idx + 1).value = 'snow depth'
            wb[sheetname].cell(2, col_idx + 2).value = 'quality'
            wb[sheetname].cell(2, col_idx + 3).value = 'comment'
            wb[sheetname].cell(3, col_idx).value = 'g'
            wb[sheetname].cell(3, col_idx + 1).value = 'cm'
            wb[sheetname].cell(3, col_idx + 2).value = '[0-9]'
            wb[sheetname].cell(3, col_idx + 3).value = '-'
            merge_range = openpyxl.utils.get_column_letter(col_idx) + '1:' + openpyxl.utils.get_column_letter(
                col_idx + col_n) + '1'
            wb[sheetname].merge_cells(merge_range)

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['C', 'I', 'L', 'O' 'R']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

            # extra formatting for snow
            bkg_col = [col_tmp_idx - 1, col_nuts - 1, col_eco_idx - 1, col_exs_idx - 1, col_smp_idx - 1,
                       col_swe_idx - 1]
            for row_idx in range(1, max_row):
                for col_idx in bkg_col:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'

            for row_idx in range(5, max_row):
                for col_idx in range(col_swe_idx, col_swe_idx + 4):
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'

            for row_idx in range(10, max_row):
                for col_idx in range(col_smp_idx, col_smp_idx + 3):
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'

        # - DENSITY-VOLUME
        sheetname = 'density-volume'
        if 'Density-volume' in wb.sheetnames:
            wb['Density-volume'].title = sheetname
            if 'density-volume1' in wb.sheetnames:
                wb['density-volume1'].title = sheetname
        if sheetname in wb.sheetnames:
            # unmerge all cell
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)

            # remove first line
            row_n = 1
            row_delete_idx = 1
            delete_row_with_merge(wb[sheetname], row_delete_idx, row_n)
            removeExistingCellDataValidation(wb[sheetname], 'C1')
            removeExistingCellDataValidation(wb[sheetname], 'E1')

            # remove field column (C) if empty
            col_n = 1
            col_delete_idx = 3
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # insert 2 column to the right of C
            col_n = 2
            col_insert_idx = 4
            wb[sheetname].insert_cols(col_insert_idx, col_n)

            # move density column L, 8 to the left
            wb[sheetname].move_range('L1:L' + str(wb[sheetname].max_row), rows=0, cols=-8, translate=True)

            wb[sheetname]['C1'].value = 'density'
            wb[sheetname]['C2'].value = 'ID'
            wb[sheetname]['D2'].value = 'value'
            wb[sheetname]['E2'].value = 'quality'
            wb[sheetname]['C3'].value = 'value'
            wb[sheetname]['D3'].value = 'kg / m3'
            wb[sheetname]['E3'].value = '[0-9]'
            wb[sheetname].merge_cells('C1:E1')

            # -- thickness
            wb[sheetname].merge_cells('F1:J1')

            # remove column K:
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('K')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # -- mass
            wb[sheetname]['K1'].value = 'mass'
            wb[sheetname]['K2'].value = 'value'
            wb[sheetname]['K3'].value = 'kg'

            wb[sheetname]['L1'].value = 'comment'
            wb[sheetname]['L2'].value = 'value'
            wb[sheetname]['L3'].value = '-'

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['C', 'F', 'K', 'L']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - DENSITY-DENSIMETRY
        sheetname = 'density-densimetry'
        if 'Density-densimetry' in wb.sheetnames:
            wb['Density-densimetry'].title = sheetname
            if 'density-densimetry1' in wb.sheetnames:
                wb['density-densimetry1'].title = sheetname
        if sheetname in wb.sheetnames:
            # remove all data validation
            row_idx = wb[sheetname].max_row
            for col in range(1, wb[sheetname].max_column):
                removeExistingCellDataValidation(wb[sheetname], wb[sheetname].cell(row_idx, col))
            # unmerge all cell
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)
            # remove first line
            delete_row_with_merge(wb[sheetname], 1, 1)

            # remove field column (C) if empty
            col_n = 1
            col_delete_idx = 3
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # insert 2 column to the right of C
            col_n = 2
            col_insert_idx = 4
            wb[sheetname].insert_cols(col_insert_idx, col_n)

            # move density column L, 8 to the left
            wb[sheetname].move_range('H1:H' + str(wb[sheetname].max_row), rows=0, cols=-4, translate=True)
            wb[sheetname]['C1'].value = 'density'
            wb[sheetname]['C2'].value = 'ID'
            wb[sheetname]['D2'].value = 'value'
            wb[sheetname]['E2'].value = 'quality'
            wb[sheetname]['C3'].value = 'value'
            wb[sheetname]['D3'].value = 'kg / m3'
            wb[sheetname]['E3'].value = '[0-9]'
            wb[sheetname].merge_cells('C1:E1')

            wb[sheetname]['F1'].value = 'mass'
            wb[sheetname]['F2'].value = 'air'
            wb[sheetname]['G2'].value = 'liquid'
            wb[sheetname]['F3'].value = 'kg / m3'
            wb[sheetname]['G3'].value = 'kg / m3'
            wb[sheetname].merge_cells('F1:G1')

            # remove column H:
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('H')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n)

            # - comment
            wb[sheetname]['H1'].value = 'comment'
            wb[sheetname]['H2'].value = 'value'
            wb[sheetname]['H3'].value = '-'

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['C', 'F', 'H']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - SEAWATER
        sheetname = 'seawater'
        if sheetname in wb.sheetnames:
            # unmerge all cells
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)

            # remove all data validation
            row_idx = wb[sheetname].max_row
            for col in range(1, wb[sheetname].max_column):
                removeExistingCellDataValidation(wb[sheetname], wb[sheetname].cell(row_idx, col))

            # remove first line
            row_n = 1
            row_delete_idx = 1
            delete_row_with_merge(wb[sheetname], row_delete_idx, row_n)

            # insert a row for the subheader if they are missing
            if not (wb[sheetname]['A2'].value == 'value' and wb[sheetname]['B2'].value == 'value'):
                row_n = 1
                row_idx = 2
                insert_row_with_merge(wb[sheetname], row_idx, row_n)

            # if column D is field Sample ID and column E is salinity
            if wb[sheetname].cell(1, 4).value == 'field' and wb[sheetname].cell(1, 5).value == 'salinity':
                # if column C is empty, delete
                flag_empty = True
                col_idx = openpyxl.utils.column_index_from_string('C')
                for row_idx in range(4, wb[sheetname].max_row):
                    if wb[sheetname].cell(row_idx, col_idx).value not in [None, '']:
                        cell = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
                        logger.error('%s - %s: cell %s is not empty impossible to delete column' % (
                            wb['metadata-core']['C1'].value, sheetname, cell))
                        flag_empty = False
                        break
                    else:
                        pass
                if flag_empty:
                    delete_col_with_merge(wb[sheetname], 3, 1)
            else:
                # if column C is empty, use it for ID
                flag_empty = True
                col_idx = openpyxl.utils.column_index_from_string('C')
                for row_idx in range(4, wb[sheetname].max_row):
                    if wb[sheetname].cell(row_idx, col_idx).value not in [None, '']:
                        cell = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
                        logger.error('%s - %s: cell %s is not empty impossible to delete column' % (
                            wb['metadata-core']['C1'].value, sheetname, cell))
                        flag_empty = False
                        break
                    else:
                        pass
                if not flag_empty:
                    col_idx = 4
                    col_n = 1
                    insert_col_with_merge(wb[sheetname], col_idx, col_n)
            wb[sheetname].cell(1, col_idx).value = 'salinity'
            wb[sheetname].cell(1, col_idx + 1).value = None
            wb[sheetname].cell(2, col_idx).value = 'ID'
            wb[sheetname].cell(2, col_idx + 1).value = 'value'
            wb[sheetname].cell(3, col_idx).value = '-'
            wb[sheetname].cell(3, col_idx + 1).value = '-'

            # move temperature column (E) to before d18O (I)
            col_idx = find_str_in_row(wb[sheetname], 'temperature')
            if len(col_idx) > 0:
                col_idx = openpyxl.utils.column_index_from_string('I')
                col_border = [3, col_idx]
                insert_col_with_merge(wb[sheetname], col_idx, 1)
                move_range = 'E1:E' + str(wb[sheetname].max_row)
                wb[sheetname].move_range(move_range, rows=0, cols=col_idx - openpyxl.utils.column_index_from_string('E'))
                wb[sheetname]['I2'].value = 'value'
            else:
                col_border = [3]
            # if column E is empty or not conductivty, use it for salinity_quality
            col_idx = openpyxl.utils.column_index_from_string('E')
            flag_empty = True
            if wb[sheetname].cell(1, col_idx).value == 'conductivity':
                flag_empty = False
            else:
                for row_idx in range(4, wb[sheetname].max_row):
                    if wb[sheetname].cell(row_idx, col_idx).value not in [None, '']:
                        cell = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
                        logger.error('%s - %s: cell %s is not empty impossible to delete column' % (
                            wb['metadata-core']['C1'].value, sheetname, cell))
                        flag_empty = False
                        break
                    else:
                        pass
            if not flag_empty:
                insert_col_with_merge(wb[sheetname], 5, 1)
            wb[sheetname].cell(1, col_idx).value = None
            wb[sheetname].cell(2, col_idx).value = 'quality'
            wb[sheetname].cell(3, col_idx).value = '[0-9]'

            wb[sheetname].merge_cells('C1:E1')

            wb[sheetname]['F2'].value = 'value'
            wb[sheetname]['H2'].value = 'value'

            wb[sheetname]['G1'].value = None
            wb[sheetname]['G2'].value = 'sample temp'
            wb[sheetname]['G3'].value = '˚C'
            wb[sheetname].merge_cells('F1:G1')

            # insert ID and quality column for d18O, respectively before and after:
            col_n = 1
            col_id_idx = find_str_in_row(wb[sheetname], 'd18O', 1)[0]
            col_border.append(col_id_idx)
            if not wb[sheetname].cell(2, col_id_idx).value == 'sample ID':
                insert_col_with_merge(wb[sheetname], col_id_idx, col_n)
            wb[sheetname].cell(1, col_id_idx).value = 'd18O'
            wb[sheetname].cell(2, col_id_idx).value = 'ID'
            wb[sheetname].cell(2, col_id_idx + 1).value = 'value'
            wb[sheetname].cell(3, col_id_idx).value = '-'
            col_qual_idx = col_id_idx + 2
            if not wb[sheetname].cell(2, col_qual_idx).value == 'quality':
                insert_col_with_merge(wb[sheetname], col_qual_idx, col_n)
            wb[sheetname].cell(2, col_qual_idx).value = 'quality'
            wb[sheetname].cell(3, col_qual_idx).value = '[0-9]'
            col_id = openpyxl.utils.get_column_letter(col_id_idx)
            col_qual = openpyxl.utils.get_column_letter(col_qual_idx)
            merge_range = col_id + str(1) + ':' + col_qual + str(1)
            wb[sheetname].merge_cells(merge_range)

            # insert ID and quality column for dD, respectively before and after:
            col_n = 1
            col_id_idx = find_str_in_row(wb[sheetname], 'dD', 1)[0]
            col_border.append(col_id_idx)
            if not wb[sheetname].cell(2, col_id_idx).value == 'sample ID':
                insert_col_with_merge(wb[sheetname], col_id_idx, col_n)
            wb[sheetname].cell(1, col_id_idx).value = 'dD'
            wb[sheetname].cell(2, col_id_idx).value = 'ID'
            wb[sheetname].cell(2, col_id_idx + 1).value = 'value'
            wb[sheetname].cell(3, col_id_idx).value = '-'
            col_qual_idx = col_id_idx + 2
            if not wb[sheetname].cell(2, col_qual_idx).value == 'quality':
                insert_col_with_merge(wb[sheetname], col_qual_idx, col_n)
            wb[sheetname].cell(2, col_qual_idx).value = 'quality'
            wb[sheetname].cell(3, col_qual_idx).value = '[0-9]'
            col_id = openpyxl.utils.get_column_letter(col_id_idx)
            col_qual = openpyxl.utils.get_column_letter(col_qual_idx)
            merge_range = col_id + str(1) + ':' + col_qual + str(1)
            wb[sheetname].merge_cells(merge_range)

            # remove density column if empty
            col_n = 1
            col_idx = find_str_in_row(wb[sheetname], 'density', 1)
            if len(col_idx) > 0:
                col_delete_idx = col_idx[0]
                delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # remove sample ID column if empty
            col_n = 1
            col_idx = find_str_in_row(wb[sheetname], 'Sample #', 1)
            if len(col_idx) > 0:
                col_delete_idx = col_idx[0]
                delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # - comment
            col_idx = find_str_in_row(wb[sheetname], 'comment', 1)[0]
            col_border.append(col_idx)
            wb[sheetname].cell(2, col_idx).value = 'value'
            wb[sheetname].cell(3, col_idx).value = '-'

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = [openpyxl.utils.get_column_letter(col_idx) for col_idx in col_border]
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - SEDIMENT
        sheetname = 'sediment'
        if sheetname in wb.sheetnames:
            # unmerge all cells
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)
            # remove all data validation
            row_idx = wb[sheetname].max_row
            for col in range(1, wb[sheetname].max_column):
                removeExistingCellDataValidation(wb[sheetname], wb[sheetname].cell(row_idx, col))

            # remove first line
            delete_row_with_merge(wb[sheetname], 1, 1)

            # insert row for subheader
            if not (wb[sheetname]['A2'].value == 'value' and wb[sheetname]['B2'].value == 'value'):
                wb[sheetname].insert_rows(2, 1)
                wb[sheetname]['A2'].value = 'value'
                wb[sheetname]['B2'].value = 'value'

            # remove depth center (C) if empty
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('C')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # remove salinity if empty
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('C')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 4)

            # - melted volume: add 2 columns
            col_n = 1
            col_idx = openpyxl.utils.column_index_from_string('C')
            wb[sheetname].insert_cols(col_idx, col_n)
            wb[sheetname].cell(1, col_idx).value = 'melted volume'
            wb[sheetname].cell(1, col_idx + 1).value = None
            wb[sheetname].cell(2, col_idx).value = 'ID'
            wb[sheetname].cell(2, col_idx + 1).value = 'value'
            wb[sheetname].cell(3, col_idx).value = '-'
            wb[sheetname].cell(3, col_idx + 1).value = 'l'

            # - sediment mass: add 1 column
            col_n = 1
            col_idx = openpyxl.utils.column_index_from_string('E')
            wb[sheetname].insert_cols(col_idx, col_n)
            wb[sheetname].cell(1, col_idx).value = 'sediment mass'
            wb[sheetname].cell(1, col_idx + 1).value = None
            wb[sheetname].cell(2, col_idx).value = 'value'
            wb[sheetname].cell(2, col_idx + 1).value = 'quality'
            wb[sheetname].cell(3, col_idx).value = 'g'
            wb[sheetname].cell(3, col_idx + 1).value = '[0-9]'
            cell_range = openpyxl.utils.get_column_letter(col_idx) + '1:' + openpyxl.utils.get_column_letter(
                col_idx + 1) + '1'
            wb[sheetname].merge_cells(cell_range)

            # - particulate mass: add 1 column
            col_n = 1
            col_idx = openpyxl.utils.column_index_from_string('G')
            wb[sheetname].insert_cols(col_idx, col_n)
            wb[sheetname].cell(1, col_idx).value = 'particulate mass'
            wb[sheetname].cell(1, col_idx + 1).value = None
            wb[sheetname].cell(2, col_idx).value = 'value'
            wb[sheetname].cell(2, col_idx + 1).value = 'quality'
            wb[sheetname].cell(3, col_idx).value = 'g'
            wb[sheetname].cell(3, col_idx + 1).value = '[0-9]'
            cell_range = openpyxl.utils.get_column_letter(col_idx) + '1:' + openpyxl.utils.get_column_letter(
                col_idx + 1) + '1'
            wb[sheetname].merge_cells(cell_range)

            # - comment
            wb[sheetname]['I2'].value = 'value'
            wb[sheetname]['I3'].value = '-'

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['C', 'E', 'G', 'I']
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # - CT
        sheetname = 'ct'
        if 'CT' in wb.sheetnames:
            wb['CT'].title = '_temp'
            wb['_temp'].title = sheetname
        if sheetname in wb.sheetnames:
            # unmerge all cells
            crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
            while crs_list:
                crs = crs_list[0]
                wb[sheetname].unmerge_cells(crs.coord)
                crs_list.remove(crs)
            # remove all data validation
            row_idx = wb[sheetname].max_row
            for col in range(1, wb[sheetname].max_column):
                removeExistingCellDataValidation(wb[sheetname], wb[sheetname].cell(row_idx, col))
            # remove first line
            delete_row_with_merge(wb[sheetname], 1, 1)

            # insert topheader row
            wb[sheetname].insert_rows(1, 1)

            # remove depth center (C) if empty
            col_n = 1
            col_delete_idx = openpyxl.utils.column_index_from_string('C')
            delete_col_with_merge(wb[sheetname], col_delete_idx, col_n, True, 5)

            # -
            topheader_list = ['', 'tar', 'sample', 'sample', 'ice', 'brine']
            header_list = ['ID', 'mass', 'mass with tar', 'mass', 'mass', 'mass']
            subheader_list = ['value', 'value', 'value', 'value', 'value', 'value']
            unit_list = ['-', 'kg', 'kg', 'kg', 'kg', 'kg']
            for ii, col_idx in enumerate(range(3, 9)):
                wb[sheetname].cell(1, col_idx).value = topheader_list[ii]
                wb[sheetname].cell(2, col_idx).value = header_list[ii]
                wb[sheetname].cell(3, col_idx).value = subheader_list[ii]
                wb[sheetname].cell(4, col_idx).value = unit_list[ii]

            # - ice: add column after salinity_value
            wb[sheetname]['I1'].value = 'ice'
            wb[sheetname]['I2'].value = 'salinity'
            wb[sheetname]['I3'].value = 'ID'
            wb[sheetname]['J3'].value = 'value'
            wb[sheetname]['K3'].value = 'quality'
            wb[sheetname]['K4'].value = '[0-9]'
            wb[sheetname]['L2'].value = 'conductivity'
            wb[sheetname]['L3'].value = 'value'
            wb[sheetname]['M3'].value = 'sample temperature'
            wb[sheetname]['N2'].value = 'specific conductance'
            wb[sheetname]['N3'].value = 'value'
            wb[sheetname].merge_cells('I1:N1')
            wb[sheetname].merge_cells('I2:K2')
            wb[sheetname].merge_cells('L2:M2')

            # - brine
            wb[sheetname]['O1'].value = 'brine'
            wb[sheetname]['O2'].value = 'salinity'
            wb[sheetname]['O3'].value = 'ID'
            wb[sheetname]['P3'].value = 'value'
            wb[sheetname]['Q3'].value = 'quality'
            wb[sheetname]['Q4'].value = '[0-9]'
            wb[sheetname]['R2'].value = 'conductivity'
            wb[sheetname]['R3'].value = 'value'
            wb[sheetname]['S3'].value = 'sample temperature'
            wb[sheetname]['T2'].value = 'specific conductance'
            wb[sheetname]['T3'].value = 'value'
            wb[sheetname].merge_cells('O1:T1')
            wb[sheetname].merge_cells('O2:Q2')
            wb[sheetname].merge_cells('R2:S2')

            # - comment
            wb[sheetname]['U1'].value = 'comment'
            wb[sheetname]['U3'].value = 'comments'
            wb[sheetname]['U4'].value = '-'

            # formatting
            max_row = 3 + data_row_n + 1
            l_border_col = ['F', 'I', 'O', 'U']
            style_d = {1: p_header_style, 2: p_header_style, 3: p_subheader_style, 4: p_unit_style}
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col, style_d)

        # - LOCATIONS
        sheetname = 'locations'
        if sheetname in wb.sheetnames:
            list_sheet_index = wb.worksheets.index(wb[sheetname])
            del wb[sheetname]
        else:
            list_sheet_index = len(wb.worksheets)
        wb.create_sheet(sheetname, list_sheet_index)
        wb_source = openpyxl.load_workbook(
            os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYMMDD-P-1.4.1-lists.xlsx'),
            data_only=True)
        copy_cells(wb_source[sheetname], wb[sheetname])  # copy all the cel values and styles
        copy_sheet_attributes(wb_source[sheetname], wb[sheetname])

        wb['metadata-station']['C1'].value = '1.4.1'
        wb.active = wb['metadata-station']

        # clean all worksheet
        for sheetname in wb.sheetnames:
            clean_worksheet(wb[sheetname])
        logger.debug('\t %s updated from 1.3.0 to 1.4.1' % wb['metadata-core']['C1'].value)
    else:
        logger.info("\t%s: already update to version %s "
                    % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    wb = fix_merged_cells(wb)
    return wb


def version_1_4_1_to_1_4_2(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 2:
        # TEX
        sheetname = 'tex'
        # look for column header
        header = 'stratigraphy'
        col_idx = find_str_in_row(wb[sheetname], header)[0]
        cr = find_merged_cell(wb[sheetname], 1, col_idx)
        if cr:
            wb[sheetname].unmerge_cells(cr.coord)
        new_cr = openpyxl.utils.get_column_letter(cr.left[0][1]) + str(cr.left[0][0]) + ':' + \
                 openpyxl.utils.get_column_letter(cr.right[0][1] - 1) + str(cr.right[0][0])
        wb[sheetname].merge_cells(new_cr)
        wb[sheetname].cell(cr.right[0][0], cr.right[0][1]).value = 'Core'

        # formatting
        wb[sheetname].cell(1, cr.right[0][1]).style = 'p_header_style'
        for row in range(1, wb[sheetname].max_row):
            wb[sheetname].cell(row, cr.right[0][1]).border = l_border

        # update version number
        wb['metadata-station']['C1'].value = '1.4.2'
        wb.active = wb['metadata-station']
        logger.debug('\t %s updated from 1.4.1 to %s'
                     % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    else:
        logger.info("\t%s: already update to version %s "
                    % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    wb = fix_merged_cells(wb)
    return wb


def version_1_4_2_to_1_4_3(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 3:
        # - METADATA-STATION
        sheetname = 'metadata-station'
        # add water temperature in row 33
        insert_row_with_merge(wb[sheetname], 33, 1)
        wb['metadata-station']['A33'].value = 'water salinity'
        wb['metadata-station']['B33'].value = 'PSU'
        wb['metadata-station']['D33'].value = 'under ice'
        wb['metadata-station']['D32'].value = 'under ice'

        # update "wind orientation" to "'"wind direction"
        wb['metadata-station']['A43'].value = 'wind direction'

        # formatting
        wb['metadata-station']['A33'].style = 'm_header_l_style'
        wb['metadata-station']['B33'].style = 'm_unit_r_style'

        wb['metadata-station']['D33'].style = 'm_comment_style'
        wb['metadata-station']['E33'].style = 'm_comment_style'

        for col in range(openpyxl.utils.column_index_from_string('A'),
                         openpyxl.utils.column_index_from_string('E') + 1):
            wb['metadata-station'].cell(32, col).border = b_border
            wb['metadata-station'].cell(34, col).style = 'm_bkg_style'  # repaint unmerged cell
            wb['metadata-station'].cell(40, col).style = 'm_bkg_style'  # repaint unmerged cell

        for col in range(openpyxl.utils.column_index_from_string('F'),
                         openpyxl.utils.column_index_from_string('Z') + 1):
            wb['metadata-station'].cell(33, col).style = 'm_bkg_style'

        # - SACKHOLE: add sackhole property tab after snow
        sackhole_sheet_index = wb.worksheets.index(wb['snow']) + 1
        wb.create_sheet("sackhole", sackhole_sheet_index)
        wb_source = openpyxl.load_workbook(
            os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYMMDD-P-1.4.3-sackhole_tab.xlsx'),
            data_only=True)
        copy_cells(wb_source['sackhole'], wb['sackhole'])  # copy all the cel values and styles
        copy_sheet_attributes(wb_source['sackhole'], wb['sackhole'])

        # update version number
        wb['metadata-station']['C1'].value = '1.4.3'
        wb.active = wb['metadata-station']
        logger.debug('\t %s updated from 1.4.2 to %s'
                     % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
        # clean all worksheet
        for sheetname in wb.sheetnames:
            clean_worksheet(wb[sheetname])
    else:
        logger.info("\t%s: already update to version %s "
                    % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    wb = fix_merged_cells(wb)
    return wb


def version_1_4_3_to_1_4_4(wb):
    global cell
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 4:
        # - TM
        TM_sheet_idx = wb.worksheets.index(wb['snow'])
        wb.create_sheet("TM", TM_sheet_idx)
        wb_source = openpyxl.load_workbook(
            os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYXXZZ-N_P-1.4.4-TM_sheet.xlsx'),
            data_only=True)
        copy_cells(wb_source['TM'], wb['TM'])  # copy all the cel values and styles
        copy_sheet_attributes(wb_source['TM'], wb['TM'])

        # DENSITY-VOLUME
        sheetname = 'density-volume'
        flag_m = {'diameter': 0}
        header_after = 'mass'
        for cell in openpyxl.utils.cols_from_range(
                'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
            if wb[sheetname][cell[0]].value is not None:
                for key in flag_m.keys():
                    if key in wb[sheetname][cell[0]].value:
                        flag_m[key] = 1

        for key in flag_m.keys():
            if flag_m[key] == 0:
                col_number = 3
                for cell in openpyxl.utils.cols_from_range(
                        'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
                    if wb[sheetname][cell[0]].value == header_after:
                        break

                col_start = openpyxl.utils.coordinate_to_tuple(cell[0])[1]
                for cr in wb[sheetname].merged_cells.ranges:
                    if cr.left[0][1] >= col_start:
                        cr.shift(col_number, 0)
                wb[sheetname].insert_cols(col_start, col_number)
                new_merge = openpyxl.utils.get_column_letter(col_start) + '1:' + openpyxl.utils.get_column_letter(
                    col_start + col_number - 1) + '1'
                wb[sheetname].merge_cells(new_merge)

                #  Add diameter measurement in density-volume sheet:
                wb[sheetname].cell(1, col_start).value = 'diameter'
                wb[sheetname].cell(2, col_start).value = 'value1'
                wb[sheetname].cell(2, col_start + 1).value = 'value2'
                wb[sheetname].cell(2, col_start + 2).value = 'average'
                wb[sheetname].cell(3, col_start).value = 'm'
                wb[sheetname].cell(3, col_start + 1).value = 'm'
                wb[sheetname].cell(3, col_start + 2).value = 'm'

                # Formatting
                col_end = col_start + 2
                style_header_painter(wb[sheetname], col_start, col_end)
                style_bottom_painter(wb[sheetname], col_start, col_end)
                for col in [col_start, col_start + 3]:
                    for row in range(1, wb[sheetname].max_row):
                        wb[sheetname].cell(row, col).border = l_border

        # SACKHOLE
        sheetname = 'sackhole'
        header_after = 'nutrient'

        # add d18O and dD in front of nutrient
        flag_m = {'d18O': 0, 'dD': 0}
        for cell in openpyxl.utils.cols_from_range(
                'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
            if wb[sheetname][cell[0]].value is not None:
                for key in flag_m.keys():
                    if key in wb[sheetname][cell[0]].value:
                        flag_m[key] = 1

        for key in flag_m.keys():
            if flag_m[key] == 0:
                for cell in openpyxl.utils.cols_from_range(
                        'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
                    if wb[sheetname][cell[0]].value == header_after:
                        break

                col_start = openpyxl.utils.coordinate_to_tuple(cell[0])[1]
                for cr in wb[sheetname].merged_cells.ranges:
                    if cr.left[0][1] >= col_start:
                        cr.shift(3, 0)
                wb[sheetname].insert_cols(col_start, 3)
                new_merge = openpyxl.utils.get_column_letter(col_start) + '1:' + openpyxl.utils.get_column_letter(
                    col_start + 2) + '1'
                wb[sheetname].merge_cells(new_merge)
                wb[sheetname].cell(1, col_start).value = key
                wb[sheetname].cell(2, col_start).value = 'ID'
                wb[sheetname].cell(2, col_start + 1).value = 'value'
                wb[sheetname].cell(2, col_start + 2).value = 'quality'
                wb[sheetname].cell(3, col_start).value = '-'
                wb[sheetname].cell(3, col_start + 1).value = '-'
                wb[sheetname].cell(3, col_start + 2).value = '[0-9]'

                # formatting
                col_end = col_start + 2
                style_header_painter(wb[sheetname], col_start, col_end)
                style_bottom_painter(wb[sheetname], col_start, col_end)
                for col in [col_start, col_start + 3]:
                    for row in range(1, wb[sheetname].max_row):
                        wb[sheetname].cell(row, col).border = l_border
            else:
                logging.error('%s NOT IMPLEMENTED YET' % key)

        # remove nutrient_volume column
        col_h_idx = find_str_in_row(wb[sheetname], 'nutrient')[0]
        col_sh_idx = find_str_in_row(wb[sheetname], 'volume', 2)
        col_sh_idx = col_sh_idx[np.where(col_sh_idx >= col_h_idx)][0]
        cr = find_merged_cell(wb[sheetname], 1, col_h_idx)
        if cr:
            try:
                wb[sheetname].unmerge_cells(cr.coord)
            except KeyError:
                # sometiems openpyxl failed to update merged cells, and it is necessary to save and reload the workbook
                wb.save('/tmp/test.xlsx')
                wb = openpyxl.load_workbook('/tmp/test.xlsx')
                wb[sheetname].unmerge_cells(cr.coord)
            else:
                pass
        delete_col_with_merge(wb[sheetname], col_sh_idx, 1, True, 4)
        new_cr = openpyxl.utils.get_column_letter(cr.left[0][1]) + str(cr.left[0][0]) + ':' \
                 + openpyxl.utils.get_column_letter(cr.right[0][1] - 1) + str(cr.right[0][0])
        wb[sheetname].merge_cells(new_cr)

        # - SNOW
        sheetname = 'snow'
        # if density is not merge on 3 column, merge it
        for cell in openpyxl.utils.cols_from_range(
                'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
            if wb[sheetname][cell[0]].value == 'density':
                col_idx = openpyxl.utils.coordinate_to_tuple(cell[0])[1]
                if wb[sheetname].cell(2, col_idx + 2).value == 'quality':
                    if cell[0] in wb[sheetname].merged_cells:
                        # lookup the cell range:
                        for cr in wb[sheetname].merged_cells:
                            if cell[0] in cr:
                                break
                        if cr.right[0][1] < col_idx + 2:
                            wb[sheetname].unmerge_cells(cr.coord)
                            new_range = openpyxl.utils.get_column_letter(
                                col_idx) + '1:' + openpyxl.utils.get_column_letter(
                                col_idx + 2) + '1'
                            wb[sheetname].merge_cells(new_range)
                            # formatting
                            wb[sheetname].cell(1, col_idx).style = 'm_header_style'
                            wb[sheetname].cell(1, col_idx).border = noBorder
                # formatting
                max_row = wb[sheetname].max_row
                for row_idx in range(1, max_row):
                    wb[sheetname].cell(row_idx, col_idx).border = l_border
                    wb[sheetname].cell(row_idx, col_idx + 3).border = l_border
            elif wb[sheetname][cell[0]].value == 'eco sample':
                wb[sheetname][cell[0]].value = 'ECO data example'

        # - SEAWATER
        sheetname = 'seawater'
        if sheetname in wb.sheetnames:
            # look for temperature header
            col_temp_idx = find_str_in_row(wb[sheetname], 'temperature', 1)

            if len(col_temp_idx) == 0:
                col_temp_idx = find_str_in_row(wb[sheetname], 'd18O', 1)[0]
                insert_col_with_merge(wb[sheetname], col_temp_idx, 2)
                wb[sheetname].cell(1, col_temp_idx).value = 'temperature'
                wb[sheetname].cell(2, col_temp_idx).value = 'value'
                wb[sheetname].cell(3, col_temp_idx).value = '˚C'
                wb[sheetname].cell(2, col_temp_idx + 1).value = 'quality'
                wb[sheetname].cell(3, col_temp_idx + 1).value = '[0-9]'
            else:
                col_temp_idx = col_temp_idx[0]
                if not is_merged(wb[sheetname], 1, col_temp_idx):
                    # insert a column
                    insert_col_with_merge(wb[sheetname], col_temp_idx + 1, 1)
                    wb[sheetname].cell(2, col_temp_idx + 1).value = 'quality'
                    wb[sheetname].cell(3, col_temp_idx + 1).value = '[0-9]'
            merge_range = openpyxl.utils.get_column_letter(col_temp_idx) + str(1) + ':' + \
                          openpyxl.utils.get_column_letter(col_temp_idx + 1) + str(1)
            wb[sheetname].merge_cells(merge_range)

            # formatting
            col_comment_idx = find_str_in_row(wb[sheetname], 'comment', 1)[0]
            max_row = wb[sheetname].max_row
            l_border_col = [find_str_in_row(wb[sheetname], 'salinity', 1)[0],
                            find_str_in_row(wb[sheetname], 'd18O', 1)[0],
                            find_str_in_row(wb[sheetname], 'dD', 1)[0],
                            find_str_in_row(wb[sheetname], 'temperature', 1)[0],
                            find_str_in_row(wb[sheetname], 'comment', 1)[0]]
            l_border_col = [openpyxl.utils.get_column_letter(col_idx) for col_idx in l_border_col]
            worksheetDataFormatting(wb[sheetname], max_row, l_border_col)

        # update version number
        wb['metadata-station']['C1'].value = '1.4.4'
        wb.active = wb['metadata-station']
        # clean all worksheet
        for sheetname in wb.sheetnames:
            clean_worksheet(wb[sheetname])
        logger.debug(
            '\t %s updated from 1.4.3 to %s' % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    else:
        logger.info("\t%s: already update to version %s " % (
            wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    wb = fix_merged_cells(wb)
    return wb


def version_1_4_4_to_1_4_5(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 5:
        # TODO :remove
        for style in m_styles_list:
            if style not in wb.style_names:
                wb.add_named_style(eval(style))

        # - METADATA-STATION
        sheetname = 'metadata-station'
        wb[sheetname]['D42'].value = None
        wb[sheetname]['D43'].value = None
        wb[sheetname]['D44'].value = None
        wb[sheetname]['B42'].value = 'm/s'
        wb[sheetname]['B43'].value = 'degree'
        wb[sheetname]['B44'].value = '/8'
        wb[sheetname]['B45'].value = 'μmol/m2s'
        wb[sheetname]['B46'].value = 'μmol/m2s'

        # formatting
        # style B42:B46 not bold, align right
        col_idx = openpyxl.utils.cell.column_index_from_string('B')
        for row_idx in range(42, 46 + 1):
            wb[sheetname].cell(row_idx, col_idx).style = 'm_unit_r_style'
        # Border formatting
        # C-H
        for col_idx in range(openpyxl.utils.column_index_from_string('C'),
                             openpyxl.utils.column_index_from_string('H') + 1):
            wb['metadata-station'].cell(9, col_idx).border = b_border
            wb['metadata-station'].cell(11, col_idx).border = noBorder
            wb['metadata-station'].cell(40, col_idx).border = noBorder
        # A-C
        for col_idx in range(openpyxl.utils.column_index_from_string('A'),
                             openpyxl.utils.column_index_from_string('D') + 1):
            wb['metadata-station'].cell(42, col_idx).border = b_border
            wb['metadata-station'].cell(43, col_idx).border = b_border
            wb['metadata-station'].cell(44, col_idx).border = b_border
            wb['metadata-station'].cell(45, col_idx).border = b_border

        # C14:C16, bottom border style: hair
        for _row in openpyxl.utils.cell.rows_from_range('C14:C16'):
            for _cell in _row:
                wb['metadata-station'][_cell].border = b_border

        # C16:D16, bottom border style: None
        for _row in openpyxl.utils.cell.rows_from_range('C16:D16'):
            for _cell in _row:
                wb['metadata-station'][_cell].border = noBorder

        # add specific vertical border
        wb['metadata-station']['F9'].border = bl_border
        wb['metadata-station']['D14'].border = bl_border
        wb['metadata-station']['D16'].border = l_border

        # update version number
        wb['metadata-station']['C1'].value = '1.4.5'
        wb.active = wb['metadata-station']
        logger.debug('\t%s: updated from 1.4.4 to %s'
                     % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
        return wb
    else:
        logger.info('\t%s: already update to version %s'
                    % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))


def version_1_4_5_to_1_4_6(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 6:
        # - METADATA-CORE
        # Clean metadata-core
        sheetname = 'metadata-core'
        pysic.io.clean.metadata_core(wb)
        # Formatting
        wb[sheetname]['D1'].style = 'm_bkg_style'
        wb[sheetname]['D2'] = None
        wb[sheetname]['D2'].style = 'm_bkg_style'

        # - METADATA-STATION
        sheetname = 'metadata-station'
        wb[sheetname]['E45'].value = '2pi sensor'
        wb[sheetname]['E46'].value = '4pi sensor'

        # - SALO18
        sheetname = 'salo18'
        # -- Reorganize d18O, dD colums and add d_excess, if d18o_id match dd_id
        crs = list(wb[sheetname].merged_cells)
        for merge_cell in crs:
            if merge_cell.left[0][0] == 1:
                col_start = merge_cell.left[0][1]
                col_end = merge_cell.right[0][1]
                header = merge_cell.start_cell.value
                # find header
                if header in ['d18O', 'dD']:
                    wb[sheetname].unmerge_cells(merge_cell.coord)
                    for col_idx in range(col_start, col_end + 1):
                        if wb[sheetname].cell(2, col_idx).value == 'value':
                            # remember ID for d18O and dD
                            if header == 'd18O':
                                d18O_value_idx = col_idx
                            elif header == 'dD':
                                dd_value_idx = col_idx
                        elif wb[sheetname].cell(2, col_idx).value == 'ID':
                            # remember ID for d18O and dD
                            if header == 'd18O':
                                d18o_id_idx = col_idx
                            elif header == 'dD':
                                dd_id_idx = col_idx
                        elif wb[sheetname].cell(2, col_idx).value == 'quality':
                            # remember ID for d18O and dD
                            if header == 'd18O':
                                d18o_qual_idx = col_idx
                            elif header == 'dD':
                                dd_qual_idx = col_idx
        # -- Reorganize d18O, dD columns and add d_excess, if d18o_id match dd_id
        max_row = wb[sheetname].max_row
        d18o_id = [wb[sheetname].cell(row=ii, column=d18o_id_idx).value for ii in range(4, max_row)]
        dd_id = [wb[sheetname].cell(row=ii, column=dd_id_idx).value for ii in range(4, max_row)]
        d18o_qual = [wb[sheetname].cell(row=ii, column=d18o_qual_idx).value for ii in range(4, max_row)]
        dd_qual = [wb[sheetname].cell(row=ii, column=dd_qual_idx).value for ii in range(4, max_row)]

        if dd_id == d18o_id:
            # Insert d_excess column after dD:
            wb[sheetname].insert_cols(dd_value_idx + 1, 1)
            wb[sheetname].cell(1, d18o_id_idx).value = None
            wb[sheetname].cell(1, d18O_value_idx).value = 'd18O'
            wb[sheetname].cell(1, dd_value_idx).value = 'dD'
            wb[sheetname].cell(1, dd_value_idx + 1).value = 'd_excess'
            wb[sheetname].cell(2, dd_value_idx + 1).value = 'value'
            wb[sheetname].cell(3, dd_value_idx + 1).value = '‰'
            # remove colum of dd_idy
            wb[sheetname].delete_cols(dd_id_idx, 1)

            if d18o_qual != dd_qual:
                logging.error(
                    '\t%s - %s NOT IMPLEMENTED: d18o_qual != dd_qual' % (wb['metadata-core']['C1'].value, sheetname))
            else:
                # remove colum for d18o_qual
                wb[sheetname].delete_cols(d18o_qual_idx, 1)

            # formatting
            col_start = d18o_id_idx
            style_header_painter(wb[sheetname], col_start, col_end)
            style_bottom_painter(wb[sheetname], col_start, col_end)
            if header in sheetname in prop_associated_tab and header in prop_associated_tab[sheetname]:
                pass
            else:
                comment_col_idx = find_str_in_row(wb[sheetname], 'comment', 1)[0]
                for col in [d18o_id_idx, comment_col_idx]:
                    for row in range(1, wb[sheetname].max_row):
                        wb[sheetname].cell(row, col).border = l_border
        else:
            logger.info('\t%s - %s d18O ID different from dD ID: merging not possible' % (
                wb['metadata-core']['C1'].value, sheetname))

        # ECO
        sheetname = 'eco'
        max_row = wb[sheetname].max_row

        # store merged cell ranges for d18O and dD
        stored_headers = ['chl-a', 'phaeo']
        property_range = {}
        crs = wb[sheetname].merged_cells.ranges
        for merge_cell in crs:
            if merge_cell.start_cell.value in stored_headers:
                for prop in stored_headers:
                    if merge_cell.start_cell.value == prop:
                        property_range[prop] = merge_cell

        # Unmerge header cells for nutrient
        unmerge_header_row(wb[sheetname])

        # update necessary for MOSAiC-UTQ dataset
        # correct typo 'eco. data sampe' with 'eco. data sample'
        for _col in range(1, wb[sheetname].max_column):
            if wb[sheetname].cell(1, _col).value == 'eco. data sampe':
                wb[sheetname].cell(1, _col).value == 'eco. data sample'

        # add phaeo column, and/or merge chl-a and phaeo ID
        if 'chl-a' in property_range and 'phaeo' in property_range:

            # found column id for ID and qual for chl-a and phaeo
            sh_idx = {}
            for prop in property_range:
                if prop not in sh_idx:
                    sh_idx[prop] = {}
                merge_cell = property_range[prop]
                col_start = merge_cell.left[0][1]
                col_end = merge_cell.right[0][1]
                for col_idx in range(col_start, col_end + 1):
                    subheader = wb[sheetname].cell(2, col_idx).value
                    sh_idx[prop][subheader] = col_idx

            # Check if IDs match:
            chla_id = [wb[sheetname].cell(row=ii, column=sh_idx['chl-a']['ID']).value for ii in range(4, max_row)]
            phaeo_id = [wb[sheetname].cell(row=ii, column=sh_idx['phaeo']['ID']).value for ii in range(4, max_row)]
            chla_qual = [wb[sheetname].cell(row=ii, column=sh_idx['chl-a']['quality']).value for ii in
                         range(4, max_row)]
            phaeo_qual = [wb[sheetname].cell(row=ii, column=sh_idx['phaeo']['quality']).value for ii in
                          range(4, max_row)]
            chla_vol = [wb[sheetname].cell(row=ii, column=sh_idx['chl-a']['volume']).value for ii in range(4, max_row)]
            phaeo_vol = [wb[sheetname].cell(row=ii, column=sh_idx['phaeo']['volume']).value for ii in range(4, max_row)]

            if chla_id == phaeo_id and chla_vol == phaeo_vol:
                # remove colum of dd_idy
                wb[sheetname].cell(1, sh_idx['chl-a']['ID']).value = None
                wb[sheetname].cell(1, sh_idx['chl-a']['value']).value = 'chl-a'
                wb[sheetname].cell(1, sh_idx['phaeo']['value']).value = 'phaeo'
                delete_col_with_merge(wb[sheetname], sh_idx['phaeo']['volume'], 1, True, 4)
                delete_col_with_merge(wb[sheetname], sh_idx['phaeo']['ID'], 1, True, 4)
                if chla_qual != phaeo_qual:
                    logging.error('NOT IMPLEMENTED: d18o_qual != dd_qual')
                    col_end = sh_idx['phaeo']['quality'] - 2
                else:
                    # remove colum for d18o_qual
                    delete_col_with_merge(wb[sheetname], sh_idx['chl-a']['quality'], 1, True, 4)
                    col_end = sh_idx['phaeo']['quality'] - 3

                # rename 'volume' subheader to 'filtered volume':
                col_start = sh_idx['chl-a']['ID']
                col_idx = find_str_in_row(wb[sheetname], 'volume', 2)
                col_idx = col_idx[(col_start <= col_idx) & (col_idx <= col_end)]
                wb[sheetname].cell(2, col_idx[0]).value = 'filtered volume'
                wb[sheetname].cell(3, col).value = 'L'

            else:
                logger.info(
                    '%s\t\tdifferent IDs, quality or volume: merging not possible' % wb['metadata-core']['C1'].value)
                # rename 'volume' to 'filtered volume' if needed
                logger.error('%s\t\tTODO: rename volume to filtered volume or filtered volume column'
                             % wb['metadata-core']['C1'].value)
                # insert 'filtered volume if needed

        elif 'chl-a' in property_range:
            # look for chl-a value column
            for col in range(4, wb[sheetname].max_column):
                if wb[sheetname].cell(1, col).value == 'chl-a':
                    break
            # insert phaeo value column after chl-a value column
            wb[sheetname].insert_cols(col + 1, 1)
            wb[sheetname].cell(1, col + 1).value = 'phaeo'
            wb[sheetname].cell(2, col + 1).value = 'value'
            wb[sheetname].cell(3, col + 1).value = 'mg/m3'

            # check if a volume column exist for 'chl-a'
            col_idx = find_str_in_row(wb[sheetname], 'volume', 2)
            col_idx = col_idx[(col - 1 <= col_idx) & (col_idx <= col + 2)]
            if len(col_idx) == 1:
                wb[sheetname].cell(2, col_idx[0]).value = 'filtered volume'
                wb[sheetname].cell(3, col).value = 'L'
            elif len(col_idx) == 0:
                # insert 'filtered volume' column before chl-a value column
                wb[sheetname].insert_cols(col, 1)
                wb[sheetname].cell(2, col).value = 'filtered volume'
                wb[sheetname].cell(3, col).value = 'L'
            else:
                logger.error("\t\t: more than one column with 'volume' in subheader")

            # formatting
            col_start = col - 2
            col_end = col + 2
        style_header_painter(wb[sheetname], col_start, col_end)
        style_bottom_painter(wb[sheetname], col_start, col_end)
        for col in [col_start, col_end + 1]:
            for row in range(1, wb[sheetname].max_row):
                wb[sheetname].cell(row, col).border = l_border

        # SNOW
        sheetname = 'snow'
        crs = list(wb[sheetname].merged_cells)
        # Store property range for d18O and dD
        stored_headers = ['d18O', 'dD']
        property_range = {}
        for merge_cell in crs:
            if merge_cell.start_cell.value in stored_headers:
                for prop in stored_headers:
                    if merge_cell.start_cell.value == prop:
                        property_range[prop] = merge_cell

        for cr in crs:
            if cr.start_cell.value in ['eco sample', 'eco. data example', 'ECO data example']:
                cr.start_cell.value = 'ECO data example'

                # reorganize dD and d18O
                max_row = wb[sheetname].max_row

        # found column id for ID and qual of d18O and dD
        sh_idx = {}
        for prop in property_range:
            if prop not in sh_idx:
                sh_idx[prop] = {}
            merge_cell = property_range[prop]
            col_start = merge_cell.left[0][1]
            col_end = merge_cell.right[0][1]
            for col_idx in range(col_start, col_end + 1):
                subheader = wb[sheetname].cell(2, col_idx).value
                sh_idx[prop][subheader] = col_idx
            wb[sheetname].unmerge_cells(merge_cell.coord)

        if 'd18O' in sh_idx and 'dD' in sh_idx:
            # Check if d18o_id match dd_id:
            d18o_id = [wb[sheetname].cell(row=ii, column=sh_idx['d18O']['ID']).value for ii in range(4, max_row)]
            dd_id = [wb[sheetname].cell(row=ii, column=sh_idx['dD']['ID']).value for ii in range(4, max_row)]
            d18o_qual = [wb[sheetname].cell(row=ii, column=sh_idx['d18O']['quality']).value for ii in range(4, max_row)]
            dd_qual = [wb[sheetname].cell(row=ii, column=sh_idx['dD']['quality']).value for ii in range(4, max_row)]

            if dd_id == d18o_id:
                # remove colum of dd_idy
                wb[sheetname].cell(1, sh_idx['d18O']['ID']).value = None
                wb[sheetname].cell(1, sh_idx['d18O']['value']).value = 'd18O'
                wb[sheetname].cell(1, sh_idx['dD']['value']).value = 'dD'
                delete_col_with_merge(wb[sheetname], sh_idx['dD']['ID'], 1, True, 4)
                if d18o_qual != dd_qual:
                    logging.error('NOT IMPLEMENTED: d18o_qual != dd_qual')
                else:
                    # remove colum for d18o_qual
                    delete_col_with_merge(wb[sheetname], sh_idx['d18O']['quality'], 1, True, 4)
            else:
                logger.info(
                    '%s\t\td18O ID different from dD ID: merging not possible' % wb['metadata-core']['C1'].value)
            # formatting
            style_header_painter(wb[sheetname], sh_idx['d18O']['ID'], sh_idx['d18O']['ID'] + 4)
            style_bottom_painter(wb[sheetname], sh_idx['d18O']['ID'], sh_idx['d18O']['ID'] + 4)
            for row in range(1, max_row):
                wb[sheetname].cell(row, sh_idx['d18O']['ID']).border = l_border

        # update version number
        wb['metadata-station']['C1'].value = '1.4.6'
        wb.active = wb['metadata-station']
        logger.debug('\t %s updated from 1.4.5 to 1.4.6' % (wb['metadata-core']['C1'].value))
    else:
        logger.info("\t%s: already update to version %s " % (
            wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))

    wb = fix_merged_cells(wb)
    return wb


def version_1_4_6_to_1_4_7(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 7:
        # - SALO18
        sheetname = 'salo18'
        # -- Unmerge header cells
        if sheetname in wb.sheetnames:
            unmerge_header_row(wb[sheetname])

        # - TEMP
        sheetname = 'temp'
        # -- Unmerge header cell
        crs = list(wb[sheetname].merged_cells)
        if sheetname in wb.sheetnames:
            unmerge_header_row(wb[sheetname])

        # - TEX:
        sheetname = 'tex'
        ## --unmerge header
        if sheetname in wb.sheetnames:
            unmerge_header_row(wb[sheetname])

        # - TM:
        sheetname = 'TM'
        if sheetname in wb.sheetnames:
            # move TM before eco
            sheet_to_move_idx = wb.worksheets.index(wb[sheetname])
            before_sheet__idx = wb.worksheets.index(wb['eco'])
            wb.active = sheet_to_move_idx
            wb.move_sheet(wb.active, offset=-(sheet_to_move_idx - before_sheet__idx))

        # -ECO:
        sheetname = 'eco'
        if sheetname in wb.sheetnames:
            # Unmerge header
            unmerge_header_row(wb[sheetname], ['melted volume'])

            # Implement incorrect merge cell
            chla_value_idx = find_str_in_row(wb[sheetname], 'chl-a', 1)
            if len(chla_value_idx) > 0:
                new_range = openpyxl.utils.get_column_letter(chla_value_idx[0] - 2) + '1:' + \
                            openpyxl.utils.get_column_letter(chla_value_idx[0] - 1) + '1'
                wb[sheetname].merge_cells(new_range)

        # - ECO-POOL
        sheetname = 'eco-pool'
        # Insert 'eco-pool' tab
        eco_sheet_index = wb.worksheets.index(wb['eco'])
        wb.create_sheet(sheetname, eco_sheet_index + 1)
        wb_source = openpyxl.load_workbook(
            os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYXXZZ-N_P-1.4.7-eco-pool_lists.xlsx'),
            data_only=True)
        copy_cells(wb_source[sheetname], wb[sheetname])  # copy all the cel values and styles
        copy_sheet_attributes(wb_source[sheetname], wb[sheetname])

        # - SNOW
        sheetname = 'snow'
        if sheetname in wb.sheetnames:
            # move TM before eco
            sheet_to_move_idx = wb.worksheets.index(wb[sheetname])
            if 'eco-pool' in wb.sheetnames:
                after_sheet__idx = wb.worksheets.index(wb['eco-pool'])
            else:
                after_sheet__idx = wb.worksheets.index(wb['eco'])
            wb.active = sheet_to_move_idx
            wb.move_sheet(wb.active, offset=after_sheet__idx - sheet_to_move_idx + 1)

            # Rename 'temperature' subheader to 'value' for 'temperature' header
            col_idx = find_str_in_row(wb[sheetname], 'temperature', 1)[0]
            if is_merged(wb[sheetname], 1, col_idx):
                cr = find_merged_cell(wb[sheetname], 1, col_idx)
                col_start = cr.left[0][1]
                col_end = cr.right[0][1]
                col_value_idx = find_str_in_row(wb[sheetname], 'temperature', 2)
                if len(col_value_idx) > 0:
                    col_value_idx = col_value_idx[(col_start <= col_value_idx) & (col_value_idx <= col_end)][0]
                    wb[sheetname].cell(2, col_value_idx).value = 'value'
                else:
                    logger.error(
                        "%s unknown column trying to  change 'temperature' subheader to 'value' for 'temperature' header" % (
                            wb['metadata-core']['C1'].value, sheetname))
            else:
                logger.error(
                    "%s  change 'temperature' subheader to 'value' for 'temperature' header not implemented for unmerged header" % (
                        wb['metadata-core']['C1'].value, sheetname))
            ## --unmerge header
            try:
                unmerge_header_row(wb[sheetname])
            except KeyError:
                wb.save('/tmp/tmp.xlsx')
                wb = openpyxl.load_workbook('/tmp/tmp.xlsx')
                unmerge_header_row(wb[sheetname])
            # Rename 'ECO data sample' rename to 'eco property' or insert 'eco property' columns
            col_idx = find_str_in_row(wb[sheetname], 'ECO data example', 1)
            if len(col_idx) == 0:
                # Add eco property block after temperature and nutrient
                col_comment = find_str_in_row(wb[sheetname], 'comment', 1)[0]
                temp_flag = find_str_in_row(wb[sheetname], 'temperature', 1)
                nuts_flag = find_str_in_row(wb[sheetname], 'nutrient', 1)

                col_idx = col_comment + 2
                if len(temp_flag) > 0:
                    col_idx = col_idx + 4
                if len(nuts_flag) > 0:
                    col_idx = col_idx + 4

                # - eco sample: insert 3 column
                col_n = 4
                insert_col_with_merge(wb[sheetname], col_idx, col_n)
                wb[sheetname].cell(1, col_idx).value = 'ECO data example'
                wb[sheetname].cell(2, col_idx).value = 'ID'
                wb[sheetname].cell(2, col_idx + 1).value = 'value'
                wb[sheetname].cell(2, col_idx + 2).value = 'quality'
                wb[sheetname].cell(3, col_idx).value = '-'
                wb[sheetname].cell(3, col_idx + 1).value = '-'
                wb[sheetname].cell(3, col_idx + 2).value = '[0-9]'
                merge_range = openpyxl.utils.get_column_letter(col_idx) + '1:' + openpyxl.utils.get_column_letter(
                    col_idx + 2) + '1'
                wb[sheetname].merge_cells(merge_range)

                # formatting
                style_header_painter(wb[sheetname], col_idx, col_idx + 2)
                style_bottom_painter(wb[sheetname], col_idx, col_idx + 2)
                for row_idx in range(1, wb[sheetname].max_row + 1):
                    wb[sheetname].cell(row_idx, col_idx + 3).style = 'm_bkg_style'
            else:
                col_idx = find_str_in_row(wb[sheetname], 'ECO data example', 1)[0]
                wb[sheetname].cell(1, col_idx).value = 'eco property'

            ## remove leftborder for temperature, nutrient, eco property, Snow micro pen SMP, SWE
            col_idx = find_str_in_row(wb[sheetname], 'comment', 1)[0]
            max_row = wb[sheetname].max_row
            max_col = wb[sheetname].max_column
            for row_idx in range(1, max_row):
                for col_idx in range(col_idx + 1, max_col):
                    wb[sheetname].cell(row_idx, col_idx).border = noBorder

        # - SACKHOLE
        sheetname = 'sackhole'
        if sheetname in wb.sheetnames:
            # move TM before eco
            sheet_to_move_idx = wb.worksheets.index(wb[sheetname])
            after_sheet__idx = wb.worksheets.index(wb['snow'])
            wb.active = sheet_to_move_idx
            wb.move_sheet(wb.active, offset=after_sheet__idx - sheet_to_move_idx + 1)

            # store merged cell ranges for d18O and dD
            stored_headers = ['d18O', 'dD']
            property_range = {}
            crs = wb[sheetname].merged_cells.ranges
            for merge_cell in crs:
                if merge_cell.start_cell.value in stored_headers:
                    for prop in stored_headers:
                        if merge_cell.start_cell.value == prop:
                            property_range[prop] = merge_cell
            # --unmerge header
            unmerge_header_row(wb[sheetname])

            # clean worksheet to remove empty row at the end
            clean_worksheet(wb[sheetname])

            # -- reorganize d18O and dD, add d_excess
            max_row = wb[sheetname].max_row

            # found column id for ID and qual of d18O and dD
            sh_idx = {}
            for prop in property_range:
                if prop not in sh_idx:
                    sh_idx[prop] = {}
                merge_cell = property_range[prop]
                col_start = merge_cell.left[0][1]
                col_end = merge_cell.right[0][1]
                for col_idx in range(col_start, col_end + 1):
                    subheader = wb[sheetname].cell(2, col_idx).value
                    sh_idx[prop][subheader] = col_idx

            # Check if d18o_id match dd_id:
            d18o_id = [wb[sheetname].cell(row=ii, column=sh_idx['d18O']['ID']).value for ii in range(4, max_row)]
            dd_id = [wb[sheetname].cell(row=ii, column=sh_idx['dD']['ID']).value for ii in range(4, max_row)]
            d18o_qual = [wb[sheetname].cell(row=ii, column=sh_idx['d18O']['quality']).value for ii in range(4, max_row)]
            dd_qual = [wb[sheetname].cell(row=ii, column=sh_idx['dD']['quality']).value for ii in range(4, max_row)]

            if dd_id == d18o_id:
                # Insert d_excess column after dD:
                wb[sheetname].insert_cols(sh_idx['dD']['value'] + 1, 1)
                wb[sheetname].cell(1, sh_idx['dD']['value'] + 1).value = 'd_excess'
                wb[sheetname].cell(2, sh_idx['dD']['value'] + 1).value = 'value'
                wb[sheetname].cell(3, sh_idx['dD']['value'] + 1).value = '‰'
                wb[sheetname].cell(3, sh_idx['dD']['value']).value = '‰'
                wb[sheetname].cell(3, sh_idx['d18O']['value']).value = '‰'
                style_header_painter(wb[sheetname], sh_idx['dD']['value'] + 1, sh_idx['dD']['value'] + 1)
                style_bottom_painter(wb[sheetname], sh_idx['dD']['value'] + 1, sh_idx['dD']['value'] + 1)

                # remove colum of dd_idy
                wb[sheetname].delete_cols(sh_idx['dD']['ID'], 1)
                if d18o_qual != dd_qual:
                    logging.error('NOT IMPLEMENTED: d18o_qual != dd_qual')
                else:
                    # remove colum for d18o_qual
                    wb[sheetname].delete_cols(sh_idx['d18O']['quality'], 1)

            # Add eco property before TM
            flag_m = {'eco property': 0}
            header_after = ['TM', 'TM ligand', 'comment']
            col_n = 4
            for cell in openpyxl.utils.cols_from_range(
                    'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
                if wb[sheetname][cell[0]].value is not None:
                    for key in flag_m.keys():
                        if key in wb[sheetname][cell[0]].value:
                            flag_m[key] = 1
            for key in flag_m.keys():
                if flag_m[key] == 0:
                    for h in header_after:
                        col_idx = find_str_in_row(wb[sheetname], h, 1)
                        if len(col_idx) > 0:
                            col_idx = col_idx[0]
                            flag_no_header = False
                            break
                        elif len(col_idx) == 0:
                            flag_no_header = True
                    if flag_no_header:
                        logger.error("%s - %s: no header available" % (wb['metadata-core']['C1'].value, sheetname, h))
                    while wb[sheetname].cell(2, col_idx - 1).value not in ['quality'] and col_idx > 1:
                        col_idx = col_idx - 1
                    insert_col_with_merge(wb[sheetname], col_idx, col_n)

                    #  Add header, subheader and unit for 'eco property'
                    wb[sheetname].cell(1, col_idx + 2).value = 'eco property'
                    wb[sheetname].cell(2, col_idx).value = 'ID'
                    wb[sheetname].cell(2, col_idx + 1).value = 'volume'
                    wb[sheetname].cell(2, col_idx + 2).value = 'value'
                    wb[sheetname].cell(2, col_idx + 3).value = 'quality'
                    wb[sheetname].cell(3, col_idx).value = '-'
                    wb[sheetname].cell(3, col_idx + 1).value = 'L'
                    wb[sheetname].cell(3, col_idx + 2).value = '-'
                    wb[sheetname].cell(3, col_idx + 3).value = '[0-9]'

                    # Formatting
                    style_header_painter(wb[sheetname], col_idx, col_idx + col_n)
                    style_bottom_painter(wb[sheetname], col_idx, col_idx + col_n)
                    for col in [col_idx, col_idx + col_n]:
                        for row in range(1, wb[sheetname].max_row):
                            wb[sheetname].cell(row, col).border = l_border

        # - SEAWATER
        sheetname = 'seawater'
        if sheetname in wb.sheetnames:
            # correct position of sheetname
            sheet_to_move_idx = wb.worksheets.index(wb[sheetname])
            after_sheet__idx = wb.worksheets.index(wb['sackhole'])
            wb.active = sheet_to_move_idx
            wb.move_sheet(wb.active, offset=after_sheet__idx - sheet_to_move_idx + 1)

            # TODO: fix seawater salinity column errors
            # Add eco property before TM, TM ligand, or comment
            flag_m = {'eco property': 0}
            header_after = ['TM', 'TM ligand', 'comment']
            col_n = 4
            # check if property is already present in header
            for cell in openpyxl.utils.cols_from_range(
                    'A1:' + openpyxl.utils.get_column_letter(wb[sheetname].max_column) + str(1)):
                if wb[sheetname][cell[0]].value is not None:
                    for key in flag_m.keys():
                        if key in wb[sheetname][cell[0]].value:
                            flag_m[key] = 1
            for key in flag_m.keys():
                if flag_m[key] == 0:
                    for h in header_after:
                        col_idx = find_str_in_row(wb[sheetname], h, 1)
                        if len(col_idx) > 0:
                            col_idx = col_idx[0]
                            flag_no_header = False
                            break
                        elif len(col_idx) == 0:
                            flag_no_header = True
                    if flag_no_header:
                        logger.error("%s - %s: no header available" % (wb['metadata-core']['C1'].value, sheetname, h))

                    while wb[sheetname].cell(2, col_idx - 1).value not in ['quality'] and col_idx > 1:
                        col_idx = col_idx - 1
                    insert_col_with_merge(wb[sheetname], col_idx, col_n)

                    #  Add header, subheader and unit for 'eco property'
                    wb[sheetname].cell(1, col_idx + 2).value = 'eco property'
                    wb[sheetname].cell(2, col_idx).value = 'ID'
                    wb[sheetname].cell(2, col_idx + 1).value = 'volume'
                    wb[sheetname].cell(2, col_idx + 2).value = 'value'
                    wb[sheetname].cell(2, col_idx + 3).value = 'quality'
                    wb[sheetname].cell(3, col_idx).value = '-'
                    wb[sheetname].cell(3, col_idx + 1).value = 'L'
                    wb[sheetname].cell(3, col_idx + 2).value = '-'
                    wb[sheetname].cell(3, col_idx + 3).value = '[0-9]'

                    # Formatting
                    style_header_painter(wb[sheetname], col_idx, col_idx + col_n)
                    style_bottom_painter(wb[sheetname], col_idx, col_idx + col_n)
                    for col in [col_idx, col_idx + col_n]:
                        for row in range(1, wb[sheetname].max_row):
                            wb[sheetname].cell(row, col).border = l_border

        # - DENSITY-DENSIMETRY
        sheetname = 'density-densimetry'
        if sheetname in wb.sheetnames:
            # -- move sheet to correct location
            sheet_to_move_idx = wb.worksheets.index(wb[sheetname])
            before_sheet__idx = wb.worksheets.index(wb['density-volume'])
            wb.active = sheet_to_move_idx
            wb.move_sheet(wb.active, offset=before_sheet__idx - sheet_to_move_idx)

            # --unmerge header
            unmerge_header_row(wb[sheetname])

        # - DENSITY-VOLUME
        sheetname = 'density-volume'
        if sheetname in wb.sheetnames:
            crs = wb[sheetname].merged_cells.ranges.copy()

            # -- rename subheader for thickness
            for merge_cell in crs:
                if merge_cell.start_cell.value == 'thickness':
                    col_idx = merge_cell.left[0][1]
                    wb[sheetname].cell(2, col_idx).value = 'measurement1'
                    wb[sheetname].cell(2, col_idx + 1).value = 'measurement2'
                    wb[sheetname].cell(2, col_idx + 2).value = 'measurement3'
                    wb[sheetname].cell(2, col_idx + 3).value = 'measurement4'
                    wb[sheetname].cell(2, col_idx + 4).value = 'value'
                if merge_cell.start_cell.value == 'diameter':
                    col_idx = merge_cell.left[0][1]
                    wb[sheetname].cell(2, col_idx).value = 'measurement1'
                    wb[sheetname].cell(2, col_idx + 1).value = 'measurement2'
                    wb[sheetname].cell(2, col_idx + 2).value = 'value'
                    for row_idx in range(1, wb[sheetname].max_row):
                        wb[sheetname].cell(row_idx, col_idx + 2).border = noBorder
            # --unmerge header
            unmerge_header_row(wb[sheetname])

        # - SEDIMENT
        sheetname = 'sediment'
        if sheetname in wb.sheetnames:
            crs = list(wb[sheetname].merged_cells)
            # Store property range for later
            property_range = {}
            stored_header = ['sediment mass']
            for merge_cell in crs:
                if merge_cell.start_cell.value in stored_header:
                    for prop in stored_header:
                        if merge_cell.start_cell.value == prop:
                            property_range[prop] = merge_cell
            # --unmerge header
            unmerge_header_row(wb[sheetname])
            # -- merge stored header
            for header in stored_header:
                wb[sheetname].merge_cells(property_range[header].coord)
                start_col = property_range[header].left[0][1]
                end_col = property_range[header].right[0][1]
                style_header_painter(wb[sheetname], start_col, end_col)

        # - CT
        sheetname = 'ct'
        if sheetname in wb.sheetnames:
            unmerge_header_row(wb[sheetname], 1, style_d={1: p_header_style, 2: p_header_style, 3: p_subheader_style})
            unmerge_header_row(wb[sheetname], 2)
            import gc
            gc.get_count()
            gc.collect()
            # correct for 'salinity'
            for col_idx in find_str_in_row(wb[sheetname], 'salinity', 2):
                cr = find_merged_cell(wb[sheetname], 2, col_idx)
                if cr:
                    wb[sheetname].unmerge_cells(cr.coord)
                wb[sheetname].cell(2, col_idx + 1).value = 'salinity'
                wb[sheetname].cell(2, col_idx).value = None

        # - LISTS
        sheetname = 'lists'
        # Override latest `list` sheet
        if sheetname in wb.sheetnames:
            list_sheet_index = wb.worksheets.index(wb['lists'])
            del wb['lists']
        wb.create_sheet("lists", list_sheet_index)
        wb_source = openpyxl.load_workbook(
            os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYXXZZ-N_P-1.4.7-eco-pool_lists.xlsx'),
            data_only=True)
        copy_cells(wb_source['lists'], wb['lists'])  # copy all the cel values and styles
        copy_sheet_attributes(wb_source['lists'], wb['lists'])

        # update version number
        version = wb['metadata-station']['C1'].value
        wb['metadata-station']['C1'].value = '1.4.7'
        wb.active = wb['metadata-station']
        logger.debug('\t %s updated from %s to %s' % (
            wb['metadata-core']['C1'].value, version, wb['metadata-station']['C1'].value))

    else:
        logger.info("\t%s: already update to version %s " % (
            wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))

    wb = fix_merged_cells(wb)
    return wb


def version_1_4_7_to_1_4_8(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 8:
        # - METADATA-CORE
        sheetname = 'metadata-core'
        wb[sheetname]['B2'].value = 'YYYY-MM-DD'
        wb[sheetname]['B2'].style = 'm_unit_r_style'
        wb[sheetname]['B2'].border = b_border
        wb[sheetname]['B3'].value = 'hh:mm | TZ'
        wb[sheetname]['B3'].style = 'm_unit_r_style'
        wb[sheetname]['B3'].border = b_border

        wb[sheetname].merge_cells('C1:H1')
        wb[sheetname]['C1'].border = noBorder
        wb[sheetname]['C2'].border = tb_border
        if is_merged(wb[sheetname], 4, 3):
            wb[sheetname].unmerge_cells(find_merged_cell(wb[sheetname], 4, 3).coord)
        wb[sheetname].merge_cells('C4:H4')

        row_idx = find_str_in_col(wb[sheetname], 'INSTRUMENTS', 1)[0]
        wb[sheetname].cell(row_idx, 1).value = 'INSTRUMENTS'
        wb[sheetname].cell(row_idx, 3).value = 'brand, specification (type, diameter, accuracy…)'
        wb[sheetname].cell(row_idx, 3).font = Font(name='Geneva', charset=1, family=2, sz=9, bold=True,
                                                   italic=False, color='FF000000')
        wb[sheetname].cell(row_idx, 5).value = None
        wb[sheetname].cell(row_idx, 8).value = 'one by line'
        wb[sheetname].cell(row_idx, 8).font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False,
                                                   italic=False, color='FF000000')

        # - METADATA-STATION
        sheetname = 'metadata-station'
        wb[sheetname]['D1'].value = "to be used with PySIC python module (https://github.com/megavolts/pysic)"

        # Insert line above 'sampling'
        insert_row_with_merge(wb[sheetname], 2, 1)
        for col_idx in range(1, 27):
            wb[sheetname].cell(2, col_idx).style = 'm_bkg_style'
        # Set all caps block title
        for title in ['sampling', 'position', 'date and time', 'ice geometry', 'ice temperature', 'sampling event',
                      'weather information', 'general comments']:
            row_idx = find_str_in_col(wb[sheetname], title, 1)
            if len(row_idx) > 0:
                row_idx = row_idx[0]
                cr = find_merged_cell(wb[sheetname], row_idx, 1)
                if cr:
                    try:
                        wb[sheetname].unmerge_cells(cr.coord)
                    except KeyError:
                        wb.save('/tmp/tmp.xlsx')
                        wb = openpyxl.load_workbook('/tmp/tmp.xlsx')
                        wb[sheetname].unmerge_cells(cr.coord)
                    else:
                        pass
            for col_idx in range(cr.left[0][1], cr.right[0][1] + 1):
                wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'
                wb[sheetname].cell(row_idx, 1).value = title.upper()
                wb[sheetname].cell(row_idx, 1).style = 'm_title_style'

        # Collapse title and header for 'POSITION' and 'DATE AND TIME'
        for title in ['POSITION', 'DATE AND TIME']:
            row_idx = find_str_in_col(wb[sheetname], title, 1)[0]
            wb[sheetname].cell(row_idx + 1, 1).value = title
            wb[sheetname].cell(row_idx + 1, 1).style = 'm_title_style'
            delete_row_with_merge(wb[sheetname], row_idx, 1)

        # Rename 'ICE TEMPERATURE' to 'ENVIRONMENTAL CONDITIONS'
        row_idx = find_str_in_col(wb[sheetname], 'ICE TEMPERATURE', 1)[0]
        wb[sheetname].cell(row_idx, 1).value = 'ENVIRONMENTAL CONDITIONS'

        # Change salinity unit from PSU to - (unitless)
        row_idx = find_str_in_col(wb[sheetname], 'water temperature', 1)[0]
        wb[sheetname].cell(row_idx, 1).value = 'seawater temperature'
        wb[sheetname].cell(row_idx + 1, 1).value = 'seawater salinity'
        wb[sheetname].cell(row_idx + 1, 2).value = '-'
        wb[sheetname].cell(row_idx + 1, 1).style = 'm_header_l_style'
        wb[sheetname].cell(row_idx + 1, 2).style = 'm_unit_r_style'
        wb[sheetname].cell(row_idx + 1, 3).style = 'm_data_r_style'
        wb[sheetname].cell(row_idx + 1, 4).style = 'm_comment_style'

        row_idx = find_str_in_col(wb[sheetname], 'associated cores', 1)[0]
        if wb[sheetname].cell(row_idx, 9).value in [None, '']:
            wb[sheetname].cell(row_idx, 9).style = 'm_bkg_style'

        # Remove bold font in data
        wb[sheetname]['C3'].font = m_data_r_style.font
        wb[sheetname]['C4'].font = m_data_r_style.font
        for row_idx in openpyxl.utils.rows_from_range('C9:H10'):
            for cell in row_idx:
                wb[sheetname][cell].font = m_data_r_style.font
        # No bottom border for associated cores after I-column
        row_idx = find_str_in_col(wb[sheetname], 'associated cores', 1)[0]
        for col_idx in range(9, wb[sheetname].max_column):
            if wb[sheetname].cell(row_idx, col_idx).value is not None:
                wb[sheetname].cell(row_idx, col_idx).border = noBorder

        # - SALO18
        sheetname = 'salo18'
        wb[sheetname].row_dimensions[1].height = 25
        wb[sheetname].row_dimensions[2].height = 25
        if 'Depth' in wb[sheetname]['B1'].value:
            wb[sheetname]['B1'].value = 'depth 2'
        wb[sheetname]['D3'].value = '-'

        # - TEMP
        sheetname = 'temp'
        # -- rename 'depth center' to 'depth':
        wb[sheetname]['A1'] = 'depth'

        # - TEX
        sheetname = 'tex'
        col_idx = find_str_in_row(wb[sheetname], 'Core', 1)[0]
        wb[sheetname].cell(1, col_idx).value = 'core section'
        wb[sheetname].cell(2, col_idx).value = 'value'

        # rename tex tab to stratigraphy
        wb[sheetname].title = 'stratigraphy'

        # - ECO
        sheetname = 'eco'
        wb[sheetname].row_dimensions[1].height = 25
        wb[sheetname].row_dimensions[2].height = 25
        unmerge_header_row(wb[sheetname])

        # - ECO-POOL
        sheetname = 'eco-pool'
        if sheetname in wb.sheetnames:
            wb[sheetname].row_dimensions[1].height = 25
            wb[sheetname].row_dimensions[2].height = 25
            col_idx = find_str_in_row(wb[sheetname], 'melted sample', 1)[0]
            if wb[sheetname].cell(2, col_idx).value in ['total volume']:
                col_new_idx = find_str_in_row(wb[sheetname], 'value', 2)
                col_new_idx = col_new_idx[col_idx <= col_new_idx][0]
                wb[sheetname].cell(1, col_new_idx).value = wb[sheetname].cell(1, col_idx).value
                wb[sheetname].cell(1, col_idx).value = None

            for cr in wb[sheetname].merged_cells.ranges:
                if cr.start_cell.value == 'field sample':
                    cr.start_cell.value = None
                elif cr.start_cell.value == 'ice section height':
                    col_idx = find_str_in_row(wb[sheetname], 'mean', 2)
                    if len(col_idx) == 1:
                        col_idx = col_idx[0]
                        wb[sheetname].cell(2, col_idx).value = 'value'
                    elif len(col_idx) > 1:
                        logger.error('%s - two many values for col_idx' % wb['metadata-core']['C1'].value)

            # Unmerge
            unmerge_header_row(wb[sheetname])

            # add vertical border in front of ECOx
            col_idx = find_str_in_row(wb[sheetname], 'ECO', 2)[0]
            for row_idx in range(1, wb[sheetname].max_row):
                wb[sheetname].cell(row_idx, col_idx).border = l_border

            # add vertical border after 'depth 2' of ECOx
            col_idx = find_str_in_row(wb[sheetname], 'depth 2', 1)[0] + 1
            for row_idx in range(1, wb[sheetname].max_row):
                wb[sheetname].cell(row_idx, col_idx).border = l_border

        # - TM
        sheetname = 'TM'
        if sheetname in wb.sheetnames:
            unmerge_header_row(wb[sheetname])

        # - SNOW:
        sheetname = 'snow'
        wb[sheetname]['D3'].value = '-'
        col_idx = find_str_in_row(wb[sheetname], 'swow micropen', 1)
        if len(col_idx) > 0:
            wb[sheetname].cell(1, col_idx[0]).value = 'snow micropen SMP'
        col_idx = find_str_in_row(wb[sheetname], 'Snow micro pen SMP', 1)
        if len(col_idx) > 0:
            wb[sheetname].cell(1, col_idx[0]).value = 'snow micropen SMP'
        unmerge_header_row(wb[sheetname])

        # -- Change SMP subheader orders to ID, Reading, Action
        col_idx = find_str_in_row(wb[sheetname], 'snow micropen SMP', 1)[0]
        col_reading_idx = find_str_in_row(wb[sheetname], 'Reading', 2)[0]
        col_action_idx = find_str_in_row(wb[sheetname], 'Action', 2)[0]
        col_start = min(col_idx, col_reading_idx, col_action_idx)
        col_n = 3
        insert_col_with_merge(wb[sheetname], col_start, col_n)
        move_range = openpyxl.utils.get_column_letter(col_idx + col_n) + '1:' + \
                     openpyxl.utils.get_column_letter(col_idx + col_n) + str(wb[sheetname].max_row)
        wb[sheetname].move_range(move_range, 0, col_start - col_idx - col_n)
        move_range = openpyxl.utils.get_column_letter(col_reading_idx + col_n) + '1:' + \
                     openpyxl.utils.get_column_letter(col_reading_idx + col_n) + str(wb[sheetname].max_row)
        wb[sheetname].move_range(move_range, 0, col_start - col_reading_idx - col_n + 1)
        move_range = openpyxl.utils.get_column_letter(col_action_idx + col_n) + '1:' + \
                     openpyxl.utils.get_column_letter(col_action_idx + col_n) + str(wb[sheetname].max_row)
        wb[sheetname].move_range(move_range, 0, col_start - col_action_idx - col_n + 2)
        delete_col_with_merge(wb[sheetname], col_start + 3, col_n)
        # -- formatting
        style_header_painter(wb[sheetname], col_start, col_start + 2)
        style_bottom_painter(wb[sheetname], col_idx, col_idx + 1)

        # remove any leftover border after comment
        col_start = find_str_in_row(wb[sheetname], 'comment', 1)[0]
        for row_idx in range(1, wb[sheetname].max_row):
            for col_idx in range(col_start, wb[sheetname].max_row):
                wb[sheetname].cell(row_idx, col_idx).border = noBorder

        if len(find_str_in_row(wb[sheetname], 'd_excess', 1)) == 0:
            col_idx = find_str_in_row(wb[sheetname], 'dD', 1)[0] + 1
            insert_col_with_merge(wb[sheetname], col_idx, 1)
            wb[sheetname].cell(1, col_idx).value = 'd_excess'
            wb[sheetname].cell(2, col_idx).value = 'value'
            wb[sheetname].cell(3, col_idx).value = '‰'
            # formatting
            style_header_painter(wb[sheetname], col_idx, col_idx + 1)
            style_bottom_painter(wb[sheetname], col_idx, col_idx + 1)

        # - SACKHOLE
        sheetname = 'sackhole'
        wb[sheetname]['A1'].value = 'depth'
        wb[sheetname]['C3'].value = '-'

        # - SEAWATER
        sheetname = 'seawater'

        # -- remove 'depth 2' column, and rename 'depth 1' to 'depth'. Collapse column if same depht.
        max_row = wb[sheetname].max_row
        depth1_val = [wb[sheetname].cell(row=ii, column=1).value for ii in range(4, max_row)]
        depth2_val = [wb[sheetname].cell(row=ii, column=2).value for ii in range(4, max_row)]
        if depth1_val == depth2_val:
            delete_col_with_merge(wb[sheetname], 2, 1)
            wb[sheetname]['A1'] = 'depth'

        else:
            cell = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
            logging.error('\t\t-%s: cell %s is not empty impossible to delete column'
                          % (wb[sheetname].title, cell))

        # Merge d180 and dD, add d-excess
        stored_headers = ['d18O', 'dD']
        property_range = {}
        crs = wb[sheetname].merged_cells.ranges.copy()
        for merge_cell in crs:
            if merge_cell.start_cell.value in stored_headers:
                for prop in stored_headers:
                    if merge_cell.start_cell.value == prop:
                        property_range[prop] = merge_cell
        # --unmerge header
        try:
            unmerge_header_row(wb[sheetname])
        except KeyError:
            wb.save('/tmp/temp.xlsx')
            wb = openpyxl.load_workbook('/tmp/temp.xlsx')
            unmerge_header_row(wb[sheetname])
        else:
            pass
        # clean worksheet to remove empty row at the end
        clean_worksheet(wb[sheetname])

        # Reorganize d18O, dD columns and add d_excess
        max_row = wb[sheetname].max_row
        # found column id for ID and quality flag of d18O and dD
        sh_idx = {}
        for prop in property_range:
            if prop not in sh_idx:
                sh_idx[prop] = {}
            merge_cell = property_range[prop]
            col_start = merge_cell.left[0][1]
            col_end = merge_cell.right[0][1]
            for col_idx in range(col_start, col_end + 1):
                subheader = wb[sheetname].cell(2, col_idx).value
                sh_idx[prop][subheader] = col_idx

        # Check if d18o_id match dd_id:
        d18o_id = [wb[sheetname].cell(row=ii, column=sh_idx['d18O']['ID']).value for ii in range(4, max_row)]
        dd_id = [wb[sheetname].cell(row=ii, column=sh_idx['dD']['ID']).value for ii in range(4, max_row)]
        d18o_qual = [wb[sheetname].cell(row=ii, column=sh_idx['d18O']['quality']).value for ii in range(4, max_row)]
        dd_qual = [wb[sheetname].cell(row=ii, column=sh_idx['dD']['quality']).value for ii in range(4, max_row)]

        if dd_id == d18o_id:
            # remove colum of dd_idy
            wb[sheetname].delete_cols(sh_idx['dD']['ID'], 1)
            if d18o_qual != dd_qual:
                logging.error('NOT IMPLEMENTED: d18o_qual != dd_qual')
            else:
                # remove colum for d18o_qual
                wb[sheetname].delete_cols(sh_idx['d18O']['quality'], 1)

        # rename column header
        wb[sheetname].cell(1, sh_idx['d18O']['ID']).value = None
        wb[sheetname].cell(1, sh_idx['d18O']['ID'] + 1).value = 'd180'
        wb[sheetname].cell(1, sh_idx['d18O']['ID'] + 2).value = 'dD'
        wb[sheetname].cell(2, sh_idx['d18O']['ID'] + 1).value = 'value'
        wb[sheetname].cell(2, sh_idx['d18O']['ID'] + 2).value = 'value'
        wb[sheetname].cell(3, sh_idx['d18O']['ID'] + 1).value = '‰'
        wb[sheetname].cell(3, sh_idx['d18O']['ID'] + 2).value = '‰'

        if len(find_str_in_row(wb[sheetname], 'd_excess', 1)) == 0:
            col_idx = find_str_in_row(wb[sheetname], 'dD', 1)[0] + 1
            insert_col_with_merge(wb[sheetname], col_idx, 1)
            wb[sheetname].cell(1, col_idx).value = 'd_excess'
            wb[sheetname].cell(2, col_idx).value = 'value'
            wb[sheetname].cell(3, col_idx).value = '‰'
            # formatting
            style_header_painter(wb[sheetname], col_idx, col_idx + 1)
            style_bottom_painter(wb[sheetname], col_idx, col_idx + 1)

        # - DENSITY-DENSIMETRY
        sheetname = 'density-densimetry'
        wb[sheetname]['F1'].border = l_border
        wb[sheetname]['G1'].border = noBorder
        wb[sheetname]['H1'].border = l_border

        # - CT
        sheetname = 'ct'
        wb[sheetname]['U1'].value = None
        wb[sheetname]['U2'].value = 'comment'
        wb[sheetname]['U3'].value = 'value'
        wb[sheetname]['J4'].value = '-'
        wb[sheetname]['P4'].value = '-'
        unmerge_header_row(wb[sheetname])

        if not has_data(wb[sheetname]):
            sheet_index = wb.worksheets.index(wb[sheetname])
            del wb[sheetname]
            wb.create_sheet(sheetname, sheet_index)
            wb_source = openpyxl.load_workbook(
                os.path.join(pysic_fp, 'resources/updated_sheets/AAA_BB-YYYYXXZZ-N_P-1.4.8-ct.xlsx'),
                data_only=True)
            copy_cells(wb_source[sheetname], wb[sheetname])  # copy all the cel values and styles
            copy_sheet_attributes(wb_source[sheetname], wb[sheetname])

        # - sediment
        sheetname = 'sediment'
        unmerge_header_row(wb[sheetname])

        # apply formatting style
        wb = spreadsheet_style(wb)

        # update version number
        version = wb['metadata-station']['C1'].value
        wb['metadata-station']['C1'].value = '1.4.8'
        wb.active = wb['metadata-station']
        logger.debug('\t %s updated from %s to %s' % (
            wb['metadata-core']['C1'].value, version, wb['metadata-station']['C1'].value))
    else:
        logger.info("\t%s: already update to version %s "
                    % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))
    wb = set_datavalidation(wb)
    wb = fix_merged_cells(wb)
    return wb


def version_1_4_8_to_1_4_9(wb):
    logger = logging.getLogger(__name__)
    version = version2int(wb['metadata-station']['C1'].value)
    if version[0] < 2 and version[1] < 5 and version[2] < 9:
        # - METADATA-STATION
        sheetname = 'metadata-station'
        # Add entry line for ice surface ablation
        row_idx = 22
        insert_row_with_merge(wb[sheetname], row_idx, 1)
        wb[sheetname].cell(row_idx, 1).value = 'ice surface ablation'
        wb[sheetname].cell(row_idx, 2).value = 'm'
        wb[sheetname].cell(row_idx, 3).value = 'N/A'
        wb[sheetname].cell(row_idx, 4).value = 'if measured during melt season'

        # Format
        for col_idx in range(1, wb[sheetname].max_column):
            wb[sheetname].cell(22, col_idx).style = 'm_bkg_style'
        wb['metadata-station'].cell(row_idx, 1).style = 'm_header_l_style'
        wb['metadata-station'].cell(row_idx, 2).style = 'm_unit_r_style'
        wb['metadata-station'].cell(row_idx, 3).style = 'm_data_r_style'
        wb['metadata-station'].cell(row_idx, 4).style = 'm_comment_style'
        wb['metadata-station'].cell(row_idx, 5).style = 'm_comment_style'
        for col_idx in range(openpyxl.utils.column_index_from_string('A'),
                             openpyxl.utils.column_index_from_string('C') + 1):
            wb['metadata-station'].cell(row_idx, col_idx).border = b_border
        # update version number
        wb['metadata-station']['C1'].value = '1.4.9'
        wb.active = wb['metadata-station']
        logger.debug('\t %s updated from 1.4.8 to 1.4.9' % (wb['metadata-core']['C1'].value))
    else:
        logger.info("\t%s: already update to version %s "
                    % (wb['metadata-core']['C1'].value, wb['metadata-station']['C1'].value))

    wb = set_datavalidation(wb)
    wb = fix_merged_cells(wb)
    return wb

# Format
def spreadsheet_style(wb, row_offset=1, col_offset=3):
    logger = logging.getLogger(__name__)

    # def old_format(wb)
    header_d = {'metadata-core': ['THICKNESS', 'REFERENCE', 'INSTRUMENTS', 'VERSION', 'brand'],
                'temperature': ['temperature'],
                'stratigraphy': ['texture', 'core section'],
                'seawater': ['temperature'],
                'sackhole': ['temperature'],
                'density-densimetry': ['mass'],
                'density-volume': ['mass'],
                'sediment': ['sediment mass', 'particulate mass'],
                }
    subheader_d = {'snow': 'snow weight', 'density-densimetry': ['air'], 'density-volume': ['measurement1']}
    sheetnames_l = ['lists', 'locations']
    version = wb['metadata-station']['C1'].value
    version_int = version2int(version)
    # TODO Throw error for early version
    if version_int[0] <= 1 and version_int[1] <= 4 and version_int[2] < 7:
        logger.error('Unable to update style')
    else:
        sheetname = 'metadata-core'
        sheetnames_l.append(sheetname)
        lr_cell = findLowerRightCell(wb[sheetname])
        # background for all cells:
        for row_idx in range(1, lr_cell.row + 1):
            for col_idx in range(1, lr_cell.column + 1):
                wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'
                wb[sheetname].cell(row_idx, col_idx).border = noBorder

        # --CORE INFORMATION
        for row_idx in range(1, 5):
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_unit_r_style'

        wb[sheetname]['C1'].style = 'm_data_l_style'
        wb[sheetname]['C2'].style = 'm_data_r_style'
        wb[sheetname]['C3'].style = 'm_data_r_style'
        wb[sheetname]['D3'].style = 'm_data_l_style'
        wb[sheetname]['C4'].style = 'm_data_l_style'
        for col_idx in range(1, 3):
            wb[sheetname].cell(2, col_idx).border = tb_border
        for col_idx in range(1, 4):
            wb[sheetname].cell(3, col_idx).border = tb_border
        wb[sheetname].cell(3, 4).border = b_border

        # --THICKNESS
        title_row = find_str_in_col(wb[sheetname], 'THICKNESS', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'

        start_row = title_row + 1
        end_row = start_row + 3
        for row_idx in range(start_row, end_row + 1):
            # -- format
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_unit_r_style'
            if row_idx == start_row + 2:
                wb[sheetname].cell(row_idx, 3).style = 'm_cdata_r_style'
            else:
                wb[sheetname].cell(row_idx, 3).style = 'm_data_r_style'
            wb[sheetname].cell(row_idx, 4).style = 'm_comment_style'

            # -- border
            col_idx = 1
            while wb[sheetname].cell(row_idx, col_idx).value not in [None, ''] or \
                    wb[sheetname].cell(row_idx, col_idx).fill.fgColor.rgb in [light_grey, '00000000']:
                if wb[sheetname].cell(row_idx + 1, col_idx).value not in [None, ''] or \
                        is_merged(wb[sheetname], row_idx + 1, col_idx) or \
                        wb[sheetname].cell(row_idx + 1, col_idx).fill.fgColor.rgb in [light_grey, '00000000']:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border
                col_idx += 1

        # REFERENCE DATA
        title_row = find_str_in_col(wb[sheetname], 'REFERENCE', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        wb[sheetname].cell(title_row, 3).style = 'm_title_style'
        wb[sheetname].cell(title_row, 8).style = 'm_title_style'
        wb[sheetname].cell(title_row, 8).font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False,
                                                     italic=False, color='FF000000')
        start_row = title_row + 1
        end_row = start_row + 5
        for row_idx in range(start_row, end_row + 1):
            # -- format
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_subheader_l_style'
            wb[sheetname].cell(row_idx, 3).style = 'm_unit_r_style'
            wb[sheetname].cell(row_idx, 4).style = 'm_data_l_style'
            wb[sheetname].cell(row_idx, 5).style = 'm_comment_style'
            # -- border
            if row_idx < end_row and wb[sheetname].cell(row_idx + 1, 1).value not in [None, '']:
                for col_idx in range(1, 9):
                    wb[sheetname].cell(row_idx, col_idx).border = b_border
        # -- number format
        wb[sheetname].cell(2, 3).number_format = 'yyyy-dd-mm'
        wb[sheetname].cell(3, 3).number_format = 'hh:mm'

        # - INSTRUMENTS
        title_row = find_str_in_col(wb[sheetname], 'INSTRUMENTS', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        wb[sheetname].cell(title_row, 3).style = 'm_title_style'

        row_start = title_row + 1
        row_end = find_str_in_col(wb[sheetname], 'VERSION', 1)[0] - 2
        for row_idx in range(row_start, row_end + 1):
            for col_idx in range(1, 9):
                # -- format
                if col_idx < 3:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_subheader_l_style'
                else:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_l_style'
                # -- border
                if row_idx < row_end:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border

        # - VERSION
        title_row = find_str_in_col(wb[sheetname], 'VERSION', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'

        row_idx = title_row + 1
        for col_idx in range(1, 9):
            # -- format
            if col_idx < 3:
                wb[sheetname].cell(row_idx, col_idx).style = 'm_subheader_r_style'
            else:
                wb[sheetname].cell(row_idx, col_idx).style = 'm_subheader_l_style'

        row_start = title_row + 2
        row_end = lr_cell.row
        for row_idx in range(row_start, row_end + 1):
            for col_idx in range(1, 9):
                # -- format
                if col_idx < 3:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_r_style'
                else:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_l_style'
                # -- border
                if row_idx < row_end:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border

                # -- number format
                if col_idx == 2:
                    wb[sheetname].cell(row_idx, col_idx).number_format = 'yyyy-dd-mm'

        # Set row height
        for row_idx in range(1, lr_cell.row):
            wb[sheetname].row_dimensions[1].height = 12.75

        # Set column width
        wb['metadata-core'].column_dimensions['A'].width = 17

        # - METADATA-STATION
        sheetname = 'metadata-station'
        sheetnames_l.append(sheetname)
        lr_cell = findLowerRightCell(wb[sheetname])

        # Remove background and border for all cells:
        for row_idx in range(1, lr_cell.row + 1):
            for col_idx in range(1, lr_cell.column + 1):
                wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'
                wb[sheetname].cell(row_idx, col_idx).border = noBorder

        # -- ICE DATA SHEET
        # Format
        wb[sheetname]['A1'].style = 'm_title_style'
        wb[sheetname]['B1'].style = 'm_title_style'
        wb[sheetname]['C1'].style = 'm_title_style'
        wb[sheetname]['C1'].font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=True,
                                        color='FF000000')
        wb[sheetname]['D1'].style = 'm_title_style'
        wb[sheetname]['D1'].font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, italic=False,
                                        color='FF000000')

        # -- SAMPLING
        title_row = find_str_in_col(wb[sheetname], 'SAMPLING', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'

        row_start = title_row + 1
        row_end = row_start + 1
        for row_idx in range(row_start, row_end + 1):
            for col_idx in range(1, 6):
                # --- format
                if col_idx < 3:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_header_l_style'
                else:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_l_style'
                # --- border
                if row_idx < row_end:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border

        # POSITION
        # - title
        title_row = find_str_in_col(wb[sheetname], 'POSITION', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        # - header
        for col_idx in range(3, 9):
            wb[sheetname].cell(title_row, col_idx).style = 'm_title_c_style'
            wb[sheetname].cell(title_row + 1, col_idx).style = 'm_title_c_style'
            wb[sheetname].cell(title_row + 1, col_idx).font = Font(name='Geneva', charset=1, family=2, sz=9, bold=False, \
                                                                   italic=True, color='FF000000')
        row_start = title_row + 2
        row_end = row_start + 1
        for row_idx in range(row_start, row_end + 1):
            for col_idx in range(1, 9):
                # -- format
                if col_idx == 1:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_header_l_style'
                elif col_idx == 2:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_unit_r_style'
                else:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_l_style'
                # -- border
                if row_idx < row_end:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border

        # -- border
        col_idx = 6
        for row_idx in range(row_start - 2, row_end + 1):
            if row_idx in [row_start]:
                wb[sheetname].cell(row_idx, col_idx).border = bl_border
            else:
                wb[sheetname].cell(row_idx, col_idx).border = l_border

        # DATE AND TIME
        title_row = find_str_in_col(wb[sheetname], 'DATE AND TIME', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        # Set header
        for col_idx in range(3, 5):
            wb[sheetname].cell(title_row, col_idx).style = 'm_title_c_style'

        # Set subheader and data
        row_start = title_row + 1
        row_end = row_start + 2
        for row_idx in range(row_start, row_end + 1):
            for col_idx in range(1, 5):
                # Format
                if col_idx == 1:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_header_l_style'
                elif col_idx == 2:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_unit_r_style'
                else:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_r_style'
                # Border
                if row_idx < row_end:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border
        # Date format
        wb[sheetname].cell(row_start, 3).number_format = 'yyyy-dd-mm'
        wb[sheetname].cell(row_start, 4).number_format = 'yyyy-dd-mm'
        wb[sheetname].cell(row_start + 1, 3).number_format = 'hh:mm'
        wb[sheetname].cell(row_start + 1, 4).number_format = 'hh:mm'

        # Border
        col_idx = 4
        for row_idx in range(row_start - 1, row_end + 1):
            if row_idx >= row_start and wb[sheetname].cell(row_idx + 1, col_idx).value not in [None, '']:
                wb[sheetname].cell(row_idx, col_idx).border = bl_border
            else:
                wb[sheetname].cell(row_idx, col_idx).border = l_border

        # ICE GEOMETRY
        title_row = find_str_in_col(wb[sheetname], 'ICE GEOMETRY', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        row_start = title_row + 1
        row_end = find_str_in_col(wb[sheetname], 'ENVIRONMENTAL CONDITIONS', 1)[0] - 2
        for row_idx in range(row_start, row_end + 1):
            # -- format
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_unit_r_style'
            if row_idx < row_start + 2:
                wb[sheetname].cell(row_idx, 3).style = 'm_cdata_r_style'
            else:
                wb[sheetname].cell(row_idx, 3).style = 'm_data_l_style'
            if row_idx == row_start:
                for col_idx in range(4, 9):
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_r_style'
            # -- border
            if row_idx < row_end:
                for col_idx in range(1, 4):
                    wb[sheetname].cell(row_idx, col_idx).border = b_border

        # ICE TEMPERATURE
        title_row = find_str_in_col(wb[sheetname], 'ENVIRONMENTAL CONDITIONS', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        row_start = title_row + 1
        row_end = row_start + 4
        for row_idx in range(row_start, row_end + 1):
            # Format
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_unit_r_style'
            wb[sheetname].cell(row_idx, 3).style = 'm_data_r_style'
            wb[sheetname].cell(row_idx, 4).style = 'm_comment_style'

            # Border
            for col_idx in range(1, 5):
                if wb[sheetname].cell(row_idx + 1, col_idx).value not in [None, ''] or \
                        wb[sheetname].cell(row_idx, col_idx).fill.fgColor.rgb in [light_grey, '00000000']:
                    if wb[sheetname].cell(row_idx + 1, col_idx).value not in [None, ''] or \
                            wb[sheetname].cell(row_idx + 1, col_idx).fill.fgColor.rgb in [light_grey, '00000000'] or \
                            is_merged(wb[sheetname], row_idx + 1, col_idx):
                        wb[sheetname].cell(row_idx, col_idx).border = b_border

        # SAMPLING EVENT
        title_row = find_str_in_col(wb[sheetname], 'SAMPLING EVENT', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        row_start = title_row + 1
        row_end = row_start + 3
        for row_idx in range(row_start, row_end + 1):
            # -- format
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 3).style = 'm_data_l_style'
            wb[sheetname].cell(row_idx, 4).style = 'm_comment_style'

            if row_idx in [row_start + 1, row_start + 2]:
                col_idx = 4
                while wb[sheetname].cell(row_idx, col_idx).value not in [None, '']:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_l_style'
                    col_idx += 1

            # -- border
            if row_idx < row_end:
                for col_idx in range(1, 4):
                    wb[sheetname].cell(row_idx, col_idx).border = b_border
                col_idx = 3
                while wb[sheetname].cell(row_idx, col_idx).value not in [None, '']:
                    if wb[sheetname].cell(row_idx + 1, col_idx).value not in [None, ''] or \
                            is_merged(wb[sheetname], row_idx + 1, col_idx) or \
                            wb[sheetname].cell(row_idx + 1, col_idx).fill == noFill:
                        wb[sheetname].cell(row_idx, col_idx).border = b_border
                    col_idx += 1

        # WEATHER INFORMATION
        title_row = find_str_in_col(wb[sheetname], 'WEATHER INFORMATION', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        row_start = title_row + 1
        row_end = row_start + 4
        for row_idx in range(row_start, row_end + 1):
            # -- format
            wb[sheetname].cell(row_idx, 1).style = 'm_header_l_style'
            wb[sheetname].cell(row_idx, 2).style = 'm_unit_r_style'
            wb[sheetname].cell(row_idx, 3).style = 'm_data_r_style'
            wb[sheetname].cell(row_idx, 4).style = 'm_comment_style'
            if wb[sheetname].cell(row_idx, 5).value not in [None, '']:
                wb[sheetname].cell(row_idx, 5).style = 'm_comment_style'
            if row_idx in [row_start + 1, row_start + 2]:
                col_idx = 4
                while wb[sheetname].cell(row_idx, col_idx).value not in [None, '']:
                    wb[sheetname].cell(row_idx, col_idx).style = 'm_data_l_style'
                    col_idx += 1
            # -- border
            col_idx = 1
            while wb[sheetname].cell(row_idx, col_idx).value not in [None, ''] or \
                    wb[sheetname].cell(row_idx, col_idx).fill.fgColor.rgb in [light_grey, '00000000']:
                if wb[sheetname].cell(row_idx + 1, col_idx).value not in [None, ''] or \
                        is_merged(wb[sheetname], row_idx + 1, col_idx) or \
                        wb[sheetname].cell(row_idx + 1, col_idx).fill.fgColor.rgb in [light_grey, '00000000']:
                    wb[sheetname].cell(row_idx, col_idx).border = b_border
                col_idx += 1

        # GENERAL COMMENTS
        title_row = find_str_in_col(wb[sheetname], 'GENERAL COMMENTS', 1)[0]
        wb[sheetname].cell(title_row, 1).style = 'm_title_style'
        row_start = title_row + 1
        row_end = row_start
        for row_idx in range(row_start, row_end + 1):
            # -- format
            wb[sheetname].cell(row_idx, 1).style = 'm_data_l_style'

        # Row height
        for row_idx in range(1, lr_cell.row):
            wb[sheetname].row_dimensions[1].height = 12.75

        # - PROPERTY TABS
        for sheetname in wb.sheetnames:
            if sheetname not in sheetnames_l:
                lr_cell = findLowerRightCell(wb[sheetname], col_offset=0, row_offset=0)

                # Header & Data
                if wb[sheetname].title in ['ct']:
                    style_d = {1: p_header_style, 2: p_header_style, 3: p_subheader_style, 4: p_unit_style}
                    start_data_row = 5
                else:
                    style_d = {1: p_header_style, 2: p_subheader_style, 3: p_unit_style}
                    start_data_row = 4
                start_col = 1
                end_col = lr_cell.column
                max_row = lr_cell.row
                style_header_painter(wb[sheetname], start_col, end_col, start_row=1, style_d=style_d)
                style_data_painter(wb[sheetname], start_col, end_col, data_row=start_data_row, max_row=max_row)

                # Add background
                col_offset = 3
                max_row = lr_cell.row + 1
                style_bottom_painter(wb[sheetname], start_col, end_col + col_offset, row_idx=max_row)
                for col_idx in range(lr_cell.column + 1, lr_cell.column + 1 + col_offset):
                    for row_idx in range(1, max_row + 1):
                        wb[sheetname].cell(1, col_idx).style = 'm_bkg_style'
                    # set column width: 1.76 in, fudge_factor = 13.97; 0.5"
                    col = openpyxl.utils.get_column_letter(col_idx)
                    wb[sheetname].column_dimensions[col].width = 0.5 * 14

                if wb[sheetname].title in ['snow']:
                    for col_idx in range(1, lr_cell.column + 1):
                        if all([wb[sheetname].cell(ii, col_idx).value is None for ii in range(1, start_data_row)]):
                            col = openpyxl.utils.get_column_letter(col_idx)
                            for row_idx in range(1, lr_cell.row + 1):
                                wb[sheetname].cell(row_idx, col_idx).style = 'm_bkg_style'
                            # set column width: 0.5 in, fudge_factor = 13.97; 0.5"
                            wb[sheetname].column_dimensions[col].width = 0.1 * 13.97

                        max_border_col = find_str_in_row(wb[sheetname], 'comment', 1)[0]
                else:
                    max_border_col = lr_cell.column

                # Find border location
                border_col = []
                for col_idx in range(1, max_border_col + 1):
                    if wb[sheetname].cell(start_data_row - 2, col_idx).value == 'ID':
                        border_col.append(col_idx)
                    elif any([wb[sheetname].cell(ii, col_idx).value == 'comment' for ii in range(1, start_data_row)]):
                        border_col.append(col_idx)
                    elif sheetname in header_d and \
                            str(wb[sheetname].cell(1, col_idx).value) in header_d[sheetname] and \
                            wb[sheetname].cell(2, col_idx).value == 'value':
                        border_col.append(col_idx)
                    elif sheetname in subheader_d and \
                            str(wb[sheetname].cell(2, col_idx).value) in subheader_d[sheetname]:
                        border_col.append(col_idx)

                # Remove all border, and apply border at defined border location
                for col_idx in range(1, lr_cell.column + 1):
                    for row_idx in range(1, max_row):
                        wb[sheetname].cell(row_idx, col_idx).border = noBorder
                        if col_idx in border_col:
                            wb[sheetname].cell(row_idx, col_idx).border = l_border
                    wb[sheetname].cell(row_idx, max_row + 1).border = noBorder

                # Set row heights
                wb[sheetname].row_dimensions[1].height = 25
                wb[sheetname].row_dimensions[2].height = 25
                for row_idx in range(3, lr_cell.row):
                    wb[sheetname].row_dimensions[row_idx].height = 12.75

                # Set column widths [.78" == 13]
                for col_idx in range(1, lr_cell.column):
                    col = openpyxl.utils.get_column_letter(col_idx)
                    wb[sheetname].column_dimensions[col].width = 13

                # set quality width to 6 if header is 0
                for col_idx in find_str_in_row(wb[sheetname], 'quality', 2):
                    if wb[sheetname].cell(1, col_idx).value in [None, '']:
                        col = openpyxl.utils.get_column_letter(col_idx)
                        wb[sheetname].column_dimensions[col].width = 6
                for col_idx in find_str_in_row(wb[sheetname], 'quality', 3):
                    if wb[sheetname].cell(1, col_idx).value in [None, ''] and \
                            wb[sheetname].cell(2, col_idx).value in [None, '']:
                        col = openpyxl.utils.get_column_letter(col_idx)
                        wb[sheetname].column_dimensions[col].width = 6

                # set width to 1" for
                for header in ['specific conductance', 'conductivity', 'stratigraphy']:
                    for col_idx in find_str_in_row(wb[sheetname], header, 1):
                        col = openpyxl.utils.get_column_letter(col_idx)
                        wb[sheetname].column_dimensions[col].width = 13 / .78  # 1")
                    for col_idx in find_str_in_row(wb[sheetname], header, 2):
                        col = openpyxl.utils.get_column_letter(col_idx)
                        wb[sheetname].column_dimensions[col].width = 13 / .78  # 1")
                # set width to 1.75" for
                for header in ['inclusion', 'description', 'comment']:
                    for col_idx in find_str_in_row(wb[sheetname], header, 1):
                        col = openpyxl.utils.get_column_letter(col_idx)
                        wb[sheetname].column_dimensions[col].width = 13 / .78  # 1")
                    for col_idx in find_str_in_row(wb[sheetname], header, 2):
                        col = openpyxl.utils.get_column_letter(col_idx)
                        wb[sheetname].column_dimensions[col].width = 1.75 * 13 / .78  # 1")

                # set wrapping text for
                for header in ['particulate mass', 'specific conductance', 'core section', 'snow micropen SMP',
                               'melted volume', 'eco property']:
                    for col_idx in find_str_in_row(wb[sheetname], header, 1):
                        wb[sheetname].cell(1, col_idx).alignment = Alignment(horizontal='center', wrapText=True,
                                                                             vertical='center')
                    for col_idx in find_str_in_row(wb[sheetname], header, 2):
                        wb[sheetname].cell(1, col_idx).alignment = Alignment(horizontal='center', wrapText=True,
                                                                             vertical='center')
    return wb


def remove_datavalidation(ws):
    dv_to_remove = []
    for dv in ws.data_validations.dataValidation:
        dv_to_remove.append(dv)
    for dv in dv_to_remove:
        ws.data_validations.dataValidation.remove(dv)

def set_datavalidation(wb):

    # remove data validation from all sheets
    for sheetname in wb.sheetnames:
        remove_datavalidation(wb[sheetname])

    # add data validation
    sheetname = 'metadata-core'
    if sheetname in wb.sheetnames:
        dv_range = 'D3'
        wb[sheetname].add_data_validation(dv_lists_dict['tz_dv'])
        dv_lists_dict['tz_dv'].add(dv_range)

        wb[sheetname].add_data_validation(vert_ref_list)
        vert_ref_list.add('D13')
        vert_ref_list.add('D15')
        vert_ref_list.add('D17')

        wb[sheetname].add_data_validation(vert_dir_list)
        vert_dir_list.add('D14')
        vert_dir_list.add('D16')
        vert_dir_list.add('D18')

    sheetname = 'metadata-station'
    if sheetname in wb.sheetnames:
        wb[sheetname].add_data_validation(dv_lists_dict['tz_dv'])
        dv_lists_dict['tz_dv'].add('C15')
        dv_lists_dict['tz_dv'].add('D15')

        row_idx = find_str_in_col(wb[sheetname], 'surface condition', 1)[0]
        wb[sheetname].add_data_validation(dv_lists_dict['age_dv'])
        dv_lists_dict['age_dv'].add('C' + str (row_idx - 3))
        wb[sheetname].add_data_validation(dv_lists_dict['topography_dv'])
        dv_lists_dict['topography_dv'].add('C' + str(row_idx - 2))
        wb[sheetname].add_data_validation(dv_lists_dict['environement_dv'])
        dv_lists_dict['environement_dv'].add('C' + str(row_idx - 1))
        wb[sheetname].add_data_validation(dv_lists_dict['surface_dv'])
        dv_lists_dict['surface_dv'].add('C' + str(row_idx))

        row_idx = find_str_in_col(wb[sheetname], 'cloud cover', 1)[0]
        wb[sheetname].add_data_validation(dv_lists_dict['cloud'])
        dv_lists_dict['cloud'].add('C' + str (row_idx))

    sheetname = 'stratigraphy'
    if sheetname in wb.sheetnames:
        max_row = wb[sheetname].max_row
        dv_range = 'C4:C' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['texture_dv'])
        dv_lists_dict['texture_dv'].add(dv_range)
        dv_range = 'E4:E' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['inclusion_dv'])
        dv_lists_dict['inclusion_dv'].add(dv_range)
        dv_range = 'F4:F' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['description_dv'])
        dv_lists_dict['description_dv'].add(dv_range)
        dv_range = 'H4:H' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['ice_type_dv'])
        dv_lists_dict['ice_type_dv'].add(dv_range)

    sheetname = 'tex'
    if sheetname in wb.sheetnames:
        max_row = wb[sheetname].max_row
        dv_range = 'C4:C' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['texture_dv'])
        dv_lists_dict['texture_dv'].add(dv_range)
        dv_range = 'E4:E' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['inclusion_dv'])
        dv_lists_dict['inclusion_dv'].add(dv_range)
        dv_range = 'F4:F' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['description_dv'])
        dv_lists_dict['description_dv'].add(dv_range)
        dv_range = 'H4:H' + str(max_row - 1)
        wb[sheetname].add_data_validation(dv_lists_dict['ice_type_dv'])
        dv_lists_dict['ice_type_dv'].add(dv_range)

    return wb

# Helper function
def find_str_in_row(ws, header, row_idx=1):
    """
    find column index for a given header at a specific row
    :param ws: worksheet to search in
    :param header: header to look for
    :param row_idx: row index to look in, default 1
    :return: int, column index where header is found in row_idx row
    """
    col_idx_list = []
    for col_idx in range(1, ws.max_column):
        if header in str(ws.cell(row_idx, col_idx).value):
            col_idx_list.append(col_idx)
    return np.array(col_idx_list)


def find_str_in_col(ws, string, col_idx):
    """
    find row index with a given string in a specific column
    :param ws: worksheet to search in
    :param string: string to search for
    :param col_idx: index of column to search in
    :return:
    """
    row_idx_list = []
    for row_idx in range(1, ws.max_row):
        if ws.cell(row_idx, col_idx).value is not None and string in str(ws.cell(row_idx, col_idx).value):
            row_idx_list.append(row_idx)
    return np.array(row_idx_list)


def is_merged(ws, row_idx, col_idx):
    """
    Test if a cell is merged
    :param ws: worksheet containing the cell
    :param row_idx: row index of the cell
    :param col_idx: column index of the cell
    :return: True if merged; False otherwise
    """
    cell = ws.cell(row_idx, col_idx)
    for mergedCell in ws.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False


def find_merged_cell(ws, row_idx, col_idx):
    """
    find if a specific cell is merged and return the merged cell parameter
    :param ws: worksheet containing the cell
    :param row_idx: row index of the cell
    :param col_idx: column index of the cell
    :return: mergedCell from Openpyxl
    """
    cell = ws.cell(row_idx, col_idx)
    for mergedCell in ws.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return mergedCell
    return False


def unmerge_header_row(ws, header_row=1, nomerge_header=None, style_d=None):
    if nomerge_header is None:
        nomerge_header = []
    if style_d is None:
        style_d = {1: p_header_style, 2: p_subheader_style, 3: p_unit_style}
    crs_d = {cr.coord: cr for cr in ws.merged_cells.ranges}
    while len(crs_d) > 0:
        key = list(crs_d.keys())[0]
        merge_cell = crs_d[key]
        if merge_cell.left[0][0] == header_row:
            col_start = merge_cell.left[0][1]
            col_end = merge_cell.right[0][1]

            # unmerged cell
            header = merge_cell.start_cell.value
            if header in nomerge_header:
                crs_d.pop(key)
            else:
                ws.unmerge_cells(merge_cell.coord)
                # find header
                value_idx = find_str_in_row(ws, 'value', 2)
                value_idx = value_idx[(col_start <= value_idx) & (value_idx <= col_end)]
                if len(value_idx) > 0:
                    col_idx = value_idx[0]
                else:
                    id_idx = find_str_in_row(ws, 'id', 2)
                    if len(id_idx) == 0:
                        id_idx = find_str_in_row(ws, 'ID', 2)
                    id_idx = id_idx[(col_start <= id_idx) & (id_idx <= col_end)]
                    if len(id_idx) > 0:
                        col_idx = id_idx[0]
                    else:
                        col_idx = col_start
                ws.cell(header_row, col_start).value = None
                ws.cell(header_row, col_idx).value = header

                style_header_painter(ws, col_start, col_end, start_row=header_row, style_d=style_d)

                sheetname = ws.title
                if header is not None and sheetname in prop_associated_tab and header in prop_associated_tab[sheetname]:
                    pass
                else:
                    for row in range(1, ws.max_row):
                        ws.cell(row, col_start).border = l_border
            crs_d.pop(key)
        else:
            crs_d.pop(key)


def has_data(ws):
    row_idx = 1
    flag_depth = False
    while not flag_depth and row_idx <= 4:
        if len(find_str_in_row(ws, 'depth', row_idx)) > 0:
            flag_depth = True
        else:
            row_idx += 1
    depth = []
    max_row = ws.max_row
    col_idx = find_str_in_row(ws, 'depth', row_idx)

    data_depth = [[ws.cell(row=ii, column=jj).value for ii in range(row_idx + 3, max_row)] for jj in col_idx]
    data_depth = np.array(data_depth).astype(float)
    # copy default ct page if there is no data
    if np.isnan(data_depth).all():
        return False
    else:
        return True


def insert_row_with_merge(worksheet, row_insert_idx, row_n=1):
    # move merged cells downwards below row_insert_idx row
    for cr in worksheet.merged_cells.ranges:
        if cr.right[0][0] >= row_insert_idx:
            cr.shift(0, row_n)
    # insert row_n rows starting at row_insert_idx row
    worksheet.insert_rows(row_insert_idx, row_n)


def delete_row_with_merge(worksheet, row_insert_idx, row_n=1):
    # move merged cells downwards below row_insert_idx row
    for cr in worksheet.merged_cells.ranges:
        if cr.right[0][0] > row_insert_idx:
            cr.shift(0, -row_n)
    # remove data validation
    for row_idx in range(row_insert_idx, row_insert_idx + row_n):
        for col_idx in range(1, worksheet.max_column):
            removeExistingCellDataValidation(worksheet, worksheet.cell(row_idx, col_idx))
    # insert row_n rows starting at row_insert_idx row
    worksheet.delete_rows(row_idx, row_n)


def insert_col_with_merge(worksheet, col_insert_idx, col_n=1):
    for cr in worksheet.merged_cells.ranges:
        if cr.left[0][1] >= col_insert_idx:
            cr.shift(col_n, 0)
    worksheet.insert_cols(col_insert_idx, col_n)


def delete_col_with_merge(worksheet, col_delete_idx, col_n=1, check_empty=False, check_row_start=1):
    flag_empty = True
    if check_empty:
        for row_idx in range(check_row_start, worksheet.max_row):
            for col_idx in range(col_delete_idx, col_delete_idx + col_n):
                if worksheet.cell(row_idx, col_idx).value not in [None, '']:
                    cell = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
                    logging.error('\t\t-%s: cell %s is not empty impossible to delete column' % (worksheet.title, cell))
                    return False
                else:
                    pass
    else:
        pass
    if flag_empty:
        for cr in worksheet.merged_cells.ranges:
            if cr.left[0][1] >= col_delete_idx:
                cr.shift(-col_n, 0)
        worksheet.delete_cols(col_delete_idx, col_n)


def correctDataRowNumber(worksheet, max_row):
    """
    Add or remove rows  for data entry to match the target number. Only remove row if empty
    """

    # add or remove extra data row as needed
    row_n = max_row - worksheet.max_row
    if row_n > 0:
        row_insert_idx = worksheet.max_row
        insert_row_with_merge(worksheet, row_insert_idx, row_n)
    elif row_n < 0:
        row_n = -row_n
        row_insert_idx = max_row
        flag_empty = True
        # only if no data are present:
        for row_idx in range(row_insert_idx, row_insert_idx + row_n):
            for col_idx in range(1, worksheet.max_column):
                if worksheet.cell(row_idx, col_idx).value not in [None, '']:
                    print(openpyxl.utils.get_column_letter(col_idx) + str(row_idx))
                    flag_empty = False
                else:
                    pass
        if flag_empty:
            delete_row_with_merge(worksheet, row_insert_idx, row_n)


def stylePainter(worksheet, col_comment_idx, max_col, max_row, style_d=None):
    if style_d is None:
        style_d = {1: p_header_style, 2: p_subheader_style, 3: p_unit_style}
    for row in range(1, max_row):
        for col in range(1, col_comment_idx + 1):
            if row in style_d.keys():
                worksheet.cell(row, col).style = style_d[row].name
            else:
                # TODO add other text entry
                if worksheet.cell(1, col).value in ['comment'] or worksheet.cell(2, col).value in ['comment']:
                    worksheet.cell(row, col).style = 'p_data_l_style'
                else:
                    worksheet.cell(row, col).style = 'p_data_r_style'
        for col in range(col_comment_idx + 1, max_col + 1):
            worksheet.cell(row, col).style = 'm_bkg_style'

    for col in range(1, max_col + 1):
        worksheet.cell(max_row, col).style = 'm_bkg_style'

    # set row height
    worksheet.row_dimensions[1].height = worksheet.row_dimensions[2].height
    for row in range(2, max_row + 1):
        worksheet.row_dimensions[row].height = 12.75


def style_header_painter(ws, start_col, end_col, start_row=1,
                       style_d={1: p_header_style, 2: p_subheader_style, 3: p_unit_style}):
    start_row -= 1
    for row in style_d.keys():
        for col in range(start_col, end_col + 1):
            ws.cell(start_row + row, col).style = style_d[row].name


def style_data_painter(ws, start_col, end_col, data_row=4, max_row=None):
    if max_row is None:
        max_row = ws.max_row
    for col_idx in range(start_col, end_col + 1):
        # detect string or number/date base on subheader value
        flag_string = False
        for row_offset in range(1, data_row):
            cell_value = ws.cell(data_row - row_offset, col_idx).value
            if isinstance(cell_value, str):
                if cell_value in ['comment', 'ID', 'texture', 'inclusion', 'description', 'character']:
                    flag_string = True
        for row_idx in range(data_row, max_row + 1):
            if flag_string:  # string data
                ws.cell(row_idx, col_idx).style = 'p_data_l_style'
            else:  # number or date #data
                ws.cell(row_idx, col_idx).style = 'p_data_r_style'


def style_bottom_painter(ws, start_col, end_col, row_idx=None):
    if row_idx is None:
        row_idx = ws.max_row
    for col_idx in range(start_col, end_col + 1):
        try:
            ws.cell(row_idx, col_idx).style = 'm_bkg_style'
        except ValueError:
            ws.cell(row_idx, col_idx).style = 'm_bkg_style'


def clean_data_worksheet(worksheet, col_comment_idx, max_col, max_row):
    logger = logging.getLogger(__name__)
    for row in range(1, worksheet.max_row):
        for col in range(max_col + 1, worksheet.max_column):
            worksheet.cell(row, col).value = None
            worksheet.cell(row, col).fill = noFill
            worksheet.cell(row, col).border = noBorder
    if col_comment_idx > max_col:
        logger.error('col_comment_idx > max_col')
    for row in range(max_row + 1, worksheet.max_row):
        for col in range(1, worksheet.max_column):
            worksheet.cell(row, col).value = None
            worksheet.cell(row, col).fill = noFill
            worksheet.cell(row, col).border = noBorder

    # delete all trailing row after max_row if rows are empty
    flag_row_empty = True
    for row in range(max_row + 1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column):
            if worksheet.cell(row, col).value not in [None, '']:
                flag_row_empty = False
    if flag_row_empty:
        worksheet.delete_rows(max_row + 1, worksheet.max_row - max_row)


def clean_worksheet(ws):
    if ws.title not in ['lists', 'locations']:
        # find max row and max_col
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row_idx, col_idx).value not in [None, '']:
                    max_col = col_idx
                    max_row = row_idx

        # correct minimum max_row
        if ws.title == 'metadata-core':
            if max_row < 27:
                max_row = 27
            if max_col < 26:
                max_col = 26
        elif ws.title in ['metadata-station', 'metadata-coring']:
            if max_row < 50:
                max_row = 50
            if max_col < 26:
                max_col = 26
        elif ws.title == 'snow':
            if max_row < 3 + data_row_n + 1:
                max_row = 3 + data_row_n + 1
            if max_col < find_str_in_row(ws, 'SWE', 1)[0] + 4:
                if is_merged(ws, 1, find_str_in_row(ws, 'SWE', 1)[0]):
                    max_col = find_str_in_row(ws, 'SWE', 1)[0] + 6
                else:
                    max_col = find_str_in_row(ws, 'SWE', 1)[0] + 3
        elif ws.title == 'ct':
            if max_row < 4 + data_row_n + 1:
                max_row = 4 + data_row_n + 1
            col_idx = 0
            if len(find_str_in_row(ws, 'comment', 1)) > col_idx:
                col_idx = find_str_in_row(ws, 'comment', 1)[0]
            if len(find_str_in_row(ws, 'comment', 2)) > col_idx:
                col_idx = find_str_in_row(ws, 'comment', 2)[0]
            if len(find_str_in_row(ws, 'comment', 3)) > col_idx:
                col_idx = find_str_in_row(ws, 'comment', 3)[0]
            if max_col < col_idx + 3:
                max_col = col_idx + 3
        else:
            if max_row < 3 + data_row_n + 1:
                max_row = 3 + data_row_n + 1
            if max_col < find_str_in_row(ws, 'comment', 1)[0] + 3:
                max_col = find_str_in_row(ws, 'comment', 1)[0] + 3
        # find max row with entry
        ws.delete_rows(max_row + 1, ws.max_row - max_row + 1)
        ws.delete_cols(max_col + 1, ws.max_column - max_col + 1)
    else:  # do nothing
        pass


def findLowerRightCell(ws, col_offset=3, row_offset=1, start_row=1, max_row=None):
    if max_row is None:
        max_row = ws.max_row
    if ws.title not in ['lists', 'locations']:
        # find max row and max_col
        max_col = 1
        max_row = 1
        for row_idx in range(start_row, ws.max_row + 1)[::-1]:
            for col_idx in range(1, ws.max_column + 1)[::-1]:
                if ws.cell(row_idx, col_idx).value not in [None, '']:
                    if max_col < col_idx:
                        max_col = col_idx
                    if max_row < row_idx:
                        max_row = row_idx

        # correct minimum max_row and max_col
        if ws.title == 'metadata-core':
            if max_col < 26:
                max_col = 26
        elif ws.title in ['metadata-station', 'metadata-coring']:
            if max_row < 50:
                max_row = 50
            if max_col < 26:
                max_col = 26
            else:
                max_col = max_col + 1
        elif ws.title == 'snow':
            if max_row < 3 + data_row_n + row_offset:
                max_row = 3 + data_row_n + row_offset
            if max_col < find_str_in_row(ws, 'SWE', 1)[0] + 3:
                if is_merged(ws, 1, find_str_in_row(ws, 'SWE', 1)[0]):
                    max_col = find_str_in_row(ws, 'SWE', 1)[0] + col_offset
                else:
                    max_col = find_str_in_row(ws, 'SWE', 1)[0] + 3 + col_offset
        elif ws.title == 'ct':
            if max_row < 4 + data_row_n + row_offset:
                max_row = 4 + data_row_n + row_offset
            col_idx = 0
            if len(find_str_in_row(ws, 'comment', 1)) > 0:
                col_idx = find_str_in_row(ws, 'comment', 1)[0]
            if len(find_str_in_row(ws, 'comment', 2)) > 0:
                col_idx = find_str_in_row(ws, 'comment', 2)[0]
            if len(find_str_in_row(ws, 'comment', 3)) > 0:
                col_idx = find_str_in_row(ws, 'comment', 3)[0]
            if max_col < col_idx + col_offset:
                max_col = col_idx + col_offset
        else:
            if max_row < 3 + data_row_n + row_offset:
                max_row = 3 + data_row_n + row_offset
            col_idx = 1
            if len(find_str_in_row(ws, 'comment', 1)) > 0:
                col_idx = find_str_in_row(ws, 'comment', 1)[0]
            elif len(find_str_in_row(ws, 'comment', 2)) > 0:
                col_idx = find_str_in_row(ws, 'comment', 2)[0]
            elif len(find_str_in_row(ws, 'comment', 3)) > 0:
                col_idx = find_str_in_row(ws, 'comment', 3)[0]
            if max_col < col_idx + col_offset:
                max_col = col_idx + col_offset
    else:
        max_row = ws.max_row
        max_col = ws.max_column
    return ws.cell(max_row, max_col)


def worksheetDataFormatting(worksheet, max_row, l_border_col=[],
                            style_d={1: p_header_style, 2: p_subheader_style, 3: p_unit_style}):
    # find last column with header
    for col in list(range(1, worksheet.max_column))[::-1]:
        if worksheet.cell(1, col).value not in [None, '']:
            col_comment_idx = col
            cell = worksheet.cell(1, col)
            for mergedCell in worksheet.merged_cells.ranges:
                if cell.coordinate in mergedCell:
                    col_comment_idx = mergedCell.max_col
            break
    max_col = int(np.ceil(col_comment_idx / 26)) * 26

    # add or remove extra data row as needed
    correctDataRowNumber(worksheet, max_row)

    # style and fill
    stylePainter(worksheet, col_comment_idx, max_col, max_row, style_d=style_d)

    # border
    for row_idx in range(1, max_row):
        for col in l_border_col:
            col_idx = openpyxl.utils.column_index_from_string(col)
            worksheet.cell(row_idx, col_idx).border = l_border

    # clear empty cell out of data
    clean_data_worksheet(worksheet, col_comment_idx, max_col, max_row)


def removeExistingCellDataValidation(worksheet, cell):
    toRemove = []

    # Append all validation rules for cell to be removed.
    for validation in worksheet.data_validations.dataValidation:
        if validation.__contains__(cell):
            toRemove.append(validation)

    # Process all data validation rules set for removal.
    for rmValidation in toRemove:
        worksheet.data_validations.dataValidation.remove(rmValidation)


###############
## Copy a sheet with style, format, layout, ect. from one Excel file to another Excel file
## Please add the ..path\\+\\file..  and  ..sheet_name.. according to your desire.

def copy_sheet(source_sheet, target_sheet):
    """
    :param source_sheet:
    :param target_sheet:
    :return:
    """
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    if isinstance(source_sheet, openpyxl.worksheet._read_only.ReadOnlyWorksheet):
        return
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of metadata in the attribute I think).
    # So we copy every row's row_dimensions. That seems to work.
    for rn in range(1, source_sheet.max_row):
        #        target_sheet.row_dimensions[rn].height = copy(source_sheet.row_dimensions[rn].height)
        target_sheet.row_dimensions[rn] = source_sheet.row_dimensions[rn]

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute, so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in
        # the targetSheet
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)
        # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the
        # issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)
        # set width for every column
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width)
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    from openpyxl.cell.read_only import EmptyCell, ReadOnlyCell
    for r, row in enumerate(source_sheet.iter_rows()):
        for c, cell in enumerate(row):
            source_cell = cell
            # skip empty cell
            if isinstance(source_cell, EmptyCell):
                continue
            target_cell = target_sheet.cell(column=c + 1, row=r + 1)
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if not isinstance(source_cell, ReadOnlyCell) and source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if not isinstance(source_cell, ReadOnlyCell) and source_cell.comment:
                target_cell.comment = copy(source_cell.comment)


def fix_merged_cells(wb):
    # check health
    for sheetname in wb.sheetnames:
        crs_list = sorted(list(wb[sheetname].merged_cells.ranges))
        for crs in crs_list:
            cr = crs
            cells = cr.cells
            next(cells)  # skip first cell
            for row, col in cells:
                # check if the cell is of type MergedCell, if not
                if (row, col) not in wb[sheetname]._cells:
                    wb[sheetname].merge_cells(cr.coord)
                elif not isinstance(wb[sheetname]._cells[(row, col)], openpyxl.cell.cell.MergedCell):
                    wb[sheetname].merge_cells(cr.coord)
    return wb
