#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
pysic.io.icxl.py : function to import import ice core data from xlsx spreadsheet
"""

__name__ = "core"
__author__ = "Marc Oggier"
__license__ = "GPL"

__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "moggier@alaska.edu"
__status__ = "dev"
__date__ = "2017/09/13"
__comment__ = "loadxl.py contained function to import ice core data from xlsx spreadsheet"
__CoreVersion__ = '1.4.6'

import logging
import os
import openpyxl
import datetime
import dateutil
import pandas as pd
import numpy as np
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

from pysic import TOL
from pysic.core.Core import Core
from pysic.core.Profile import Profile
from pysic.property import prop_parameter_dict
from pysic import inverse_dict

# __all__ = ["import_ic_path", "import_ic_list", "import_ic_sourcefile", "list_ic", "list_ic_path", "make_ic_sourcefile"]


# Cell style
# unit
unit_style = Font(name='Geneva', charset=1, family=2.0, b=False, i=True, strike=None, outline=None,
                  shadow=None, condense=None, color=None, extend=None, sz=9.0, u=None, vertAlign=None,
                  scheme=None)
unit_alignment = Alignment(horizontal='right')

# Border style
# bottom border
b_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# bottom_left_border
bl_border = Border(left=Side(border_style='hair', color='00000000'), right=Side(border_style=None),
                   top=Side(border_style=None), bottom=Side(border_style='hair', color='00000000'))
# left_border
l_border = Border(left=Side(border_style='hair', color='00000000'), right=Side(border_style=None),
                  top=Side(border_style=None), bottom=Side(border_style=None))
# no_border
no_border = Border(left=Side(border_style=None), right=Side(border_style=None),
                   top=Side(border_style=None), bottom=Side(border_style=None))

MAX_ROW = 1000
# variable_2_sheet = {'temperature': 'T_ice',
#                     'salinity': 'S_ice',
#                     'conductivity': 'S_ice',
#                     'specific conductance': 'S_ice',
#                     'd18O': 'S_ice',
#                     'dD': 'S_ice',
#                     'Vf_oil': 'Vf_oil', 'oil volume fraction': 'Vf_oil',  # MOSIDEO project
#                     'Wf_oil': 'Wf_oil', 'oil weight fraction': 'Vf_oil',  # MOSIDEO project
#                     'oil content': 'oil_content',  # CMI project
#                     'oil mass': 'Vf_oil', 'm_oil': 'Vf_oil',
#                     'brine': 'sackhole',
#                     'sackhole': 'sackhole',
#                     'seawater': 'seawater',
#                     'snow': 'snow'
#                     # 'eco':
#                     # 'sediment': 'sediment',
#                     # 'Chla': 'eco',
#                     # 'chlorophyl a': 'eco',
#                     # 'Phae': 'eco'
#                     }

# ## Default values:
# v_ref = 'top'
# verbose = False

# DEBUG:
logging.basicConfig(level=logging.DEBUG)

ic_path = '/home/megavolts/git/pysic/test/ice core/BRW_CS-20210519/BRW_CS-20210519-85_SALO18.xlsx'


# TODO: modifiy reading to account for false bottom
def import_ic_path(ic_path, property=None, drop_empty=False, fill_missing=True):
    """
    :param ic_path:
        string, path to the ice core data. Ice core data should be stored in defined spreadsheet
    :param property:
        list of string, properties to import. If not specified all properties are imported.
    :param drop_empty:
        boolean. True drop empty profile
    :param fill_missing:
        boolean, default True. If true, missing section are filled with np.nan/none
    :return:
    """

    logger = logging.getLogger(__name__)

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet

    ws_summary = wb['metadata-station']
    ws_metadata_core = wb['metadata-core']

    name = ws_metadata_core['C1'].value
    if ws_summary['C1'].value:
        version = ws_summary['C1'].value
        if version != __CoreVersion__:
            logger.warning("%s: ice core data is updated from %s to %s" % (name, version, __CoreVersion__))
            # TODO: update function pysic.core.update_data_spreadsheet
            update_data_spreadsheet(ic_path)
            import_ic_path(ic_path, property, drop_empty=drop_empty)
    else:
        logger.error("(%s) ice core spreadsheet version not unavailable" % name)
        wb.close()

    comments = []

    logger.info("importing data for %s" % ic_path)

    # Import datetime with tz
    date = parse_datetimetz(ws_metadata_core['C2'].value, ws_metadata_core['C3'].value, ws_metadata_core['D3'].value)

    # project
    campaign = ws_summary['C3'].value
    site = ws_summary['C4'].value

    # Location
    lat_start_deg = parse_coordinate(ws_summary['C9'].value, ws_summary['D9'].value, ws_summary['E9'].value)
    lon_start_deg = parse_coordinate(ws_summary['C10'].value, ws_summary['D10'].value, ws_summary['E10'].value)
    lat_end_deg = parse_coordinate(ws_summary['F9'].value, ws_summary['G9'].value, ws_summary['H9'].value)
    lon_end_deg = parse_coordinate(ws_summary['F10'].value, ws_summary['G10'].value, ws_summary['H10'].value)

    # Station time
    station_date_start = parse_datetimetz(ws_summary['C14'].value, ws_summary['C15'].value, ws_summary['C16'].value)
    station_date_end = parse_datetimetz(ws_summary['D14'].value, ws_summary['D15'].value, ws_summary['D16'].value)

    ## Snow Depth
    n_snow = 1
    row_snow = 19
    snow_depth = []
    while ws_summary.cell(row=row_snow, column=3+n_snow).value is not None:
        snow_depth.append(ws_summary.cell(row=row_snow, column=3+n_snow).value)
        n_snow += 1
    snow_depth = np.array(snow_depth)
    if any(snow_depth) > 1:  # check for metric unit
        logger.warning('%s: check if snow depth are reported in cm rather than m' % name)

    # Snow depth measurement
    if len(snow_depth) > 0:
        if isinstance(snow_depth[-1], str):
            comments.append(snow_depth[-1])
            snow_depth = pd.to_numeric(snow_depth[:-1], errors='coerce')
        else:
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')

    # Snow average
    if isinstance(ws_summary.cell(row=row_snow, column=3).value, (float, int)):
        snow_depth_avg = ws_summary.cell(row=row_snow, column=3).value
    elif not np.isnan(snow_depth).all():
        snow_depth_avg = np.nanmean(snow_depth)
    else:
        snow_depth_avg = [np.nan]
    if len(snow_depth) == 0:
        snow_depth = snow_depth_avg

    # Ice Thickness
    try:
        h_i = ws_metadata_core['D7'].value
    except:
        logger.info('(%s) no ice thickness information ' % name)
        h_i = np.nan
    else:
        if h_i == 'n/a':
            comments.append('ice thickness not available')
            logger.info('(%s) ice thickness is not available (n/a)' % name)
            h_i = np.nan
        elif not isinstance(h_i, (int, float)):
            logger.info('%s ice thickness is not a number' % name)
            comments.append('ice thickness not available')
            h_i = np.nan
    if h_i > 10:  # check for metric unit
        logger.warning('%s: check if ice thickness is reported in cm rather than m' % name)

    # Ice Draft
    try:
        h_d = ws_metadata_core['D8'].value
    except:
        logger.info('(%s) no ice draft information ' % name)
        h_d = np.nan
    else:
        if h_d == 'n/a':
            comments.append('ice draft not available')
            logger.info('(%s) ice draft is not available (n/a)' % name)
            h_d = np.nan
        elif not isinstance(h_d, (int, float)):
            logger.info('%s ice draft is not a number' % name)
            comments.append('ice draft not available')
            h_d = np.nan
    if h_d > 10:  # check for metric unit
        logger.warning('%s: check if ice draft is reported in cm rather than m' % name)

    # Ice freeboard
    if not (np.isnan(h_d) and np.isnan(h_i)):
        h_f = h_i - h_d
    else:
        try:
            h_f = ws_metadata_core['D9'].value
        except:
            logger.info('(%s) no ice draft information ' % name)
            h_f = np.nan
        else:
            if h_f == 'n/a':
                comments.append('ice draft not available')
                logger.info('(%s) ice draft is not available (n/a)' % name)
                h_f = np.nan
            elif not isinstance(h_f, (int, float)):
                logger.info('%s ice draft is not a number' % name)
                comments.append('ice draft not available')
                h_f = np.nan
    if h_f > 10:  # check for metric unit
        logger.warning('%s: check if ice freeboard is reported in cm rather than m' % name)

    # Core Length l_c (measured in with ruler)
    try:
        l_c = ws_metadata_core['D10'].value
    except:
        logger.info('(%s) no ice core length' % name)
        l_c = np.nan
    else:
        if l_c == 'n/a':
            comments.append('ice core length not available')
            logger.info('(%s) ice core length is not available (n/a)' % name)
            l_c = np.nan
        elif not isinstance(l_c, (int, float)):
            logger.info('%s ice core length is not a number' % name)
            comments.append('ice core length not available')
            l_c = np.nan
    if l_c > 10:  # check for metric unit
        logger.warning('%s: check if ice freeboard is reported in cm rather than m' % name)

    core = Core(name, date, campaign, lat_start_deg, lon_start_deg, l_c, h_i, h_f, snow_depth)

    # Temperature values
    if ws_summary['C29'].value:
        core.t_air = ws_summary['C29'].value
    if ws_summary['C30'].value:
        core.t_snow_surface = ws_summary['C30'].value
    if ws_summary['C31'].value:
        core.t_ice_surface = ws_summary['C31'].value
    if ws_summary['C32'].value:
        core.t_water = ws_summary['C32'].value
    if ws_summary['C33'].value:
        core.s_water = ws_summary['C33'].value

    # Sampling event
    if ws_summary['C36'].value:
        core.station = ws_summary['C36'].value

    # Sampling protocol
    if ws_summary.cell(3, 39).value:
        core.protocol = ws_summary.cell(3, 39).value
    else:
        core.protocol = 'N/A'

    # Sampling instrument
    instrument_d = {}
    row_instrument = 21
    while ws_metadata_core.cell(row_instrument, 1).value is not None:
        instrument_d[ws_metadata_core.cell(row_instrument, 1).value] = ws_metadata_core.cell(row_instrument, 3).value
        row_instrument += 1
    core.instrument = instrument_d

    # Core collection
    m_col = 3
    row_collection = 37
    while ws_summary.cell(row_collection, m_col).value:
        core.add_to_collection(ws_summary.cell(row_collection, m_col).value)
        m_col += 1

    # comment
    if ws_summary['A49'].value is not None:
        comments.append(ws_summary['A42'].value)
    core.add_comment('; '.join(comments))

    # weather
    # TODO: read weather information

    # References
    reference_d = {}
    if ws_metadata_core['D13'].value is not None:
        if ws_metadata_core['D14'].value is not None:
            reference_d['ice'] = [ws_metadata_core['D13'].value, ws_metadata_core['D14'].value]
        else:
            reference_d['ice'] = [ws_metadata_core['D13'].value, 'up']
    if ws_metadata_core['D15'].value is not None:
        if ws_metadata_core['D16'].value is not None:
            reference_d['snow'] = [ws_metadata_core['D15'].value, ws_metadata_core['D16'].value]
        else:
            reference_d['snow'] = [ws_metadata_core['D15'].value, 'up']
    if ws_metadata_core['D17'].value is not None:
        if ws_metadata_core['D18'].value is not None:
            reference_d['seawater'] = [ws_metadata_core['D17'].value, ws_metadata_core['D18'].value]
        else:
            reference_d['seawater'] = [ws_metadata_core['D18'].value, 'up']
    core.reference.update(reference_d)


    # import property profile
    sheets = wb.sheetnames
    if property is None:
        # TODO: special import for TEX, snow
        sheets = [sheet for sheet in sheets if (sheet not in ['tex', 'TM',  'summary', 'abreviation', 'locations',
                                                              'lists', 'Vf_oil_calculation', 'metadata-core',
                                                              'metadata-station', 'density-volume', 'sediment',
                                                              'ct']) and
                  (sheet.lower().find('fig') == -1)]
        for sheet in sheets:
            ws_property = wb[sheet]
            if sheet == 'snow':
                profile = read_snow_profile(ws_property, property=None, version=version, reference_d=reference_d, fill_missing=True)
                matter = 'snow'
            elif sheet == 'seawater':
                profile = read_generic_profile(ws_property, property=None, version=version, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
                matter = 'seawater'
            elif sheet == 'sackhole' or sheet == 'brine':
                profile = read_generic_profile(ws_property,property=None, version=version, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
                matter = 'brine'
            else:
                profile = read_generic_profile(ws_property, property=None, version=version, reference_d=reference_d, core_length=core.length, fill_missing=fill_missing)
                matter = 'ice'


#             profile['matter'] = matter
#             if matter == 'ice' and profile.get_property() is not None and 'temperature' in profile.get_property():
#                 headers = ['y_mid', 'temperature_value', 'comment', 'variable', 'v_ref', 'matter']
#                 v_ref = profile.v_ref.unique()
#                 if len(v_ref) == 1:
#                     v_ref = v_ref[0]
#                     if v_ref == 'top':
#                         if not -1 in profile.y_mid.values:
#                             # air temperature 1 m above snow surface
#                             if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
#                                 data_surface = [-1, core.t_air, 'Air temperature', 'temperature', 'top', 'air']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                         if not 0 in profile.y_mid.values:
#                             # ice surface
#                             if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
#                                 data_surface = [0, core.t_ice_surface, 'Ice surface temperature', 'temperature', 'top', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                         if core.length - profile.y_mid.max() > TOL:
#                             # ice bottom / seawtaer
#                             if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
#                                 data_bottom = [core.length, core.t_water, 'Ice bottom temperature', 'temperature', 'top', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                                 data_bottom = [core.length, core.t_water, 'Seawater temperature', 'temperature', 'top', 'seawater']
#                                 profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                     elif v_ref == 'bottom':
#                         # air temperature
#                         if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
#                             data_surface = [1, core.t_air, 'Air temperature', 'temperature', 'bottom', 'air']
#                             profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                         if not 0 in profile.y_mid.values:
#                             # ice bottom
#                             if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
#                                 data_bottom = [0, core.t_water, 'Seawater temperature', 'temperature', 'bottom', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                         if np.abs(core.length - profile.y_mid.max()) < TOL:
#                             # ice surface
#                             if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
#                                 data_surface = [core.length, core.t_ice_surface, 'Ice surface temperature', 'temperature', 'bottom', 'ice']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                 else:
#                     logger.error('%s - %s: vertical references mixed up ' %(core.name, sheet))
#                 profile = profile.sort_values(by='y_mid')
#             profile['name'] = name
#             if drop_empty:
#                 profile.drop_empty_property()
#
#             if not profile.empty:
#                 core.add_profile(profile)
#                 logger.info('(%s) data imported with success: %s' % (core.name, ", ".join(profile.get_property())))
#             else:
#                 logger.info('(%s) no data to import from %s ' % (core.name, sheet))
#     else:
#         if not isinstance(variables, list):
#             if variables.lower().find('state variable')+1:
#                 variables = ['temperature', 'salinity']
#             else:
#                 variables = [variables]
#
#         _imported_variables = []
#         for variable in variables:
#             if variable_2_sheet[variable] in ws_name and variable not in _imported_variables:
#                 sheet = variable_2_sheet[variable]
#                 ws_property = wb[sheet]
#                 variable2import = [var for var in variables if var in inverse_dict(variable_2_sheet)[sheet]]
#                 if sheet == 'snow':
#                     profile = read_profile_MOSAiC_UTQ_snow(ws_property, variables=None, version=version, reference_d=reference_d)
#                     matter = 'snow'
#                 elif sheet == 'seawater':
#                     matter = 'seawater'
#                     profile = read_profile_MOSAiC_UTQ(ws_property, variables=None, version=version, reference_d=reference_d)
#                 elif sheet == 'sackhole' or sheet == 'brine':
#                     matter = 'brine'
#                     profile = read_profile_MOSAiC_UTQ(ws_property, variables=None, version=version, reference_d=reference_d)
#                 else:
#                     matter = 'ice'
#                     profile = read_profile_MOSAiC_UTQ(ws_property, variables=None, version=version, reference_d=reference_d)
#                 profile['matter'] = matter
#
#                 # Add temperature at ice surface for temperautre profile
#                 if matter == 'ice' and profile.get_property() is not None and 'temperature' in profile.get_property():
#                     headers = ['y_mid', 'temperature_value', 'comment', 'variable', 'v_ref', 'matter']
#                     v_ref = profile.v_ref.unique()
#                     if len(v_ref) == 1:
#                         v_ref = v_ref[0]
#                         if v_ref == 'top':
#                             if not -1 in profile.y_mid.values:
#                                 # air temperature 1 m above snow surface
#                                 if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
#                                     data_surface = [-1, core.t_air, 'Air temperature', 'temperature', 'top', 'air']
#                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                             if not 0 in profile.y_mid.values:
#                                 # ice surface
#                                 if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
#                                     data_surface = [0, core.t_ice_surface, 'Ice surface temperature', 'temperature',
#                                                     'top', 'ice']
#                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                             if core.length - profile.y_mid.max() > TOL:
#                                 # ice bottom / seawtaer
#                                 if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
#                                     data_bottom = [core.length, core.t_water, 'Ice bottom temperature', 'temperature',
#                                                    'top', 'ice']
#                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                                     data_bottom = [core.length, core.t_water, 'Seawater temperature', 'temperature',
#                                                    'top', 'seawater']
#                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                         elif v_ref == 'bottom':
#                             # air temperature
#                             if isinstance(core.t_air, (float, int)) and not np.isnan(core.t_air):
#                                 data_surface = [1, core.t_air, 'Air temperature', 'temperature', 'bottom', 'air']
#                                 profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                             if not 0 in profile.y_mid.values:
#                                 # ice bottom
#                                 if isinstance(core.t_water, (float, int)) and not np.isnan(core.t_water):
#                                     data_bottom = [0, core.t_water, 'Seawater temperature', 'temperature', 'bottom',
#                                                    'ice']
#                                     profile = profile.append(pd.DataFrame([data_bottom], columns=headers))
#                             if np.abs(core.length - profile.y_mid.max()) < TOL:
#                                 # ice surface
#                                 if isinstance(core.t_ice_surface, (float, int)) and not np.isnan(core.t_ice_surface):
#                                     data_surface = [core.length, core.t_ice_surface, 'Ice surface temperature',
#                                                     'temperature', 'bottom', 'ice']
#                                     profile = profile.append(pd.DataFrame([data_surface], columns=headers))
#                     else:
#                         logger.error('%s - %s: vertical references mixed up ' % (core.name, sheet))
#                     profile = profile.sort_values(by='y_mid')
#                 profile['name'] = name
#                 if drop_empty:
#                     profile.drop_empty_property()
#
#                 if not profile.empty:
#                     core.add_profile(profile)
#                     logger.info('(%s) data imported with success: %s' % (core.name, ", ".join(profile.get_property())))
#                 else:
#                     logger.info('(%s) no data to import from %s ' % (core.name, sheet))
#                 _imported_variables += variable2import
#     return core
#
#
#
# def import_ic_list(ic_list, variables=variables, v_ref=v_ref, verbose=verbose, drop_empty=drop_empty):
#     """
#     :param ic_list:
#             array, array contains absolute filepath for the cores
#     :param variables:
#     :param v_ref:
#         top, or bottom
#     """
#     logger = logging.getLogger(__name__)
#
#     ic_dict = {}
#     inexisting_ic_list = []
#     for ic_path in ic_list:
#         if verbose:
#             print('Importing data from %s' % ic_path)
#         if not os.path.exists(ic_path):
#             logger.warning("%s does not exists in core directory" % ic_path.split('/')[-1])
#             inexisting_ic_list.append(ic_path.split('/')[-1].split('.')[0])
#         else:
#             ic_data = import_ic_path(ic_path, variables=variables, v_ref=v_ref, drop_empty=drop_empty)
#             if not ic_data.variables():
#                 inexisting_ic_list.append(ic_path.split('/')[-1].split('.')[0])
#                 logger.warning("%s have no properties profile" % (ic_data.name))
#             else:
#                 ic_dict[ic_data.name] = ic_data
#
#     logging.info("Import ice core lists completed")
#     if inexisting_ic_list.__len__() > 0:
#         logger.info("%s core does not exits. Removing from collection" % ', '.join(inexisting_ic_list))
#
#     for ic in inexisting_ic_list:
#         for ic2 in ic_dict.keys():
#             if ic in ic_dict[ic2].collection:
#                 ic_dict[ic2].del_from_collection(ic)
#                 logger.info("remove %s from %s collection" % (ic, ic2))
#     return ic_dict
#
#
# def import_ic_sourcefile(f_path, variables=None, ic_dir=None, v_ref='top', drop_empty=False):
#     """
#     :param filepath:
#             string, absolute path to the file containing either the absolute path of the cores (1 path by line) or the
#             core names (1 core by line). In this last case if core_dir is None core_dir is the directory contianing the
#             file.
#     :param variables:
#
#     :param v_ref:
#         top, or bottom
#     """
#     logger = logging.getLogger(__name__)
#     logger.info('Import ice core from source file: %s' % f_path)
#
#     if ic_dir is not None:
#         with open(f_path) as f:
#             ics = sorted([os.path.join(ic_dir, line.strip()) for line in f if not line.strip().startswith('#')])
#     else:
#         with open(f_path) as f:
#             ics = sorted([line.strip() for line in f if not line.strip().startswith('#')])
#
#     print(ics)
#
#     return import_ic_list(ics, variables=variables, v_ref=v_ref, drop_empty=drop_empty)
#
#
# # read profile
#
# def read_profile(ws_property, variables=None, version=__CoreVersion__, v_ref='top'):
#     """
#     :param ws_property:
#         openpyxl.worksheet
#     :param variables:
#     :param version:
#     :param v_ref:
#         top, or bottom
#     """
#     logger = logging.getLogger(__name__)
#
#     if version == 1:
#         row_data_start = 6
#         row_header = 4
#     elif version == 1.1:
#         row_data_start = 8
#         row_header = 5
#         if ws_property['C4'].value:
#             v_ref = ws_property['C4'].value
#     else:
#         logger.error("ice core spreadsheet version not defined")
#
#     sheet_2_data = {'S_ice': [row_data_start, 'ABC', 'DEFG', 'J'],
#                     'T_ice': [row_data_start, 'A', 'B', 'C'],
#                     'Vf_oil': [row_data_start, 'ABC', 'DEFG', 'H']}
#     # TODO: add other sheets for seawater, sediment, CHla, Phae, stratigraphy
#     #                'stratigraphy': [row_data_start, 'AB', 'C', 'D'],
#     #                'seawater': [row_data_start, 'A', 'DEFGF', 'G']}
#
#     # define section
#     headers_depth = ['y_low', 'y_mid', 'y_sup']
#
#     if not ws_property.title in sheet_2_data:  # if the sheet does not exist, return an empty profile
#         profile = pysic.core.profile.Profile()
#     else:
#         name = ws_property['C1'].value
#         # Continuous profile
#         if sheet_2_data[ws_property.title][1].__len__() == 1:
#             y_mid = np.array([ws_property[sheet_2_data[ws_property.title][1] + str(row)].value
#                               for row in range(sheet_2_data[ws_property.title][0], ws_property.max_row + 1)]).astype(float)
#             y_low = np.nan * np.ones(y_mid.__len__())
#             y_sup = np.nan * np.ones(y_mid.__len__())
#
#         # Step profile
#         elif sheet_2_data[ws_property.title][1].__len__() >= 2:
#             y_low = np.array([ws_property[sheet_2_data[ws_property.title][1][0] + str(row)].value
#                               for row in range(sheet_2_data[ws_property.title][0], ws_property.max_row+1)]).astype(float)
#             y_sup = np.array([ws_property[sheet_2_data[ws_property.title][1][1] + str(row)].value
#                               for row in range(sheet_2_data[ws_property.title][0], ws_property.max_row+1)]).astype(float)
#             y_mid = np.array([ws_property[sheet_2_data[ws_property.title][1][2] + str(row)].value
#                               for row in range(sheet_2_data[ws_property.title][0], ws_property.max_row+1)]).astype(float)
#
#             # check if y_mid are not nan:
#             if np.isnan(y_mid).any():
#                 y_mid = (y_low + y_sup) / 2
#                 logger.info('(%s - %s ) not all y_mid exits, calculating y_mid = (y_low+y_sup)/2'
#                             % (name, ws_property.title))
#             elif np.any(np.abs(((y_low + y_sup)/ 2) - y_mid > 1e-12)):
#                     logger.error('(%s - %s ) y_mid are not mid point between y_low and y_sup. \\'
#                                  'Replacing with y_mid = (y_low+y_sup)/2'
#                                  % (name, ws_property.title))
#             else:
#                 logger.info('(%s - %s ) y_low, y_mid and y_sup read with success'
#                             % (name, ws_property.title))
#
#
#         # read data
#         min_col = sheet_2_data[ws_property.title][2][0]
#         min_col = openpyxl.utils.column_index_from_string(min_col)
#
#         max_col = ws_property.max_column
#
#         min_row = sheet_2_data[ws_property.title][0]
#         max_row = min_row + y_mid.__len__()-1
#
#         # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
#         #          for row in ws_property.iter_cols(min_col, max_col, min_row, max_row)]
#         _data = [[cell.value if isinstance(cell.value, (float, int, str)) else np.nan for cell in row]
#                  for row in ws_property.iter_cols(min_col, max_col, min_row, max_row)]
#
#         data = np.array([y_low, y_mid, y_sup])
#         data = np.vstack([data, np.array(_data)])
#
#         variable_headers = [ws_property.cell(row_header, col).value for col in range(min_col, max_col+1)]
#
#         # # fill missing section with np.nan
#         # if fill_missing:
#         #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
#         #     for ii_idx in idx:
#         #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
#         #         empty += [np.nan] * (variable_headers.__len__()+1)
#         #     data = np.vstack([data, empty])
#
#         # assemble profile dataframe
#         profile = pd.DataFrame(data.transpose(), columns=headers_depth + variable_headers)
#
#         # drop empty variable header
#         if None in profile.columns:
#             profile = profile.drop(labels=[None], axis=1)
#
#         # # convert string to float:
#         # float_header = [h for h in profile.columns if h not in ['comments']]
#         # profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')
#
#         # drop property with all nan value
#         profile = profile.dropna(axis=1, how='all')
#
#         # remove empty line if all element of depth are nan:
#         subset = [col for col in ['y_low', 'y_mid', 'y_sup'] if col in profile.columns]
#         profile = profile.dropna(axis=0, subset=subset, how='all')
#
#         # singularize comments
#         if 'comments' in profile.columns:
#             profile.rename(columns={'comments': "comment"}, inplace=True)
#         # add comment column if it does not exist
#         if 'comment' not in profile.columns:
#             profile['comment'] = None
#         else:
#             profile['comment'] = profile['comment'].astype(str).replace({'nan': None})
#
#         # get all property variable (e.g. salinity, temperature, ...)
#         property = [var for var in profile.columns if var not in ['comment'] + headers_depth]
#
#         # remove subvariable (e.g. conductivity temperature measurement for conductivity
#         property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]
#
#         # set variable to string of property
#         profile['variable'] = [', '.join(property)]*len(profile.index)
#
#         # ice core length
#         try:
#             length = float(ws_property['C2'].value)
#         except:
#             logger.info('(%s) no ice core length' % name)
#             length = np.nan
#         else:
#             if length == 'n/a':
#                 profile['comment'] = 'ice core length not available'
#                 logger.info('(%s) ice core length is not available (n/a)' % name)
#                 length = np.nan
#             elif not isinstance(length, (int, float)):
#                 logger.info('%s ice core length is not a number' % name)
#                 profile['comment'] = 'ice core length not available'
#                 length = np.nan
#         profile['length'] = [length]*len(profile.index)
#
#         # set vertical references
#         profile['v_ref'] = [v_ref]*len(profile.index)
#
#         # set ice core name for profile
#         profile['name'] = [name]*len(profile.index)
#
#         # set columns type
#         col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
#         col_date = ['date']
#         col_float = [h for h in profile.columns if h not in col_string and h not in col_date]
#         profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
#         c_string = [h for h in col_string if h in profile.columns]
#         profile[c_string] = profile[c_string].astype(str).replace({'nan': None})
#
#         profile = pysic.core.profile.Profile(profile)
#         # remove variable not in variables
#         if variables is not None:
#             for property in profile.properties():
#                 if property not in variables:
#                     profile.delete_property(property)
#
#     return profile
#
#
def read_generic_profile(ws_property, property=None, version=__CoreVersion__, reference_d={'ice': ['ice surface', 'down']}, core_length=np.nan, fill_missing=True):
    """
    :param ws_property:
        openpyxl.worksheet, worksheet property to import
    :param property:
        str array, containing property to import. Default None import all properties available
    :param version:
        ice core data version
    :param reference_d:
        dict, contain information of vertical reference system
    :param core_length (optional):
        float, default np.nan. Ice core length
    """
    version_int = version2int(version)

    logger = logging.getLogger(__name__)

    # find last column number with column header and/or subheaders
    n_row = 1  # header row
    n_col_min = 1  # start column
    n_col = n_col_min
    empty_header = 0
    headers = []
    subheaders = []
    units = []
    # stop when header is empty or max column number is reached
    while not empty_header and n_col < ws_property.max_column:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value  # header
            hs_ = ws_property.cell(n_row + 1, n_col).value  # subheader
            headers.append(h_)
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            n_col += 1
        elif isinstance(ws_property.cell(n_row+1, n_col).value, str):
            hs_ = ws_property.cell(n_row + 1, n_col).value   # subheader
            headers.append(h_)
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row+2, n_col).value)
            n_col += 1
        else:
            empty_header = 1
    n_col_max = n_col - 1

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Read center depth for continuous profile and step profile (if available)
    if 'depth center' in headers:
        loc3 = headers.index('depth center')
        headers[loc3] = 'y_mid'

        # look for 1st numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc3 + 1).value, (float, int)):
                min_row_3 = ii_row
                break

        y_mid = np.array(
            [ws_property.cell(row, loc3+1).value for row in range(min_row_3, max_row)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_mid))[::-1] if np.isnan(y_mid[ii])]
        y_mid = np.delete(y_mid, y_nan_loc)
    else:
        y_mid = []
        min_row_3 = np.nan

    # Read lower and upper section depth for step profile or set to nan for continuous profile
    if 'depth 1' in headers and 'depth 2' in headers:
        # find column with depth 1
        loc1 = headers.index('depth 1')
        headers[loc1] = 'y_low'
        loc2 = headers.index('depth 2')
        headers[loc2] = 'y_sup'

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc1 + 1).value, (float, int)):
                min_row_1 = ii_row
                break
            else:
                min_row_2 = np.nan
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc2 + 1).value, (float, int)):
                min_row_2 = ii_row
                break
            else:
                min_row_2 = np.nan

        min_row_12 = min(min_row_1, min_row_2)
        y_low = np.array([ws_property.cell(ii_row, loc1+1).value for ii_row in range(min_row_12, max_row + 1)]).astype(float)
        y_sup = np.array([ws_property.cell(ii_row, loc2+1).value for ii_row in range(min_row_12, max_row + 1)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_low))[::-1] if np.isnan(y_low[ii]) and np.isnan(y_sup[ii])]
        y_low = np.delete(y_low, y_nan_loc)
        y_sup = np.delete(y_sup, y_nan_loc)

        if len(y_mid) == len(y_low) and (min_row_3 == min_row_12 or np.isnan(min_row_3)):
            if np.any(np.abs(((y_low + y_sup) / 2) - y_mid > TOL)):
                logger.error('\t\t%s: y_mid are not mid points of y_low and y_sup. Computing y_mid = (y_low+y_sup)/2'
                             % (ws_property.title))
                y_mid = (y_low + y_sup) / 2
            else:
                logger.info('\t\t%s: y_mid are mid points of y_low and y_sup. Do nothing' % (ws_property.title))
        elif np.isnan(y_mid).any() or min_row_3 != min_row_12:
            y_mid = (y_low + y_sup) / 2
            logger.warning('\t\t%s: not all y_mid exit. Computing y_mid = (y_low+y_sup)/2'
                        % (ws_property.title))
        else:
            y_mid = (y_low + y_sup) / 2
            logger.info('\t\t%s: y_mid does not exist. Computing y_mid = (y_low+y_sup)/2'
                        % (ws_property.title))
    elif 'depth 1' in headers:
        loc1 = headers.index('depth 1')
        headers[loc1] = 'y_low'

        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc1 + 1).value, (float, int)):
                min_row_1 = ii_row
                break
        y_low = np.array([ws_property.cell(ii_row, loc1+1).value for ii_row in range(min_row_1, max_row + 1)]).astype(float)

        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_low))[::-1] if np.isnan(y_low[ii])]
        y_low = np.delete(y_low, y_nan_loc)

        # Fill y_sup
        logger.warning('\t\t%s: y_sup do not exist. Attempting to infer y_sup from y_low'
                       % (ws_property.title))
        if not np.isnan(core_length):
            logger.info('\t\t%s: using core length, y_sup[-1] = l_c' % (ws_property.title))
            y_sup = np.concatenate([y_low[1:], [core_length]])
        elif min_row_1 == min_row_3 and len(y_low) == len(y_mid):
            logger.warning('\t\t%s: using y_mid and y_low,  y_sup[-1] = y_low[-1] + 2 * (y_mid[-1] - y_low[-1])'
                           % (ws_property.title))
            dy = 2 * (y_mid[-1] - y_low[-1])
            y_sup = np.concatenate([y_low[1:], [y_low[-1]+dy]])
        else:
            logger.warning('\t\t%s: core length not available, y_sup[-1] = y_low[-1] + (y_low[-1]-y_low[-2])'
                           % (ws_property.title))
            dy = np.diff(y_low[-2:])
            y_sup = np.concatenate([y_low[1:], [y_low[-1]+dy]])
    elif 'depth 2' in headers:
        loc2 = headers.index('depth 2')
        headers[loc2] = 'y_sup'
        # find first numerical value
        for ii_row in range(min_row, max_row):
            if isinstance(ws_property.cell(ii_row, loc2 + 1).value, (float, int)):
                min_row_2 = ii_row
                break
        y_sup = np.array([ws_property.cell(ii_row, loc2+1).value for ii_row in range(min_row_2, max_row + 1)]).astype(float)

        logger.warning('\t\t%s: y_low do not exist. Attempting to infer y_low from y_sup'% (ws_property.title))
        # discard trailing nan value starting at the end
        y_nan_loc = [ii for ii in np.arange(1, len(y_sup))[::-1] if np.isnan(y_sup[ii])]
        y_sup = np.delete(y_sup, y_nan_loc)

        # Fill y_low
        if min_row_2 == min_row_3 and len(y_low) == len(y_mid):
            dy = 2 * (y_sup[0] - y_mid[0])
            y_sup0 = y_sup[0] - dy
            if y_sup0 >= 0:
                logger.warning('\t\t%s: using y_mid and y_low,  y_low[0] = y_sup[0] - 2 * (y_sup[0]-y_mid[0])'
                               % (ws_property.title))
                y_low = np.concatenate([[y_sup0], y_sup[1:]])
            else:
                logger.info('\t\t%s: For lower y_low, using y_low[0] = 0' % (ws_property.title))
                y_low = np.concatenate([[0], y_sup[:-1]])
        else:
            logger.info('\t\t%s: For lower y_low, using y_low[0] = 0' % (ws_property.title))
            y_low = np.concatenate([[0], y_sup[:-1]])
    else:
        logger.info('\t\t%s: y_low and y_sup do not exist. Creating nan array of same size as y_mid'% (ws_property.title))
        y_low = np.nan * np.ones(y_mid.__len__())
        y_sup = np.nan * np.ones(y_mid.__len__())

    # Check length consistency
    if len(y_low) != len(y_mid):
        logger.error('\t\t%s: y_low/y_sup and y_mid of different size'% (ws_property.title))

    # Read data, according to depth value
    # look up for first numeric or standard entry value
    row_min = min(min_row_12, min_row_3)
    row_max = row_min + len(y_mid) -1

    # Drop column with depth:
    # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
    #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
    _data = [[cell.value for cell in row] for row in ws_property.iter_rows(row_min, row_max, n_col_min, n_col_max)]

    # concatenate header and subheader
    if property is None:
        variable_prefix = ''
        matter = 'sea_ice'
    elif any(map(property.__contains__, ['brine', 'sackhole'])):
        variable_prefix = 'brine_'
        matter = 'brine'
    elif any(map(property.__contains__, ['seawater'])):
        variable_prefix = 'seawater_'
        matter = 'seawater'
    elif any(map(property.__contains__, ['snow'])):
        variable_prefix = 'snow_'
        matter = 'snow'
    else:
        variable_prefix = ''
        matter = 'seaice'

    # concatenate header and subheader to generate header
    subheaders = [sh if sh is not None else '' for sh in subheaders]
    profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid', 'comment']) else h
                       for ii, h in enumerate(headers)]



    profile = pd.DataFrame(_data, columns=profile_headers)

    # Add 'y_mid' column if does not exist, and fill it
    if 'y_mid' not in profile.keys():
        profile['y_mid'] = profile['y_mid'] = (profile.y_low + profile.y_sup)/2
    elif any(profile.y_mid.isna()):
        profile.loc[profile.y_mid.isna(), 'y_mid'] = profile.loc[profile.y_mid.isna(), ['y_low', 'y_sup']].sum(axis=1)/2

    # Add 'comment' column if does not exist and fill it with none value
    if 'comment' not in profile.columns:
        profile['comment'] = None

    # Fill row with np.nan/None for missing section
    if fill_missing:
        for ii_row in np.where(profile.y_sup.values[:-1] - profile.y_low.values[1:] < -TOL)[0]:
            y_low = profile.loc[profile.index == ii_row, 'y_sup'].values
            y_sup = profile.loc[profile.index == ii_row + 1, 'y_low'].values
            empty_row = pd.DataFrame([[np.nan] * len(profile_headers)], columns=profile_headers)
            empty_row['y_low'] = y_low
            empty_row['y_sup'] = y_sup
            empty_row['y_mid'] = (y_low + y_sup) / 2
            profile = pd.concat([profile, empty_row]).reset_index(drop=True)
            logger.info('\t\t%s: filling missing section (%.3f - %.3f) with nan/none values'
                        % (ws_property.title, y_low, y_sup))
        if any(np.isnan(profile.y_mid)):
            profile = profile.drop('y_mid', axis=1)
            profile = profile.sort_values('y_low').reset_index(drop=True)
        else:
            profile = profile.sort_values('y_mid').reset_index(drop=True)

    # Clean profile:
    # Drop empty headers
    if None in profile.columns:
        profile = profile.drop(labels=[None], axis=1)

    # generate list of headers, with str type
    string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

    # generate list of headers, with float type, and set columnt type to float
    float_header = [h for h in profile.columns if h not in string_header]
    profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

    # drop property with all nan value
    profile = profile.dropna(axis=1, how='all')

    # remove empty line if all element of depth are nan:
    subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
    profile = profile.dropna(axis=0, subset=subset, how='all')

    # get profile property from headers (e.g. salinity, temperature, ...)
    property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
    property = [prop.split('_')[0] for prop in property]
    property = list(set(property))

    # remove property parameters (e.g. conductivity temperature measurement for conductivity
    property = [prop for prop in property if prop not in inverse_dict(prop_parameter_dict)]

    # set property
    # TODO: enter property by row
    # TODO: do not remove property column if there is a sample ID
    profile['variable'] = [', '.join(property)] * len(profile.index)

    # set vertical references
    # TODO: improve vertical reference
    if reference_d['ice'][0] == 'ice surface':
        v_ref = 'top'
    elif reference_d['ice'][0] == 'ice/water interface':
        v_ref = 'bottom'
    else:
        logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
    profile['v_ref'] = [v_ref] * len(profile.index)

    # set columns type
    col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
    col_date = ['date']
    col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
    col_string = col_string + [h for h in profile.columns if 'ID' in h]
    profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
    c_string = [h for h in col_string if h in profile.columns]
    profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

    profile = pysic.core.profile.Profile(profile)
    # remove variable not in variables
    if variables is not None:
        for property in profile.properties():
            if property not in variables:
                profile.delete_property(property)

    return profile


def read_snow_profile(ws_property, variables=None, version=__CoreVersion__, reference_d={'ice': ['ice surface', 'down']}):
    """
    :param ws_property:
        openpyxl.worksheet
    :param variables:
    :param version:
    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)


    # Read Salinity block
    # read headers:
    n_row = 1
    n_col = 1
    n_col_min = n_col
    cell_flag = 2
    headers = []
    subheaders = []
    units = []
    while cell_flag >= 1:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value
            headers.append(ws_property.cell(n_row, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(ws_property.cell(n_row + 1, n_col).value)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            cell_flag = 2
            n_col += 1
        elif isinstance(ws_property.cell(n_row+1, n_col).value, str):
            headers.append(h_)
            subheaders.append(ws_property.cell(n_row+1, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            units.append(ws_property.cell(n_row+2, n_col).value)
            n_col += 1
        else:
            cell_flag = 0
            cell_mark = n_col + 1
    n_col_max = n_col - n_col_min

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Check for step or continuous profiles:
    if 'depth center' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth center'][0] + 1
        headers[loc1-1] = 'y_mid'
        y_mid = np.array(
            [ws_property.cell(row, loc1).value for row in range(min_row, max_row)]).astype(float)

        # discard trailing nan value from the end up
        # find nan value in y_low and y_sup
        y_nan_loc = np.where(np.isnan(y_mid))[0]
        # discard trailing nan value starting at the end
        if len(y_nan_loc) > 0 and len(y_mid) > 1 and y_nan_loc[-1] == len(y_mid)-1:
            y_nan_loc = [len(y_mid)-1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if val == y_nan_loc[::-1][ii]-1]
            y_nan_loc = y_nan_loc[::-1]
            y_mid = [y for ii, y in enumerate(y_mid) if ii not in y_nan_loc]

    if 'depth 1' in headers and 'depth 2' in headers:
        step_flag = 1

        # find column with depth 1
        # TODO find a better way to find the location
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 1'][0]+1
        headers[loc1 - 1] = 'y_low'
        loc2 = [ii for ii, h in enumerate(headers) if h == 'depth 2'][0]+1
        headers[loc2 - 1] = 'y_sup'
        # TODO: remove 'depth center'
        y_low = np.array([ws_property.cell(row, loc1).value for row in range(min_row, max_row + 1)]).astype(float)
        y_sup = np.array([ws_property.cell(row, loc2).value for row in range(min_row, max_row + 1)]).astype(float)

        # discard trailing nan value from the end up
        # find nan value in y_low and y_sup
        y_nan_loc = [ii for ii in np.where(np.isnan(y_sup))[0] if ii in np.where(np.isnan(y_low))[0]]

        # discard trailing nan value starting at the end
        if len(y_sup) > 0 and len(y_nan_loc) > 0 and y_nan_loc[-1] == len(y_sup)-1:
            y_nan_loc = [len(y_sup)-1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if val == y_nan_loc[::-1][ii]-1]
            y_nan_loc = y_nan_loc[::-1]
            y_low = np.array([y for ii, y in enumerate(y_low) if ii not in y_nan_loc])
            y_sup = np.array([y for ii, y in enumerate(y_sup) if ii not in y_nan_loc])

            # TODO: replace missing y_low and y_sup with y_mid if possible

        if len(y_low) == 0:
            profile = pysic.core.profile.Profile()
            y_mid = []
        elif 'y_mid' in headers:
            if np.isnan(y_mid).any() or len(y_mid) == 0:
                y_mid = (y_low + y_sup) / 2
                logger.info('(%s ) not all y_mid exits, calculating y_mid = (y_low+y_sup)/2'
                            % (ws_property.title))
            elif np.any(np.abs(((y_low + y_sup) / 2) - y_mid > 1e-12)):
                logger.error('(%s ) y_mid are not mid point between y_low and y_sup. \\'
                             'Replacing with y_mid = (y_low+y_sup)/2'
                             % (ws_property.title))
                y_mid = (y_low + y_sup) / 2
            else:
                logger.info('(%s ) y_low, y_mid and y_sup read with success'
                            % (ws_property.title))
        else:
            y_mid = (y_low + y_sup) / 2
    elif 'depth 1' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 1'][0]+1
        headers[loc1 - 1] = 'y_low'
        # TODO : fill y_sup
    elif 'depth 2' in headers:
        loc1 = [ii for ii, h in enumerate(headers) if h == 'depth 2'][0]+1
        headers[loc1 - 1] = 'y_sup'
        # TODO : fill y_low
    # Continuous profile
    else:
        y_low = np.nan * np.ones(y_mid.__len__())
        y_sup = np.nan * np.ones(y_mid.__len__())

    # Read data:
    # look up for first numeric or standard entry value
    # n_row_min = 4

    if len(y_mid) > 0:
        n_row_min = 1
        n_col_min = 1
        while not isinstance(ws_property.cell(n_row_min, n_col_min).value, (float, int)):
            n_row_min += 1
            if n_row_min > 1000:
                break
        n_row_max = n_row_min + len(y_mid) - 1

        # Drop column with depth:
        # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
        #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
        _data = [[cell.value for cell in row] for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]

        # TODO:  fill missing section with np.nan
        # if fill_missing:
        #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
        #     for ii_idx in idx:
        #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
        #         empty += [np.nan] * (variable_headers.__len__()+1)
        #     data = np.vstack([data, empty])

        # concatenate header and subheader
        if variables is None:
            variable_prefix = ''
            phase = 'N/A'
        elif any(map(variables.__contains__, ['brine', 'sackhole'])):
            variable_prefix = 'brine_'
            phase = 'brine'
        elif any(map(variables.__contains__, ['seawater'])):
            variable_prefix = 'seawater_'
            phase = 'seawater'
        elif any(map(variables.__contains__, ['snow'])):
            variable_prefix = 'snow_'
            phase = 'snow'
        else:
            variable_prefix = ''
            phase = 'seaice'
        subheaders = [sh if sh is not None else '' for sh in subheaders]
        profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid']) else h
                           for ii, h in enumerate(headers)]
        # TODO: double header for dataframe with header and subheader

        profile = pd.DataFrame(_data, columns=profile_headers)
        if 'y_mid' not in profile.keys():
            profile['y_mid'] = y_mid

        # drop empty variable header
        if None in profile.columns:
            profile = profile.drop(labels=[None], axis=1)

        # sample ID columns is string:
        string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

        # convert string to float:
        float_header = [h for h in profile.columns if h not in string_header]
        profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

        # drop property with all nan value
        profile = profile.dropna(axis=1, how='all')

        # remove empty line if all element of depth are nan:
        subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
        profile = profile.dropna(axis=0, subset=subset, how='all')

        # singularize comments
        if 'comments' in profile.columns:
            profile.rename(columns={'comments': "comment"}, inplace=True)
        # add comment column if it does not exist
        if 'comment' not in profile.columns:
            profile['comment'] = None
        else:
            profile['comment'] = profile['comment'].astype(str).replace({'nan': None})

        # get all property variable (e.g. salinity, temperature, ...)
        property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
        property = [prop.split('_')[0] for prop in property]
        property = list(set(property))

        # remove subvariable (e.g. conductivity temperature measurement for conductivity
        property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

        # set variable to string of property
        profile['variable'] = [', '.join(property)] * len(profile.index)

        # set vertical references
        # TODO: improve vertical reference
        if reference_d['snow'][0] == 'ice surface':
            v_ref = 'bottom'
        elif reference_d['snow'][0] == 'snow interface':
            v_ref = 'bottom'
        else:
            logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
        profile['v_ref'] = [v_ref] * len(profile.index)

        # set columns type
        col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
        col_date = ['date']
        col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
        col_string = col_string + [h for h in profile.columns if 'ID' in h]
        profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
        c_string = [h for h in col_string if h in profile.columns]
        profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

        profile = pysic.core.profile.Profile(profile)
        # remove variable not in variables
        if variables is not None:
            for property in profile.properties():
                if property not in variables:
                    profile.delete_property(property)

    if variables is not None and 'salinity' not in variables:
        profile_S = pysic.core.profile.Profile()
    else:
        profile_S = profile

    cell_mark = 26

    del profile
    # Read Temperature block
    # read headers:
    n_row = 1
    n_col = cell_mark
    n_col_min = n_col
    cell_flag = 2
    headers = []
    subheaders = []
    units = []
    while cell_flag >= 1:
        if isinstance(ws_property.cell(n_row, n_col).value, str):
            h_ = ws_property.cell(n_row, n_col).value
            headers.append(ws_property.cell(n_row, n_col).value)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(ws_property.cell(n_row + 1, n_col).value)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            cell_flag = 2
            n_col += 1
        elif isinstance(ws_property.cell(n_row + 1, n_col).value, str):
            headers.append(h_)
            hs_ = ws_property.cell(n_row + 1, n_col).value
            subheaders.append(hs_)
            units.append(ws_property.cell(n_row + 2, n_col).value)
            n_col += 1
        else:
            cell_flag = 0
            cell_mark = n_col + 1
    n_col_max = n_col

    if ws_property.max_row < MAX_ROW:
        max_row = ws_property.max_row
    else:
        max_row = MAX_ROW
    min_row = 4

    # Check for step or continuous profiles:
    #loc1 = [ii for ii, h in enumerate(headers) if h == 'depth'][0] + 1
    headers[loc1 - 1] = 'depth'
    subheaders[loc1 - 1] = 'y_mid'
    headers = ['y_mid', 'temperature', 'temperature']
    subheaders = ['', 'value', 'quality']
    y_mid = np.array(
        [ws_property.cell(row, n_col_min).value for row in range(min_row, max_row)]).astype(float)

    # discard trailing nan value from the end up
    # find nan value in y_low and y_sup
    y_nan_loc = np.where(np.isnan(y_mid))[0]
    # discard trailing nan value starting at the end
    if len(y_nan_loc) > 0 and len(y_mid) > 1 and y_nan_loc[-1] == len(y_mid) - 1:
        y_nan_loc = [len(y_mid) - 1] + [val for ii, val in enumerate(y_nan_loc[-2::-1]) if
                                        val == y_nan_loc[::-1][ii] - 1]
        y_nan_loc = y_nan_loc[::-1]
        y_mid = [y for ii, y in enumerate(y_mid) if ii not in y_nan_loc]

    y_low = np.nan * np.ones(y_mid.__len__())
    y_sup = np.nan * np.ones(y_mid.__len__())

    # Read data:
    # look up for first numeric or standard entry value
    # n_row_min = 4

    if len(y_mid) > 0:
        n_row_min = 1
        while not isinstance(ws_property.cell(n_row_min, n_col_min).value, (float, int)):
            n_row_min += 1
            if n_row_min > 1000:
                break
        n_row_max = n_row_min + len(y_mid) - 1

        # Drop column with depth:
        # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
        #                   for row in ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max)]
        _data = [[cell.value for cell in row] for row in
                 ws_property.iter_rows(n_row_min, n_row_max, n_col_min, n_col_max-1)]

        # TODO:  fill missing section with np.nan
        # if fill_missing:
        #     idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
        #     for ii_idx in idx:
        #         empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
        #         empty += [np.nan] * (variable_headers.__len__()+1)
        #     data = np.vstack([data, empty])

        # concatenate header and subheader
        if variables is None:
            variable_prefix = ''
            phase = 'N/A'
        elif any(map(variables.__contains__, ['brine', 'sackhole'])):
            variable_prefix = 'brine_'
            phase = 'brine'
        elif any(map(variables.__contains__, ['seawater'])):
            variable_prefix = 'seawater_'
            phase = 'seawater'
        elif any(map(variables.__contains__, ['snow'])):
            variable_prefix = 'snow_'
            phase = 'snow'
        else:
            variable_prefix = ''
            phase = 'seaice'
        subheaders = [sh if sh is not None else '' for sh in subheaders]
        profile_headers = [variable_prefix + h + '_' + subheaders[ii] if (
                    len(subheaders[ii]) > 1 and h not in ['y_low', 'y_sup', 'y_mid']) else h
                           for ii, h in enumerate(headers)]
        # TODO: double header for dataframe with header and subheader

        profile = pd.DataFrame(_data, columns=profile_headers)
        if 'y_mid' not in profile.keys():
            profile['y_mid'] = y_mid

        # drop empty variable header
        if None in profile.columns:
            profile = profile.drop(labels=[None], axis=1)

        # sample ID columns is string:
        string_header = ['comment'] + [h for h in profile.columns if 'sample ID' in h or 'ID' in h]

        # convert string to float:
        float_header = [h for h in profile.columns if h not in string_header]
        profile[float_header] = profile[float_header].apply(pd.to_numeric, errors='coerce')

        # drop property with all nan value
        profile = profile.dropna(axis=1, how='all')

        # remove empty line if all element of depth are nan:
        subset = [col for col in ['y_low', 'y_sup', 'y_mid'] if col in profile.columns]
        profile = profile.dropna(axis=0, subset=subset, how='all')

        # singularize comments
        if 'comments' in profile.columns:
            profile.rename(columns={'comments': "comment"}, inplace=True)
        # add comment column if it does not exist
        if 'comment' not in profile.columns:
            profile['comment'] = None
        else:
            profile['comment'] = profile['comment'].astype(str).replace({'nan': None})

        # get all property variable (e.g. salinity, temperature, ...)
        property = [var for var in profile.columns if var not in ['comment', 'y_low', 'y_sup', 'y_mid']]
        property = [prop.split('_')[0] for prop in property]
        property = list(set(property))

        # remove subvariable (e.g. conductivity temperature measurement for conductivity
        property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

        # set variable to string of property
        profile['variable'] = [', '.join(property)] * len(profile.index)

        # set vertical references
        # TODO: improve vertical reference
        if reference_d['snow'][0] == 'ice surface':
            v_ref = 'bottom'
        elif reference_d['snow'][0] == 'snow interface':
            v_ref = 'bottom'
        else:
            logger.error(ws_property.title + ' - Vertical references not set or not yet handled')
        profile['v_ref'] = [v_ref] * len(profile.index)

        # set columns type
        col_string = ['comment', 'v_ref', 'name', 'profile', 'variable']
        col_date = ['date']
        col_float = [h for h in profile.columns if h not in col_string and h not in col_date and 'ID' not in h]
        col_string = col_string + [h for h in profile.columns if 'ID' in h]
        profile[col_float] = profile[col_float].apply(pd.to_numeric, errors='coerce')
        c_string = [h for h in col_string if h in profile.columns]
        profile[c_string] = profile[c_string].astype(str).replace({'nan': None})

        profile = pysic.core.profile.Profile(profile)
        # remove variable not in variables
        if variables is not None:
            for property in profile.properties():
                if property not in variables:
                    profile.delete_property(property)

        if variables is not None and 'temperature' not in variables:
            profile = profile_S
        else:
            profile = profile.append(profile_S)
        profile.reset_index(drop=True, inplace=True)
    else:
        profile = pysic.core.profile.Profile()
    return profile

# create list or source
def list_folder(dirpath, fileext='.xlsx', level=0):
    """
    list all files with specific extension in a directory

    :param dirpath: str; directory to scan for ice core
    :param fileext: str, default .xlsx; file extension for ice core data
    :param level: numeric, default 0; level of recursitivy in directory search
    :return ic_list: list
        list of ice core path
    """

    if not fileext.startswith('.'):
        fileext = '.' + fileext

    _ics = []

    logger = logging.getLogger(__name__)

    def walklevel(some_dir, level=level):
        some_dir = some_dir.rstrip(os.path.sep)
        assert os.path.isdir(some_dir)
        num_sep = some_dir.count(os.path.sep)
        for root, dirs, files in os.walk(some_dir):
            yield root, dirs, files
            num_sep_this = root.count(os.path.sep)
            if num_sep + level <= num_sep_this:
                del dirs[:]

    for dirName, subdirList, fileList in walklevel(dirpath, level=level):
        _ics.extend([dirName + '/' + f for f in fileList if f.endswith(fileext)])

    ics_set = set(_ics)
    logger.info("Found %i ice core datafile in %s" % (ics_set.__len__(), dirpath))

    return ics_set


def list_corename(dirpath, fileext='.xlsx', level=0):
    """
    list all files with specific extension in a directory

    :param dirpath: str; directory to scan for ice core
    :param fileext: str, default .xlsx; file extension for ice core data
    :param level: numeric, default 0; level of recursitivy in directory search
    :return ic_list: list
        list of ice core path
    """
    _ics = []

    if not fileext.startswith('.'):
        fileext = '.' + fileext

    logger = logging.getLogger(__name__)

    ics_set = list_folder(dirpath, fileext, level)
    ics_name = []
    for ic in sorted(ics_set):
        name = ic.split('/')[-1].split(fileext)[0]
        print(name)
        ics_name.append(name)
    return ics_name
#
#
# def list_ic_path(dirpath, fileext):
#     """
#     list all files with specific extension in a directory
#
#     :param dirpath: str
#     :param fileext: str
#     :return ic_list: list
#         list of ice core path
#     """
#     logger = logging.getLogger(__name__)
#
#     ics_set = list_folder(dirpath=dirpath, fileext=fileext)
#     ic_paths_set = set([os.path.join(os.path.realpath(dirpath), f) for f in ics_set])
#     return ic_paths_set
#
#
# def make_ic_sourcefile(dirpath, fileext, source_filepath=None):
#     """
#     list all files with specific extension in a directory
#
#     :param dirpath: str
#     :param fileext: str
#     :return source_file: str
#         filepath to the text file containing ice core filepath with absolute path.
#     """
#     logger = logging.getLogger(__name__)
#
#     ic_paths_set = list_ic_path(dirpath, fileext)
#
#     if source_filepath is None:
#         source_filepath = os.path.join(os.path.realpath(dirpath), 'ic_list.txt')
#
#     with open(source_filepath, 'w') as f:
#         for ic_path in ic_paths_set:
#             f.write(ic_path + "\n")
#
#     return source_filepath
#
# #
# # def read_ic_list(file_path):
# #     """
# #
# #     :param file_path:
# #     :return:
# #     """
# #
# #     with open(file_path) as f:
# #         data = [list(map(int, row.split())) for row in f.read().split('\n\n')]
# #
# #     return ic_list
#
# TODO: cleaner function to (1) remove trailing and heading space in instrument

def version2int(version):
    version_int =[int(v) for v in version.split('.')]
    if len(version_int) < 3:
        version_int.append(0)
    return version_int


def update_data_spreadsheet(ic_path, backup=True):
    """
    :param ic_path:
        path; Filepath to the data spreadsheet to update
    :param backup:
        boolean, default True; Make a backup copy in a subfolder
    """
    import shutil

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet
    try:
        ws_summary = wb['metadata-coring']  # load the data from the summary sheet
    except KeyError:
        ws_summary = wb['metadata-station']  # load the data from the summary sheet
    else:
        pass
    version = ws_summary['C1'].value
    version_int = version2int(version)

    # Backup old version
    if backup:
        backup_dir = os.path.join(os.path.dirname(ic_path), 'bkp-ic_version_' + str(version))
        ic_bkp = os.path.join(backup_dir, os.path.basename(ic_path))
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        shutil.copy(ic_path, ic_bkp)

    while version_int[1] < 5 and version_int[2] < 6:
        # Update to 1.4.6
        if version_int[2] < 6:
            version_int[2] = 6
            wb['metadata-station']['D42'].value = None
            wb['metadata-station']['D43'].value = None
            wb['metadata-station']['D44'].value = None
            wb['metadata-station']['B42'].value = 'm/s'
            wb['metadata-station']['B43'].value = 'degree'
            wb['metadata-station']['B44'].value = '/8'
            wb['metadata-station']['B45'].value = 'μmol/m2s'
            wb['metadata-station']['B46'].value = 'μmol/m2s'

            # style B42:B46 not bold, align right
            _col = openpyxl.utils.cell.column_index_from_string('B')
            for ii_row in range(42, 46+1):
                wb['metadata-station'].cell(ii_row, _col).font = unit_style
                wb['metadata-station'].cell(ii_row, _col).alignment = unit_alignment

            # Border formatting
            # C9:H9, bottom border sytle: hair

            for _row in openpyxl.utils.cell.rows_from_range('C9:H9'):
                for _cell in _row:
                    wb['metadata-station'][_cell].border = b_border

            # C14:C16, bottom border sytle: hair
            for _row in openpyxl.utils.cell.rows_from_range('C14:C16'):
                for _cell in _row:
                    wb['metadata-station'][_cell].border = b_border
            # C16:D16, bottom border sytle: None
            for _row in openpyxl.utils.cell.rows_from_range('C16:D16'):
                for _cell in _row:
                    wb['metadata-station'][_cell].border = no_border

            # add specific vertical border
            wb['metadata-station']['F9'].border = bl_border
            wb['metadata-station']['D14'].border = bl_border
            wb['metadata-station']['D16'].border = l_border

    wb['metadata-station']['C1'] = __CoreVersion__
    wb.save(ic_path)
    wb.close()

    # while version_int[1] < 3 and version[2] == 0 :
    #     # backup old version
    #     if backup:
    #         backup_dir = os.path.join(os.path.dirname(ic_path), 'bkp-ic_version_' + str(version))
    #         ic_bkp = os.path.join(backup_dir, os.path.basename(ic_path))
    #
    #         if not os.path.exists(backup_dir):
    #             os.makedirs(backup_dir)
    #         shutil.move(ic_path, ic_bkp)
    #
    #     if version_int[1] == 2:
    #         old_version = version
    #         version = '1.3'
    #         ws_summary['C1'] = version
    #
    #         if "TEMP" in wb.sheetnames:
    #             ws = wb["TEMP"]
    #
    #             # add reference row=4
    #             ws.insert_rows(3)
    #             ws['A3'] = 'value'
    #             ws['B3'] = 'value'
    #             ws['C3'] = '-'
    #
    #         if "Density-volume" in wb.sheetnames:
    #             ws = wb["Density-volume"]
    #
    #             # add reference row=4
    #             ws.insert_rows(3)
    #             ws['A3'] = 'value'
    #             ws['B3'] = 'value'
    #             ws['C2'] = 'field'
    #             ws['C3'] = 'sample ID'
    #             ws['D3'] = 'value'
    #             ws['E2'] = 'thickness'
    #             ws['E3'] = 'value1'
    #             ws['F2'] = 'thickness'
    #             ws['F3'] = 'value2'
    #             ws['G2'] = 'thickness'
    #             ws['G3'] = 'value3'
    #             ws['H2'] = 'thickness'
    #             ws['H3'] = 'value4'
    #             ws['I2'] = 'thickness'
    #             ws['I3'] = 'average'
    #             ws['J3'] = 'value'
    #             ws['K2'] = 'density'
    #             ws['K3'] = 'value'
    #             ws['L3'] = '-'
    #
    #         if "Density-densimetry" in wb.sheetnames:
    #             ws = wb["Density-densimetry"]
    #
    #             # add reference row=4
    #             ws.insert_rows(3)
    #             ws['A3'] = 'value'
    #             ws['B3'] = 'value'
    #             ws['C2'] = 'field'
    #             ws['C3'] = 'sample ID'
    #             ws['D3'] = 'value'
    #             ws['E3'] = 'value'
    #             ws['F3'] = 'value'
    #             ws['G3'] = 'value'
    #             ws['H3'] = '-'
    #
    #         if "TEX" in wb.sheetnames:
    #             ws = wb["TEX"]
    #
    #             # add reference row=4
    #             ws.insert_rows(3)
    #             ws['A3'] = 'value'
    #             ws['B3'] = 'value'
    #             ws['C2'] = 'depth center'
    #             ws['C3'] = 'value'
    #             ws['D3'] = 'value'
    #             ws['F3'] = '-'
    #
    #         if "snow" in wb.sheetnames:
    #             ws = wb["snow"]
    #
    #             # add reference row=4
    #             ws.insert_rows(1)
    #             ws['A1'] = 'Reference'
    #             ws['B1'] = 'zero vertical'
    #             ws['C1'] = 'ice surface'
    #             ws['C2'] = 'field'
    #             ws['C3'] = 'sample ID'
    #             ws['D1'] = 'direction'
    #             ws['E1'] = 'up'
    #
    #         if "metadata-core" in wb.sheetnames:
    #             ws = wb["metadata-core"]
    #
    #             max_row = ws.max_row
    #             n_row = 1
    #             while n_row <= max_row:
    #                 if ws.cell(n_row, 1).value != 'DATA VERSION':
    #                     n_row += 1
    #                 else:
    #                     break
    #
    #             if n_row-1 == max_row:
    #                 n_row = 12
    #                 while ws.cell(n_row, 1).value is not None:
    #                     n_row += 1
    #
    #                 ws.cell(n_row+1, 1).value = 'DATA VERSION'
    #                 ws.cell(n_row+2, 1).value = 1.0
    #                 ws.cell(n_row+2, 2).value = ws_summary['C12'].value
    #                 ws.cell(n_row+2, 3).value = 'initial data entry'
    #
    # # TODO: update to version 1.3.1
    #
    # if version == '2.3':
    #     new_version = '1.3.1'
    #     print('Update to future version ' + new_version)
    #     ws_summary['C1'] = new_version
    #
    #     if "metadata-core" in wb.sheetnames:
    #         ws = wb["metadata-core"]
    #
    #         max_row = ws.max_row
    #         n_row = 1
    #         while n_row <= max_row:
    #             if ws.cell(n_row, 1).value == 'DATA VERSION':
    #                 n_row += 1
    #             else:
    #                 break
    #
    #         if n_row-1 == max_row:
    #             n_row = 12
    #             while ws.cell(n_row, 1).value is not None:
    #                 n_row += 1
    #
    #             ws.cell(n_row+1, 1).value = 'DATA VERSION'
    #             ws.cell(n_row+2, 1).value = 2.0
    #             ws.cell(n_row+2, 2).value = datetime.date.today()
    #             ws.cell(n_row+2, 3).value = 'Typo in data corrected. Updated spreadsheet to v. 1.3.1'
    #
    #     if 'ECO' in wb.sheetnames:
    #         ws = wb["ECO"]
    #         if ws['F2'].value == 'Nutrient':
    #             ws['F2'] = 'nutrient'
    #
    #         # TODO: update nutrient with 5 nutrient
    #
    #     if 'snow' in wb.sheetnames:
    #         ws = wb["snow"]
    #         ws['L2'].value = 'snow'
    #         ws['L3'].value = 'weight'
    #         ws['M3'].value = 'density'
    #         ws['M3'].value = 'snow'
    #         ws['M3'].value = 'density'
    #         ws['N3'].value = 'comment'
    #         ws['N3'].value = '-'
    #         ws['N4'].value = '-'
    #         ws['O3'].value = 'temperature'
    #         ws['P3'].value = 'temperature'
    #         ws['P3'].value = 'value'
    #
    # if v_release == 1:
    #     if v_major == 3:
    #         if v_minor == 1:
    #             print('Latest version')
    #         else:
    #             print('Future version does not exist.')
    #     else:
    #         print('Future version does not exist')
    # else:
    #     print('Future version does not exist.')


def parse_datetimetz(c_date_v, c_hour_v, c_tz_v):
    """
    :param c_date_v:
        string, value of date cell
    :param c_hour_v:
        string, value of hour cell
    :param c_tz_v:
        string, value of timezone cell
    :return:
        datetime.date or datetime.datetime (aware or naive) as function of the input
    """
    import pytz

    logger = logging.getLogger(__name__)

    if isinstance(pd.to_datetime(c_date_v, format='%Y-%m-%d').date(), datetime.date):
        _d = pd.to_datetime(c_date_v)
        _d = _d.date()

        # time
        if pd.to_datetime(c_hour_v, format='%H:%M:%S') is None:
            logger.info("\ttime and timezone unavailable")
            return _d
        elif isinstance(pd.to_datetime(c_hour_v, format='%H:%M:%S').time(), datetime.time):
            _t = pd.to_datetime(c_hour_v, format='%H:%M:%S').time()
            _dt = datetime.datetime.combine(_d, _t)
        else:
            logger.info("\ttime and timezone unavailable")
            return _d

        # timezone
        if c_tz_v is not None:
            # format 'Country/City'
            if c_tz_v in pytz.all_timezones :
                _tz = pytz.all_timezones
                _dt.replace(tzinfo=dateutil.tz.gettz(_tz))
            elif c_tz_v.split(' ')[0] in pytz.all_timezones:
                _tz = pytz.all_timezones
                _dt.replace(tzinfo=dateutil.tz.gettz(_tz))
            elif c_tz_v.startswith('UTC'):
                _tz = c_tz_v.split(' ')[0]
                _utc_offset = _tz.split('UTC')[-1]
                # convert +HHMM or +HH:MM to timeoffset
                if ':' in _utc_offset:
                    _tz_h = int(_utc_offset.split(':')[0])
                    _tz_m = int(_utc_offset.split(':')[1])
                elif len(_utc_offset[1:]) == 4:
                    _tz_h = int(_utc_offset[0:3])
                    _tz_m = int(_utc_offset[3:])
                elif len(_utc_offset[1:]) == 3:
                    _tz_h = int(_utc_offset[0:2])
                    _tz_m = int(_utc_offset[2:])
                else:
                    _tz_h = int(_utc_offset[0:2])
                    _tz_m = 0
                _utc_offset = datetime.timedelta(hours=_tz_h, minutes=_tz_m)
                _tz = datetime.timezone(_utc_offset, name=_tz)

                # offset is in seconds
                _dt = _dt.replace(tzinfo=_tz)
            else:
                _tz = None
                # TODO: guess timezone base of date and location
                logger.info("\t Timezone unavailable.")
        else:
            _tz = None
            # TODO: guess timezone base of date and location
            logger.info("\tTimezone unavailable.")
    else:
        logger.warning("\tDate unavailable")
        _dt = None
    return _dt


def parse_coordinate(v_cell_deg, v_cell_min, v_cell_sec):
    """
    :param v_cell_deg:
        float or int, value of degree cell
    :param v_cell_min:
        float or int, value of minute cell
    :param v_cell_sec:
        float or int, value of second cell
    :return:
        float, coordinate in degree
    """
    logger = logging.getLogger(__name__)

    if isinstance(v_cell_deg, (float, int)):
        degree = v_cell_deg
        if isinstance(v_cell_min, (float, int)):
            minute = v_cell_min
            if isinstance(v_cell_sec, (float, int)):
                second = v_cell_sec
            else:
                second = 0
        else:
            minute = 0
            second = 0
        coordinate = degree + minute / 60 + second / 3600
    else:
        coordinate = np.nan

    # Add control for
    return coordinate
