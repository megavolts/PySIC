#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
icecoredata.py: icecoredata is a toolbox to import ice core data file from xlsx spreadsheet. Xlsx spreadsheet should be
formatted according to the template provided by the Sea Ice Group of the Geophysical Institute of University of Alaska,
Fairbanks.
"""
__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "2.0.0"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "marc.oggier@gi.alaska.edu"
__status__ = "development"
__date__ = "2014/11/25"

__all__ = ['Core', 'CoreSet']

import numpy as np
import logging
from decimal import Decimal
from seaice import icdtools

#from . import icdtools

LOG_FILENAME='ICimport.log'
LOG_LEVELS = {'debug': logging.DEBUG,
              'info': logging.INFO,
              'warning': logging.WARNING,
              'error': logging.ERROR,
              'critical': logging.CRITICAL}

log_level = 'warning'
missvalue = float('NaN')

class profile:
    def __init__(self, x, y, note=None, length=None):
        self.x = x
        self.y = y
        self.note = note
        if length is None:
            self.length = x[-1]=x[0]
        else:
            self.length = length

class CoreV2:
    def __init__(self, name, date, location, ice_thickness, snow_thickness):
        self.name = name
        self.date = date
        self.location = location
        self.corenames = [name]
        self.temperature = None
        self.salinity = None
        self.sediment = None
        self.chla = None
        self.D18O = None
        self.D2H = None
        self.pfaeo = None

    def add_corenames(self, corename):
        if corename not in self.corenames:
            self.corenames.append(corename)

    def salinity (self, x, y, note = None, length = None):
        self.salinity = profile(x, y, note, length)

class CoreProperties:
    def __init__(self, name, date, location, ice_thickness, snow_depth):
        self.name = name
        self.date = date
        self.location = location
        self.corenames = [name]
        self.ice_thickness = ice_thickness
        self.snow_depth = snow_depth
        self.t_air = None
        self.t_snow = None
        self.t_ice0 = None
        self.t_water = None
        self.d180 = None
        self.dO2 = None
        self.Chla = None
        self.Phaeo = None
        self.sediment = None
        self.t = None
        self.t_length = None
        self.s = None
        self.s_length = None
        self.comment = None

    def add_corenames(self, corename):
        if corename not in self.corenames:
            self.corenames.append(corename)
        self.corenames = [sorted(self.corenames)[0]]

class Core:
    def __init__(self, name, date, location, ice_thickness, snow_depth):
        self.name = name
        self.date = date
        self.location = location
        self.corenames = [name]
        self.ice_thickness = ice_thickness
        self.snow_depth = snow_depth
        self.t_air = None
        self.t_snow = None
        self.t_ice0 = None
        self.t_water = None

    def add_corenames(self, corename):
        if corename not in self.corenames:
            self.corenames.append(corename)

    def error(message):
        import sys
        sys.stderr.write("error: %s\n" % message)
        sys.exit(1)

    def t_profile(self, t, yt, lengtht, t_note):
        self.t = t
        self.yt = yt
        self.lengtht = lengtht
        self.t_note = t_note

    def s_profile(self, s, ys, lengths, s_note):
        self.s = s
        self.ys = ys
        self.lengths = lengths
        self.s_note = s_note

    def sediment_profile(self, s, ys, lengths, s_note):
        self.sediment_profile = s
        self.sedimenty  = ys
        self.lengthsediment = lengths
        self.sediment_note = s_note

    def algae_profile(self, s, ys, lengths, s_note):
        self.algae_profile = s
        self.algae_y  = ys
        self.length_algae = lengths
        self.algae_note = s_note

    def __getattr__(self, key):
        return None

    def __getstate__(self):
        return self.__dict__

    def __setstate__(self, d):
        self.__dict__.update(d)

def getfilepath(datadir, dataext, subdir='no'):
    import os
    filepath = []
    if subdir == 'yes':
        for path, subdirs, files in os.walk(datadir):
            subdirs.sort()
            files.sort()
            for name in files:
                if name.endswith(dataext):
                    f = os.path.join(path, name)
                    filepath.append(f)
    else:
        files = os.listdir(datadir)
        for name in files:
            if name.endswith(dataext):
                f = os.path.join(datadir, name)
                filepath.append(f)
    return filepath


def generatesrc(datadir, dataext):
    list = getfilepath(datadir, dataext)
    with open(datadir+'/ic_list.txt','w') as f:
        for ii in range(0, len(list)):
            f.write(list[ii]+"\n")

## import specific ice core
def importprop(ic_path):
    import openpyxl
    import datetime

    wb = openpyxl.load_workbook(filename=ic_path, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()

    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    # name
    temp_cell = ws_summary['C21']
    ic_name = temp_cell.value
    print(ic_name)


    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    logging.info('Processing ' + ic_name + '...')

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30)+datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value


    # location
    temp_cell = ws_summary['C5']
    ic_loc = temp_cell.value

    # snow thickness
    temp_cell = ws_summary['C9']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_snow_depth = missvalue
    else:
        ic_snow_depth = temp_cell.value

    # ice thickness
    temp_cell = ws_summary['C11']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_ice_thickness = missvalue
    else:
        ic_ice_thickness = temp_cell.value

    imported_core = CoreProperties(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value
    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value
    temp_cell = ws_summary['C17']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value
    temp_cell = ws_summary['C18']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)] is not None and ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)].value is not None :
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)].value)
        ii_col += 1

    # comment
    try:
        temp_cell = ws_summary['A33'].value
    except IndexError:
        temp_cell = None
    else:
        imported_core.comment = temp_cell

    # salinity spreadsheet
    if 'S_ice' in ws_name:
        ws_s = wb.get_sheet_by_name('S_ice')  # load data from the temperature sheet
        flag = 1
        jj = 6
        while flag == 1:
            try:
                x1 = ws_s.cell(column=1, row=jj).value
                x2 = ws_s.cell(column=2, row=jj).value
            except IndexError:
                try:
                    x3 = ws_s.cell(column=3, row=jj).value
                except IndexError:
                    break
                else:
                    if isinstance(x3, (float, int)):
                         xmid = x3
            else:
                if isinstance(x1, (float, int)):
                    xmid = (x1 + x2) / 2
            if jj == 6 and flag :
                x = np.array(xmid)
            else:
                x = np.append(x, xmid)
            jj += 1

        # core length
        if x2 is not None:
            imported_core.s_length = x2
        else:
            if ws_s['C2'] is not None:
                imported_core.s_length = ws_s['C2'].value
        # s
        s = np.nan*np.empty(len(x))
        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=4, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length
        if ~np.isnan(s).all():
            imported_core.s = 'X'

        # d18O
        s = np.nan*np.empty(len(x))
        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=6, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length
        if ~np.isnan(s).all():
            imported_core.d180 = 'X'
        # dO2
        s = np.nan*np.empty(len(x))
        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=7, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length
        if ~np.isnan(s).all():
            imported_core.dO2 = 'X'

    # Temperature
    if 'T_ice' in ws_name:
        ws_t = wb.get_sheet_by_name('T_ice')  # load data from the temperature sheet
        flag = 1
        jj = 6

        # y coordinate
        while flag == 1:
            try:
                ytemp = ws_t.cell(column=1, row=jj).value
            # TODO: check if need AttributeError
            except AttributeError and IndexError:
                flag = 0
            else:
                if jj == 6:
                    y = np.array(ytemp)
                else:
                    y = np.append(y, ytemp)
                jj += 1

        # remove trailing None Cell
        temp  =[]
        for ii in y[::-1]:
            if ii is not None:
                temp.append(ii)
        y = np.array(temp[::-1])
        del temp


        t = np.nan*np.empty(len(y))
        for jj in range(6, 6 + len(y)):
            ws_s_t = ws_t.cell(column=2, row=jj).value
            if isinstance(ws_s_t, (float, int)):
                t[jj - 6] = ws_s_t
            jj += 1  # core length
        if ~np.isnan(t).all():
            imported_core.t = 'X'

        # core length
        temp_cell = ws_t['C2'].value
        if isinstance(temp_cell, (float, int)):
            imported_core.t_length = temp_cell

    # sediment
    if 'sediment' in ws_name:
        ws_s = wb.get_sheet_by_name('sediment')  # load data from the temperature sheet
        flag = 1
        jj = 6
        while flag == 1:
            try:
                x1 = ws_s.cell(column=1, row=jj).value
                x2 = ws_s.cell(column=2, row=jj).value
            except IndexError:
                try:
                    x3 = ws_s.cell(column=3, row=jj).value
                except IndexError:
                    break
                else:
                    if isinstance(x3, (float, int)):
                         xmid = x3
            else:
                if isinstance(x1, (float, int)):
                    xmid = (x1 + x2) / 2
            if jj == 6 and flag :
                x = np.array(xmid)
            else:
                x = np.append(x, xmid)
            jj += 1

        s = np.nan*np.empty(len(x))
        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=6, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length
        if ~np.isnan(s).all():
            imported_core.sediment = 'X'

    # algae
    if 'algal_pigment' in ws_name:
        ws_s = wb.get_sheet_by_name('algal_pigment')  # load data from the temperature sheet
        flag = 1
        jj = 6
        while flag == 1:
            try:
                x1 = ws_s.cell(column=1, row=jj).value
                x2 = ws_s.cell(column=2, row=jj).value
            except IndexError:
                try:
                    x3 = ws_s.cell(column=3, row=jj).value
                except IndexError:
                    break
                else:
                    if isinstance(x3, (float, int)):
                         xmid = x3
            else:
                if isinstance(x1, (float, int)):
                    xmid = (x1 + x2) / 2
            if jj == 6 and flag :
                x = np.array(xmid)
            else:
                x = np.append(x, xmid)
            jj += 1

        s = np.nan*np.empty(len(x))
        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=5, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length
        if ~np.isnan(s).all():
            imported_core.Chla = 'X'

        s = np.nan*np.empty(len(x))
        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=6, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length
        if ~np.isnan(s).all():
            imported_core.Phaeo = 'X'
    return imported_core

def importcore(ic_path, section_thickness=0.05, missvalue=missvalue, log_level='warning' ):
    import openpyxl
    import datetime

    wb = openpyxl.load_workbook(filename=ic_path, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()

    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    # name
    temp_cell = ws_summary['C21']
    ic_name = temp_cell.value
    print(ic_name)


    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    logging.info('Processing ' + ic_name + '...')

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30)+datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value

    # location
    temp_cell = ws_summary['C5']
    ic_loc = temp_cell.value

    # snow thickness
    temp_cell = ws_summary['C9']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_snow_depth = missvalue
    else:
        ic_snow_depth = temp_cell.value

    # ice thickness
    temp_cell = ws_summary['C11']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_ice_thickness = missvalue
    else:
        ic_ice_thickness = temp_cell.value

    imported_core = Core(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value
    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value
    temp_cell = ws_summary['C17']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value
    temp_cell = ws_summary['C18']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)] is not None and ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)].value is not None :
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)].value)
        ii_col += 1

    # TEMPERATURE
    if 'T_ice' in ws_name:
        logging.info('\ttemperature profile present')
        ws_t = wb.get_sheet_by_name('T_ice')  # load data from the temperature sheet
        flag = 1
        jj = 6

        # y coordinate
        while flag == 1:
            try:
                ytemp = ws_t.cell(column=1, row=jj).value
            # TODO: check if need AttributeError
            except AttributeError and IndexError:
                flag = 0
            else:
                if jj == 6:
                    y = np.array(ytemp)
                else:
                    y = np.append(y, ytemp)
                jj += 1

        # remove trailing None Cell
        temp  =[]
        for ii in y[::-1]:
            if ii is not None:
                temp.append(ii)
        y = np.array(temp[::-1])
        del temp

        # core length
        temp_cell = ws_t['C2'].value
        if isinstance(temp_cell, (float, int)):
            tlength = temp_cell
        else:
            tlength = y[-1]

        # note
        temp_cell = ws_t['C3']
        if temp_cell is None:
            t_note = None
        else:
            t_note = temp_cell.value
        t = np.empty(len(y))
        t[:] = np.nan
        for jj in range(6, 6 + len(y)):
            ws_s_t = ws_t[openpyxl.cell.get_column_letter(2)+str('%.0f' % jj)]
            if ws_s_t is None:
                temp = missvalue
            elif isinstance(ws_s_t.value, (float, int)):
                temp = ws_s_t.value
            else:
                temp = missvalue
            t[jj - 6] = temp
            jj += 1

        # profile writing with dx section.
        # the bottom section could be smaller than the section thickness
        yt = np.arange(section_thickness / 2, tlength, section_thickness)
        if (tlength+len(yt)*section_thickness)/2 < tlength:
            yt = np.append(yt, (tlength+len(yt)*section_thickness)/2)

        t = np.interp(yt, y[~np.isnan(y)], t[~np.isnan(t)])
        imported_core.t_profile(t, yt, tlength, t_note)

    # SALINITY S
    if 'S_ice' in ws_name:
        logging.info('\tsalinity profile present')

        ws_s = wb.get_sheet_by_name('S_ice')  # load data from the salinity sheet


        # note
        temp_cell = ws_s['C3']
        if temp_cell is None:
            s_note = None
        else:
            s_note = temp_cell.value

        # x coordinate
        flag = 1
        jj = 6

        while flag == 1:
            try:
                x1 = ws_s.cell(column=1, row=jj).value
                x2 = ws_s.cell(column=2, row=jj).value
            except IndexError:
                try:
                    x3 = ws_s.cell(column=3, row=jj).value
                except IndexError:
                    break
                else:
                    if isinstance(x3, (float, int)):
                         xmid = x3
            else:
                if isinstance(x1, (float, int)):
                    xmid = (x1 + x2) / 2
            if jj == 6 and flag :
                x = np.array(xmid)
            else:
                x = np.append(x, xmid)
            jj += 1

        s = np.empty(len(x))
        s[:] = missvalue

        for jj in range(6, 6 + len(x)):
            ws_s_s = ws_s.cell(column=4, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1  # core length


        if ~np.isnan(s).all():
            # core length
            temp_cell = ws_s['C2']
            if temp_cell is None or temp_cell.value in ['n/m', 'n/a']:
                slength = x2.value
            else:
                slength = temp_cell.value

            # profile writing with dx section
            ys = np.arange(section_thickness / 2, slength, section_thickness)
            if (slength+len(ys)*section_thickness)/2 < slength:
                ys = np.append(ys, (slength+len(ys)*section_thickness)/2)
            # interp
            s = np.interp(ys, x[~np.isnan(s)], s[~np.isnan(s)])

            # profile writing with dx section, S constant in every bin
            imported_core.s_profile(s, ys, slength, s_note)
    else:
        logging.info('\tsalinity profile missing')
    return imported_core



def importcorebin(ic_path, section_thickness=0.05, missvalue=missvalue, log_level='warning'):
    import openpyxl
    import datetime

    wb = openpyxl.load_workbook(filename=ic_path, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()

    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    # name
    temp_cell = ws_summary['C21']
    ic_name = temp_cell.value
    print(ic_name)

    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    logging.info('Processing ' + ic_name + '...')

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30)+datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value

    # location
    temp_cell = ws_summary['C5']
    ic_loc = temp_cell.value

    # snow thickness
    temp_cell = ws_summary['C9']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_snow_depth = missvalue
    else:
        ic_snow_depth = temp_cell.value

    # ice thickness
    temp_cell = ws_summary['C11']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_ice_thickness = missvalue
    else:
        ic_ice_thickness = temp_cell.value

    imported_core = Core(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value
    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value
    temp_cell = ws_summary['C17']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value
    temp_cell = ws_summary['C18']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)] is not None and ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)].value is not None :
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col)+str('%.0f' % ii_row)].value)
        ii_col += 1

    # SALINITY S
    if 'S_ice' in ws_name:
        logging.info('\tsalinity profile present')

        ws_s = wb.get_sheet_by_name('S_ice')  # load data from the salinity sheet


        # note
        temp_cell = ws_s['C3']
        if temp_cell is None:
            s_note = None
        else:
            s_note = temp_cell.value

        # vertical coordinate
        flag = 1
        jj = 6

        xbin_sup = []
        xbin_down = []
        while flag == 1:
            try:
                x1 = ws_s.cell(column=1, row=jj).value
                x2 = ws_s.cell(column=2, row=jj).value
            except IndexError:
                if len(xbin_sup) is 0:
                    print('no bin present: contact developer to generate bin from midpoint')
                break

                #try:
                # x3 = ws_s.cell(column=3, row=jj).value
                #except IndexError:
                #     break
                # else:
                #     if isinstance(x3, (float, int)):
                #          xmid = x3
                #         # to do write the end of the x3
            else:
                xbin_sup.append(x1)
                xbin_down.append(x2)
            jj += 1

        # generate the x_bin
        xbin = np.append(xbin_sup, xbin_down[-1])

        # reading salinity value

        # using the bottom section value for the bin
        s = np.nan*np.empty(len(xbin_down))

        for jj in range(6, 6 + len(xbin_down)):
            ws_s_s = ws_s.cell(column=4, row=jj).value
            if isinstance(ws_s_s, (float, int)):
                s[jj - 6] = ws_s_s
            jj += 1

        # core length
        if ~np.isnan(s).all():
            # core length
            temp_cell = ws_s['C2']
            if temp_cell is None or temp_cell.value in ['n/m', 'n/a']:
                slength = x2.value
            else:
                slength = temp_cell.value

            # profile writing with dx section
            # profile writing with dx section, S constant in every bin
            imported_core.s_profile(np.array(s), np.array(xbin), slength, s_note)
    else:
        logging.info('\tsalinity profile missing')

    # TEMPERATURE
    if 'T_ice' in ws_name:
        logging.info('\ttemperature profile present')
        ws_t = wb.get_sheet_by_name('T_ice')  # load data from the temperature sheet
        flag = 1
        jj = 6
        # y coordinate
        while flag == 1:
            try:
                ytemp = ws_t.cell(column=1, row=jj).value
            # TODO: check if need AttributeError
            except AttributeError and IndexError:
                flag = 0
            else:
                if jj == 6:
                    y = np.array(ytemp)
                else:
                    y = np.append(y, ytemp)
                jj += 1

        # remove trailing None Cell
        temp  =[]
        for ii in y[::-1]:
            if ii is not None:
                temp.append(ii)
        y = np.array(temp[::-1])
        del temp

        # core length
        temp_cell = ws_t['C2'].value
        if isinstance(temp_cell, (float, int)):
            tlength = temp_cell
        else:
            tlength = y[-1]

        # note
        temp_cell = ws_t['C3']
        if temp_cell is None:
            t_note = None
        else:
            t_note = temp_cell.value

        t = np.nan*np.empty(len(y))
        for jj in range(6, 6 + len(y)):
            ws_s_t = ws_t[openpyxl.cell.get_column_letter(2)+str('%.0f' % jj)]
            if ws_s_t is None:
                temp = missvalue
            elif isinstance(ws_s_t.value, (float, int)):
                temp = ws_s_t.value
            else:
                temp = missvalue
            t[jj - 6] = temp
            jj += 1
        imported_core.t_profile(np.array(t), np.array(y), tlength, t_note)
    return imported_core


## import all ice core data which path are given in a source text file
def importsrc(txtfilepath, section_thickness=0.05, missvalue=float('nan'), log_level=log_level):

    print('Ice core data importation in progress ...')

    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])

    a = open(txtfilepath)
    filepath = [line.strip() for line in a]
    filepath = sorted(filepath)

    ic_dict = {}
    for ii in range(0, len(filepath)):
        if not filepath[ii][0] == '#':
            ic_data = importcore(filepath[ii], section_thickness, missvalue, log_level=log_level)
            ic_dict[ic_data.name]=ic_data

    logging.info('Ice core importation complete')
    print('done')

    return ic_dict


def importlist(ics_list, section_thickness=0.05, missvalue=float('nan'), log_level='warning'):

    print('Ice core data importation in progress ...')

    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])

    ic_dict = {}
    for ii in range(0, len(ics_list)):
        print(ics_list[ii])
        ic_data = importcore(ics_list[ii], section_thickness, missvalue, log_level=log_level)
        ic_dict[ic_data.name]=ic_data

    logging.info('Ice core importation complete')
    print('done')

    return ic_dict


class CoreSet:
    def __init__(self, name, date):
        self.name = name
        self.datestamp = [date]

        self.ice_thickness = []
        self.snow_depth = []
        self.corenames = [name]

        self.s_profile = np.array([[]])
        self.sy_profile = np.array([[]])
        self.s_length = []
        self.s_legend = []

        self.t_profile = np.array([[]])
        self.st_profile = np.array([[]])
        self.t_length = []
        self.t_legend = []

        self.t_air = []
        self.t_snow = []
        self.t_ice0 = []
        self.t_water = []

    def date(self):
        if len(self.datestamp)<2:
            return self.datestamp[0]
        else:
            return self.datestamp

    def error(message):
        import sys
        sys.stderr.write("error: %s\n" % message)
        sys.exit(1)

    def t_avg(self, core_to_ignore=None):
        if core_to_ignore is not None:
            if isinstance(core_to_ignore, str):
                core_to_ignore = [core_to_ignore]
            t_bkp = [self.t_profile, self.t_length]
            for c in core_to_ignore:
                self.del_t_profile(c)
        t_avg_profile = np.nanmean(self.t_profile, axis=1)
        t_avg_length = np.nanmax(self.t_length)

        if core_to_ignore is not None:
            self.t_profile = t_bkp[0]
            self.t_length = t_bkp[1]

        return t_avg_profile, t_avg_length, 't-avg'

    def add_s_profile(self, s, s_length, core_name, dt=None):
        self.s_profile = icdtools.column_merge([self.s_profile, s])
        self.s_length.append(s_length)
        self.s_legend.append(core_name)
        if core_name not in self.corenames:
            self.corenames.append(core_name)
        if dt not in self.datestamp and dt is not None:
            self.date.append(dt)

    def add_t_profile(self, t, t_length, core_name, dt=None):
        self.t_profile = icdtools.column_merge([self.t_profile, t])
        self.t_length.append(t_length)
        self.t_legend.append(core_name)
        if core_name not in self.corenames:
            self.corenames.append(core_name)
        if dt not in self.datestamp and dt is not None:
            self.date.append(dt)

    @property
    def length(self):
        all_length = self.t_length + self.s_length
        return all_length

#    @property
#    def _nanmean_(self, key):
#        return np.nanmean(self._keys)

    def del_s_profile(self, s_name):
        if s_name in self.s_legend:
            index = self.s_legend.index(s_name)
            self.s_legend.pop(index)
            self.s_length.pop(index)
            np.delete(self.s_profile, index, axis=1)
            if s_name not in self.t_legend:
                self.corenames.pop(s_name)
        else:
            print('selected profile does not exist')

    def del_t_profile(self, t_name):
        if t_name in self.t_legend:
            index = self.t_legend.index(t_name)
            self.t_legend.pop(index)
            self.t_length.pop(index)
            np.delete(self.t_profile, index, axis=1)
            if t_name not in self.s_legend:
                self.corenames.pop(t_name)
        else:
            print('selected profile does not exist')


    def add_t_air(self, t):
        self.t_air.append(t)

    def add_t_snow(self, t):
        self.t_snow.append(t)

    def add_t_ice0(self, t):
        self.t_ice0.append(t)

    def add_t_water(self, t):
        self.water.append(t)

    def __getattr__(self, key):
        return None

    def __getstate__(self):
        return self.__dict__

    def __setstate__(self, d):
        self.__dict__.update(d)