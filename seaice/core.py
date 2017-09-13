#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
core.py: ice core data is a toolbox to import ice core data file from xlsx spreadsheet, formatted according to the
 template developped by the Sea Ice Group of the Geophysical Institute of the University of Alaska, Fairbanks.
 core.py integrate the module panda into the module core version 2.1 to simplify the operation and decrease
 computation time. Core profiles are considered as collection of point in depth, time and properties (salinity,
 temperature or other variable)

"""
import datetime
import logging
import os
import time

import dateutil
import matplotlib as mpl
import matplotlib.cm as cm
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import seaice.properties

__name__ = "core"
__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "1.1"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "moggier@alaska.edu"
__status__ = "dev"
__date__ = "2017/05/06"
__comment__ = "core.py contained classes and function destinated to analyezed sea ice core data "
__CoreVersion__ = 1.1


# create logger
module_logger = logging.getLogger(__name__)
TOL =1e-6
variable_2_sheet = {'temperature': 'T_ice',
                    'salinity': 'S_ice',
                    'conductivity': 'S_ice',
                    'specific conductance': 'S_ice',
                    'd18O': 'S_ice',
                    'dD': 'S_ice',
                    'Vf_oil': 'Vf_oil', 'oil volume fraction': 'Vf_oil', # MOSIDEO project
                    'Wf_oil': 'Wf_oil', 'oil weight fraction': 'Vf_oil', # MOSIDEO project
                    'oil content': 'oil_content',    # CMI project
                    'm_oil': 'Vf_oil', 'oil mass': 'Vf_oil'
                    # 'seawater': 'seawater',
                    # 'sediment': 'sediment',
                    # 'Chla': 'algal_pigment',
                    # 'chlorophyl a': 'algal_pigment',
                    # 'Phae': 'algal_pigment'
                    }


def timing(f):
    """
    :param f:
    :return:
    """
    def wrap(*args):
        """
        :param args:
        :return:
        """
        time1 = time.time()
        ret = f(*args)
        time2 = time.time()
        print('%s function took %0.3f ms' % (f.func_name, (time2-time1)*1000.0))
        return ret
    return wrap


class Core:
    """
    Core
    """

    def __getstate__(self):
        d = self.__dict__.copy()
        if 'logger' in d.keys():
            d['logger'] = d['logger'].name
        return d

    def __setstate__(self, d):
        if 'logger' in d.keys():
            d['logger'] = logging.getLogger(d['logger'])
        self.__dict__.update(d)

    def __init__(self, name, date, origin=np.nan, lat=np.nan, lon=np.nan, ice_thickness=np.nan, freeboard=np.nan, snow_depth=np.nan):
        """
        :param name:
            string, name of the ice core
        :param date:
            datetime.datetime, date of coring YYYY-mm-dd hh:mm
        :param origin:
            string, location name of the sampling
        :param lat:
            float, latitude of sampling
        :param lon:
            float, longitude of sampling
        :param ice_thickness:
            np.array, ice thickness as measured in the core hole, following elements are ice thickness
            measured in the vicinity of the core hole.
        :param snow_depth:
            np.array, ice thickness measured at the location of coring
        :return:
        """
        self.logger = logging.getLogger(__name__)
        self.logger.info('Creating an instance of Core %s' %name)
        self.name = name
        self.date = date
        self.origin = origin
        self.lat = lat
        self.lon = lon
        self.snow_depth = snow_depth
        self.freeboard = freeboard
        self.ice_thickness = ice_thickness
        # todo length should be prefer over ice core length, as we remove the core of name, collection, ...
        self.ice_core_length = np.array([])
        self.collection = [name]
        self.comment = None
        self.profile = pd.DataFrame([])
        self.t_air = np.nan
        self.t_snow_surface = np.nan
        self.t_ice_surface = np.nan
        self.t_water = np.nan
        self.protocol = None
        self.variables = None

    def add_ice_core_length(self, ice_core_length):
        """
        :param ice_core_length:
        :return:
        """
        self.ice_core_length = np.concatenate((self.ice_core_length, [ice_core_length]))

    def add_to_collection(self, core_list):
        """
        :param core_list:
        :return:
        """
        if isinstance(core_list, list):
            for core in core_list:
                if core not in self.collection:
                    self.collection.append(core)
        else:
            if core_list not in self.collection:
                self.collection.append(core_list)

    def remove_core(self, core):
        """
        :param core:
            string, core to remove from the collection
        :return:
        """
        self.del_from_collection(core)
        self.del_profile(core)

    def del_from_collection(self, core_list):
        """
        :param core_list:
        :return:
        """
        if isinstance(core_list, list):
            for core in core_list:
                if core in self.collection:
                    self.collection.remove(core)
        else:
            if core_list in self.collection:
                self.collection.remove(core_list)

    def add_comment(self, comment):
        """
        :param comment:
        :return:
        """
        if comment is not None:
            if self.comment is None:
                self.comment = comment
            # add comment only if the comment is different to any other comment
            elif comment not in self.comment.split('; '):
                self.comment += '; '+ comment

    def add_variable(self, variable):
        """
        :param variable:
        :return:
        """
        if self.variables is None:
            self.variables = [variable]
        elif variable not in self.variables:
            self.variables.append(variable)

    def del_variable(self, variable):
        """
        :param variable:
            str, variable to delete
        :return:
        """
        self.variables.remove(variable)
        self.profile = self.profile[~self.profile.variable.str.contains(variable)]

    def del_profile(self, core):
        """
        Delete all profile belonging to the core CORE
        :param core:
            string, name of the core to delete the profile
        :return:
        """
        self.profile = self.profile[~self.profile.name.str.contains(core)]

    def summary(self):
        """
        :return:
        """
        print("#---------------------------------------------------------------#")
        print("# SUMMARY FOR ICE CORE : %s" % self.name)
        print("#---------------------------------------------------------------#")
        print('date:\t %s' % self.date.strftime('%y-%b-%d %H:%S (UTC%z)'))
        print('ice thickness\t\th_i = %.2f m' % self.ice_thickness[0])
        if self.ice_thickness.__len__() > 1:
            print('\t\t\taverage\t\t   %.2f ± %.2f m (n = %d)' % (self.ice_thickness.mean(), self.ice_thickness.mean(),
                                                                  self.ice_thickness.__len__()))
        if self.ice_core_length.__len__() > 1:
            print('average ice core length\th_c = %.2f ± %.2f m (n = %d)' % (self.ice_core_length.mean(), self.ice_core_length.mean(),
                                                                        self.ice_core_length.__len__()))
        elif self.snow_depth.__len__() == 1:
            print('ice core length\t\th_c = %.2f m' % self.ice_core_length.mean())

        print('freeboard\t\t\th_f = %.2f m' % self.freeboard[0])
        if self.freeboard.__len__() > 1:
            print('\t\t\taverage\t\t   %.2f ± %.2f m (n = %d)' % (self.freeboard.mean(), self.freeboard.mean(),
                                                                  self.freeboard.__len__()))
        if self.snow_depth.__len__() > 1:
            print('average snow depth\th_s = %.2f ± %.2f m (n = %d)' % (self.snow_depth.mean(), self.snow_depth.mean(),
                                                                        self.snow_depth.__len__()))
        elif self.snow_depth.__len__() == 1:
            print('snow depth\t\t\th_s = %.2f m' % self.snow_depth.mean())

        print('variables:')
        if self.variables is not None:
            for variable in self.variables:
                print('\t%s' % variable)
        else:
            print('\tNO VARIABLE')

        print('comment: %s' % self.comment)

    def add_snow_depth(self, snow_depth):
        """
        :param snow_depth:
        :return:
        """
        self.snow_depth = np.concatenate(self.snow_depth, [snow_depth])

    def add_profile(self, profile):
        """
        :param profile:
            pd.DataFrame, profile to add
        :return:
        """
        self.profile = self.profile.append(profile)
        self.profile.reset_index(inplace=True, drop=True)

    def plot_variables(self, variables=None, ax=None, param_dict=None):
        """
        :param ax:
        :param variable:
        :param param_dict:
        :return:
        """
        # check variables :
        if variables is None:
            variables = self.variables
        if not isinstance(variables, list):
            variables = [variables]

        if ax is None:
            ax = [plt.subplot(1, variables.__len__(), ii) for ii in range(1, variables.__len__()+1)]
        elif not ax.__len__() == variables.__len__():
            module_logger.error("lenght of ax and variables should be identical")

        plt.figure()
        n_ax = 0
        #TODO : automate color splitting according to the number of variable
        variable_color = ['r', 'b', 'g']
        for variable in variables:
            profile = self.profile[self.profile.variable == variable]
            if param_dict is None or 'color' not in param_dict:
                param_dict = {'color':variable_color[n_ax]}
            plot_profile(profile, ax=ax[n_ax], param_dict=param_dict)
            n_ax += 1
        return ax

    # # def calc_prop(self, property):
    # #     """
    # #     :param property:
    # #     :return:
    # #     """
    # #     # check properties variables
    # #     if property not in si_prop_list.keys():
    # #         logging.warning('property %s not defined in the ice core property module' % property)
    # #         return None
    # #     elif 'salinity' not in self.profiles:
    # #         logging.warning('ice core %s is missing salinity profile for further calculation' % self.name)
    # #         return None
    # #     elif 'temperature' not in self.profiles:
    # #         logging.warning('ice core %s is missing temperature profile for further calculation' % self.name)
    # #         return None
    # #     else:
    # #         import seaice.properties
    # #         variable = si_prop_list[property]
    # #         function = getattr(seaice.properties, variable.replace(" ", "_"))
    # #         profilet = self.profiles['temperature']
    # #         profiles = self.profiles['salinity']
    # #         y = np.sort(np.unique(np.concatenate((profilet.y, profiles.y))))
    # #         y = y[np.where(y <= max(max(profilet.y), max(profiles.y)))]
    # #         y = y[np.where(y >= max(min(profilet.y), min(profiles.y)))]
    # #         y_mid = y[0:-1] + np.diff(y) / 2
    # #         length = max(y) - min(y)
    # #
    # #         xt = np.interp(y_mid, profilet.y, profilet.x, left=np.nan, right=np.nan)
    # #         xs = np.interp(y_mid, profiles.y[:-1] + np.diff(profiles.y) / 2, profiles.x, left=np.nan, right=np.nan)
    # #         x = function(xt, xs)
    # #
    # #         note = 'computed from ' + self.profiles['temperature'].profile_label + ' temperature profile and ' + \
    # #                self.profiles['salinity'].profile_label + ' salinity profile'
    # #         profile_label = self.name
    # #         prop_profile = Profile(x, y, profile_label, variable, comment=None, note=note, length=length)
    # #         self.add_profile(prop_profile, variable)
    # #         self.add_comment('computed %s' % variable)

    # def plot_state_variable(self, flag_figure_number=None, param_dict=None):
    #     """
    #     :param flag_figure_number:
    #     :param param_dict:
    #     :return:
    #     """
    #     if flag_figure_number is None:
    #         fig = plt.figure()
    #         ax1 = fig.add_subplot(1, 2, 1)
    #         ax2 = fig.add_subplot(1, 2, 2, sharey=ax1)
    #     else:
    #         fig = plt.figure(flag_figure_number)
    #         ax1 = fig.add_subplot(1, 2, 1)
    #         ax2 = fig.add_subplot(1, 2, 2, sharey=ax1)
    #
    #     if 'salinity' in self.profiles.keys():
    #         self.plot_variable(ax1, 'salinity', param_dict)
    #         ax1.set_ylabel('depth [m]')
    #     else:
    #         logging.warning('salinity profile missing for %s' % self.name)
    #
    #     if 'temperature' in self.profiles.keys():
    #         self.plot_variable(ax2, 'temperature', param_dict)
    #     else:
    #         logging.warning('temperature profile missing for %s' % self.name)
    #
    #     ax_fig = [ax1, ax2]
    #     return fig, ax_fig
    #
    # def get_profile_variable(self):
    #     return sorted(self.profiles.keys())
    #
    # def rescale(self, variable=None, section_thickness=0.05):
    #     return make_section(self, variable, section_thickness)


class CoreStack(pd.DataFrame):
    """
    CoreStack
    """

    def __getstate__(self):
        d = self.__dict__.copy()
        if 'logger' in d.keys():
            d['logger'] = d['logger'].name
        return d

    def __setstate__(self, d):
        if 'logger' in d.keys():
            d['logger'] = logging.getLogger(d['logger'])
        self.__dict__.update(d)

    def __init__(self, *args, **kwargs):
        super(CoreStack, self).__init__(*args, **kwargs)
        self.logger = logging.getLogger(__name__)
        self.logger.info('creating an instance of CoreStack')

    def add_profile(self, profile):
        self = self.append(profile)
        return CoreStack(self)

    def delete_profile(self, variable_dict):
        return CoreStack(delete_profile(self, variable_dict))

    def add_profiles(self, ic_data):
        """
        :param ic_data:
        :return:
        """
        if ic_data.variables is not None:
            self.logger.info("Adding %s profiles for core %s" % (", ".join(ic_data.variables), ic_data.name))
            print(ic_data.name)
            profile = ic_data.profile
            profile['name'] = ic_data.name
            profile['ice_core_length'] = ic_data.ice_core_length[~np.isnan(ic_data.ice_core_length)].mean()
            if isinstance(ic_data.ice_thickness, (int, float)):
                profile['ice_thickness'] = ic_data.ice_thickness
            else:
                profile['ice_thickness'] = ic_data.ice_thickness[~np.isnan(ic_data.ice_thickness)].mean()

            if isinstance(ic_data.freeboard, (int, float)):
                profile['freeboard'] = ic_data.freeboard
            else:
                profile['freeboard'] = np.nanmean(ic_data.freeboard)

            if isinstance(ic_data.snow_depth, (int, float)):
                profile['snow_depth'] = ic_data.snow_depth
            else:
                profile['snow_depth'] = np.nanmean(ic_data.snow_depth)

            profile['date'] = ic_data.date
            profile['collection'] = ', '.join(ic_data.collection)
            temp = self.append(profile).reset_index(drop=True)
            return CoreStack(temp)
        else:
            return CoreStack(self)

    def remove_profile_from_core(self, core, variable=None):
        if variable is None:
            temp = self[self.name != core]
        elif isinstance(variable, list):
            for ii_variable in core:
                temp = self[(self.name != core) & (self.variable != ii_variable)]
        else:
            temp = self[(self.name != core) & (self.variable != variable)]
        return CoreStack(temp)

    # def discretize(self, y_bins=None, y_mid=None, variables=None, comment='y', display_figure='n'):
    #     """
    #     :param y_bins:
    #     :param y_mid:
    #     :param variables:
    #     :param comment: {0, 1}, Default 0
    #     :param display_figure: {0, 1}, Default 0
    #     :return:
    #     """
    #
    #     if y_bins is None and y_mid is None:
    #         y_bins = pd.Series(self.y_low.dropna().tolist()+self.y_sup.dropna().tolist()).sort_values().unique()
    #         y_mid = self.y_mid.dropna().sort_values().unique()
    #
    #     elif y_mid is None:
    #         y_mid = self.y_mid.dropna().sort_values().unique()
    #
    #     for ii_core in sorted(self.core_name.unique().tolist()):
    #         if comment == 'y':
    #             print(ii_core)
    #         ic_data = self[self.core_name == ii_core]
    #         ic_data = discretize_profile(ic_data, y_bins=y_bins, y_mid=y_mid, variables=variables,
    #                                      comment=comment, display_figure=display_figure)
    #         self = self[(self.core_name != ii_core)]
    #         self = self.append(ic_data)
    #     return CoreStack(self.reset_index(drop=True))
    #
    def grouped_stat(self, variables, stats, bins_DD, bins_y, comment='n'):
        y_cuts = pd.cut(self.y_mid, bins_y, labels=False)
        t_cuts = pd.cut(self.DD, bins_DD, labels=False)

        if not isinstance(variables, list):
            variables = [variables]
        if not isinstance(stats, list):
            stats = [stats]

        temp_all = pd.DataFrame()
        for ii_variable in variables:
            if comment == 'y':
                print('\ncomputing %s' % ii_variable)
            data = self[self.variable == ii_variable]
            data_grouped = data.groupby([t_cuts, y_cuts])

            for ii_stat in stats:
                if comment == 'y':
                    print('computing %s' % ii_stat)
                func = "groups['" + ii_variable + "']." + ii_stat + "()"
                stat_var = np.nan * np.ones((bins_DD.__len__() - 1, bins_y.__len__()))
                core_var = [[[None] for x in range(bins_y.__len__())] for y in range(bins_DD.__len__() - 1)]
                for k1, groups in data_grouped:
                    stat_var[int(k1[0]), int(k1[1])] = eval(func)
                    core_var[int(k1[0])][int(k1[1])] = [list(groups.dropna(subset=[ii_variable])
                                                             ['core_name'].unique())]
                for ii_bin in range(stat_var.__len__()):
                    temp = pd.DataFrame(stat_var[ii_bin], columns=[ii_variable])
                    temp = temp.join(pd.DataFrame(core_var[ii_bin], columns=['core_collection']))
                    DD_label = 'DD-' + str(bins_DD[ii_bin]) + '_' + str(bins_DD[ii_bin + 1])
                    data = [str(bins_DD[ii_bin]), str(bins_DD[ii_bin + 1]), DD_label, int(ii_bin), ii_stat,
                            ii_variable, self.v_ref.unique()[0]]
                    columns = ['DD_min', 'DD_max', 'DD_label', 'DD_index', 'stats', 'variable', 'v_ref']
                    index = np.array(temp.index.tolist())  #[~np.isnan(temp[ii_variable].tolist())]
                    temp = temp.join(pd.DataFrame([data], columns=columns, index=index))
                    temp = temp.join(pd.DataFrame(index, columns=['y_index'], index=index))
                    for row in temp.index.tolist():
                        print('test')
                        temp.loc[temp.index == row, 'n'] = temp.loc[temp.index == row, 'core collection'].__len__()
                    columns = ['y_low', 'y_sup', 'y_mid']
                    t2 = pd.DataFrame(columns=columns)
                    # For step profile, like salinity
                    # if ii_variable in ['salinity']:
                    if not self[self.variable == ii_variable].y_low.isnull().any():
                        for ii_layer in index[:-1]:
                            data = [bins_y[ii_layer], bins_y[ii_layer + 1],
                                    (bins_y[ii_layer] + bins_y[ii_layer + 1]) / 2, DD_label + str('-%03d' % ii_layer)]
                            t2 = t2.append(pd.DataFrame([data], columns=columns, index=[ii_layer]))
                    # For linear profile, like temperature
                    # if ii_variable in ['temperature']:
                    elif self[self.variable == ii_variable].y_low.isnull().all():
                        for ii_layer in index[:-1]:
                            data = [np.nan, np.nan, (bins_y[ii_layer] + bins_y[ii_layer + 1]) / 2,
                                    DD_label + str('-%03d' % ii_layer)]
                            t2 = t2.append(pd.DataFrame([data], columns=columns, index=[ii_layer]))

                    if temp_all.empty:
                        temp_all = temp.join(t2)
                    else:
                        temp_all = temp_all.append(temp.join(t2), ignore_index=True)

        data_grouped = self.groupby([t_cuts, self['variable']])

        grouped_dict = {}
        for var in variables:
            grouped_dict[var] = [[] for ii_DD in range(bins_DD.__len__()-1)]

        for k1, groups in data_grouped:
            if k1[1] in variables:
                grouped_dict[k1[1]][int(k1[0])] = groups['core'].unique().tolist()

        return CoreStack(temp_all.reset_index(drop=True)), grouped_dict

    # def plot_core(self, core_dict, ax=None, param_dict=None):
    #     for ii_core in core_dict:
    #         ic_data = self.select_profile(core_dict)[0]
    #         ax = plot_profile(ic_data, core_dict['variable'], ax=ax, param_dict=param_dict)
    #     return ax
    #
    # def plot_core_profile(self, core, ax=None, variable=None, param_dict=None):
    #     """
    #     :param core:
    #     :param ax:
    #     :param variable:
    #     :param param_dict:
    #     :return:
    #     """
    #     ic_data = self.select_profile({'variable': variable, 'core_name':core})[0].reset_index()
    #     if ic_data[ic_data.variable == variable].__len__() != 0:
    #         ax = plot_profile(ic_data, variable, ax = ax, param_dict=param_dict)
    #     ax.axes.set_xlabel(variable + ' ' + si_prop_unit[variable])
    #     ax.axes.set_ylim(max(ax.get_ylim()), 0)
    #     return ax
    #
    # def plot_stat_mean(self, ax, variable, bin_index):
    #     x_mean = self.select_profile({'stats': 'mean', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #     x_max = self.select_profile({'stats': 'max', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #     x_min  = self.select_profile({'stats': 'min', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #     x_std = self.select_profile({'stats': 'std', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #
    #     if x_mean[variable].__len__() !=0:
    #         plot_profile(x_max, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'r', 'label': 'min'})
    #         plot_profile(x_min, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'b', 'label': 'max'})
    #         plot_profile(x_mean, variable, ax=ax,  param_dict={'linewidth': 3, 'color': 'k', 'label': 'mean'})
    #
    #         if x_std.__len__() < x_mean.__len__():
    #             index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
    #             x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))
    #
    #         if variable in ['salinity']:
    #             y_low = x_mean['y_low']
    #             y_sup = x_mean['y_sup']
    #             x_std_l = x_mean[variable][0] - x_std[variable][0]
    #             x_std_h = x_mean[variable][0] + x_std[variable][0]
    #             y_std = y_low[0]
    #             for ii in range(1, len(x_mean)):
    #                 x_std_l = np.append(x_std_l, x_mean[variable][ii - 1] - x_std[variable][ii - 1])
    #                 x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
    #                 x_std_h = np.append(x_std_h, x_mean[variable][ii - 1] + x_std[variable][ii - 1])
    #                 x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
    #                 y_std = np.append(y_std, y_low[ii])
    #                 y_std = np.append(y_std, y_low[ii])
    #             if len(x_mean) == 1:
    #                 ii = 0
    #             x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
    #             x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
    #             y_std = np.append(y_std, y_sup[ii])
    #         elif variable in ['temperature']:
    #             y_std = x_mean['y_mid']
    #             x_std_l = []
    #             x_std_h = []
    #             for ii in range(0, len(x_mean)):
    #                 x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
    #                 x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
    #         ax = plt.fill_betweenx(y_std, x_std_l, x_std_h, facecolor='black', alpha=0.3,
    #                                label=str(r"$\pm$"+"std dev"))
    #         ax.axes.set_xlabel(variable + ' ' + si_prop_unit[variable])
    #         ax.axes.set_ylim([max(ax.axes.get_ylim()), min(ax.axes.get_ylim())])
    #     return ax
    #
    # def plot_stat_median(self, ax, variable, bin_index):
    #     x_mean = self.select_profile({'stats': 'median', 'variable': variable,
    #                                   'DD_index': bin_index})[0].reset_index()
    #     x_max = self.select_profile({'stats': 'max', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #     x_min = self.select_profile({'stats': 'min', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #     x_std = self.select_profile({'stats': 'mad', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
    #
    #     if x_mean[variable].__len__() != 0:
    #         plot_profile(x_max, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'r', 'label':'min'})
    #         plot_profile(x_min, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'b', 'label':'max'})
    #         plot_profile(x_mean, variable, ax=ax,  param_dict={'linewidth': 3, 'color': 'k', 'label':'median'})
    #
    #         if x_std.__len__() < x_mean.__len__():
    #             index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
    #             x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))
    #
    #         if variable in ['salinity']:
    #             y_low = x_mean['y_low']
    #             y_sup = x_mean['y_sup']
    #             x_std_l = x_mean[variable][0] - x_std[variable][0]
    #             x_std_h = x_mean[variable][0] + x_std[variable][0]
    #             y_std = y_low[0]
    #             for ii in range(1, len(x_mean)):
    #                 x_std_l = np.append(x_std_l, x_mean[variable][ii - 1] - x_std[variable][ii - 1])
    #                 x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
    #                 x_std_h = np.append(x_std_h, x_mean[variable][ii - 1] + x_std[variable][ii - 1])
    #                 x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
    #                 y_std = np.append(y_std, y_low[ii])
    #                 y_std = np.append(y_std, y_low[ii])
    #             if len(x_mean) == 1:
    #                 ii = 0
    #             x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
    #             x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
    #             y_std = np.append(y_std, y_sup[ii])
    #         elif variable in ['temperature']:
    #             y_std = x_mean['y_mid']
    #             x_std_l = []
    #             x_std_h = []
    #             for ii in range(0, len(x_mean)):
    #                 x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
    #                 x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
    #         ax = plt.fill_betweenx(y_std, x_std_l, x_std_h, facecolor='black', alpha=0.1,
    #                                         label=str(r"$\pm$"+"mad"))
    #         ax.axes.set_xlabel(variable + ' ' + si_prop_unit[variable])
    #         ax.axes.set_ylabel('ice thickness [m]')
    #         ax.axes.set_ylim([max(ax.axes.get_ylim()), min(ax.axes.get_ylim())])
    #     return ax
    #
    # def core_set(self):
    #     return list(set(seaice.toolbox.flatten_list(self.core_collection.tolist())))
    #
    # def compute_physical_property(self, si_prop, s_profile_shape='linear', comment='n'):
    #     # look for all core belonging to a coring event:
    #     temp_core_processed = []
    #     ic_prop = seaice.core.CoreStack()
    #     for f_core in sorted(self.core_name.unique()):
    #         ic = self[self.core_name == f_core]
    #         ic_data = seaice.core.CoreStack()
    #         if comment == 'y':
    #             print('\n')
    #         if f_core not in temp_core_processed:
    #             for ff_core in list(set(seaice.toolbox.flatten_list(ic.core_collection.tolist()))):
    #                 if comment == 'y':
    #                     print(ff_core)
    #                 ic_data = ic_data.add_profiles(self[self.core_name == ff_core])
    #             ic_prop = ic_prop.append(seaice.core.calc_prop(ic_data, si_prop, s_profile_shape=s_profile_shape))
    #             temp_core_processed.append(ff_core)
    #
    #     ics_stack = seaice.core.CoreStack(self)
    #     ic_prop = seaice.core.CoreStack(ic_prop)
    #
    #     ics_stack = ics_stack.add_profiles(ic_prop)
    #     return CoreStack(ics_stack)

    def discretize(self, y_bins=None, y_mid=None, variables=None, display_figure='n'):
        if variables is None:
            variables = self.variable.unique().tolist()

        data_binned = pd.DataFrame()
        for core in self.name.unique():
            data_binned = data_binned.append(
                discretize_profileV2(self[self.name == core], y_bins=y_bins, y_mid=y_mid, variables=variables,
                                       display_figure=display_figure))
        return CoreStack(data_binned)

    def compute_phys_prop(self, inplace=True):


        if inplace is True:
            return self

    def set_reference(self, v_ref, comment=False):
        temp = CoreStack()
        for f_core in self.name.unique():
            ic_data = self[self.name == f_core]
            ic_data = set_profile_orientation(ic_data, v_ref=v_ref)
            temp = temp.append(ic_data)
        return CoreStack(temp)


# UPDATER
# TODO : finish spreadsheet updater
def update_spreadhseet(ic_path, v_ref='top',  verbose=logging.WARNING):
    """

    :param ic_path:
    :param verbose:
    :return:
    """
    import shutil

    module_logger.setLevel(verbose)

    if not os.path.exists(ic_path):
        module_logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])
        return 0
    else:
        module_logger.info("\t ice core file path %s" % ic_path)

    wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
    else:
        module_logger.error("\tice core spreadsheet version not unavailable")
    if version < __CoreVersion__:
        module_logger.info("Updating core data to latest version %.1f" % __CoreVersion__)

    # backup old version
    backup_dir = os.path.join(os.path.dirname(ic_path), 'old_version-'+str(version))
    try:
        os.stat(backup_dir)
    except:
        os.makedirs(backup_dir)

    shutil.copyfile(ic_path, os.path.join(backup_dir, os.path.basename(ic_path)))

    def add_row(ws, row_number):
        max_row = ws.max_row
        for row in range(row_number, ws.max_row + 1):
            new_row = row_number + max_row + 1 - row
            old_row = row_number + max_row - row
            for col in range(1, ws.max_column+1):
                ws.cell(row=new_row, column=col).value = ws.cell(row=old_row, column=col).value
        for col in range(1, ws.max_column+1):
            ws.cell(row=row_number, column=col).value = ""
        return ws

    def delete_row(ws, row_number):
        max_row = ws.max_row
        for row in range(row_number, max_row):
            new_row = row
            old_row = row+1
            for col in range(1, ws.max_column+1):
                ws.cell(row=new_row, column=col).value = ws.cell(row=old_row, column=col).value
        for col in range(1, ws.max_column+1):
            ws.cell(row=max_row+1, column=col).value = ""
        return ws

    def delete_column(ws, target_col, start_row=None, end_row=None):
        if start_row is None:
            start_row = ws.min_row
        if end_row is None:
            end_row = ws.max_row
        if not isinstance(target_col, int):
            target_col = openpyxl.utils.column_index_from_string(target_col)

        max_col = ws.max_column
        if np.alltrue([ws.cell(row=row, column=ws.max_column).value == None for row in range(row_start, ws.max_row)]):
            max_col = max_col - 1

        if not target_col == max_col:
            for col in range(target_col, max_col):
                new_col = col
                old_col = col + 1
                for row in range(start_row, end_row + 1):
                    ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=old_col).value
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=max_col).value = ""

        return ws

    def move_column(ws, target_col, source_col, start_row=None, end_row=None):
        if start_row is None:
            start_row = ws.min_row
        if end_row is None:
            end_row = ws.max_row
        if not isinstance(target_col, int):
            target_col = openpyxl.utils.column_index_from_string(target_col)
        if not isinstance(source_col, int):
            source_col = openpyxl.utils.column_index_from_string(source_col)

        max_col = ws.max_column
        if np.alltrue([ws.cell(row=row, column=ws.max_column).value == None for row in range(row_start, ws.max_row)]):
            max_col = max_col - 1

        # insert column in target column
        for col in range(target_col, max_col+1):
            new_col = target_col + max_col - col + 1
            old_col = target_col + max_col - col
            # print(openpyxl.utils.get_column_letter(old_col), openpyxl.utils.get_column_letter(new_col))
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=old_col).value

        # copy source col to target column
        if target_col < source_col:
            source_col = source_col + 1
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=target_col).value = ws.cell(row=row, column=source_col).value

        ws = delete_column(ws, source_col, start_row, end_row)

        return ws

    def print_column(ws, col):
        if ~isinstance(col, int):
            col = openpyxl.utils.column_index_from_string(col)
        for row in range(1, ws.max_row + 1):
            print(row, ws.cell(row=row, column=col).value)

    def print_row(ws, row):
        col_string = ''
        for col in range(1, ws.max_column + 1):
            col_string += openpyxl.utils.get_column_letter(col) + ':' + str(
                ws.cell(row=row, column=col).value) + '\t'
        print(col_string)

    while(version < __CoreVersion__):
        # update from 1.0 to 1.1
        if version == 1:
            version = 1.1
            ws_summary['C3'] = version
            ws_summary = delete_row(ws_summary, 22)

            # loop through variables
            if "S_ice" in wb.get_sheet_names():
                ws = wb.get_sheet_by_name("S_ice")
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref
                #print_column('A')

                # add notation row=6
                ws = add_row(ws, 6)
                # print_column('A')

                row_start = 4
                ws = delete_column(ws, 'E', row_start, ws.max_row)
                ws['D5']='salinity'
                ws['D6']='S'
                # print_row(5)
                ws = move_column(ws, 'E', 'J', row_start, ws.max_row)
                ws['E5'] = 'conductivity'
                ws['E6'] = 'σ'
                ws = move_column(ws, 'F', 'K', row_start, ws.max_row)
                ws['F5'] = 'conductivity measurement temperature'
                ws['F6'] = 'T_σ'
                ws = move_column(ws, 'G', 'J', row_start, ws.max_row)
                ws['G5'] = 'specific conductance'
                ws['G6'] = 'κ'
                ws = delete_column(ws, 'K', row_start, ws.max_row)
                #print_row(ws, row_start+1)

            if "T_ice" in wb.get_sheet_names():
                ws = wb.get_sheet_by_name("T_ice")
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                ws['A6'] = 'd'
                ws['B6'] = 'T'


            if "oil_content" in wb.get_sheet_names():
                ws = wb.get_sheet_by_name("oil_content")
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                ws['A6'] = 'd_1'
                ws['B6'] = 'd_2'
                ws['C6'] = 'd'
                ws['D6'] = 'V_i'
                ws['E6'] = 'h_menisc'
                ws['F6'] = 'd_menisc'
                ws['G6'] = 'd_center'
        wb.save(ic_path)
    return None


# IMPORTATION
def import_core(ic_path, variables=None, v_ref='top', verbose=logging.WARNING):
    """
    :param ic_path:
        string, path to the xlsx ice core spreadsheet
    :param variables:
        list of string, variables to import. If not defined, all variable will be imported.
    :param v_ref:
        'top' or 'bottom', vertical reference. top for ice/snow or ice/air surface, bottom for ice/water interface
    """

    if not os.path.exists(ic_path):
        module_logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])
        return 0
    module_logger.setLevel(verbose)
    module_logger.info("Ice core file path %s" % ic_path)

    wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    name = ws_summary['C21'].value

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
    else:
        module_logger.error("(%s) ice core spreadsheet version not unavailable" %name)

    if version < __CoreVersion__:
        update_spreadhseet(ic_path, v_ref=v_ref, verbose=verbose)
        module_logger.error("Updating ice core spreadsheet %s to last version (%s)" % (name, str(__CoreVersion__)))
        wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
        ws_name = wb.get_sheet_names()
        ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet
        version = ws_summary['C3'].value

    n_row_collection = 22
    module_logger.info("(%s) importing data" %name)

    if isinstance(ws_summary['C2'].value, datetime.datetime):
        if isinstance(ws_summary['D2'].value, datetime.time):
            date = datetime.datetime.combine(ws_summary['C2'].value, ws_summary['D2'].value)
            if ws_summary['E2'].value:
                try:
                    tz = dateutil.tz.gettz(ws_summary['E2'].value)
                except Exception:
                    module_logger.error("\t(%s) timezone unavailable" %name)
                else:
                    date = date.replace(tzinfo=tz)
        else:
            date = ws_summary['C2'].value
            if ws_summary['D2'].value:
                try:
                    tz = dateutil.tz.gettz(ws_summary['D2'].value)
                except Exception:
                    module_logger.error("\t(%s) timezone unavailable" %name)
                else:
                    date = date.replace(tzinfo=tz)
    else:
        module_logger.warning("\t(%s) date unavailable" %name)
        date = None

    origin = ws_summary['C5'].value

    lat = np.nan
    lon = np.nan
    if isinstance(ws_summary['C6'].value, (float, int)) or isinstance(ws_summary['D6'], (float, int)):
        lat = ws_summary['C6'].value
        lon = ws_summary['D6'].value
    elif ws_summary['C6'].value and ws_summary['D6'].value:
        module_logger.warning("\t(%s) lat/lon not defined in decimal degree" % name)
    else:
        module_logger.info("\t(%s) lat/lon unknown" % name)

    if isinstance(ws_summary['C9'].value, (float, int)):
        snow_depth = np.array([ws_summary['C9'].value]).astype(float)
        n_snow = 1
        while ws_summary.cell(row=9, column=3+n_snow).value is not None:
            snow_depth = np.concatenate((snow_depth, np.array([ws_summary.cell(row=9, column=3+n_snow).value])))
            n_snow +=1
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')
    else:
        snow_depth = np.array([np.nan])

    if isinstance(ws_summary['C10'].value, (float, int)):
        freeboard = np.array([ws_summary['C10'].value])
        n_temp = 1
        while ws_summary.cell(row=10, column=3+n_temp).value is not None:
            if isinstance(ws_summary.cell(row=10, column=3+n_temp).value, (float, int)):
                freeboard = np.concatenate((freeboard, np.array([ws_summary.cell(row=10, column=3+n_snow).value])))
            else:
                module_logger.info("(%s) \tfreeboard cell %s not a float" % (name, openpyxl.utils.get_column_letter(3+n_temp)+str(9)))
            n_temp += 1
        freeboard = pd.to_numeric(freeboard, errors='coerce')
    else:
        freeboard = np.array([np.nan])

    if isinstance(ws_summary['C11'].value, (float, int)):
        ice_thickness = np.array([ws_summary['C11'].value])
        n_temp = 1
        while ws_summary.cell(row=11, column=3+n_temp).value:
            if isinstance(ws_summary.cell(row=11, column=3+n_temp).value, (float, int)):
                ice_thickness = np.concatenate((ice_thickness, np.array([ws_summary.cell(row=11, column=3+n_snow).value])))
            else:
                module_logger.info("\t(%s) ice_thickness cell %s not a float" % (name, openpyxl.utils.get_column_letter(3+n_temp)+str(9)))
            n_temp += 1
        ice_thickness = pd.to_numeric(ice_thickness, errors='coerce')
    else:
        ice_thickness = np.array([np.nan])

    core = Core(name, date, origin, lat, lon, ice_thickness, freeboard, snow_depth)

    # temperature
    if ws_summary['C15'].value:
        core.t_air = ws_summary['C15'].value
    if ws_summary['C16'].value:
        core.t_snow_surface = ws_summary['C16'].value
    if ws_summary['C17'].value:
        core.t_ice_surface = ws_summary['C17'].value
    if ws_summary['C18'].value:
        core.t_water = ws_summary['C18'].value

    # sampling protocol
    m_col = 3
    if ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0i' %(n_row_collection+2))].value is not None:
        core.protocol = ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0i' %(n_row_collection+2))].value
    else:
        core.protocol = 'N/A'

    # core collection
    while (ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)] is not None and
           ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)].value is not None):
        core.add_to_collection(ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)].value)
        m_col += 1

    # comment
    if ws_summary['C33'].value is not None:
        core.add_comment(ws_summary['C33'].value)

    # variable
    if variables is None:
        sheets = [sheet for sheet in ws_name if (sheet not in ['summary', 'abreviation', 'locations', 'lists', 'Vf_oil_calculation']) and
                     (sheet.lower().find('fig') == -1)]
        for sheet in sheets:
            ws_variable = wb.get_sheet_by_name(sheet)
            core_flag = 0
            profile = read_variable(ws_variable, variables=None, version=version, v_ref=v_ref)
            for variable in profile.keys():
                if not profile[variable][1] == core.name:
                    module_logger.error('\t(%s) core name %s and profile name %s does not match'
                                        % (ic_path, core.name, profile[variable][1]))
                else:
                    core.add_profile(profile[variable][0])
                    core.add_variable(variable)
                    if core_flag == 0 and ~np.isnan(profile[variable][3]):
                        core.add_ice_core_length(profile[variable][3])
                        core_flag = 1
                    core.add_comment(profile[variable][2])

        if core.variables is None:
            module_logger.info('(%s) no variable to import' %(name))
        elif core.variables.__len__() > 1:
            module_logger.info('(%s) variables %s imported with success' %(name, ", ".join(core.variables)))
        else:
            module_logger.info('(%s) variable %s imported with success' % (name, ", ".join(core.variables)))
    else:
        if not isinstance(variables, list):
            if variables.lower().find('state variable')+1:
                variables = ['temperature', 'salinity']
            else:
                variables = [variables]
        module_logger.info("(%s) Variables are %s" % (name, ', '.join(variables)))

        core_flag = []
        for variable in variables:
            if variable_2_sheet[variable] in ws_name:
                ws_variable = wb.get_sheet_by_name(variable_2_sheet[variable])
                profile = read_variable(ws_variable, variables=variable, version=version, v_ref=v_ref)
                if profile.keys().__len__() == 0:
                    module_logger.warning('\t(%s) no data exist for %s' % (core.name, variable))
                elif profile[variable][1] is not core.name:
                    module_logger.error('\t(%s) core name %s and profile name %s does not match'
                                        % (ic_path, core.name, profile[variable][1]))
                else:
                    core.add_profile(profile[variable][0])
                    core.add_variable(variable)
                    if core.name not in core_flag and profile[variable][3] is not None and not np.isnan(profile[variable][3]):
                        core.add_ice_core_length(profile[variable][3])
                        core_flag.append(core.name)
                    core.add_comment(profile[variable][2])
            else:
                module_logger.info('\tsheet %s does not exists' % variable)

    # weather
    # TODO:adding a weather class and reading the information

    if verbose < 30:
        core.summary()

    return core


def import_list(ics_list, variables=None, v_ref='top', verbose=logging.WARNING):
    """
    :param ics_list:
            array, array contains absolute filepath for the cores
    :param variables:

    :param v_ref:
        top, or bottom
    """
    module_logger.setLevel(verbose)
    module_logger.info('Import ice core lists:')
    ic_dict = {}
    inexisting_ics_list = []
    for ic_path in ics_list:
        module_logger.warning("%s" % ic_path)
        if not os.path.exists(ic_path):
            module_logger.warning("%s does not exists in core directory" % ic_path.split('/')[-1])
            inexisting_ics_list.append(ic_path.split('/')[-1].split('.')[0])
        else:
            ic_data = import_core(ic_path, variables=variables, v_ref=v_ref)
            if ic_data.variables is None:
                inexisting_ics_list.append(ic_path.split('/')[-1].split('.')[0])
                module_logger.warning("%s have no profile" % ic_data.name)
            else:
                ic_dict[ic_data.name] = ic_data
            logging.debug("done")
            module_logger.info('importation completed')

    module_logger.info(
        "%s core does not exits. Removing from collection" % ', '.join(inexisting_ics_list))
    for ic in inexisting_ics_list:
        for ic2 in ic_dict.keys():
            if ic in ic_dict[ic2].collection:
                ic_dict[ic2].del_from_collection(ic)
                module_logger.info("remove %s from %s collection" % (ic, ic2))
    return ic_dict


def import_src(ic_src, core_dir=None, variables=None, v_ref='top', verbose=logging.WARNING):
    """
    :param ic_src:
            string, absolute path to the file containing either the absolute path of the cores (1 path by line) or the
            core names (1 core by line). In this last case if core_dir is None core_dir is the directory contianing the
            file.
    :param variables:

    :param v_ref:
        top, or bottom
    :param verbose:
        number, verbosity of the log
    """

    ics_list = []

    module_logger.setLevel(verbose)
    module_logger.info('Import ice core from list files: %s' % ic_src)

    a = open(ic_src)
    ics = sorted([line.strip() for line in a])

    for ic_path in ics:
        if os.path.dirname(ic_path) == '':
            if core_dir is None:
                core_dir = os.path.dirname(ic_src)
                module_logger.info(
                    'Core directory not define. Using by default directory of the source file %s' % core_dir)
            ic_path = os.path.join(core_dir, ic_path)
        ics_list.append(ic_path)

    return import_list(ics_list, variables=variables, v_ref=v_ref, verbose=verbose)


def read_variable(ws_variable, variables=None, version=1.1, v_ref='top'):
    """
    :param ws_variable:
        openpyxl.worksheet
    :param v_ref:
        top, or bottom
    """

    if version == 1:
        row_data_start = 6
    elif version == 1.1:
        row_data_start = 8
        if ws_variable['C4'].value:
            v_ref = ws_variable['C4'].value
    else:
        module_logger.error("ice core spreadsheet version not defined")

    variable_2_data = {'temperature': [row_data_start, 'A', 'B', 'C'],
                       'salinity': [row_data_start, 'ABC', 'D', 'J'],
                       'conductivity': [row_data_start, 'ABC', 'EF', 'J'],
                       'specific conductance': [row_data_start, 'ABC', 'G', 'J'],
                       'd18O': [row_data_start, 'ABC', 'H', 'J'],
                       'dD': [row_data_start, 'ABC', 'I', 'J'],
                       'oil weight fraction': [row_data_start, 'ABC', 'G', 'H'],
                       'oil volume fraction': [row_data_start, 'ABC', 'F', 'H'],
                       'oil content': [row_data_start, 'ABC', 'DEFGH', 'I'],
                       'oil mass': [row_data_start, 'ABC', 'ED', 'H']
                       # 'seawater': [row_data_start, 'ABC', 'D', 'E'],
                       # 'sediment': [row_data_start, 'ABC', 'D', 'E'],
                       # 'Chla': [row_data_start, 'ABC', 'D', 'E'],
                       # 'Phae': [row_data_start, 'ABC', 'D', 'E']
                       }

    if variables is None:
        variables = [key for key in variable_2_data.keys() if variable_2_sheet[key] == ws_variable.title]
    if not isinstance(variables, list):
        variables = [variables]

    name = ws_variable['C1'].value

    profile = {}
    for variable in variables:
        module_logger.info('(%s) importing %s' %(name, variable))

        columns_string = ['comment', 'variable']
        # step profile
        if variable_2_data[variable][1].__len__() == 3:
            columns_float = ['y_low', 'y_sup', 'y_mid']
            # y_data = np.array([ws_variable[variable_2_data[variable][2][0] + str(row)].value
            #                  for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])

            y_low = np.array([ws_variable[variable_2_data[variable][1][0]+str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])
            y_sup = np.array([ws_variable[variable_2_data[variable][1][1]+str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])
            y_mid = np.array([ws_variable[variable_2_data[variable][1][2] + str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])

            if not np.array([isinstance(element, (float, int)) for element in y_mid]).any():
                if (np.array([isinstance(element, (float, int)) for element in y_low]).any() or
                        np.array([isinstance(element, (float, int)) for element in y_sup]).any()):
                    y_mid = (y_low+y_sup)/2
                    module_logger.info(
                        '\t(%s : %s) y_mid does not exit, calculating y_mid from section depth with success' % (
                        name, variable))
                else:
                    module_logger.warning(
                        '\t(%s : %s) y_mid does not exit, not able to calculate y_mid from section depth. Section depth maybe not numeric' % (name, variable))
            data = np.array([y_low, y_sup, y_mid])

            for col_variable in range(0, variable_2_data[variable][2].__len__()):
                data = np.vstack((data, np.array([ws_variable[variable_2_data[variable][2][col_variable] + str(row)].value
                                                  for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])))
                if version is 1:
                    if col_variable is 0:
                        columns_float.append(variable)
                    else:
                        columns_float.append(
                            ws_variable[variable_2_data[variable][2][col_variable] + str(row_data_start - 3)].value)

                elif version >= 1.1:
                    columns_float.append(ws_variable[variable_2_data[variable][2][col_variable] + str(row_data_start-3)].value)

            data = np.vstack((data, np.array([ws_variable[variable_2_data[variable][3] + str(row)].value
                                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])))

            variable_profile = pd.DataFrame(data.transpose(), columns=columns_float + ['comment'])

        # continuous variable_profile
        else:
            columns_float = ['y_mid']
            y_mid = np.array([ws_variable[variable_2_data[variable][1]+str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])

            for col_variable in range(0, variable_2_data[variable][2].__len__()):
                data = np.vstack((y_mid, np.array([ws_variable[variable_2_data[variable][2][col_variable] + str(row)].value
                                                   for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])))
                if version is 1:
                    if col_variable is 0:
                        columns_float.append(variable)
                    else:
                        columns_float.append(
                            ws_variable[variable_2_data[variable][2][col_variable] + str(row_data_start - 3)].value)

                elif version >= 1.1:
                    columns_float.append(ws_variable[variable_2_data[variable][2][col_variable] + str(row_data_start-3)].value)

            data = np.vstack((data, np.array([ws_variable[(variable_2_data[variable][3] + str(row))].value for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])))

            variable_profile = pd.DataFrame(data.transpose(), columns=columns_float + ['comment'])

        # convert numeric to float
        key_temp = [key for key in variable_profile.keys() if key not in ['comment', 'v_ref']]
        for key in key_temp:
            variable_profile[key] = pd.to_numeric(variable_profile[key], errors='coerce')
        variable_profile['v_ref'] = v_ref

        key_temp = [key for key in key_temp if key not in ['y_low', 'y_sup', 'y_mid']]
        l_empty = 0
        for key in key_temp:
            if not np.isnan(variable_profile[key]).all():
                l_empty += 1
        if l_empty > 0:
            variable_profile['variable'] = variable

            for col in columns_float:
                if col not in variable_profile.keys():
                    variable_profile[col] = np.nan
                else:
                    variable_profile[col] = pd.to_numeric(variable_profile[col])
            for col in columns_string:
                if col not in variable_profile.keys():
                    variable_profile[col] = None

            note = ws_variable['C3'].value
            length = ws_variable['C2'].value
            if not isinstance(length, (float, int)):
                length = np.nan

            if v_ref == 'top':
                variable_profile['v_ref'] = v_ref

            profile[variable] = [variable_profile, name, note, length]
            module_logger.info('\t(%s : %s) variable imported with success' %(name, variable))
        else:
            module_logger.warning('\t(%s : %s) variable not defined' %(name, variable))
    return profile


# PLOTTING
def plot_profile(profile, ax=None, param_dict=None):
    """
    :param profile:
    :param ax:
    :param param_dict:
    :return:
    """

    var = profile.variable.unique()[0]
    if var is None or isinstance(var, list):
        module_logger.error("only one variable should be selected")
        return 0

    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)

    # step variable
    if not profile[profile.variable == var].y_low.isnull().all() and not profile[profile.variable == var].y_sup.isnull().all():
        x = []
        y = []
        for ii in profile[profile.variable == var].index.tolist():
            y.append(profile['y_low'][ii])
            y.append(profile['y_sup'][ii])
            x.append(profile[var][ii])
            x.append(profile[var][ii])
    # continuous variable (temperature)
    else:
        x = profile[var].values
        y = profile.y_mid.values
    if param_dict is None:
        ax.plot(x, y)
    else:
        ax.plot(x, y, **param_dict)
    return ax

def semilogx_profile(profile, ax=None, param_dict=None):
    """
    :param profile:
    :param ax:
    :param param_dict:
    :return:
    """

    var = profile.variable.unique()[0]
    if var is None or isinstance(var, list):
        module_logger.error("only one variable should be selected")
        return 0

    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)

    # step variable
    if not profile[profile.variable == var].y_low.isnull().all() and not profile[profile.variable == var].y_sup.isnull().all():
        x = []
        y = []
        for ii in profile[profile.variable == var].index.tolist():
            y.append(profile['y_low'][ii])
            y.append(profile['y_sup'][ii])
            x.append(profile[var][ii])
            x.append(profile[var][ii])
    # continuous variable (temperature)
    else:
        x = profile[var].values
        y = profile.y_mid.values

    if param_dict is None:
        ax.semilogx(x, y)
    else:
        ax.semilogx(x, y, **param_dict)
    return ax



# Ice core operation
def stack_core(ics_dict):
    """"
    :param ics_dict:
        dictionnary of core
    :return ics_stack:
        panda.DataFrame()
    """
    ics_stack = CoreStack()
    for key in ics_dict.keys():
        ics_stack = ics_stack.add_profiles(ics_dict[key])
    return CoreStack(ics_stack)


def select_profile(ics_stack, variable_dict):
    str_select = '('
    ii_var = []
    ii = 0
    for ii_key in variable_dict.keys():
        if ii_key in ics_stack.columns.values:
            ii_var.append(variable_dict[ii_key])
            str_select = str_select + 'ics_stack.' + ii_key + '==ii_var[' + str('%d' % ii) + ']) & ('
            ii += 1
    str_select = str_select[:-4]
    return CoreStack(ics_stack.loc[eval(str_select)])


def delete_profile(ics_stack, variable_dict):
    str_select = '('
    ii_var = []
    ii = 0
    for ii_key in variable_dict.keys():
        if ii_key in ics_stack.columns.values:
            ii_var.append(variable_dict[ii_key])
            str_select = str_select + 'ics_stack.' + ii_key + '!=ii_var[' + str('%d' % ii) + ']) | ('
            ii += 1
    str_select = str_select[:-4]
    return CoreStack(ics_stack.loc[eval(str_select)])


def compute_phys_prop_from_core(s_profile, t_profile, si_prop, si_prop_format='step', resize_core=None, display_figure=False):
    """
    :param s_profile:
    :param t_profile:
    :param si_prop:
    :param si_prop_format: 'linear' or 'step':
    :param main_core: 'S', 'T', default 'None':
    :return:
    """

    # parameters check
    if isinstance(si_prop_format, dict):
        for prop in si_prop:
            if prop not in si_prop_format.keys():
                print("no format for prop %s, please define linear or step" % prop)
                si_prop.remove(prop)
        si_prop_dict = si_prop_format
    elif not isinstance(si_prop, dict):
        if not isinstance(si_prop, list):
            si_prop = [si_prop]
        if not isinstance(si_prop_format, list):
            si_prop_format=[si_prop_format]
        if si_prop_format.__len__() > 1 and not si_prop_format.__len__() == si_prop.__len__():
            module_logger.error("length of si_prop format does not match length of si_prop. si_prop should be length 1 or should match length si_prop")
        si_prop_dict = {}
        for ii in range(0, si_prop.__len__()):
            si_prop_dict[si_prop[ii]] = si_prop_format[ii]
    else:
        si_prop_dict = si_prop

    # initialisation
    prop_profile = pd.DataFrame()

    # check parameters
    if 'salinity' not in s_profile.keys() or not s_profile['salinity'].notnull().any():
        print("no salinity data")
        return prop_profile
    else:
        S_core_name = s_profile.name.values[0]
        s_profile['salinity'] = pd.to_numeric(s_profile['salinity'])

    if 'temperature' not in t_profile.keys() or not t_profile['temperature'].notnull().any():
        print("no temperature data")
        return prop_profile
    else:
        T_core_name = t_profile.name.values[0]
        t_profile['temperature'] = pd.to_numeric(t_profile['temperature'])


    if resize_core in ['S', S_core_name]:
        if s_profile.ice_core_length.notnull().all():
            profile_length = s_profile.ice_core_length.unique()[0]
        elif s_profile.ice_core_length.notnull().all():
            profile_length = s_profile.ice_thickness.unique()[0]
            print("ice core length unknown, using ice thickness instead")
        else:
            profile_length = max(s_profile.y_low.min(), s_profile.y_sup.max())[0]
            print("todo: need warning text")
        if not t_profile.ice_core_length.unique() == profile_length:
            t_profile = scale_profile(t_profile, profile_length)

    if resize_core in ['T', T_core_name]:
        if t_profile.ice_core_length.notnull().all():
            profile_length = t_profile.ice_core_length.unique()[0]
        elif t_profile.ice_core_length.notnull().all():
            profile_length = t_profile.ice_thickness.unique()[0]
            print("ice core length unknown, using ice thickness instead")
        else:
            profile_length = max(t_profile.y_low.min(), t_profile.y_sup.max())[0]
            print("todo: need warning text")

        if not t_profile.ice_core_length.unique() == profile_length:
            s_profile = scale_profile(s_profile, profile_length)

    # interpolate temperature profile to match salinity profile
    y_mid = s_profile.y_mid.dropna().values
    if y_mid.__len__() < 1:
        y_mid = (s_profile.y_low/2.+s_profile.y_sup/2).dropna().astype(float)

    interp_data = pd.concat([t_profile, pd.DataFrame(y_mid, columns=['y_mid'])])
    interp_data = interp_data.set_index('y_mid').sort_index().interpolate(method='index').reset_index().drop_duplicates(subset='y_mid')

    if 'temperature' in s_profile.keys():
        s_profile = s_profile.drop('temperature', axis=1)
    s_profile = pd.merge(s_profile, interp_data[['temperature', 'y_mid']], on=['y_mid'])


    # compute properties
    for f_prop in si_prop_dict.keys():
        if f_prop not in seaice.properties.si_prop_list.keys():
            print('property %s not defined in the ice core property module' % property)

        prop = seaice.properties.si_prop_list[f_prop]
        function = getattr(seaice.properties, prop.replace(" ", "_"))
        prop_data = function(t_profile['temperature'], s_profile['salinity'])

        prop_data = pd.DataFrame(np.vstack((prop_data, s_profile['y_mid'])).transpose(), columns=[prop, 'y_mid'])
        prop_data['name'] = list(set(s_profile.name))[0]
        comment_core = 'physical properties computed from ' + S_core_name + '(S) and ' + T_core_name + '(T)'
        prop_data['variable'] = prop

        var_drop = [var for var in ['salinity', 'temperature', 'variable', f_prop, 'name', 'core'] if var in s_profile.keys()]
        core_frame = s_profile.drop(var_drop, axis=1)

        if si_prop_dict[f_prop] == 'linear':
            core_frame[['y_low', 'y_sup']] = np.nan
        prop_data = pd.merge(prop_data, core_frame, how='inner', on=['y_mid'])

        for index in prop_data.index:
            if 'comment' in prop_data.keys():
                if prop_data.loc[prop_data.index == index, 'comment'].isnull().all():
                    prop_data.loc[prop_data.index == index, 'comment'] = comment_core
                else:
                    prop_data.loc[prop_data.index == index, 'comment'] += ';'+ comment_core
            else:
                prop_data.loc[prop_data.index == index, 'comment'] = comment_core

        if display_figure:
            ax = plot_profile_variable(prop_data, {'name': S_core_name, 'variable': prop},
                                  ax=None, param_dict=None)
            ax.set_xlabel(prop)
            ax.set_ylabel('ice thickness)')
            ax.set_title(S_core_name)
        prop_profile = prop_profile.append(prop_data, ignore_index=True, verify_integrity=False)

    return prop_profile


def compute_phys_prop_from_core_name(ics_stack, S_core_name, T_core_name, si_prop, si_prop_format='step', resize_core=None, inplace=True, display_figure=False):
    """
    :param ics_stack:
    :param S_core_name:
    :param T_core_name:
    :param si_prop:
    :param si_prop_format: 'linear' or 'step':
    :param resize_core: 'S', 'T', default 'None':
    :return:
    """

    # parameters check
    if isinstance(si_prop_format, dict):
        if isinstance(si_prop, list):
            for prop in si_prop:
                if prop not in si_prop_format.keys():
                    print("no format for prop %s, please define linear or step" % prop)
                    si_prop.remove(prop)
            si_prop_dict = si_prop_format
        else:
            si_prop_dict = si_prop
    elif not isinstance(si_prop, dict):
        if not isinstance(si_prop, list):
            si_prop = [si_prop]
        if not isinstance(si_prop_format, list):
            si_prop_format=[si_prop_format]
        if si_prop_format.__len__() > 1 and not si_prop_format.__len__() == si_prop.__len__():
            module_logger.error("length of si_prop format does not match length of si_prop. si_prop should be length 1 or should match length si_prop")
        si_prop_dict = {}
        for ii in range(0, si_prop.__len__()):
            si_prop_dict[si_prop[ii]] = si_prop_format[ii]
    else:
        si_prop_dict = si_prop

    # check parameters
    if S_core_name not in ics_stack.name.unique():
        print("%s core not present in data" % S_core_name)
        return pd.DataFrame();
    elif 'salinity' not in ics_stack.loc[ics_stack.name == S_core_name, 'variable'].unique():
        print("salinity data not existing for %s " % S_core_name)
        return pd.DataFrame();
    else:
        s_profile = ics_stack[(ics_stack.name == S_core_name) & (ics_stack.variable == 'salinity')]

    if T_core_name not in ics_stack.name.unique():
        print("%s core not present in data" % T_core_name)
        return pd.DataFrame();
    elif 'temperature' not in ics_stack.loc[ics_stack.name == T_core_name, 'variable'].unique():
        print("temperature data not existing for %s " % T_core_name)
        return pd.DataFrame();
    else:
        t_profile = ics_stack[(ics_stack.name == T_core_name) & (ics_stack.variable == 'temperature')]

    prop_profile = compute_phys_prop_from_core(s_profile, t_profile, si_prop=si_prop_dict, si_prop_format=si_prop_format, resize_core=resize_core, display_figure=display_figure)

    if inplace is True:
        for f_prop in prop_profile.variable.unique():
            if not ics_stack[(ics_stack.name == S_core_name) & (ics_stack.variable == f_prop)].empty:
                ics_stack = ics_stack[(ics_stack.name != S_core_name) | (ics_stack.variable != f_prop)]
            ics_stack = ics_stack.append(prop_profile, ignore_index=True)
        return ics_stack
    else:
        return prop_profile


def set_profile_orientation(profile, v_ref, hi=None, comment=False):
    profile = CoreStack(profile)
    for variable in profile.variable.unique():
        data = CoreStack(profile[profile.variable == variable])
        # look for ice thickness:
        if hi is None:
            if not np.isnan(profile.ice_thickness.astype(float)).all():
                hi = profile.ice_thickness.astype(float).dropna().unique()
            elif not np.isnan(profile.ice_core_length.astype(float)).all():
                hi = profile.ice_core_length.astype(float).dropna().unique()
            else:
                module_logger.warning("%s ice core length and ice thickness not available for the profile" % profile.name.unique())
                return CoreStack()
        if comment is True:
            print(profile.name.unique()[0], variable, hi)
        if data.v_ref.unique().__len__() > 1:
            module_logger.error("vertical reference for profile are not consistent")
            return CoreStack()
        elif not data.v_ref.unique()[0] == v_ref:
            data['y_low'] = hi - data['y_low']
            data['y_mid'] = hi - data['y_mid']
            data['y_sup'] = hi - data['y_sup']
            data['v_ref'] = v_ref
        profile = profile.delete_profile({'name':profile.name.unique()[0], 'variable': variable})
        profile = profile.append(data)
    return CoreStack(profile)


def set_profile_vertical_reference_depth(profile, h_ref):
    profile['y_low'] = profile['y_low'] - h_ref
    profile['y_mid'] = profile['y_mid'] - h_ref
    profile['y_sup'] = profile['y_sup'] - h_ref
    return CoreStack(profile)


def set_vertical_reference_for_profile(profile, h_ref=0, new_v_ref=None):
    """

    :param profile:
    :param h_ref:
    :param new_v_ref: default, same as profile origin
    :return:
    """

    if new_v_ref is None:
        if profile.v_ref.unique().__len__() > 1:
            module_logger.error("vertical reference for profile are not consistent")
            return CoreStack()
        else:
            new_v_ref = profile.v_ref.unique()[0]

    # look for ice thickness:
    if not np.isnan(profile.ice_thickness.astype(float)).all():
        hi = profile.ice_thickness.astype(float).dropna().unique()
    elif not np.isnan(profile.ice_core_length.astype(float)).all():
        hi = profile.ice_core_length.astype(float).dropna().unique()
    else:
        module_logger.warning("ice core length and ice thickness not available for the profile")
        return CoreStack()


    if not new_v_ref == profile.v_ref.unique()[0]:
        profile['y_low'] = hi - profile['y_low']
        profile['y_mid'] = hi - profile['y_mid']
        profile['y_sup'] = hi - profile['y_sup']

    if not h_ref == 0:
        profile['y_low'] = profile['y_low']-h_ref
        profile['y_mid'] = profile['y_mid']-h_ref
        profile['y_sup'] = profile['y_sup']-h_ref


# DEPRECATED bottom_reference, use ics_stack = ics_stack.set_reference('bottom') instead
def bottom_reference(ics_stack, comment='n', status='DEPRECATED'):
    logging.error('DEPRECATION ERROR bottom reference is deprecated, use ics_stack.set_reference("bottom") instead')
    return None
# def bottom_reference(ics_stack, comment = 'n'):
#     logging.error('DEPRECATION ERROR bottom reference is deprecated, use ics_stack.set_reference("bottom") instead')
#     for f_core in ics_stack.name.unique():
#         ic_data = ics_stack[ics_stack.name == f_core]
#
#         hi = []
#         if not np.isnan(ic_data.ice_thickness.astype(float)).all():
#             hi.append(ic_data.ice_thickness.astype(float).dropna().unique()[0])
#         elif not np.isnan(ic_data.ice_core_length.astype(float)).all():
#             hi.append(ic_data.ice_core_length.astype(float).dropna().unique()[0])
#         else:
#             for ff_core in ic_data.collection.iloc[0]:
#                 if not np.isnan(ics_stack[ics_stack.name == ff_core].ice_thickness.astype(float)).all():
#                     hi.append(ics_stack[ics_stack.name == ff_core].ice_thickness.astype(float).dropna().unique()[0])
#                 if not np.isnan(ics_stack[ics_stack.name == ff_core].ice_core_length.astype(float)).all():
#                     hi.append(ics_stack[ics_stack.name == ff_core].ice_core_length.astype(float).dropna().unique()[0])
#         if not hi == []:
#             hi = np.nanmax(hi)
#         else:
#             hi = np.nan
#
#         if not np.isnan(hi):
#             for f_variable in ic_data.variable.unique():
#                 ics_stack.loc[(ics_stack.name == f_core) & (ics_stack.variable == f_variable), 'y_low'] = hi - ics_stack.loc[(ics_stack.name == f_core) & (ics_stack.variable == f_variable), 'y_low']
#                 ics_stack.loc[(ics_stack.name == f_core) & (ics_stack.variable == f_variable), 'y_mid'] = hi - ics_stack.loc[(ics_stack.name == f_core) & (ics_stack.variable == f_variable), 'y_mid']
#                 ics_stack.loc[(ics_stack.name == f_core) & (ics_stack.variable == f_variable), 'y_sup'] = hi - ics_stack.loc[(ics_stack.name == f_core) & (ics_stack.variable == f_variable), 'y_sup']
#         else:
#             ics_stack = ics_stack.remove_profiles(f_core)
#         if comment == 'y':
#             print(f_core, hi)
#         if 'index' in ics_stack.columns:
#             ics_stack.drop('index', axis=1)
#     return ics_stack


def discretize_profile(profile, y_bins=None, y_mid=None, variables=None, display_figure='y', fill_gap=False):
    """
    :param profile:
    :param y_bins:
    :param y_mid:
    :param variables:
    :param display_figure:
    :param fill_gap:
    :return:
    """
    v_ref = profile.v_ref.unique()[0]

    # VARIABLES CHECK
    if y_bins is None and y_mid is None:
        y_bins = pd.Series(profile.y_low.dropna().tolist() + profile.y_sup.dropna().tolist()).sort_values().unique()
        y_mid = profile.y_mid.dropna().sort_values().unique()
    elif y_bins is None:
        if y_mid is not None:
            y_mid = y_mid.sort_values().values
            dy = np.diff(y_mid)/2
            y_bins = np.concatenate([[y_mid[0]-dy[0]], y_mid[:-1] + dy, [y_mid[-1]+dy[-1]]])
            if y_bins[0] < 0:
                y_bins[0] = 0
    elif y_mid is None:
        if y_bins is not None:
            y_mid = np.diff(y_bins)/2+y_bins[:-1]
        else:
            y_mid = profile.y_mid.dropna().sort_values().unique()

    if variables is None:
        variables = [variable for variable in profile.variable.unique().tolist() if variable in profile.keys()]

    if not isinstance(variables, list):
        variables = [variables]


    for variable in variables:
        # continuous profile (temperature-like)
        profile[variable] = pd.to_numeric(profile[variable])
        if (profile[profile.variable == variable].y_low.isnull().all() and
                    profile[profile.variable == variable].y_low.__len__() > 0):
            yx = profile[profile.variable == variable].set_index('y_mid').sort_index()[[variable]]
            y2x = yx.reindex(y_mid)
            for index in yx.index:
                y2x.loc[abs(y2x.index - index) < 1e-6, variable] = yx.loc[yx.index == index, variable].values
            if np.isnan(y2x[variable].astype(float)).all():
                dat_temp = np.interp(y2x.index, yx.index, yx[variable].astype(float), left=np.nan, right=np.nan)
                y2x = pd.DataFrame(dat_temp, index=y2x.index, columns=[variable])
            else:
                y2x.ix[(y2x.index <= max(yx.index)) & (min(yx.index) <= y2x.index)] = y2x.interpolate(method='index')[(y2x.index <= max(yx.index)) & (min(yx.index) <= y2x.index)]
            temp = pd.DataFrame(columns=profile.columns.tolist(), index=range(y_mid.__len__()))
            temp.update(y2x.reset_index())

            profile_prop = profile.head(1)
            profile_prop = profile_prop.drop(variable, 1)
            profile_prop = profile_prop.drop('y_low', 1)
            profile_prop = profile_prop.drop('y_mid', 1)
            profile_prop = profile_prop.drop('y_sup', 1)
            temp.update(pd.DataFrame([profile_prop.iloc[0].tolist()], columns=profile_prop.columns.tolist(),
                                     index=temp.index.tolist()))
            temp['date'] = temp['date'].astype('datetime64[ns]')

            if display_figure == 'y' or display_figure == 'c':
                if display_figure == 'c':
                    plt.close()
                plt.figure()
                yx = yx.reset_index()
                plt.plot(yx[variable], yx['y_mid'], 'k')
                plt.plot(temp[variable], temp['y_mid'], 'xr')
                plt.title(profile_prop.name.unique()[0] + ' - ' + variable)

        # step profile (salinity-like)
        elif (not profile[profile.variable == variable].y_low.isnull().any() and
              profile[profile.variable == variable].y_low.__len__() > 0):
            if v_ref == 'bottom':
                yx = profile[profile.variable == variable].set_index('y_mid', drop=False).sort_index().as_matrix(
                   ['y_sup', 'y_low', variable])
            else:
                yx = profile[profile.variable == variable].set_index('y_mid', drop=False).sort_index().as_matrix(
                    ['y_low', 'y_sup', variable])
            x_step = []
            y_step = []
            ii_bin = 0
            if yx[0, 0] < y_bins[0]:
                ii_yx = np.where(yx[:, 0] <= y_bins[0])[0][-1]
            else:
                ii_yx = 0
                while y_bins[ii_bin] < yx[ii_yx, 0]:
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(np.nan)
                    x_step.append(np.nan)
                    ii_bin += 1
                    y_bins[ii_bin]

            while ii_bin < y_bins.__len__() - 1:
                while y_bins[ii_bin + 1] <= yx[ii_yx, 1]:
                    S = s_nan(yx, ii_yx, fill_gap)
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(S)
                    x_step.append(S)
                    ii_bin += 1

                    if ii_bin == y_bins.__len__() - 1:
                        break

                L = (yx[ii_yx, 1] - y_bins[ii_bin])
                S = (yx[ii_yx, 1] - y_bins[ii_bin]) * s_nan(yx, ii_yx, fill_gap)

                while ii_yx < len(yx[:, 1]) - 1 and yx[ii_yx + 1, 1] <= y_bins[ii_bin + 1]:
                    L += (yx[ii_yx + 1, 1] - yx[ii_yx + 1, 0])
                    S += (yx[ii_yx + 1, 1] - yx[ii_yx + 1, 0]) * s_nan(yx, ii_yx + 1, fill_gap)
                    ii_yx += 1
                    if ii_yx == yx[:, 1].__len__() - 1:
                        break

                if ii_bin + 1 == y_bins.__len__():
                    break

                if yx[ii_yx, 1] <= y_bins[ii_bin + 1] and ii_yx + 1 < yx.__len__():
                    L += (y_bins[ii_bin + 1] - yx[ii_yx + 1, 0])
                    S += (y_bins[ii_bin + 1] - yx[ii_yx + 1, 0]) * s_nan(yx, ii_yx + 1, fill_gap)
                S = S / L
                if S != 0:
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(S)
                    x_step.append(S)
                    ii_bin += 1
                ii_yx += 1
                if y_bins[ii_bin] >= yx[-1, 1]:
                    while ii_bin + 1 < y_bins.__len__():
                        y_step.append(y_bins[ii_bin])
                        y_step.append(y_bins[ii_bin + 1])
                        x_step.append(np.nan)
                        x_step.append(np.nan)
                        ii_bin += 1

            temp = pd.DataFrame(columns=profile.columns.tolist(), index=range(y_bins[:-1].__len__()))
            temp.update(pd.DataFrame(np.vstack((y_bins[:-1], y_bins[:-1] + np.diff(y_bins) / 2, y_bins[1:],
                                                [x_step[2 * ii] for ii in
                                                 range(int(x_step.__len__() / 2))])).transpose(),
                                     columns=['y_low', 'y_mid', 'y_sup', variable], index=temp.index))

            # properties
            profile_prop = profile.head(1)
            profile_prop = profile_prop.drop(variable, 1)
            profile_prop = profile_prop.drop('y_low', 1)
            profile_prop = profile_prop.drop('y_mid', 1)
            profile_prop = profile_prop.drop('y_sup', 1)
            temp.update(pd.DataFrame([profile_prop.iloc[0].tolist()], columns=profile_prop.columns.tolist(),
                                     index=temp.index.tolist()))
            temp['date'] = temp['date'].astype('datetime64[ns]')

            if display_figure == 'y' or display_figure == 'c':
                if display_figure == 'c':
                    plt.close()
                plt.figure()
                x = []
                y = []
                for ii in range(yx[:, 0].__len__()):
                    y.append(yx[ii, 0])
                    y.append(yx[ii, 1])
                    x.append(yx[ii, 2])
                    x.append(yx[ii, 2])
                plt.step(x, y, 'bx')
                plt.step(x_step, y_step, 'ro')
                plt.title(profile_prop.name.unique()[0] + ' - ' + variable)

        profile = profile[(profile.name != profile.name.unique().tolist()[0]) | (profile.variable != variable)]
        profile = profile.append(temp)

        if 'index' in profile.columns:
            profile.drop('index', axis=1)
    return CoreStack(profile)


def discretize_profileV2(profile, y_bins=None, y_mid=None, variables=None, display_figure='y', fill_gap=True):
    """
    :param profile:
    :param y_bins:
    :param y_mid:
    :param variables:
    :param display_figure:
    :param fill_gap:
    :return:
    """

    if profile.empty:
        return CoreStack(profile)

    v_ref = profile.v_ref.unique()[0]

    # VARIABLES CHECK
    if y_bins is None and y_mid is None:
        y_bins = pd.Series(profile.y_low.dropna().tolist() + profile.y_sup.dropna().tolist()).sort_values().unique()
        y_mid = profile.y_mid.dropna().sort_values().unique()
    elif y_bins is None:
        if y_mid is not None:
            y_mid = y_mid.sort_values().values
            dy = np.diff(y_mid)/2
            y_bins = np.concatenate([[y_mid[0]-dy[0]], y_mid[:-1] + dy, [y_mid[-1]+dy[-1]]])
            if y_bins[0] < 0:
                y_bins[0] = 0
    elif y_mid is None:
        if y_bins is not None:
            y_mid = np.diff(y_bins)/2+y_bins[:-1]
        else:
            y_mid = profile.y_mid.dropna().sort_values().unique()

    y_bins = np.array(y_bins)
    y_mid = np.array(y_mid)

    if variables is None:
        variables = [variable for variable in profile.variable.unique().tolist() if variable in profile.keys()]

    if not isinstance(variables, list):
        variables = [variables]

    discretized_profile = pd.DataFrame()

    module_logger.debug("Processing %s" %profile.name.unique()[0])
    #print("Processing %s" %profile.name.unique()[0])

    for variable in variables:
        profile[variable] = pd.to_numeric(profile[variable])
        temp = pd.DataFrame()

        if profile[profile.variable == variable].empty:
            module_logger.debug("no %s data" % (variable))
        else:
            module_logger.debug("%s data discretized" % variable)
            #print("\t%s data discretized" % (variable))
        # continuous profile (temperature-like)
        if (profile[profile.variable == variable].y_low.isnull().all() and
                    profile[profile.variable == variable].y_low.__len__() > 0):
            yx = profile[profile.variable == variable].set_index('y_mid').sort_index()[[variable]]
            y2x = yx.reindex(y_mid)
            for index in yx.index:
                y2x.loc[abs(y2x.index - index) < 1e-6, variable] = yx.loc[yx.index == index, variable].values
            #if np.isnan(y2x[variable].astype(float)).all():
            dat_temp = np.interp(y2x.index, yx.index, yx[variable].astype(float), left=np.nan, right=np.nan)
            y2x = pd.DataFrame(dat_temp, index=y2x.index, columns=[variable])
            #else:
            #    y2x.ix[(y2x.index <= max(yx.index)) & (min(yx.index) <= y2x.index)] = y2x.interpolate(method='index')[(y2x.index <= max(yx.index)) & (min(yx.index) <= y2x.index)]
            temp = pd.DataFrame(columns=profile.columns.tolist(), index=range(y_mid.__len__()))
            temp.update(y2x.reset_index())

            profile_prop = profile.head(1)
            profile_prop = profile_prop.drop(variable, 1)
            profile_prop['variable'] = variable
            profile_prop = profile_prop.drop('y_low', 1)
            profile_prop = profile_prop.drop('y_mid', 1)
            profile_prop = profile_prop.drop('y_sup', 1)
            temp.update(pd.DataFrame([profile_prop.iloc[0].tolist()], columns=profile_prop.columns.tolist(),
                                     index=temp.index.tolist()))
            temp['date'] = temp['date'].astype('datetime64[ns]')

            if display_figure == 'y' or display_figure == 'c':
                if display_figure == 'c':
                    plt.close()
                plt.figure()
                yx = yx.reset_index()
                plt.plot(yx[variable], yx['y_mid'], 'k')
                plt.plot(temp[variable], temp['y_mid'], 'xr')
                plt.title(profile_prop.name.unique()[0] + ' - ' + variable)

        # step profile (salinity-like)
        elif (not profile[profile.variable == variable].y_low.isnull().all() and
              profile[profile.variable == variable].y_low.__len__() > 0):
            if v_ref == 'bottom':
                yx = profile[profile.variable == variable].set_index('y_mid', drop=False).sort_index().as_matrix(
                   ['y_sup', 'y_low', variable])
                if yx[0, 0] > yx [0, 1]:
                    yx = profile[profile.variable == variable].set_index('y_mid', drop=False).sort_index().as_matrix(
                        ['y_low', 'y_sup', variable])
            else:
                yx = profile[profile.variable == variable].set_index('y_mid', drop=False).sort_index().as_matrix(
                    ['y_low', 'y_sup', variable])
            x_step = []
            y_step = []
            ii_bin = 0
            if yx[0, 0] < y_bins[0]:
                ii_yx = np.where(yx[:, 0] - y_bins[0] <= TOL)[0][-1]
            else:
                ii_bin = np.where(y_bins - yx[0, 0] <= TOL)[0][-1]
                ii_yx = 0
                ii = 0
                while ii < ii_bin:
                     y_step.append(y_bins[ii])
                     y_step.append(y_bins[ii + 1])
                     x_step.append(np.nan)
                     x_step.append(np.nan)
                     ii += 1

            while ii_bin < y_bins.__len__() - 1 :
                while ii_bin+1 < y_bins.__len__() and y_bins[ii_bin + 1] - yx[ii_yx, 1] <= TOL :
                    S = s_nan(yx, ii_yx, fill_gap)
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(S)
                    x_step.append(S)
                    ii_bin += 1
                    #plt.step(x_step, y_step, 'ro')
                    #if ii_bin == y_bins.__len__() - 1:
                    #    break

                if not yx[-1, 1] - y_bins[ii_bin] <= TOL:
                    L = 0
                    S = 0
                    if ii_yx < yx[:, 0].__len__()-1:
                        while ii_yx < yx[:, 0].__len__()-1 and yx[ii_yx, 1] - y_bins[ii_bin+1] <= TOL:
                            L += (yx[ii_yx, 1] - y_bins[ii_bin])
                            S += (yx[ii_yx, 1] - y_bins[ii_bin]) * s_nan(yx, ii_yx, fill_gap)
                            ii_yx +=1

                    # ABOVE
                    #while ii_yx < len(yx[:, 1]) - 1 and yx[ii_yx + 1, 1] - y_bins[ii_bin + 1] <= TOL:
                    #    L += (yx[ii_yx + 1, 1] - yx[ii_yx + 1, 0])
                    #    S += (yx[ii_yx + 1, 1] - yx[ii_yx + 1, 0]) * s_nan(yx, ii_yx + 1, fill_gap)
                    #    ii_yx += 1
                    #    if ii_yx == yx[:, 1].__len__() - 1:
                    #        break
                    #   break
                        if yx[ii_yx, 0] - y_bins[ii_bin+1] <= TOL :
                            S += (y_bins[ii_bin+1] - yx[ii_yx, 0]) * s_nan(yx, ii_yx, fill_gap)
                            L += y_bins[ii_bin+1] - yx[ii_yx, 0]
                        if L > TOL:
                            S = S / L
                        else:
                            S = np.nan

                    else:
                        S = yx[-1, -1]
                        #y_step.append(y_bins[ii_bin])
                        #y_step.append(y_bins[ii_bin + 1])
                        #x_step.append(S)
                        #x_step.append(S)
                        #ii_bin += 1
                    # ABOVE
                    # if yx[ii_yx, 1] - y_bins[ii_bin + 1] <= TOL and ii_yx + 1 < yx.__len__():
                    #     if np.isnan(s_nan(yx, ii_yx + 1, fill_gap)) and not np.isnan(S) and y_bins[ii_bin + 1] - yx[ii_yx+1, 1] < TOL:
                    #         S += S/L*(y_bins[ii_bin + 1] - yx[ii_yx + 1, 0])
                    #     else:
                    #         S += (y_bins[ii_bin + 1] - yx[ii_yx + 1, 0]) * s_nan(yx, ii_yx + 1, fill_gap)
                    #     L += (y_bins[ii_bin + 1] - yx[ii_yx + 1, 0])

                    #if S != 0 : #and y_bins[ii_bin] - yx[ii_yx, 1] < TOL:
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(S)
                    x_step.append(S)
                    ii_bin += 1
                    #plt.step(x_step, y_step, 'ro')

                else:
                    while ii_bin + 1 < y_bins.__len__():
                        y_step.append(y_bins[ii_bin])
                        y_step.append(y_bins[ii_bin + 1])
                        x_step.append(np.nan)
                        x_step.append(np.nan)
                        ii_bin += 1

            temp = pd.DataFrame(columns=profile.columns.tolist(), index=range(np.unique(y_step).__len__()-1))
            temp.update(pd.DataFrame(np.vstack((np.unique(y_step)[:-1], np.unique(y_step)[:-1] + np.diff(np.unique(y_step)) / 2, np.unique(y_step)[1:],
                                                [x_step[2 * ii] for ii in
                                                 range(int(x_step.__len__() / 2))])).transpose(),
                                     columns=['y_low', 'y_mid', 'y_sup', variable], index=temp.index[0:np.unique(y_step).__len__()-1]))

            # properties
            profile_prop = profile.head(1)
            profile_prop = profile_prop.drop(variable, 1)
            profile_prop['variable'] = variable
            profile_prop = profile_prop.drop('y_low', 1)
            profile_prop = profile_prop.drop('y_mid', 1)
            profile_prop = profile_prop.drop('y_sup', 1)
            temp.update(pd.DataFrame([profile_prop.iloc[0].tolist()], columns=profile_prop.columns.tolist(),
                                     index=temp.index.tolist()))
            temp['date'] = temp['date'].astype('datetime64[ns]')

            if display_figure == 'y' or display_figure == 'c':
                if display_figure == 'c':
                    plt.close()
                plt.figure()
                x = []
                y = []
                for ii in range(yx[:, 0].__len__()):
                    y.append(yx[ii, 0])
                    y.append(yx[ii, 1])
                    x.append(yx[ii, 2])
                    x.append(yx[ii, 2])
                plt.step(x, y, 'bx')
                plt.step(x_step, y_step, 'ro')
                plt.title(profile_prop.name.unique()[0] + ' - ' + variable)

            #profile = profile[(profile.name != profile.name.unique().tolist()[0]) | (profile.variable != variable)]
        discretized_profile = discretized_profile.append(temp)

            #if temp.empty:
            #    print(profile.name.unique())
            #else:
            #    profile = profile.append(temp)

            #if 'index' in discretized_profile.columns:
            #    discretized_profile.drop('index', axis=1)
    return CoreStack(discretized_profile)


# HELPER
def s_nan(yx, ii_yx, fill_gap=True):
    """
    :param yx:
    :param ii_yx:
    :param fill_gap:
    :return:
    """
    if np.isnan(yx[ii_yx, 2]) and fill_gap:
        ii_yx_l = ii_yx - 1
        while ii_yx_l > 0 and np.isnan(yx[ii_yx_l, 2]):
            ii_yx_l -= 1
        s_l = yx[ii_yx_l, 2]

        ii_yx_s = ii_yx
        while ii_yx_s < yx.shape[0]-1 and np.isnan(yx[ii_yx_s, 2]):
            ii_yx_s += 1
        s_s = yx[ii_yx_s, 2]

        s = (s_s + s_l) / 2
    else:
        s = yx[ii_yx, 2]
    return s

#--------------------------#

# OK
def plot_profile_variable(ic_data, variable_dict, ax=None, param_dict=None):
    """
    :param ic_data:
        pd.DataFrame
    :param variable_dict:
    :param ax:
    :param param_dict:
    :return:
    """
    if 'variable' not in variable_dict.keys():
        module_logger.warning("a variable should be specified for plotting")
        return 0

    profile = select_profile(ic_data, variable_dict)
    ax = plot_profile(profile, ax=ax, param_dict=param_dict)
    return ax

def semilogx_profile_variable(ic_data, variable_dict, ax=None, param_dict=None):
    """
    :param ic_data:
        pd.DataFrame
    :param variable_dict:
    :param ax:
    :param param_dict:
    :return:
    """
    if 'variable' not in variable_dict.keys():
        module_logger.warning("a variable should be specified for plotting")
        return 0

    profile = select_profile(ic_data, variable_dict)
    ax = semilogx_profile(profile, ax=ax, param_dict=param_dict)
    return ax


# OK
def plot_mean_envelop(ic_data, variable_dict, ax=None, param_dict=None):
    """

    :param ic_data:
    :param variable_dict:
    :param ax:
    :param param_dict:
    :return:
    """
    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)

    if param_dict is None:
        param_dict = {}

    if 'variable' not in variable_dict.keys():
        module_logger.warning("a variable should be specified for plotting")
        return 0

    ii_variable = variable_dict['variable']

    variable_dict.update({'stats': 'mean'})
    x_mean = select_profile(ic_data, variable_dict).reset_index()
    variable_dict.update({'stats': 'std'})
    x_std = select_profile(ic_data, variable_dict).reset_index()

    if x_mean.__len__() != 0:
        if x_std.__len__() < x_mean.__len__():
            index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
            x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))

        if not x_mean.y_low.isnull().all():
            y_low = x_mean['y_low']
            y_sup = x_mean['y_sup']
            y = np.concatenate((y_low.tolist(), [y_sup.tolist()[-1]]))
            x_std_l = x_mean[ii_variable] - x_std[ii_variable]
            x_std_h = x_mean[ii_variable] + x_std[ii_variable]

            x_std_l = seaice.toolbox.plt_step(x_std_l.tolist(), y).transpose()
            x_std_h = seaice.toolbox.plt_step(x_std_h.tolist(), y).transpose()
        elif x_mean.y_low.isnull().all():
            y_std = x_mean['y_mid']
            x_std_l = np.array([x_mean[ii_variable] - np.nan_to_num(x_std[ii_variable]), y_std])
            x_std_h = np.array([x_mean[ii_variable] + np.nan_to_num(x_std[ii_variable]), y_std])

        if 'facecolor' not in param_dict.keys():
            param_dict['facecolor'] = {'black'}
        if 'alpha' not in param_dict.keys():
            param_dict['alpha'] = {0.3}
        if 'label' not in param_dict.keys():
            param_dict['label'] = str(r"$\pm$"+"std dev")
        ax.fill_betweenx(x_std_l[1, :], x_std_l[0, :], x_std_h[0, :], facecolor='black', alpha=0.2,
                         label=param_dict['label'])
    return ax


def semilogx_mean_envelop(ic_data, variable_dict, ax=None, param_dict=None):
    """

    :param ic_data:
    :param variable_dict:
    :param ax:
    :param param_dict:
    :return:
    """
    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)

    if param_dict is None:
        param_dict = {}

    if 'variable' not in variable_dict.keys():
        module_logger.warning("a variable should be specified for plotting")
        return 0

    ii_variable = variable_dict['variable']

    variable_dict.update({'stats': 'mean'})
    x_mean = select_profile(ic_data, variable_dict).reset_index()
    variable_dict.update({'stats': 'std'})
    x_std = select_profile(ic_data, variable_dict).reset_index()

    if x_mean.__len__() != 0:
        if x_std.__len__() < x_mean.__len__():
            index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
            x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))

        if not x_mean.y_low.isnull().all():
            y_low = x_mean['y_low']
            y_sup = x_mean['y_sup']
            y = np.concatenate((y_low.tolist(), [y_sup.tolist()[-1]]))
            x_std_l = x_mean[ii_variable] - x_std[ii_variable]
            x_std_h = x_mean[ii_variable] + x_std[ii_variable]

            index_outlier = x_std_l[(x_std_l <= 0)].index.tolist()
            for ii in index_outlier:
                l = ''
                for key in variable_dict:
                    l += key + ': ' + variable_dict[key] + '; '
                l = l[:-2]
                module_logger.warning('%s index of %s bin modified lower value for logarithmic scale' % (ii, l))

            ii_outlier = 1
            if index_outlier.__len__() > 0:
                variable_dict.update({'stats': 'min'})
                x_min = select_profile(ic_data, variable_dict).reset_index(drop=True)
                while index_outlier.__len__() > 0:
                    #for index in index_outlier:
                    x_std_l[(x_std_l <= 0)] = x_min.loc[x_min.index.isin(index_outlier), ii_variable] - x_std.loc[x_std.index.isin(index_outlier), ii_variable]/ii_outlier
                    index_outlier = x_std_l[(x_std_l <= 0)].index.tolist()
                    ii_outlier += 1

            x_std_l = seaice.toolbox.plt_step(x_std_l.tolist(), y).transpose()
            x_std_h = seaice.toolbox.plt_step(x_std_h.tolist(), y).transpose()
        elif x_mean.y_low.isnull().all():
            y_std = x_mean['y_mid']
            x_std_l = np.array([x_mean[ii_variable] - np.nan_to_num(x_std[ii_variable]), y_std])
            x_std_h = np.array([x_mean[ii_variable] + np.nan_to_num(x_std[ii_variable]), y_std])

        if 'facecolor' not in param_dict.keys():
            param_dict['facecolor'] = {'black'}
        if 'alpha' not in param_dict.keys():
            param_dict['alpha'] = {0.3}
        if 'label' not in param_dict.keys():
            param_dict['label'] = str(r"$\pm$"+"std dev")

        ax.fill_betweenx(x_std_l[1, :], x_std_l[0, :], x_std_h[0, :], facecolor='black', alpha=0.2,
                         label=param_dict['label'])
    return ax


# OK
def plot_number(ic_data, variable_dict, ax=None, position='right', x_delta=0.1, z_delta=0.05, every=1,
                fontsize=mpl.rcParams['font.size']):
    """
    :param ic_data:
    :param variable_dict:
    :param ax:
    :param position:
    :param x_delta:
    :param z_delta:
    :param every:
    :param fontsize:
    :return:
    """
    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)

    if 'variable' not in variable_dict.keys():
        module_logger.warning("a variable should be specified for plotting")
        return 0

    ii_variable = variable_dict['variable']

    if position == 'left':
        stat = 'min'
    elif position == 'center':
        stat = 'mean'
    else:
        stat = 'max'

    depth = select_profile(ic_data, variable_dict).reset_index()['y_mid'].values
    n = select_profile(ic_data, variable_dict).reset_index()['n'].values
    variable_dict.update({'stats':stat})
    pos = select_profile(ic_data, variable_dict).reset_index()[ii_variable].values

    # check for nan value:
    depth = depth[~np.isnan(pos)]
    n = n[~np.isnan(pos)]
    pos = pos[~np.isnan(pos)]

    for ii in np.arange(0, pos.__len__(), every):
        ax.text(pos[ii]+x_delta, depth[ii]+z_delta, str('(%.0f)' % n[ii]), fontsize=fontsize)

    return ax


def ice_core_stat(ics_subset, variables, stats, ic_subset_name = 'average core'):
    """
    :param ics_subset:
    :param variables:
    :param stats:
        accept as statistical function all the main function of pandas.
    :param ic_subset_name:
    :return:
    """
    if 'y_low' not in ics_subset.keys() or np.isnan(ics_subset['y_low'].values.astype(float)).all():
        y_bins = np.unique(ics_subset.y_mid.values)
        y_bins = list(y_bins[:-1]-np.diff(y_bins)/2) + list(y_bins[-2:]+np.diff(y_bins)[-2:]/2)
    else:
        y_bins = np.unique(np.concatenate((ics_subset.y_low.values, ics_subset.y_sup.values)))
    y_cuts = pd.cut(ics_subset.y_mid, y_bins, labels=False)
    data_grouped = ics_subset.groupby([y_cuts])

    data = CoreStack()
    for ii_variable in variables:

        # core data
        columns = ['y_low', 'y_sup', 'y_mid']
        if not ics_subset[ics_subset.variable == ii_variable].y_low.isnull().any():
            data_core = [[y_bins[ii_layer], y_bins[ii_layer + 1], (y_bins[ii_layer] + y_bins[ii_layer + 1]) / 2] for
                    ii_layer in range(0, y_bins.__len__() - 1)]
        elif ics_subset[ics_subset.variable == ii_variable].y_low.isnull().all():
            data_core = [[np.nan, np.nan, (y_bins[ii_layer] + y_bins[ii_layer + 1]) / 2] for ii_layer in
                    range(0, y_bins.__len__() - 1)]
        data_core = pd.DataFrame(data_core, columns=columns)
        data_core['name'] = ic_subset_name
        data_core['variable'] = ii_variable

        # stat variable
        for ii_stat in stats:
            print('computing %s' % ii_stat)
            func = "groups['" + ii_variable + "']." + ii_stat + "()"
            data_stat = [[None, None, None] for x in range(y_bins.__len__())]
            for k1, groups in data_grouped:
                data_stat[k1][0] = eval(func)
                temp = list(groups.dropna(subset=[ii_variable])['name'].unique())
                data_stat[k1][1] = temp.__len__()
                data_stat[k1][2] =', '.join(temp)

            data_stat = pd.DataFrame(data_stat[:-1], columns=[ii_variable, 'n', 'core_collection'])
            data_stat['stat'] = ii_stat
            if data.empty:
                data = pd.concat([data_core, data_stat], axis=1)
            else:
                print(ii_stat)
                data = data.append(pd.concat([data_core, data_stat], axis=1), ignore_index=True)
        data.reset_index()
    return data


def DD_fillup(ics_stack, DD, freezup_dates):
    import datetime

    for f_day in ics_stack.date.unique():
        # look for freezup_day
        if isinstance(f_day, np.datetime64):
            f_day = datetime.datetime.utcfromtimestamp((f_day - np.datetime64('1970-01-01T00:00:00')) / np.timedelta64(1, 's'))
        f_day = datetime.datetime(f_day.year, f_day.month, f_day.day)
        if f_day < datetime.datetime(f_day.year, 9, 1):
            freezup_day = datetime.datetime.fromordinal(freezup_dates[f_day.year])
        else:
            freezup_day = datetime.datetime.fromordinal(freezup_dates[f_day.year+1])

        # look for number of freezing/thawing degree day:
        if DD[f_day][1] < 0:
            ics_stack.loc[ics_stack.date == f_day, 'DD'] = DD[f_day][1]
        else:
            ics_stack.loc[ics_stack.date == f_day, 'DD'] = DD[f_day][0]

        ics_stack.loc[ics_stack.date == f_day, 'FDD'] = DD[f_day][0]
        ics_stack.loc[ics_stack.date == f_day, 'TDD'] = DD[f_day][1]
        ics_stack.loc[ics_stack.date == f_day, 'freezup_day'] = ['a']
        ics_stack.loc[ics_stack.date == f_day, 'freezup_day'] = [freezup_day]
    return CoreStack(ics_stack)


def stack_DD_fud(ics_data, DD, freezup_dates):
    ics_data_stack = CoreStack()
    for ii_core in ics_data.keys():
        core = ics_data[ii_core]
        ics_data_stack = ics_data_stack.add_profiles(core.profiles)

    for ii_day in ics_data_stack.date.unique():
        variable_dict = {'date': ii_day}
        ii_day = pd.DatetimeIndex([ii_day])[0].to_datetime()

        # freezup day:
        if ii_day < datetime.datetime(ii_day.year, 9, 1):
            freezup_day = datetime.datetime.fromordinal(freezup_dates[ii_day.year - 1])
        else:
            freezup_day = datetime.datetime.fromordinal(freezup_dates[ii_day.year])
        # DD
        if DD[ii_day][1] < 0:
            data = [[DD[ii_day][0], DD[ii_day][1], DD[ii_day][1], np.datetime64(freezup_day)]]
        else:
            data = [[DD[ii_day][0], DD[ii_day][1], DD[ii_day][0], np.datetime64(freezup_day)]]
        data_label = ['date', 'FDD', 'TDD', 'DD', 'freezup_day']
        data = pd.DataFrame(data, columns=data_label)

        ics_data_stack = ics_data_stack.add_variable(variable_dict, data)
    return ics_data_stack





def plot_state_variable(profile_stack, ax=None, variables='state variables', color_map='core'):
    """
    :param profile_stack:
    :param ax:
    :param variables:
        default 'state variables' which plot salinity and temperature
    :param color:
    :return:
    """
    if variables == None:
        variables = np.unique(profile_stack.variable).tolist()
    elif not isinstance(variables, list):
        variables = [variables]
    elif variables == 'state variables':
        variables = ['salinity, temperature']

    if ax is None:
        fig = plt.figure()
        ax = []
        for ii in range(len(variables)):
            ax.append(plt.subplot(1, len(variables), ii+1))
    elif len(ax) != len(variables):
        logging.warning('ax (len %d) and variables (len %d) should be of same size' %(len(ax), len(variables)))
        return None

    # colors
    color = {}
    if color_map is 'core':
        n_core = np.unique(profile_stack.core_name).__len__()
        color[ii] = [cm.jet(float(ii)/n_core) for ii in n_core]
    elif color_map is 'year':
        n_year = pd.unique([ii.year for ii in profile_stack.date]).__len__()
        color[ii] = [cm.jet(float(ii)/n_year) for ii in n_year]

    for ii in len(variables):
        var = variables[ii]


def drop_profile(data, core_name, keys):
    data = data[(data.core_name != core_name) | (data.variable != keys)]
    return data

## DEPRECATED make_section, replace with discretize_profile
def make_section(core, variables=None, status='DEPRECATED', section_thickness=0.05):
    logging.error('DEPRECATION ERROR bottom reference is deprecated, use ics_stack.set_reference("bottom") instead')
    return None
# def make_section(core, variables=None, section_thickness=0.05):
#     """
#     :param core:
#     :param variables:
#     :param section_thickness:
#     """
#     if variables is None:
#         variables = sorted(core.profiles.keys())
#     if not isinstance(variables, list):
#         variables = [variables]
#
#     for ii_profile in variables:
#         profile = core.profiles[ii_profile]
#         if core.ice_thickness is not None and ~np.isnan(core.ice_thickness):
#             ice_thickness = core.ice_thickness
#         else:
#             ice_thickness = core.profiles.loc[core.profiles.variable=='salinity', 'ice_core_length'].unique()
#
#         y_mid_section = np.arange(section_thickness / 2, ice_thickness, section_thickness)
#         delta_y = (ice_thickness + len(y_mid_section) * section_thickness) / 2
#
#         if delta_y < ice_thickness:
#             y_mid_section = np.append(y_mid_section, np.atleast_1d(delta_y))
#         x = np.array(core.profiles[ii_profile])
#         y = np.array(core.profiles[ii_profile])
#
#         if len(y) is len(x) + 1:
#             y = (y[1:] + y[:-1]) / 2
#
#         x_mid_section = np.interp(y_mid_section, y[~np.isnan(y)], x[~np.isnan(y)], left=np.nan, right=np.nan)
#
#         profile.x = x_mid_section
#         profile.y = y_mid_section
#         core.add_comment(
#             'artificial section thickness computed with a vertical resolution of ' + str(section_thickness) + 'm')
#         core.del_profile(ii_profile)
#         core.add_profile(profile)
#     return core


def get_filepath(data_dir, data_ext, subdir='no'):
    """
    :param data_dir:
    :param data_ext:
    :param subdir:
    """
    import os
    filepath = []
    if subdir == 'yes':
        for path, subdirs, files in os.walk(data_dir):
            subdirs.sort()
            files.sort()
            for name in files:
                if name.endswith(data_ext):
                    f = os.path.join(path, name)
                    filepath.append(f)
    else:
        files = os.listdir(data_dir)
        for name in files:
            if name.endswith(data_ext):
                f = os.path.join(data_dir, name)
                filepath.append(f)
    return filepath


def generate_source(data_dir, data_ext):
    """
    :param data_dir:
    :param data_ext:
    """
    import glob

    files = []
    for root, dirnames, filenames in os.walk(data_dir):
        files.extend(glob.glob(root + "/*."+data_ext))
    with open(data_dir + '/ic_list.txt', 'w') as f:
        for f_file in files:
            f.write(f_file + "\n")


def scale_profile(profile, h_ice_f):
    """
    :param profile:
        CoreStack, ice core profile to scale to a target ice thickness
    :param h_ice_f:
        scalar, target ice thickness
    :return:
    """

    if profile.ice_core_length.unique().size and ~np.isnan(profile.ice_core_length.unique()[0]):
        h_ice = profile.ice_core_length.unique()[0]
    elif profile.ice_thickness.unique().size and ~np.isnan(profile.ice_thickness.unique()[0]):
        h_ice = profile.ice_thickness.unique()[0]
    else:
        logging.error("Scale: no core length or ice thickness given for %s" % profile.core_name.unique()[0])
        return 0

    r = h_ice_f / h_ice
    if r == 1:
        return profile
    profile[['y_low', 'y_mid', 'y_sup']] = r * profile[['y_low', 'y_mid', 'y_sup']]
    profile.ice_core_length = h_ice_f
    return profile


# DEPRECATED
def compute_phys_prop(ics_data, si_prop, S_core_name, T_core_name, si_prop_format='linear', resize_core=None):
    """
    :param ic_data:
        dict of ice core
    :param si_prop:
        physical properties or list of physical properties
    :param si_prop_format: 'linear' or 'step'
    :return:
    """
    print("comput_phys_prop is deprecated. use compute_phys_from_core_name instead")
    if not isinstance(si_prop, list):
        si_prop = [si_prop]

    ## function variable:
    property_stack = pd.DataFrame()

    ## check parameters
    if S_core_name not in ics_data.keys() or 'salinity' not in ics_data[S_core_name].profile.variable.unique():
        print("missing salinity core")
        return property_stack;
    else:
        s_data = ics_data[S_core_name].profile
        s_data = s_data.loc[s_data.variable == 'salinity']
    if T_core_name not in ics_data.keys() or 'temperature' not in ics_data[T_core_name].profile.variable.unique():
        print("missing temperature core")
        return property_stack;
    else:
        t_data = ics_data[T_core_name].profile
        t_data = t_data.loc[t_data.variable == 'temperature', ['y_mid', 'temperature']]

    # interpolate temperature profile to match salinity profile
    y_mid = s_data.y_mid.dropna().tolist()
    if y_mid.__len__() < 1:
        y_mid = (s_data.y_low/2+s_data.y_sup/2).tolist()

    interp_data = pd.concat([t_data, pd.DataFrame(y_mid, columns=['y_mid'])])
    interp_data = interp_data.set_index('y_mid').sort_index().interpolate(method='index').reset_index().drop_duplicates(subset='y_mid')

    data = s_data
    if 'temperature' in s_data.keys():
        data = s_data.drop('temperature', axis=1)
    data = pd.merge(data, interp_data, on=['y_mid'])


    for f_prop in si_prop:
        if f_prop not in seaice.properties.si_prop_list.keys():
            print('property %s not defined in the ice core property module' % property)

        prop = seaice.properties.si_prop_list[f_prop]
        function = getattr(seaice.properties, prop.replace(" ", "_"))
        prop_data = function(data['temperature'], data['salinity'])

        property_frame = pd.DataFrame(np.vstack((prop_data, y_mid)).transpose(), columns=[prop, 'y_mid'])
        property_frame['name'] = list(set(s_data.name))[0]
        property_frame['comment_core'] = 'physical properties computed from ' + S_core_name + '(S) and ' + T_core_name + '(T)'
        property_frame['variable'] = prop

        var_drop = [var for var in ['salinity', 'temperature', 'variable', 'name', 'core'] if var in s_data.keys()]
        core_frame = s_data.drop(var_drop, axis=1)

        if si_prop_format == 'linear':
            core_frame = core_frame.drop(['y_sup', 'y_low'], axis=1)

        prop_data = pd.merge(property_frame, core_frame, how='inner', on=['y_mid'])

        property_stack = property_stack.append(prop_data, ignore_index=True, verify_integrity=False)

    return property_stack

def grouped_stat(ics_stack, variables, stats, bins_DD, bins_y, comment=False):
    ics_stack = ics_stack.reset_index(drop=True)
    y_cuts = pd.cut(ics_stack.y_mid, bins_y, labels=False)
    t_cuts = pd.cut(ics_stack.DD, bins_DD, labels=False)

    if not isinstance(variables, list):
        variables = [variables]
    if not isinstance(stats, list):
        stats = [stats]

    temp_all = pd.DataFrame()
    for ii_variable in variables:
        if comment:
            print('\ncomputing %s' % ii_variable)
        data = ics_stack[ics_stack.variable == ii_variable]
        data_grouped = data.groupby([t_cuts, y_cuts])

        for ii_stat in stats:
            if comment:
                print('\tcomputing %s' % ii_stat)
            func = "groups['" + ii_variable + "']." + ii_stat + "()"
            stat_var = np.nan * np.ones((bins_DD.__len__() - 1, bins_y.__len__()-1))
            core_var = [[[None] for x in range(bins_y.__len__())] for y in range(bins_DD.__len__() - 1)]
            for k1, groups in data_grouped:
                stat_var[int(k1[0]), int(k1[1])] = eval(func)
                core_var[int(k1[0])][int(k1[1])] = [list(groups.dropna(subset=[ii_variable])
                                                         ['name'].unique())]
            for ii_bin in range(stat_var.__len__()):
                temp = pd.DataFrame(stat_var[ii_bin], columns=[ii_variable])
                temp = temp.join(pd.DataFrame(core_var[ii_bin], columns=['core collection']))
                DD_label = 'DD-' + str(bins_DD[ii_bin]) + '_' + str(bins_DD[ii_bin + 1])
                data = [str(bins_DD[ii_bin]), str(bins_DD[ii_bin + 1]), DD_label, int(ii_bin), ii_stat,
                        ii_variable, ics_stack.v_ref.unique()[0]]
                columns = ['DD_min', 'DD_max', 'DD_label', 'DD_index', 'stats', 'variable', 'v_ref']
                index = np.array(temp.index.tolist())  #[~np.isnan(temp[ii_variable].tolist())]
                temp = temp.join(pd.DataFrame([data], columns=columns, index=index))
                temp = temp.join(pd.DataFrame(index, columns=['y_index'], index=index))
                for row in temp.index.tolist():
                    temp.loc[temp.index == row, 'n'] = temp.loc[temp.index == row, 'core collection'].__len__()
                columns = ['y_low', 'y_sup', 'y_mid']
                t2 = pd.DataFrame(columns=columns)
                # For step profile, like salinity
                # if ii_variable in ['salinity']:
                if not ics_stack[ics_stack.variable == ii_variable].y_low.isnull().any():
                    for ii_layer in index:
                        data = [bins_y[ii_layer], bins_y[ii_layer + 1],
                                (bins_y[ii_layer] + bins_y[ii_layer + 1]) / 2]
                        t2 = t2.append(pd.DataFrame([data], columns=columns, index=[ii_layer]))
                # For linear profile, like temperature
                # if ii_variable in ['temperature']:
                elif ics_stack[ics_stack.variable == ii_variable].y_low.isnull().all():
                    for ii_layer in index:
                        data = [np.nan, np.nan, (bins_y[ii_layer] + bins_y[ii_layer + 1]) / 2]
                        t2 = t2.append(pd.DataFrame([data], columns=columns, index=[ii_layer]))

                if temp_all.empty:
                    temp_all = temp.join(t2)
                else:
                    temp_all = temp_all.append(temp.join(t2), ignore_index=True)

    data_grouped = ics_stack.groupby([t_cuts, ics_stack['variable']])

    grouped_dict = {}
    for var in variables:
        grouped_dict[var] = [[] for ii_DD in range(bins_DD.__len__()-1)]

    for k1, groups in data_grouped:
        if k1[1] in variables:
            grouped_dict[k1[1]][int(k1[0])] = groups['name'].unique().tolist()

    return CoreStack(temp_all.reset_index(drop=True)), grouped_dict

