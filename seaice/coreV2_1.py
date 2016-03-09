#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
icecoredata.py: ice core data is a toolbox to import ice core data file from xlsx spreadsheet. Xlsx spreadsheet should
be formatted according to the template provided by the Sea Ice Group of the Geophysical Institute of University of
Alaska, Fairbanks.
"""
import numpy as np
import matplotlib.pyplot as plt
import logging
import openpyxl
import datetime
from seaice.properties import si_prop_list
from seaice.properties import si_prop_unit
from seaice.properties import si_state_variable

__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "2.1.0"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "marc.oggier@gi.alaska.edu"
__status__ = "development"
__date__ = "2014/11/25"
__all__ = ['Core', 'CoreSet', 'Profile']

LOG_FILENAME = 'import.log'
LOG_LEVELS = {'debug': logging.DEBUG,
              'info': logging.INFO,
              'warning': logging.WARNING,
              'error': logging.ERROR,
              'critical': logging.CRITICAL}
# logging.basicConfig(filename='example.log',level=logging.DEBUG)
# logging.debug('This message should go to the log file')
# logging.info('So should this')
# logging.warning('And this, too')
# ic_path = '/mnt/data_lvm/seaice/core/BRW/2010_11/BRW_CS-20110122A.xlsx'
unit = {'salinity': '[PSU]', 'temperature': '[°C]', 'vb': '[%]', 'brine volume fraction': '[%]'}
nan_value = float('nan')


class Profile:
    """
    Profile
    """
    def __init__(self, x, y, profile_label, variable, comment=None, note=None, length=None):
        self.x = x
        self.y = y
        self.profile_label = profile_label
        self.variable = variable
        self.note = note
        if length is None:
            self.length = x[-1] - x[0]
        else:
            self.length = length
        self.comment = comment

    def plot(self, fig_num=None):
        """
        :param fig_num:
        :return:
        """
        if fig_num is not None:
            plt.figure(fig_num)
        plt.plot(self.x, self.y[0:len(self.x)])


class Core:
    """
    Core
    """
    def __init__(self, name, date, location, ice_thickness, snow_thickness, comment=None):
        """
        :param name:
        :param date:
        :param location:
        :param ice_thickness:
        :param snow_thickness:
        :param comment:
        :return:
        """
        self.name = name
        self.date = date
        self.location = location
        self.corenames = [name]
        self.ice_thickness = ice_thickness
        self.snow_thickness = snow_thickness
        self.profiles = {}
        self.comment = None
        if comment is not None:
            self.add_comment(comment)

    def add_profile(self, profile, variable):
        """
        :param profile:
        :param variable:
        :return:
        """
        self.profiles[variable] = profile

    def del_profile(self, variable):
        self.profiles.pop(variable)

    def add_corenames(self, corename):
        """
        :param corename:
        :return:
        """
        if corename not in self.corenames:
            self.corenames.append(corename)

    def add_comment(self, comment):
        """
        :param comment:
        :return:
        """
        if self.comment is None:
            self.comment = [comment]
        else:
            self.comment.append(comment)

    def calc_prop(self, prop):
        """
        :param prop:
        :return:
        """
        # check properties variables
        if prop not in si_prop_list.keys():
            logging.warning('property %s not defined in the ice core property module' % prop)
            return None
        elif 'salinity' not in self.profiles:
            logging.warning('ice core %s is missing salinity profile for further calculation' % self.name)
            return None
        elif 'temperature' not in self.profiles:
            logging.warning('ice core %s is missing temperature profile for further calculation' % self.name)
            return None
        else:
            import seaice.properties
            variable = si_prop_list[prop]
            function = getattr(seaice.properties, variable.replace(" ", "_"))
            profilet = self.profiles['temperature']
            profiles = self.profiles['salinity']
            y = np.sort(np.unique(np.concatenate((profilet.y, profiles.y))))
            y = y[np.where(y <= max(max(profilet.y), max(profiles.y)))]
            y = y[np.where(y >= max(min(profilet.y), min(profiles.y)))]
            y_mid = y[0:-1] + np.diff(y) / 2
            length = max(y) - min(y)

            xt = np.interp(y_mid, profilet.y, profilet.x)
            xs = np.interp(y_mid, profiles.y[:-1] + np.diff(profiles.y) / 2, profiles.x)
            x = function(xt, xs)

            note = 'computed from ' + self.profiles['temperature'].profile_label + ' temperature profile and ' + \
                   self.profiles['salinity'].profile_label + ' salinity profile'
            profile_label = self.name
            prop_profile = Profile(x, y, profile_label, variable, comment=None, note=note, length=length)
            self.add_profile(prop_profile, variable)
            self.add_comment('computed %s' % variable)

    def plot(self, ax, prop, param_dict=None):
        """
        :param ax:
        :param prop:
        :param param_dict:
        :return:
        """
        if prop in si_state_variable.keys():
            profile_label = si_state_variable[prop.lower()]
        elif prop in si_prop_list.keys():
            profile_label = si_prop_list[prop.lower()]
            if profile_label not in self.profiles.keys():
                logging.warning('computing %s' % prop)
                self.calc_prop(prop)
        else:
            logging.warning('no data available to plot')
            return None
        profile = self.profiles[profile_label]

        y = profile.y
        x = profile.x
        if profile.y.__len__() > profile.x.__len__():
            x = np.concatenate((x, np.atleast_1d(x[-1])))
            if param_dict is None:
                out = ax.step(x, y)
            else:
                out = ax.step(x, y, **param_dict)
        else:
            if param_dict is None:
                out = ax.plot(x, y)
            else:
                out = ax.plot(x, y, **param_dict)

        ax.set_xlabel(profile_label + ' ' + si_prop_unit[profile_label])
        ax.set_ylim(max(ax.get_ylim()), 0)
        return out

    def plot_state_variable(self, flag_figure_number=None, param_dict=None):
        """
        :param param_dict:
        :return:
        """
        if flag_figure_number is None:
            fig, (ax1, ax2) = plt.subplots(1, 2)
        else:
            print(flag_figure_number)
            fig = plt.figure(flag_figure_number)
            ax1 = fig.add_subplot(1, 2, 1)
            ax2 = fig.add_subplot(1, 2, 2)

        if 'salinity' in self.profiles.keys():
            self.plot(ax1, 'salinity', param_dict)
            ax1.set_ylabel('depth [m]')
        else:
            logging.warning('salinity profile missing for %s' % self.name)

        if 'temperature' in self.profiles.keys():
            ax2 = self.plot(ax2, 'temperature', param_dict)
        else:
            logging.warning('temperature profile missing for %s' % self.name)
        return fig, (ax1, ax2)


class CoreSet:
    """
    CoreSet() is a class
    """
    def __init__(self, set_name, core, comment=None):
        """
        :rtype: CoreSet
        """
        self.name = set_name
        self.core_data = {core.name: core}
        self.core = [core.name]
        self.comment = None
        self.variables = []
        self.add_variable(core)

    def add_variable(self, core):
        for ii_variable in core.profiles.keys():
            if ii_variable not in self.variables:
                self.variables.append(ii_variable)

    def add_comment(self, comment):
        self.comment.append(comment)

    def add_core(self, core):
        """
        :param core:
        :return:
        """
        self.core_data[core.name] = core
        self.core.append(core.name)
        self.add_variable(core)

    def ice_thickness(self):
        """
        :return:
        """
        hi = []
        for key in self.core_data.keys():
            if self.core_data[key].ice_thickness is not None:
                hi.append(self.core_data[key].ice_thickness)
        if hi == [] or np.isnan(hi).all():
            return None
        else:
            return hi, np.nanmean(hi), np.nanmax(hi)

    def core_length(self):
        """
        :return:
        """
        lc = []
        for key in self.core_data.keys():
            core = self.core_data[key]
            for a in dir(core):
                try:
                    temp = core.__getattribute__(a).length
                except AttributeError:
                    logging.warning('core length for %s not defined' % a)
                else:
                    lc.append(temp)
        return lc, np.nanmean(lc), np.nanmax(lc)

    # statistic
    def mean(self, var=None):
        '''
        :param variable:
        :return:
        '''
        variable = {}
        ics_data = merge_bin(self)
        for ii_core in ics_data.core:
            ic_data = ics_data.core_data[ii_core]
            if var is None:
                for ii_variable in ics_data.variables:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')
            else:
                if not isinstance(var, list):
                    var = [var]
                for ii_variable in var:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')

            for ii_variable in variable.keys():
                y = variable[ii_variable][1]
                x = np.nanmean(np.atleast_2d(variable[ii_variable][0]), axis=0)
                count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)

                profile = Profile(x, y, ic_data.name, 'mean', comment='statistic mean computed from merged bin ice cores', note=count, length=None)
                if ii_variable in ic_data.profiles.keys():
                    ic_data.del_profile(ii_variable)
                ic_data.add_profile(profile, ii_variable)
        return ic_data

    def std(self, var=None):
        '''
        :param variable:
        :return:
        '''
        variable = {}
        ics_data = merge_bin(self)
        for ii_core in ics_data.core:
            ic_data = ics_data.core_data[ii_core]
            if var is None:
                for ii_variable in ics_data.variables:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')
            else:
                if not isinstance(var, list):
                    var = [var]
                for ii_variable in var:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')

            for ii_variable in variable.keys():
                y = variable[ii_variable][1]
                x = np.nanstd(np.atleast_2d(variable[ii_variable][0]), axis=0)
                count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)

                profile = Profile(x, y, ic_data.name, 'mean', comment='statistic mean computed from merged bin ice cores', note=count, length=None)
                if ii_variable in ic_data.profiles.keys():
                    ic_data.del_profile(ii_variable)
                ic_data.add_profile(profile, ii_variable)
        return ic_data

    def min(self, var=None):
        '''
        :param variable:
        :return:
        '''
        variable = {}
        ics_data = merge_bin(self)
        for ii_core in ics_data.core:
            ic_data = ics_data.core_data[ii_core]
            if var is None:
                for ii_variable in ics_data.variables:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')
            else:
                if not isinstance(var, list):
                    var = [var]
                for ii_variable in var:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')

            for ii_variable in variable.keys():
                y = variable[ii_variable][1]
                x = np.nanmin(np.atleast_2d(variable[ii_variable][0]), axis=0)
                count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)

                profile = Profile(x, y, ic_data.name, 'mean', comment='statistic mean computed from merged bin ice cores', note=count, length=None)
                if ii_variable in ic_data.profiles.keys():
                    ic_data.del_profile(ii_variable)
                ic_data.add_profile(profile, ii_variable)
        return ic_data

    def max(self, var=None):
        '''
        :param variable:
        :return:
        '''
        variable = {}
        ics_data = merge_bin(self)
        for ii_core in ics_data.core:
            ic_data = ics_data.core_data[ii_core]
            if var is None:
                for ii_variable in ics_data.variables:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')
            else:
                if not isinstance(var, list):
                    var = [var]
                for ii_variable in var:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')

            for ii_variable in variable.keys():
                y = variable[ii_variable][1]
                x = np.nanmax(np.atleast_2d(variable[ii_variable][0]), axis=0)
                count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)

                profile = Profile(x, y, ic_data.name, 'mean', comment='statistic mean computed from merged bin ice cores', note=count, length=None)
                if ii_variable in ic_data.profiles.keys():
                    ic_data.del_profile(ii_variable)
                ic_data.add_profile(profile, ii_variable)
        return ic_data

    def count(self, var=None):
        '''
        :param variable:
        :return:
        '''
        variable = {}
        ics_data = merge_bin(self)
        for ii_core in ics_data.core:
            ic_data = ics_data.core_data[ii_core]
            if var is None:
                for ii_variable in ics_data.variables:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')
            else:
                if not isinstance(var, list):
                    var = [var]
                for ii_variable in var:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')

            for ii_variable in variable.keys():
                y = variable[ii_variable][1]
                x = np.nancount(np.atleast_2d(variable[ii_variable][0]), axis=0)

                profile = Profile(x, y, ic_data.name, 'mean', comment='statistic mean computed from merged bin ice cores', note=None, length=None)
                if ii_variable in ic_data.profiles.keys():
                    ic_data.del_profile(ii_variable)
                ic_data.add_profile(profile, ii_variable)
        return ic_data

    def statistic(self, var=None):
        '''
        :param variable:
        :return:
        '''
        ics_merge_bin_set = merge_bin(self)
        variable = {}
        for ii_core in ics_merge_bin_set.core:
            ic_data = ics_merge_bin_set.core_data[ii_core]
            if var is None:
                for ii_variable in ics_merge_bin_set.variables:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')
            else:
                if not isinstance(var, list):
                    var = [var]
                for ii_variable in var:
                    if ii_variable in ic_data.profiles.keys():
                        if ii_variable not in variable.keys():
                            variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
                        elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
                            variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
                        else:
                            logging.warning('vertical resolution is not the same between the cores')

        for ii_variable in variable.keys():
            y = variable[ii_variable][1]
            x_mean = np.nanmean(np.atleast_2d(variable[ii_variable][0]), axis=0)
            x_std = np.nanstd(np.atleast_2d(variable[ii_variable][0]), axis=0)
            x_min = np.nanmin(np.atleast_2d(variable[ii_variable][0]), axis=0)
            x_max = np.nanmax(np.atleast_2d(variable[ii_variable][0]), axis=0)
            count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
            x = [x_mean, x_std, x_min, x_max, count]

            profile = Profile(x, y, ic_data.name, ii_variable+'-stat', comment='statistic envelop computed from merged bin ice cores', note=None, length=None)

            if ii_variable in ic_data.profiles.keys():
                ic_data.del_profile(ii_variable)
            ic_data.add_profile(profile, ii_variable)
        return ic_data

    def plot_variable_stat(self, ax, var, param_dict=None):
        """
        :param ax:
        :param prop:
        :param param_dict:
        :return:
        """

        if var not in self.variables:
            logging.warning('variable %s not a statistical variable' % variable)
            return None

        ic_stat = self.statistic()

        y = ic_stat.profiles[var].y
        x = ic_stat.profiles[var].x

        if len(y) == len(x[0]):
            out = ax.plot(x[0], y, color='k', label='Mean')
            ax.plot(x[2], y, color='b', label='Min')
            ax.plot(x[3], y, color='r', label='Max')
            ax.fill_betweenx(y, x[0] - x[1], x[0] + x[1], facecolor='black', alpha=0.3, label=str(u"\c2b1") + "std dev")
            ax.set_ylim(max(ax.get_ylim()), 0)
            ax.set_xlabel(var + ' ' + si_prop_unit[var])
        else:
            out = ax.step(np.append(x[0], x[0][-1]), y, color='k', label='Mean')
            ax.step(np.append(x[2], x[2][-1]), y, color='b', label='Min')
            ax.step(np.append(x[3], x[3][-1]), y, color='r', label='Max')

            x_fill_l = [x[0][0]-x[1][0]]
            x_fill_h = [x[0][0]+x[1][0]]
            y_fill = y[0]
            for ii in range(1, len(x[0])):
                x_fill_l = np.append(x_fill_l, x[0][ii-1]-x[1][ii-1])
                x_fill_l = np.append(x_fill_l, x[0][ii]-x[1][ii])
                x_fill_h = np.append(x_fill_h, x[0][ii-1]+x[1][ii-1])
                x_fill_h = np.append(x_fill_h, x[0][ii]+x[1][ii])
                y_fill = np.append(y_fill, y[ii])
                y_fill = np.append(y_fill, y[ii])

            ax.fill_betweenx(y_fill, x_fill_l, x_fill_h, facecolor='black', alpha=0.3, label=str(u"\c2b1") + "std dev")
            ax.set_ylim(max(ax.get_ylim()), 0)
            ax.set_xlabel(var + ' ' + si_prop_unit[var])
        return out


def import_core(ic_path, missing_value=float('nan'), comment='off'):
    """
    :param ic_path:
    :param missing_value:
    :param comment: on/off, 0/1, y[n
        toggle comment display
    """
    # check comment
    if comment in ['on', 1, 'yes', 'y']:
        comment = 1
    elif comment not in ['off', 1, 'no', 'n']:
        comment = 0
    else:
        logging.warning('comment parameters not defined')

    wb = openpyxl.load_workbook(filename=ic_path, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    # extract basic information about the ice core
    # name
    ic_name = ws_summary['C21'].value
    if comment:
        print(ic_name)

    # logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    # logging.info('Processing ' + ic_name + '...')

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value

    # location
    ic_loc = ws_summary['C5'].value

    # snow thickness
    temp_cell = ws_summary['C9']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_snow_depth = missing_value
    else:
        ic_snow_depth = temp_cell.value

    # ice thickness
    temp_cell = ws_summary['C11']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_ice_thickness = missing_value
    else:
        ic_ice_thickness = temp_cell.value

    imported_core = Core(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value
    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value
    temp_cell = ws_summary['C17']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value
    temp_cell = ws_summary['C18']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknown']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)] is not None and \
                    ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value is not None:
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value)
        ii_col += 1

    # SALINITY S
    if 'S_ice' in ws_name:
        profile = read_variable(wb, sheet_name='S_ice', col_x='D', col_y='AB', col_c=8, row_start=6)
        imported_core.add_profile(profile, 'salinity')
        logging.info('\tsalinity profile imported')
    else:
        logging.info('\tsalinity profile missing')

    if 'T_ice' in ws_name:
        profile = read_variable(wb, sheet_name='T_ice', col_x='B', col_y='A', col_c=8, row_start=6)
        imported_core.add_profile(profile, 'temperature')
        logging.info('\ttemperature profile imported')
    else:
        logging.info('\ttemperature profile missing')
    return imported_core


def merge_bin(ics_set):
    variable = {}
    for ii_core in ics_set.core:
        ic_data = ics_set.core_data[ii_core]
        for ii_variable in ic_data.profiles.keys():
            y = ic_data.profiles[ii_variable].y
            if ii_variable not in variable.keys():
                variable[ii_variable] = np.array(y)
            else:
                variable[ii_variable] = np.unique(np.append(y, variable[ii_variable]))
            variable[ii_variable] = variable[ii_variable][~np.isnan(variable[ii_variable])]

    flag = 0
    for ii_core in ics_set.core:
        ic_data = ics_set.core_data[ii_core]
        print(ii_core)
        for ii_variable in ics_set.variables:
            if ii_variable in ic_data.profiles.keys():
                print(ii_variable)
                x = np.array(ic_data.profiles[ii_variable].x)
                y = np.array(ic_data.profiles[ii_variable].y)
                if len(x)==len(y):
                    print('same length')
                    y_bin = variable[ii_variable]
                    x_bin = np.interp(y_bin, y[~np.isnan(y)], x[~np.isnan(y)])
                else:
                    print('asymetrical length')
                    ii_bin = 1 # cycle from 0 to len(y)
                    ii = 0  # cycle from 0 to len(x_bin)
                    x_bin = np.nan*np.ones(len(variable[ii_variable])-1)
                    y_bin = variable[ii_variable]
                    while ii < len(x_bin):
                        if y_bin[ii] < y[ii_bin]:
                            x_bin[ii] = x[ii_bin-1]
                            ii += 1
                        elif y_bin[ii] == max(y):
                            break
                        else:
                            ii_bin += 1

                profile = Profile(x_bin, y_bin, ic_data.name, ii_variable, comment='merged bin from core set', note=None, length=None)

                ic_data.del_profile(ii_variable)
                ic_data.add_profile(profile, ii_variable)

            if flag is 0:
                ics_merged_bin_set = CoreSet(ics_set.name+'-merged_bin', ic_data)
                flag = 1
            else:
                ics_merged_bin_set.add_core(ic_data)

    return ics_merged_bin_set


def make_section(core, section_thickness=0.05):
    """
    :param core:
    :param section_thickness:
    """
    core = core
    for a in dir(core):
        if isinstance(core.__getattribute__(a), Profile) and core.__getattribute__(a) is not None:
            profile = core.__getattribute__(a)
            if profile.length is not None:
                length = profile.length
            else:
                length = core.ice_thickness
            y_mid_section = np.arange(section_thickness / 2, length, section_thickness)
            delta_y = (length + len(y_mid_section) * section_thickness) / 2
            if delta_y < length:
                y_mid_section = np.append(y_mid_section, np.atleast_1d(delta_y))
            x = profile.x
            y = np.array(profile.y)

            if len(y) is len(x) + 1:
                y = (y[1:] + y[:-1]) / 2

            x_mid_section = np.interp(y_mid_section, y[~np.isnan(y)], x[~np.isnan(y)])
            profile.x = x_mid_section
            profile.y = y_mid_section
            core.__delattr__(a)
            core.__setattr__(a, profile)
    return core


def get_filepath(data_dir, data_ext, subdir='no'):
    """
    :param data_dir:
    :param data_ext:
    :param subdir:
    """
    import os
    filepath = []
    if subdir == 'yes':
        for path, subdirs, files in os.walk(data_dir):
            subdirs.sort()
            files.sort()
            for name in files:
                if name.endswith(data_ext):
                    f = os.path.join(path, name)
                    filepath.append(f)
    else:
        files = os.listdir(data_dir)
        for name in files:
            if name.endswith(data_ext):
                f = os.path.join(data_dir, name)
                filepath.append(f)
    return filepath


def generate_source(data_dir, data_ext):
    """
    :param data_dir:
    :param data_ext:
    """
    ic_path_list = get_filepath(data_dir, data_ext)
    with open(data_dir + '/ic_list.txt', 'w') as f:
        for ii in range(0, len(ic_path_list)):
            f.write(ic_path_list[ii] + "\n")


def read_variable(wb, sheet_name, col_x, col_y, col_c, row_start):
    """
    :param wb:
    :param sheet_name:
    :param col_x:
    :param col_y:
    :param col_c:
    :param row_start:
    """
    ice_core_spreadsheet = {'T_ice': 'temperature', 'S_ice': 'salinity'}

    if sheet_name in wb.sheetnames:
        logging.info('%s sheet present, importing profile for %s' % (sheet_name, ice_core_spreadsheet[sheet_name]))

        if not isinstance(col_y, list) and len(col_y) < 2:
            col_y = [col_y]
        col_y_temp = [float('nan')] * len(col_y)
        for ii_col in range(0, len(col_y)):
            if isinstance(col_y[ii_col], str):
                col_y_temp[ii_col] = openpyxl.utils.column_index_from_string(col_y[ii_col])
        col_y = col_y_temp
        del col_y_temp
        if isinstance(col_x, str):
            col_x = openpyxl.utils.column_index_from_string(col_x)
        if isinstance(col_c, str):
            col_c = openpyxl.utils.column_index_from_string(col_c)

        ws = wb.get_sheet_by_name(sheet_name)
        profile_name = ws['C1'].value
        note = ws['C3'].value

        # import y
        row_jj = row_start
        y_bin_mid = []
        y_bin = []
        col_flag = 0
        if len(col_y) == 1:
            while row_jj > 0:
                try:
                    y1 = ws.cell(column=col_y[0], row=row_jj).value
                except IndexError:
                    break
                if y1 is not None:
                    y_bin_mid.append(y1)
                else:
                    break
                row_jj += 1
            y_bin = y_bin_mid
        elif len(col_y) == 2:
            col_flag = 1
            bin_flag = 1
            y_bin_sup = []
            y_bin_low = []
            while bin_flag > 0:
                try:
                    y1 = ws.cell(column=col_y[0], row=row_jj).value
                    y2 = ws.cell(column=col_y[1], row=row_jj).value
                except IndexError:
                    break
                if y1 is not None and y2 is not None:
                    y_bin_sup.append(y1)
                    y_bin_low.append(y2)
                    y_bin_mid.append((y1 + y2) / 2)
                else:
                    break
                row_jj += 1
            y_bin = np.concatenate((y_bin_sup, [y_bin_low[-1]]))

        # length
        length = ws['C2'].value
        y_bin_length = max(y_bin) - min(y_bin)
        if length is None or y_bin_length < length:
            length = max(y_bin) - min(y_bin)

        # read variable:
        x = [nan_value] * (len(y_bin) - col_flag)
        comment = [None] * (len(y_bin) - col_flag)
        for row_jj in range(row_start, row_start + len(x)):
            x_jj = ws.cell(column=col_x, row=row_jj).value
            if isinstance(x_jj, (float, int)):
                x[row_jj - row_start] = x_jj
            comment[row_jj - 6] = ws.cell(column=col_c, row=row_jj).value
        comment.append(ws['C3'].value)
        return Profile(x, y_bin, profile_name, ice_core_spreadsheet[sheet_name], comment, note, length)
    else:
        logging.info('profile %s missing' % ice_core_spreadsheet[sheet_name])


def import_src(txt_filepath, section_thickness=None, missing_value=float('nan'), log_level='warning', comment='off'):
    """
    import_src import all ice core data which path are given in a source text file
    :param txt_filepath:
    :param section_thickness:
    :param missing_value:
    :param log_level:
    :param comment:
    :return:
    """
    print('Ice core data importation in progress ...')
    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    a = open(txt_filepath)
    filepath = [line.strip() for line in a]
    filepath = sorted(filepath)
    ic_dict = {}
    for ii in range(0, len(filepath)):
        if not filepath[ii].startswith('#'):
            ic_data = import_core(filepath[ii], missing_value, comment=comment)
            if section_thickness is not None:
                ic_data = make_section(ic_data, section_thickness)
            ic_dict[ic_data.name] = ic_data
    logging.info('Ice core importation complete')
    print('done')
    return ic_dict


def import_list(ics_list, missing_value=float('nan'), log_level='warning', comment='n'):
    """
    :param ics_list:
    :param missing_value:
    :param log_level:
    :param comment:
    """
    print('Ice core data importation in progress ...')

    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])

    ic_dict = {}
    for ii in range(0, len(ics_list)):
        if comment:
            print(ics_list[ii])
        ic_data = import_core(ics_list[ii], missing_value, comment=comment)
        ic_dict[ic_data.name] = ic_data

    logging.info('Ice core importation complete')
    print('done')

    return ic_dict


def import_variable(ic_path, variable='Salinity', missing_value=float('nan')):
    """
    :param ic_path:
    :param variable: string or array of string
        prop contain variable to import from a spreadsheet wb
    :param missing_value:
    :return:
    """

    variable_sheet_name = {'salinity': 'S_ice', 's': 'S_ice',
                           'temperature': 'T_ice', 't': 'T_ice',
                           'd18O': 'S_ice',
                           'dO2': 'S_ice',
                           'sediment': 'sediment', 'sed': 'sediment',
                           'Chla': 'algal_pigment', 'chlorophyl a': 'algal_pigment',
                           'Phae': 'algal_pigment'}
    variable_coordinate = {'salinity': ['D', 'AB', 8, 6], 's': ['D', 'AB', 8, 6],
                           'temperature': ['B', 'A', 8, 6],
                           'd180': [6, 'AB', 8, 6],
                           'sediment': [6, 'AB', 8, 6], 'sed':[6, 'AB', 8, 6],
                           'dO2': [7, 'AB', 8, 6],
                           'Chla': [5, 'AB', 8, 6], 'chlorophyl a': [5, 'AB', 8, 6],
                           'Phae': [6, 'AB', 8, 6]}

    wb = openpyxl.load_workbook(filename=ic_path, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    # extract basic information about the ice core
    # name
    ic_name = ws_summary['C21'].value
    print(ic_name)

    # logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    # logging.info('Processing ' + ic_name + '...')

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value

    # location
    ic_loc = ws_summary['C5'].value

    # snow thickness
    temp_cell = ws_summary['C9']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_snow_depth = missing_value
    else:
        ic_snow_depth = temp_cell.value

    # ice thickness
    temp_cell = ws_summary['C11']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_ice_thickness = missing_value
    else:
        ic_ice_thickness = temp_cell.value

    imported_core = Core(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value
    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value
    temp_cell = ws_summary['C17']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value
    temp_cell = ws_summary['C18']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknown']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)] is not None and \
                    ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value is not None:
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value)
        ii_col += 1

    # import properties:
    if isinstance(variable, str):
        variable = [variable]
    for ii_variable in variable:
        print(ii_variable)
        if variable_sheet_name[ii_variable.lower()] in ws_name:
            col_x_temp = variable_coordinate[ii_variable.lower()][0]
            col_y_temp = variable_coordinate[ii_variable.lower()][1]
            col_c_temp = variable_coordinate[ii_variable.lower()][2]
            row_start_temp = variable_coordinate[ii_variable.lower()][3]
            profile = read_variable(wb, sheet_name=variable_sheet_name[ii_variable.lower()], col_x=col_x_temp, col_y=col_y_temp,
                                    col_c=col_c_temp, row_start=row_start_temp)
            imported_core.add_profile(profile, ii_variable.lower())
            logging.info('\t%s profile imported' % ii_variable.lower())
        else:
            logging.info('\t%s profile missing' % ii_variable.lower())
    if isinstance(imported_core, Core):
        return imported_core
