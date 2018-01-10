#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
seaice.io.icxl.py : function to import import ice core data from xlsx spreadsheet
"""
import datetime
import logging
import os

import dateutil
import numpy as np
import openpyxl
import pandas as pd
import seaice

__name__ = "icxl"
__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "1.1"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "moggier@alaska.edu"
__status__ = "dev"
__date__ = "2017/09/13"
__comment__ = "loadxl.py contained function to import ice core data from xlsx spreadsheet"
__CoreVersion__ = 1.1

__all__ = ["import_ic_path", "import_ic_list", "import_ic_sourcefile", "list_ic", "list_ic_path", "make_ic_sourcefile"]

variable_2_sheet = {'temperature': 'T_ice',
                    'salinity': 'S_ice',
                    'conductivity': 'S_ice',
                    'specific conductance': 'S_ice',
                    'd18O': 'S_ice',
                    'dD': 'S_ice',
                    'Vf_oil': 'Vf_oil', 'oil volume fraction': 'Vf_oil',  # MOSIDEO project
                    'Wf_oil': 'Wf_oil', 'oil weight fraction': 'Vf_oil',  # MOSIDEO project
                    'oil content': 'oil_content',  # CMI project
                    'oil mass': 'Vf_oil', 'm_oil': 'Vf_oil'
                    # 'seawater': 'seawater',
                    # 'sediment': 'sediment',
                    # 'Chla': 'algal_pigment',
                    # 'chlorophyl a': 'algal_pigment',
                    # 'Phae': 'algal_pigment'
                    }



def import_ic_path(ic_path, variables=None, v_ref='top'):
    """
    :param ic_path:
        string, path to the xlsx ice core spreadsheet
    :param variables:
        list of string, variables to import. If not defined, all variable will be imported.
    :param v_ref:
        'top' or 'bottom', vertical reference. top for ice/snow or ice/air surface, bottom for ice/water interface
    :return:
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(ic_path):
        logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])
        return 0


    wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    name = ws_summary['C21'].value

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
    else:
        logger.error("(%s) ice core spreadsheet version not unavailable" % name)

    if version < __CoreVersion__:
        update_spreadsheet(ic_path, v_ref=v_ref)
        logger.info("Updating ice core spreadsheet %s to last version (%s)" % (name, str(__CoreVersion__)))
        wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
        ws_name = wb.get_sheet_names()
        ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet
        version = ws_summary['C3'].value

    n_row_collection = 22
    logger.info("importing data for %s" % name)

    if isinstance(ws_summary['C2'].value, datetime.datetime):
        if isinstance(ws_summary['D2'].value, datetime.time):
            date = datetime.datetime.combine(ws_summary['C2'].value, ws_summary['D2'].value)
            if ws_summary['E2'].value is not None and dateutil.tz.gettz(ws_summary['E2'].value):
                tz = dateutil.tz.gettz(ws_summary['E2'].value)
                date = date.replace(tzinfo=tz)
            else:
                logger.info("\t(%s) timezone unavailable." % name)
        else:
            date = ws_summary['C2'].value
            if ws_summary['D2'].value is not None and dateutil.tz.gettz(ws_summary['D2'].value):
                tz = dateutil.tz.gettz(ws_summary['D2'].value)
                date = date.replace(tzinfo=tz)
            else:
                logger.info("\t(%s) timezone unavailable." % name)
    else:
        logger.warning("\t(%s) date unavailable" % name)
        date = None

    origin = ws_summary['C5'].value

    lat = np.nan
    lon = np.nan
    if isinstance(ws_summary['C6'].value, (float, int)) or isinstance(ws_summary['D6'], (float, int)):
        lat = ws_summary['C6'].value
        lon = ws_summary['D6'].value
    elif ws_summary['C6'].value and ws_summary['D6'].value:
        logger.error("\t(%s) lat/lon not defined in decimal degree" % name)
    else:
        logger.info("\t(%s) lat/lon unknown" % name)

    if isinstance(ws_summary['C9'].value, (float, int)):
        snow_depth = np.array([ws_summary['C9'].value]).astype(float)
        n_snow = 1
        while ws_summary.cell(row=9, column=3+n_snow).value is not None:
            snow_depth = np.concatenate((snow_depth, np.array([ws_summary.cell(row=9, column=3+n_snow).value])))
            n_snow +=1
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')
    else:
        snow_depth = np.array([np.nan])

    if isinstance(ws_summary['C10'].value, (float, int)):
        freeboard = np.array([ws_summary['C10'].value])
        n_temp = 1
        while ws_summary.cell(row=10, column=3+n_temp).value is not None:
            if isinstance(ws_summary.cell(row=10, column=3+n_temp).value, (float, int)):
                freeboard = np.concatenate((freeboard, np.array([ws_summary.cell(row=10, column=3 + n_snow).value])))
            else:
                logger.info("(%s)\tfreeboard cell %s not a float" % (name,
                                                                            openpyxl.utils.get_column_letter(3 + n_temp)+str(9)))
            n_temp += 1
        freeboard = pd.to_numeric(freeboard, errors='coerce')
    else:
        freeboard = np.array([np.nan])

    if isinstance(ws_summary['C11'].value, (float, int)):
        ice_thickness = np.array([ws_summary['C11'].value])
        n_temp = 1
        while ws_summary.cell(row=11, column=3+n_temp).value:
            if isinstance(ws_summary.cell(row=11, column=3+n_temp).value, (float, int)):
                ice_thickness = np.concatenate((ice_thickness, np.array([ws_summary.cell(row=11, column=3+n_snow).value])))
            else:
                logger.info("\t(%s) ice_thickness cell %s not a float" % (name, openpyxl.utils.get_column_letter(3+n_temp)+str(9)))
            n_temp += 1
        ice_thickness = pd.to_numeric(ice_thickness, errors='coerce')
    else:
        ice_thickness = np.array([np.nan])

    core = seaice.Core(name, date, origin, lat, lon, ice_thickness, freeboard, snow_depth)

    # temperature
    if ws_summary['C15'].value:
        core.t_air = ws_summary['C15'].value
    if ws_summary['C16'].value:
        core.t_snow_surface = ws_summary['C16'].value
    if ws_summary['C17'].value:
        core.t_ice_surface = ws_summary['C17'].value
    if ws_summary['C18'].value:
        core.t_water = ws_summary['C18'].value

    # sampling protocol
    m_col = 3
    if ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0i' %(n_row_collection+2))].value is not None:
        core.protocol = ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0i' %(n_row_collection+2))].value
    else:
        core.protocol = 'N/A'

    # core collection
    while (ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)] is not None and
           ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)].value is not None):
        core.add_to_collection(ws_summary[openpyxl.utils.cell.get_column_letter(m_col) + str('%.0f' % n_row_collection)].value)
        m_col += 1

    # comment
    if ws_summary['C33'].value is not None:
        core.add_comment(ws_summary['C33'].value)

    # variable
    if variables is None:
        sheets = [sheet for sheet in ws_name if (sheet not in ['summary', 'abreviation', 'locations', 'lists', 'Vf_oil_calculation']) and
                     (sheet.lower().find('fig') == -1)]
        core_flag = []
        for sheet in sheets:
            ws_variable = wb.get_sheet_by_name(sheet)
            profile = read_variable(ws_variable, variables=None, version=version, v_ref=v_ref)
            for variable in profile.keys():
                if not profile[variable][1] == core.name:
                    logger.error('\t(%s) core name %s and profile name %s does not match'
                                        % (ic_path, core.name, profile[variable][1]))
                else:
                    core.add_profile(profile[variable][0])
                    core.add_variable(variable)
                    if core.name not in core_flag and profile[variable][3] is not None and ~np.isnan(
                            profile[variable][3]):
                        core.add_length(profile[variable][3])
                        core_flag.append(core.name)
                    core.add_comment(profile[variable][2])

        if core.variables is None:
            logger.info('(%s) no variable to import' % name)
        else:
            logger.info('\t(%s) variables %s imported with success' % (name, ", ".join(core.variables)))
    else:
        if not isinstance(variables, list):
            if variables.lower().find('state variable')+1:
                variables = ['temperature', 'salinity']
            else:
                variables = [variables]
        logger.info("(%s) Variables are %s" % (name, ', '.join(variables)))

        core_flag = []
        for variable in variables:
            if variable_2_sheet[variable] in ws_name:
                ws_variable = wb.get_sheet_by_name(variable_2_sheet[variable])
                profile = read_variable(ws_variable, variables=variable, version=version, v_ref=v_ref)
                if profile.keys().__len__() == 0:
                    logger.warning('\t(%s) no data exist for %s' % (core.name, variable))
                elif profile[variable][1] is not core.name:
                    logger.error('\t(%s) core name %s and profile name %s does not match'
                                        % (ic_path, core.name, profile[variable][1]))
                else:
                    core.add_profile(profile[variable][0])
                    core.add_variable(variable)
                    if core.name not in core_flag and profile[variable][3] is not None and ~np.isnan(
                            profile[variable][3]):
                        core.add_length(profile[variable][3])
                        core_flag.append(core.name)
                    core.add_comment(profile[variable][2])
            else:
                logger.info('\tsheet %s does not exists' % variable)

    # weather
    # TODO:adding a weather class and reading the information

    return core


def import_ic_list(ics_list, variables=None, v_ref='top'):
    """
    :param ics_list:
            array, array contains absolute filepath for the cores
    :param variables:
    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)

    ic_dict = {}
    inexisting_ics_list = []
    for ic_path in ics_list:
        if not os.path.exists(ic_path):
            logger.warning("%s does not exists in core directory" % ic_path.split('/')[-1])
            inexisting_ics_list.append(ic_path.split('/')[-1].split('.')[0])
        else:
            ic_data = import_ic_path(ic_path, variables=variables, v_ref=v_ref)
            if ic_data.variables is None:
                inexisting_ics_list.append(ic_path.split('/')[-1].split('.')[0])
                logger.warning("%s have no properties profile" % ic_data.name)
            else:
                ic_dict[ic_data.name] = ic_data

    logging.info("Import ice core lists completed")
    if inexisting_ics_list.__len__()>0:
        logger.info("%s core does not exits. Removing from collection" % ', '.join(inexisting_ics_list))

    for ic in inexisting_ics_list:
        for ic2 in ic_dict.keys():
            if ic in ic_dict[ic2].collection:
                ic_dict[ic2].del_from_collection(ic)
                logger.info("remove %s from %s collection" % (ic, ic2))
    return ic_dict


def import_ic_sourcefile(filepath, variables=None, v_ref='top'):
    """
    :param filepath:
            string, absolute path to the file containing either the absolute path of the cores (1 path by line) or the
            core names (1 core by line). In this last case if core_dir is None core_dir is the directory contianing the
            file.
    :param variables:

    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)
    logger.info('Import ice core from source file: %s' % filepath)

    with open(filepath) as f:
        ics = sorted([line.strip() for line in f])

    return import_ic_list(ics, variables=variables, v_ref=v_ref)


# read variable
def read_variable(ws_variable, variables=None, version=__CoreVersion__, v_ref='top'):
    """
    :param ws_variable:
        openpyxl.worksheet
    :param variables:
    :param version:
    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)

    if version == 1:
        row_data_start = 6
    elif version == 1.1:
        row_data_start = 8
        if ws_variable['C4'].value:
            v_ref = ws_variable['C4'].value
    else:
        logger.error("ice core spreadsheet version not defined")
        return None

    variable_2_data = {'temperature': [row_data_start, 'A', 'B', 'C'],
                       'salinity': [row_data_start, 'ABC', 'D', 'J'],
                       'conductivity': [row_data_start, 'ABC', 'EF', 'J'],
                       'specific conductance': [row_data_start, 'ABC', 'G', 'J'],
                       'd18O': [row_data_start, 'ABC', 'H', 'J'],
                       'dD': [row_data_start, 'ABC', 'I', 'J'],
                       'oil weight fraction': [row_data_start, 'ABC', 'G', 'H'],
                       'oil volume fraction': [row_data_start, 'ABC', 'F', 'H'],
                       'oil content': [row_data_start, 'ABC', 'DEFGH', 'I'],
                       'oil mass': [row_data_start, 'ABC', 'ED', 'H']
                       # 'seawater': [row_data_start, 'ABC', 'D', 'E'],
                       # 'sediment': [row_data_start, 'ABC', 'D', 'E'],
                       # 'Chla': [row_data_start, 'ABC', 'D', 'E'],
                       # 'Phae': [row_data_start, 'ABC', 'D', 'E']
                       }

    if variables is None:
        variables = [key for key in variable_2_data.keys() if variable_2_sheet[key] == ws_variable.title]
    if not isinstance(variables, list):
        variables = [variables]

    name = ws_variable['C1'].value

    profile = {}
    for variable in variables:
        columns_string = ['comment', 'variable']
        # step profile
        if variable_2_data[variable][1].__len__() == 3:
            columns_float = ['y_low', 'y_sup', 'y_mid']
            # y_data = np.array([ws_variable[variable_2_data[variable][2][0] + str(row)].value
            #                  for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])

            y_low = np.array([ws_variable[variable_2_data[variable][1][0]+str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])
            y_sup = np.array([ws_variable[variable_2_data[variable][1][1]+str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])
            y_mid = np.array([ws_variable[variable_2_data[variable][1][2] + str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])

            if not np.array([isinstance(element, (float, int)) for element in y_mid]).any():
                if (np.array([isinstance(element, (float, int)) for element in y_low]).any() or
                        np.array([isinstance(element, (float, int)) for element in y_sup]).any()):
                    y_mid = (y_low+y_sup)/2
                    logger.info(
                        '\t(%s : %s) y_mid does not exit, calculating y_mid from section depth with success'
                        % (name, variable))
                else:
                    logger.warning(
                        '\t(%s : %s) y_mid does not exit, not able to calculate y_mid from section depth. Section'
                        'depth maybe not numeric' % (name, variable))
            data = np.array([y_low, y_sup, y_mid])

            for col_variable in range(0, variable_2_data[variable][2].__len__()):
                data = np.vstack((data, np.array([ws_variable[variable_2_data[variable][2][col_variable] +
                                                              str(row)].value
                                                  for row in range(variable_2_data[variable][0],
                                                                   ws_variable.max_row+1)])))
                if version is 1:
                    if col_variable is 0:
                        columns_float.append(variable)
                    else:
                        columns_float.append(
                            ws_variable[variable_2_data[variable][2][col_variable] + str(row_data_start - 3)].value)

                elif version >= 1.1:
                    columns_float.append(ws_variable[variable_2_data[variable][2][col_variable] +
                                                     str(row_data_start-3)].value)

            data = np.vstack((data, np.array([ws_variable[variable_2_data[variable][3] + str(row)].value
                                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])))

            variable_profile = pd.DataFrame(data.transpose(), columns=columns_float + ['comment'])

        # continuous variable_profile
        else:
            columns_float = ['y_mid']
            y_mid = np.array([ws_variable[variable_2_data[variable][1]+str(row)].value
                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])

            for col_variable in range(0, variable_2_data[variable][2].__len__()):
                data = np.vstack((y_mid, np.array([ws_variable[variable_2_data[variable][2][col_variable] +
                                                               str(row)].value
                                                   for row in range(variable_2_data[variable][0],
                                                                    ws_variable.max_row+1)])))
                if version is 1:
                    if col_variable is 0:
                        columns_float.append(variable)
                    else:
                        columns_float.append(
                            ws_variable[variable_2_data[variable][2][col_variable] + str(row_data_start - 3)].value)

                elif version >= 1.1:
                    columns_float.append(ws_variable[variable_2_data[variable][2][col_variable] +
                                                     str(row_data_start-3)].value)

            data = np.vstack((data, np.array([ws_variable[(variable_2_data[variable][3] + str(row))].value
                                              for row in range(variable_2_data[variable][0], ws_variable.max_row+1)])))

            variable_profile = pd.DataFrame(data.transpose(), columns=columns_float + ['comment'])

        # convert numeric to float
        key_temp = [key for key in variable_profile.keys() if key not in ['comment', 'v_ref']]
        for key in key_temp:
            variable_profile[key] = pd.to_numeric(variable_profile[key], errors='coerce')
        variable_profile['v_ref'] = v_ref

        key_temp = [key for key in key_temp if key not in ['y_low', 'y_sup', 'y_mid']]
        l_empty = 0
        for key in key_temp:
            if not np.isnan(variable_profile[key]).all():
                l_empty += 1
        if l_empty > 0:
            variable_profile['variable'] = variable

            for col in columns_float:
                if col not in variable_profile.keys():
                    variable_profile[col] = np.nan
                else:
                    variable_profile[col] = pd.to_numeric(variable_profile[col])
            for col in columns_string:
                if col not in variable_profile.keys():
                    variable_profile[col] = None

            note = ws_variable['C3'].value
            length = ws_variable['C2'].value
            if not isinstance(length, (float, int)):
                length = np.nan

            if v_ref == 'top':
                variable_profile['v_ref'] = v_ref

            profile[variable] = [variable_profile, name, note, length]
            logger.debug('\t(%s) %s imported with success' % (name, variable))
        else:
            logger.warning('\t(%s) %s unknown' % (name, variable))
    return profile


# create list or source
def list_ic(dirpath, fileext):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return ics_list: list
        list of ice core path
    """
    logger = logging.getLogger(__name__)

    ics_set = set([f for f in os.listdir(dirpath) if f.endswith(fileext)])
    logger.info("Found %i ice core datafile in %s" % (ics_set.__len__(), dirpath))
    return ics_set


def list_ic_path(dirpath, fileext):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return ics_list: list
        list of ice core path
    """
    logger = logging.getLogger(__name__)

    ics_set = list_ic(dirpath=dirpath, fileext=fileext)
    ic_paths_set = set([os.path.join(os.path.realpath(dirpath), f) for f in ics_set])
    return ic_paths_set


def make_ic_sourcefile(dirpath, fileext, source_filepath=None):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return source_file: str
        filepath to the text file containing ice core filepath with absolute path.
    """
    logger = logging.getLogger(__name__)

    ic_paths_set = list_ic_path(dirpath, fileext)

    if source_filepath is None:
        source_filepath = os.path.join(os.path.realpath(dirpath), 'ics_list.txt')

    with open(source_filepath, 'w') as f:
        for ic_path in ic_paths_set:
            f.write(ic_path + "\n")

    return source_filepath


# updater
def update_spreadsheet(ic_path, v_ref='top', backup=True):
    """
    update_spreadsheet update an ice core file to the latest ice core file version (__CoreVersion__).

    :param ic_path: str
        Path to the ice core file to update
    :param v_ref: int, 'top' or 'bottom', Default 'top'
        Set the zero vertical reference of the core. If 'top, zero vertical reference is set at the ice/snow or ice/air
        interface. If 'bottom, zero vertical reference is set at the ice/water interface
    :param verbose: int
        verbosity of the function.
    :param backup: bool, default True
        If True, backup the previous file version in the subdirectory 'folder-COREVERSION'
    :return nothing:

    USAGE:
        update_spreadhseet(ic_path)
    """
    logger = logging.getLogger(__name__)

    import shutil
    if not os.path.exists(ic_path):
        logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])
        return 0
    else:
        logger.info("\t ice core file path %s" % ic_path)

    wb = openpyxl.load_workbook(filename=ic_path)  # load the xlsx spreadsheet
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
        if version < __CoreVersion__:
            logger.info("Updating core data to latest version %.1f" % __CoreVersion__)
    else:
        logger.error("\t%s ice core fiel version not unavailable" % ic_path)

    # backup old version
    if backup:
        backup_dir = os.path.join(os.path.dirname(ic_path), 'version-'+str(version))
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        shutil.copyfile(ic_path, os.path.join(backup_dir, os.path.basename(ic_path)))

    while version < __CoreVersion__:
        # update from 1.0 to 1.1
        if version == 1:
            version = 1.1
            ws_summary['C3'] = version
            ws_summary = delete_row(ws_summary, 22)

            # loop through variables
            if "S_ice" in wb.get_sheet_names():
                ws = wb.get_sheet_by_name("S_ice")
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                row_start = 4
                ws = delete_column(ws, 'E', row_start, ws.max_row)
                ws['D5'] = 'salinity'
                ws['D6'] = 'S'
                ws = move_column(ws, 'E', 'J', row_start, ws.max_row)
                ws['E5'] = 'conductivity'
                ws['E6'] = 'σ'
                ws = move_column(ws, 'F', 'K', row_start, ws.max_row)
                ws['F5'] = 'conductivity measurement temperature'
                ws['F6'] = 'T_σ'
                ws = move_column(ws, 'G', 'J', row_start, ws.max_row)
                ws['G5'] = 'specific conductance'
                ws['G6'] = 'κ'
                delete_column(ws, 'K', row_start, ws.max_row)

            if "T_ice" in wb.get_sheet_names():
                ws = wb.get_sheet_by_name("T_ice")
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                ws['A6'] = 'd'
                ws['B6'] = 'T'

            if "oil_content" in wb.get_sheet_names():
                ws = wb.get_sheet_by_name("oil_content")
                # add reference row=4
                ws = add_row(ws, 4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws = add_row(ws, 6)
                ws['A6'] = 'd_1'
                ws['B6'] = 'd_2'
                ws['C6'] = 'd'
                ws['D6'] = 'V_i'
                ws['E6'] = 'h_menisc'
                ws['F6'] = 'd_menisc'
                ws['G6'] = 'd_center'
        wb.save(ic_path)


def add_row(ws, row_number):
    """
    :param ws:
    :param row_number:
    :return:
    """
    max_row = ws.max_row
    for row in range(row_number, ws.max_row + 1):
        new_row = row_number + max_row + 1 - row
        old_row = row_number + max_row - row
        for col in range(1, ws.max_column+1):
            ws.cell(row=new_row, column=col).value = ws.cell(row=old_row, column=col).value
    for col in range(1, ws.max_column+1):
        ws.cell(row=row_number, column=col).value = ""
    return ws


def delete_row(ws, row_number):
    """

    :param ws:
    :param row_number:
    :return:
    """
    max_row = ws.max_row
    for row in range(row_number, max_row):
        new_row = row
        old_row = row+1
        for col in range(1, ws.max_column+1):
            ws.cell(row=new_row, column=col).value = ws.cell(row=old_row, column=col).value
    for col in range(1, ws.max_column+1):
        ws.cell(row=max_row+1, column=col).value = ""
    return ws


def delete_column(ws, target_col, start_row=None, end_row=None):
    """
    :param ws:
    :param target_col:
    :param start_row:
    :param end_row:
    :return:
    """
    if start_row is None:
        start_row = ws.min_row
    if end_row is None:
        end_row = ws.max_row
    if not isinstance(target_col, int):
        target_col = openpyxl.utils.column_index_from_string(target_col)

    max_col = ws.max_column
    if np.alltrue([ws.cell(row=row, column=ws.max_column).value is None for row in range(start_row, ws.max_row)]):
        max_col = max_col - 1

    if not target_col == max_col:
        for col in range(target_col, max_col):
            new_col = col
            old_col = col + 1
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=old_col).value
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=max_col).value = ""

    return ws


def move_column(ws, target_col, source_col, start_row=None, end_row=None):
    """
    :param ws:
    :param target_col:
    :param source_col:
    :param start_row:
    :param end_row:
    :return:
    """

    if start_row is None:
        start_row = ws.min_row
    if end_row is None:
        end_row = ws.max_row
    if not isinstance(target_col, int):
        target_col = openpyxl.utils.column_index_from_string(target_col)
    if not isinstance(source_col, int):
        source_col = openpyxl.utils.column_index_from_string(source_col)

    max_col = ws.max_column
    if np.alltrue([ws.cell(row=row, column=ws.max_column).value is None for row in range(start_row, ws.max_row)]):
        max_col = max_col - 1

    # insert column in target column
    for col in range(target_col, max_col+1):
        new_col = target_col + max_col - col + 1
        old_col = target_col + max_col - col
        # print(openpyxl.utils.get_column_letter(old_col), openpyxl.utils.get_column_letter(new_col))
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=old_col).value

    # copy source col to target column
    if target_col < source_col:
        source_col = source_col + 1
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=target_col).value = ws.cell(row=row, column=source_col).value

    ws = delete_column(ws, source_col, start_row, end_row)

    return ws


def print_column(ws, col):
    """
    :param ws:
    :param col:
    :return:
    """
    if ~isinstance(col, int):
        col = openpyxl.utils.column_index_from_string(col)
    for row in range(1, ws.max_row + 1):
        print(row, ws.cell(row=row, column=col).value)


def print_row(ws, row):
    """
    :param ws:
    :param row:
    :return:
    """
    col_string = ''
    for col in range(1, ws.max_column + 1):
        col_string += openpyxl.utils.get_column_letter(col) + ':' + str(
            ws.cell(row=row, column=col).value) + '\t'
    print(col_string)
