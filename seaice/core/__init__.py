#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
seaice.io.icxl.py : function to import import ice core data from xlsx spreadsheet
"""

__name__ = "icxl"
__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "1.1"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "moggier@alaska.edu"
__status__ = "dev"
__date__ = "2017/09/13"
__comment__ = "loadxl.py contained function to import ice core data from xlsx spreadsheet"
__CoreVersion__ = 1.2

import datetime
import logging
import os

import dateutil
import numpy as np
import openpyxl
import pandas as pd

import seaice

__all__ = ["import_ic_path", "import_ic_list", "import_ic_sourcefile", "list_ic", "list_ic_path", "make_ic_sourcefile"]

TOL =1e-6
subvariable_dict = {'conductivity': ['conductivity measurement temperature']}

variable_2_sheet = {'temperature': 'temperature',
                    'salinity': 'salinity',
                    'conductivity': 'salinity',
                    'specific conductance': 'salinity',
                    'd18O': 'salinity',
                    'dD': 'salinity',
                    'Vf_oil': 'Vf_oil', 'oil volume fraction': 'Vf_oil',  # MOSIDEO project
                    'Wf_oil': 'Wf_oil', 'oil weight fraction': 'Vf_oil',  # MOSIDEO project
                    'oil content': 'oil_content',  # CMI project
                    'oil mass': 'Vf_oil', 'm_oil': 'Vf_oil'
                    # 'seawater': 'seawater',
                    # 'sediment': 'sediment',
                    # 'Chla': 'algal_pigment',
                    # 'chlorophyl a': 'algal_pigment',
                    # 'Phae': 'algal_pigment'
                    }

# all header type are float unless
non_float_property = {'comment': str,
                       'description': str}


def import_ic_path(ic_path, variables=None, v_ref='top', drop_empty=False):
    """
    :param ic_path:
        string, path to the xlsx ice core spreadsheet
    :param variables:
        list of string, variables to import. If not defined, all variable will be imported.
    :param v_ref:
        'top' or 'bottom', vertical reference. top for ice/snow or ice/air surface, bottom for ice/water interface
    :return:
    """
    logger = logging.getLogger(__name__)

    if not os.path.exists(ic_path):
        logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet
    ws_name = wb.sheetnames
    ws_summary = wb['summary']  # load the data from the summary sheet

    name = ws_summary['C28'].value

    if isinstance(ws_summary['C1'].value, (float, int)):
        version = ws_summary['C1'].value
    else:
        logger.error("(%s) ice core spreadsheet version not unavailable" % name)
    wb.close()

    # convert ice core spreadsheet to last version
    if version < 1.1:
        wb.close()
        update_spreadsheet(ic_path, v_ref=v_ref)
        logger.info("Updating ice core spreadsheet %s to last version (%s)" % (name, str(__CoreVersion__)))
        wb = openpyxl.load_workbook(filename=ic_path, keep_vba=True)  # load the xlsx spreadsheet
        ws_name = wb.sheetnames
        ws_summary = wb['summary']  # load the data from the summary sheet
        version = ws_summary['C3'].value

    logger.info("importing data for %s" % name)

    # date
    if isinstance(ws_summary['C10'].value, datetime.datetime):
        if isinstance(ws_summary['C11'].value, datetime.time):
            date = datetime.datetime.combine(ws_summary['C10'].value.date(), ws_summary['C11'].value)
            if ws_summary['C12'].value is not None and dateutil.tz.gettz(ws_summary['C12'].value):
                tz = ws_summary['C12'].value
                for tz_i in tz.split(' '):
                    if 'UTC' in tz_i:
                        tz = dateutil.tz.gettz(tz_i)
                        date = date.replace(tzinfo=tz)
                        print(tz)
                    #else
                        #print('find a lookup table for popular timezone')
                        # TODO: create lookup table based on spreadsheet
            else:
                logger.info("\t(%s) timezone unavailable." % name)
        else:
            date = ws_summary['C10'].value
            if ws_summary['C10'].value is not None and dateutil.tz.gettz(ws_summary['C10'].value):
                tz = dateutil.tz.gettz(ws_summary['C10'].value)
                date = date.replace(tzinfo=tz)
            else:
                logger.info("\t(%s) timezone unavailable." % name)
    else:
        logger.warning("\t(%s) date unavailable" % name)
        date = None

    # used to be origin, now site:
    # TODO: is location better ?
    origin = ws_summary['C4'].value

    # Parse lat/lon
    def parse_coordinate(DEG, MIN, SEC):
        if isinstance(DEG, (float, int)):
            # DEG DEC:
            if MIN is None:
                return DEG
            else:
                # DEG & MIN.DEC
                if SEC is None:
                    return DEG+MIN/60
                else:
                    # convert DEG MIN SEC to DEG DEC
                    return DEG+MIN/60+SEC/3600
        else:
            logger.info("\t(%s) lat/lon unknown" % name)
            return np.nan

    lat = parse_coordinate(ws_summary['C7'].value, ws_summary['D7'].value, ws_summary['E7'].value)
    lon = parse_coordinate(ws_summary['C8'].value, ws_summary['D8'].value, ws_summary['E8'].value)

    # comments:
    # TODO rename comments to note
    comments = []

    # Snow depth
    if isinstance(ws_summary['C14'].value, (float, int)):
        snow_depth = np.array([ws_summary['C14'].value]).astype(float)
        n_snow = 1
        while ws_summary.cell(row=14, column=3+n_snow).value is not None:
            snow_depth = np.concatenate((snow_depth, np.array([ws_summary.cell(row=14, column=3+n_snow).value])))
            n_snow += 1
        if isinstance(snow_depth[-1], str):
            comments.append(snow_depth[-1])
            snow_depth = pd.to_numeric(snow_depth[:-1], errors='coerce')
        else:
            snow_depth = pd.to_numeric(snow_depth, errors='coerce')

        # check if 1st value is mean of all other value
        c0 = len(snow_depth) > 1  # more than snow_depth
        c1 = snow_depth[0] == np.nanmean(snow_depth[1:])  #
        c2 = snow_depth[0] != np.nanmean(snow_depth)  #
        if c0 & c1 & c2:
            snow_depth = snow_depth[1:]
    elif ws_summary['C14'].data_type == 'f':
        try:
            import pycel
        except ModuleNotFoundError:
            # import snow depth from the cell on the right
            n_snow = 1
            snow_depth = []
            while ws_summary.cell(row=14, column=3 + n_snow).value is not None:
                snow_depth = np.concatenate((snow_depth, np.array([ws_summary.cell(row=14, column=3 + n_snow).value])))
                n_snow += 1
            if isinstance(snow_depth[-1], str):
                comments.append(snow_depth[-1])
                snow_depth = pd.to_numeric(snow_depth[:-1], errors='coerce')
            else:
                snow_depth = pd.to_numeric(snow_depth, errors='coerce')

        else:
            logger.error('Install pycel and script stuff here')
            # TODO: just read the error message
            # TODO: check if the value is the average of the other one.
    else:
        snow_depth = np.array([np.nan])

    # freeboard
    if isinstance(ws_summary['C15'].value, (float, int)):
        freeboard = np.array([ws_summary['C15'].value])
        n_temp = 1
        while ws_summary.cell(row=10, column=3+n_temp).value is not None:
            freeboard = np.concatenate((freeboard, np.array([ws_summary.cell(row=15, column=3 + n_temp).value])))
            if not isinstance(ws_summary.cell(row=15, column=3+n_temp).value, (float, int)):
                logger.info("(%s)\tfreeboard cell %s not a float" % (name, openpyxl.utils.get_column_letter(3 + n_temp)+str(9)))
            n_temp += 1

        if isinstance(freeboard[-1], str):
            comments.append(freeboard[-1])
            freeboard = pd.to_numeric(freeboard[:-1], errors='coerce')
        else:
            freeboard = pd.to_numeric(freeboard, errors='coerce')

        # check if 1st value is mean of all other value
        c0 = len(freeboard) > 1  # more than snow_depth
        c1 = freeboard[0] == np.nanmean(freeboard[1:])  #
        c2 = freeboard[0] != np.nanmean(freeboard)  #
        if c0 & c1 & c2:
            freeboard = freeboard[1:]
    else:
        freeboard = np.array([np.nan])

    # ice thickness
    if isinstance(ws_summary['C16'].value, (float, int)):
        ice_thickness = np.array([ws_summary['C16'].value])
        n_temp = 1
        while ws_summary.cell(row=16, column=3+n_temp).value:
            ice_thickness = np.concatenate(
                (ice_thickness, np.array([ws_summary.cell(row=16, column=3 + n_temp).value])))
            if not isinstance(ws_summary.cell(row=16, column=3+n_temp).value, (float, int)):
                logger.info("\t(%s) ice_thickness cell %s not a float" % (name, openpyxl.utils.get_column_letter(3+n_temp)+str(9)))
            n_temp += 1
        if isinstance(ice_thickness[-1], str):
            comments.append(ice_thickness[-1])
            ice_thickness = pd.to_numeric(ice_thickness[:-1], errors='coerce')
        else:
            ice_thickness = pd.to_numeric(ice_thickness, errors='coerce')

        # check if 1st value is mean of all other value
        c0 = len(ice_thickness) > 1  # more than snow_depth
        c1 = np.abs(ice_thickness[0]-np.nanmean(ice_thickness[1:])) < TOL  #
        c2 = ice_thickness[0] != np.nanmean(ice_thickness[0:-1])  #
        if c0 & c1 & c2:
            ice_thickness = ice_thickness[1:]
    else:
        ice_thickness = np.array([np.nan])

    core = seaice.Core(name, date, origin, lat, lon, ice_thickness, freeboard, snow_depth)

    # temperature
    if ws_summary['C23'].value:
        core.t_air = ws_summary['C23'].value
    if ws_summary['C24'].value:
        core.t_snow_surface = ws_summary['C24'].value
    if ws_summary['C25'].value:
        core.t_ice_surface = ws_summary['C25'].value
    if ws_summary['C26'].value:
        core.t_water = ws_summary['C26'].value

    # sampling protocol
    if ws_summary['C32'].value:
        core.protocol = ws_summary['C32'].value
    else:
        core.protocol = 'N/A'

    # core collection
    n_temp = 1
    while ws_summary.cell(row=16, column=3 + n_temp).value:
        ice_thickness = np.concatenate(
            (ice_thickness, np.array([ws_summary.cell(row=16, column=3 + n_temp).value])))
        n_temp += 1

    n_temp = 0
    while ws_summary.cell(row=30, column=3 + n_temp).value:
        if ws_summary.cell(row=30, column=3 + n_temp).value is not name:
            core.add_to_collection(ws_summary.cell(row=30, column=3 + n_temp).value)
        n_temp += 1

    # comment
    if ws_summary['A40'].value:
        comments.append(ws_summary['A40'].value)
    core.add_comment('; '.join(filter(None, comments)))

    # import all variables
    if variables is None:
        sheets = [sheet for sheet in ws_name if (sheet not in ['summary', 'abreviation', 'locations', 'lists',
                                                               'Vf_oil_calculation']) and
                     (sheet.lower().find('fig') == -1)]
        for sheet in sheets:
            print(sheet)
            ws_variable = wb[sheet]
            profile = read_profile(ws_variable, variables=None, version=version, v_ref=v_ref)

            if drop_empty:
                profile.drop_empty_property()

            if not profile.empty:
                if profile.get_name() is not core.name:
                    logger.error('\t(%s) core name %s and profile name %s does not match'
                                 % (ic_path, core.name, profile.get_name()))
                else:
                    core.add_profile(profile)
                    logger.info('(%s) data imported with success: %s' % (core.name, ", ".join(profile.get_property())))
            else:
                logger.info('(%s) no data to import from %s ' % (core.name, sheet))
    else:
        if not isinstance(variables, list):
            if variables.lower().find('state variable')+1:
                variables = ['temperature', 'salinity']
            else:
                variables = [variables]

        _imported_variables = []
        for variable in variables:
            if variable_2_sheet[variable] in ws_name and variable not in _imported_variables:
                sheet = variable_2_sheet[variable]
                ws_variable = wb[sheet]

                variable2import = [var for var in variables if var in inverse_dict(variable_2_sheet)[sheet]]

                profile = read_profile(ws_variable, variables=variable2import, version=version, v_ref=v_ref)

                if profile.empty:
                    logger.info('(%s) no %s to import' %(name, variable))
                elif profile.get_name() is not core.name:
                    logger.error('\t(%s) core name %s and profile name %s does not match'
                                 % (ic_path, core.name, profile.name()))
                elif not profile.empty:
                    # add snow temperature
                    if variable is 'temperature' and core.t_ice_surface is not np.nan:
                        profile_surface = profile.head(1).copy()
                        profile_surface['y_mid'] = 0
                        profile_surface['temperature'] = core.t_ice_surface
                        profile_surface['comment'] = 'ice surface'
                        profile = profile.append(profile_surface)

                        if core.t_snow_surface is not np.nan:
                            profile_snow_surface = profile.head(1).copy()
                            profile_snow_surface['y_mid'] = -np.nanmean(core.snow_depth)
                            profile_snow_surface['temperature'] = core.t_snow_surface
                            profile_surface['comment'] = 'snow surface'
                            profile = profile.append(profile_snow_surface)
                        profile = profile.sort_values(by='y_mid').reset_index(drop=True)
                    core.add_profile(profile)
                    logger.info(' (%s) data imported with success: %s' % (profile.get_name(), profile.get_property()))
                else:
                    _temp = [variable for variable in profile.get_variable() if profile[variable].isnull().all()]
                    if _temp.__len__() > 1:
                        logger.info(' (%s) no data to import: %s ' % (profile.get_name(), ", ".join(_temp)))
                    else:
                        logger.info('(%s) no variable to import' % name)

                _imported_variables += variable2import

    return core


def import_ic_list(ic_list, variables=None, v_ref='top', verbose=False, drop_empty=False):
    """
    :param ic_list:
            array, array contains absolute filepath for the cores
    :param variables:
    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)

    ic_dict = {}
    inexisting_ic_list = []
    for ic_path in ic_list:
        if verbose:
            print('Importing data from %s' % ic_path)
        if not os.path.exists(ic_path):
            logger.warning("%s does not exists in core directory" % ic_path.split('/')[-1])
            inexisting_ic_list.append(ic_path.split('/')[-1].split('.')[0])
        else:
            ic_data = import_ic_path(ic_path, variables=variables, v_ref=v_ref, drop_empty=drop_empty)
            if not ic_data.variables():
                inexisting_ic_list.append(ic_path.split('/')[-1].split('.')[0])
                logger.warning("%s have no properties profile" % (ic_data.name))
            else:
                ic_dict[ic_data.name] = ic_data

    logging.info("Import ice core lists completed")
    if inexisting_ic_list.__len__() > 0:
        logger.info("%s core does not exits. Removing from collection" % ', '.join(inexisting_ic_list))

    for ic in inexisting_ic_list:
        for ic2 in ic_dict.keys():
            if ic in ic_dict[ic2].collection:
                ic_dict[ic2].del_from_collection(ic)
                logger.info("remove %s from %s collection" % (ic, ic2))
    return ic_dict


def import_ic_sourcefile(f_path, variables=None, ic_dir=None, v_ref='top', drop_empty=False):
    """
    :param filepath:
            string, absolute path to the file containing either the absolute path of the cores (1 path by line) or the
            core names (1 core by line). In this last case if core_dir is None core_dir is the directory contianing the
            file.
    :param variables:

    :param v_ref:
        top, or bottom
    """
    logger = logging.getLogger(__name__)
    logger.info('Import ice core from source file: %s' % f_path)

    if ic_dir is not None:
        with open(f_path) as f:
            ics = sorted([os.path.join(ic_dir, line.strip()) for line in f if not line.strip().startswith('#')])
    else:
        with open(f_path) as f:
            ics = sorted([line.strip() for line in f if not line.strip().startswith('#')])

    print(ics)

    return import_ic_list(ics, variables=variables, v_ref=v_ref, drop_empty=drop_empty)


# read profile
def read_profile(ws_variable, variables=None, version=__CoreVersion__, v_ref='top', fill_missing=False):
    """
    :param ws_variable:
        openpyxl.worksheet
    :param variables:
    :param version:
    :param v_ref:
        top, or bottom
    :param fill_missing:

    """
    logger = logging.getLogger(__name__)

    if version == 1:
        row_data_start = 6
        row_header = 4
    elif version == 1.1:
        row_data_start = 8
        row_header = 5
        if ws_variable['C4'].value:
            v_ref = ws_variable['C4'].value
    elif version == 1.2:
        row_header = None
        row_data_start = None
    else:
        logger.error("ice core spreadsheet version not defined")

    sheet_2_data = {'S_ice': [row_data_start, 'ABC', 'DEFG', 'J'],
                    'T_ice': [row_data_start, 'A', 'B', 'C'],
                    'Vf_oil': [row_data_start, 'ABC', 'DEFG', 'H'],
                    'salinity': [row_data_start, 'ABC', 'DEFG', 'J'],
                    'temperature': [row_data_start, 'A', 'B', 'C'],
                    'description': [row_data_start, 'ABC', 'DEFG', 'H'],
#                    'density': [row_data_start, 'ABC', 'DEFG', 'H'],
#                    'snow': [row_data_start, 'ABC', 'DEFG', 'H'],
#                    'seawater': [row_data_start, 'ABC', 'DEFG', 'H'],
#                    'sediment': [row_data_start, 'ABC', 'DEFG', 'H'],
                    'BC': [row_data_start, 'ABC', 'DEFG', 'H']
                    }

    # TODO: add other sheets for seawater, sediment, CHla, Phae, stratigraphy
    #                'stratigraphy': [row_data_start, 'AB', 'C', 'D'],
    #                'seawater': [row_data_start, 'A', 'DEFGF', 'G']}

    # define section
    headers_depth = ['y_low', 'y_mid', 'y_sup']
    # Continuous profile
    if not ws_variable.title in sheet_2_data:
        profile = seaice.core.profile.Profile()
    else:
        name = ws_variable['C1'].value
        # Continuous profile
        if sheet_2_data[ws_variable.title][1].__len__() == 1:
            if row_data_start is None:
                rows = [row.value for row in ws_variable['A']]
                row_data_start = rows.index('depth center') + 3
                row_header = rows.index('depth center') + 1

            y_mid = np.array([ws_variable[sheet_2_data[ws_variable.title][1] + str(row)].value
                              for row in range(row_data_start, ws_variable.max_row + 1)]).astype(float)
            y_low = np.nan * np.ones(y_mid.__len__())
            y_sup = np.nan * np.ones(y_mid.__len__())

        # Step profile
        elif sheet_2_data[ws_variable.title][1].__len__() >= 2:
            if row_data_start is None:
                rows = [row.value for row in ws_variable['A']]
                row_data_start = rows.index('depth 1') + 3
                row_header = rows.index('depth 1') + 1

            y_low = np.array([ws_variable[sheet_2_data[ws_variable.title][1][0] + str(row)].value
                              for row in range(row_data_start, ws_variable.max_row+1)]).astype(float)
            y_sup = np.array([ws_variable[sheet_2_data[ws_variable.title][1][1] + str(row)].value
                              for row in range(row_data_start, ws_variable.max_row+1)]).astype(float)
            y_mid = np.array([ws_variable[sheet_2_data[ws_variable.title][1][2] + str(row)].value
                              for row in range(row_data_start, ws_variable.max_row+1)]).astype(float)

            # check if y_mid are not nan:
            if np.isnan(y_mid).any():
                y_mid = (y_low + y_sup) / 2
                logger.info('(%s - %s ) not all y_mid exits, calculating y_mid = (y_low+y_sup)/2'
                            % (name, ws_variable.title))
            elif np.any(np.abs(((y_low + y_sup)/ 2) - y_mid > 1e-12)):
                    logger.error('(%s - %s ) y_mid are not mid point between y_low and y_sup. \\'
                                 'Replacing with y_mid = (y_low+y_sup)/2'
                                 % (name, ws_variable.title))
            else:
                logger.info('(%s - %s ) y_low, y_mid and y_sup read with success'
                            % (name, ws_variable.title))

        data = np.array([y_low, y_mid, y_sup])

        # find row with valid depth entry
        # TODO fix nan check for description spreadsheet

        if np.isnan(y_mid).all():
            logger.warning('(%s - %s ) no data' % (name, ws_variable.title))
            profile = seaice.core.profile.Profile()
        else:
            index = np.where(~np.isnan(y_mid))[0]
            data = data[:, index]

            # read data
            min_col = sheet_2_data[ws_variable.title][2][0]
            min_col = openpyxl.utils.column_index_from_string(min_col)

            max_col = ws_variable.max_column

            min_row = row_data_start
            max_row = min_row + index.__len__()-1

            # data import, supporting merged cell
            _data = []
            for row in ws_variable.iter_cols(min_col, max_col, min_row, max_row):
                _data_row = []
                for cell in row:
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        _data_row.append(np.nan)
                    else:
                        _data_row.append(cell.value)
                _data.append(_data_row)
            _data = np.array(_data)
            # _data = [[cell.value if isinstance(cell.value, (float, int)) else np.nan for cell in row]
            #          for row in ws_variable.iter_cols(min_col, max_col, min_row, max_row)]

            variable_headers = []
            for col in range(min_col, max_col + 1):
                if isinstance(ws_variable.cell(row_header, col), openpyxl.cell.cell.MergedCell):
                    variable_headers.append(None)
                else:
                    variable_headers.append(ws_variable.cell(row_header, col).value)

            headers_index = [headers_i for headers_i, header in enumerate(variable_headers) if header is not None]
            variable_headers = [header for header_i, header in enumerate(variable_headers) if header is not None]
            _data = _data[headers_index]
            data = np.vstack([data, np.array(_data)])
            data = data.transpose()

            # fill missing section with np.nan
            if fill_missing:
                idx = np.where(np.abs(y_low[1:-1]-y_sup[0:-2]) > TOL)[0]
                for ii_idx in idx:
                    empty = [y_sup[ii_idx], (y_sup[ii_idx]+y_low[ii_idx+1])/2, y_low[ii_idx+1]]
                    empty += [np.nan] * (variable_headers.__len__()+1)
                data = np.vstack([data, empty])

            # assemble profile dataframe
            profile = pd.DataFrame(data, columns=headers_depth+variable_headers)

            # set type for property:
            headers_float = [h for h in variable_headers if h not in non_float_property and h not in ['comment', 'Sample #', 'Melting vessel']]
            profile[headers_float] = profile[headers_float].apply(lambda x: pd.to_numeric(x, errors='coerce'))

            # drop property with all nan value
            _profile_notnull = profile[variable_headers].dropna(axis=1, how='all')
            if 'comment' in _profile_notnull.columns:
                profile = profile[headers_depth]
            else:
                profile = profile[headers_depth + ['comment']]
            profile = pd.concat([profile, _profile_notnull], axis=1, sort=False)

            # set comment as string and replace nan value by empty string
            profile.comment = profile.comment.apply(str).replace('nan', '')

            # remove empty line:
            profile = profile.dropna(axis=0, subset=['y_low', 'y_mid', 'y_sup'], how='all')

            # get all property variable (e.g. salinity, temperature, ...)
            # TODO: define not property headers
            property = [var for var in _profile_notnull.columns if var not in ['comment', 'Sample #', 'Melting vessel']]

            if len(property) == 0:
                profile = seaice.core.profile.Profile()
            else:
                # remove subvariable (e.g. conductivity temperature measurement for conductivity
                property = [prop for prop in property if prop not in inverse_dict(subvariable_dict)]

                # set variable to string of property
                profile['variable'] = [', '.join(property)]*len(profile.index)
                if isinstance(property, list):
                    print('\t' + ', '.join(property) + ' (' + ws_variable.title +')')

                else:
                    print('\t' + property + '  in sheet ' + ' (' + ws_variable.title +')')
                # ice thickness
                try:
                    length = float(ws_variable['C8'].value)
                except:
                    logger.info('(%s) no ice core length' % name)
                    length = np.nan
                else:
                    if length == 'n/a':
                        logger.info('(%s) ice core length is not available (n/a)' % name)
                        length = np.nan
                    elif not isinstance(length, (int, float)):
                        logger.info('%s ice core length is not a number' % name)
                        length = np.nan
                profile['length'] = [length]*len(profile.index)

                # set vertical references
                profile['v_ref'] = [v_ref]*len(profile.index)

                # set ice core name for profile
                profile['name'] = [name]*len(profile.index)

                # set depth and property type to float
                headers_float = [h for h in profile.columns if h not in ['comment', 'v_ref', 'name', 'profile', 'variable']
                                 and h not in non_float_property]


                profile[headers_float] = profile[headers_float].apply(lambda x: pd.to_numeric(x, errors='coerce'))

                profile = seaice.core.profile.Profile(profile)

                # remove variable not in variables
                if variables is not None:
                    for property in profile.properties():
                        if property not in variables:
                            profile.delete_property(property)

    return profile


# create list or source
def list_folder(dirpath, fileext='.xlsx', level=0):
    """
    list all files with specific extension in a directory

    :param dirpath: str; directory to scan for ice core
    :param fileext: str, default .xlsx; file extension for ice core data
    :param level: numeric, default 0; level of recursitivy in directory search
    :return ic_list: list
        list of ice core path
    """
    _ics = []

    logger = logging.getLogger(__name__)

    def walklevel(some_dir, level=level):
        some_dir = some_dir.rstrip(os.path.sep)
        assert os.path.isdir(some_dir)
        num_sep = some_dir.count(os.path.sep)
        for root, dirs, files in os.walk(some_dir):
            yield root, dirs, files
            num_sep_this = root.count(os.path.sep)
            if num_sep + level <= num_sep_this:
                del dirs[:]

    for dirName, subdirList, fileList in walklevel(dirpath, level=level):
        _ics.extend([dirName + '/' + f for f in fileList if f.endswith(fileext)])

    ics_set = set(_ics)
    logger.info("Found %i ice core datafile in %s" % (ics_set.__len__(), dirpath))

    return ics_set


def list_ic_path(dirpath, fileext):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return ic_list: list
        list of ice core path
    """
    logger = logging.getLogger(__name__)

    ics_set = list_folder(dirpath=dirpath, fileext=fileext)
    ic_paths_set = set([os.path.join(os.path.realpath(dirpath), f) for f in ics_set])
    return ic_paths_set


def make_ic_sourcefile(dirpath, fileext, source_filepath=None):
    """
    list all files with specific extension in a directory

    :param dirpath: str
    :param fileext: str
    :return source_file: str
        filepath to the text file containing ice core filepath with absolute path.
    """
    logger = logging.getLogger(__name__)

    ic_paths_set = list_ic_path(dirpath, fileext)

    if source_filepath is None:
        source_filepath = os.path.join(os.path.realpath(dirpath), 'ic_list.txt')

    with open(source_filepath, 'w') as f:
        for ic_path in ic_paths_set:
            f.write(ic_path + "\n")

    return source_filepath


def read_ic_list(file_path):
    """

    :param file_path:
    :return:
    """

    with open(file_path) as f:
        data = [list(map(int, row.split())) for row in f.read().split('\n\n')]


    return ic_list

# updater
def update_spreadsheet(ic_path, v_ref='top', backup=True):
    """
    update_spreadsheet update an ice core file to the latest ice core file version (__CoreVersion__).

    :param ic_path: str
        Path to the ice core file to update
    :param v_ref: int, 'top' or 'bottom', Default 'top'
        Set the zero vertical reference of the core. If 'top, zero vertical reference is set at the ice/snow or ice/air
        interface. If 'bottom, zero vertical reference is set at the ice/water interface
    :param verbose: int
        verbosity of the function.
    :param backup: bool, default True
        If True, backup the previous file version in the subdirectory 'folder-COREVERSION'
    :return nothing:

    USAGE:
        update_spreadhseet(ic_path)
    """
    logger = logging.getLogger(__name__)

    import shutil
    if not os.path.exists(ic_path):
        logger.error("%s does not exists in core directory" % ic_path.split('/')[-1])
        return 0
    else:
        logger.info("\t ice core file path %s" % ic_path)

    wb = openpyxl.load_workbook(filename=ic_path, keep_vba=False)  # load the xlsx spreadsheet
    ws_summary = wb['summary']  # load the data from the summary sheet

    if isinstance(ws_summary['C3'].value, (float, int)):
        version = ws_summary['C3'].value
        if version < __CoreVersion__:
            logger.info("Updating core data to latest version %.1f" % __CoreVersion__)
    else:
        logger.error("\t%s ice core file version not unavailable" % ic_path)

    while version < __CoreVersion__:

        # backup old version
        if backup:
            backup_dir = os.path.join(os.path.dirname(ic_path), 'version-' + str(version))
            ic_bkp = os.path.join(backup_dir, os.path.basename(ic_path))

            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
            shutil.move(ic_path, ic_bkp)

         # update from 1.0 to 1.1
        if version == 1:
            # update from 1.0 to 1.1
            version = 1.1
            ws_summary['C3'] = version
            # remove rows "core number in series"
            ws_summary.delete_rows(22, 1)

            if "S_ice" in wb.sheetnames:
                ws = wb['S_ice']

                # add reference row=4
                ws.insert_rows(4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # remove "sample number" column
                if ws['M5'].value == 'sample number':
                    ws.delete_cols(openpyxl.utils.cell.column_index_from_string('M'))

                # remove salinity for temperature
                range = "F5:" + openpyxl.utils.cell.get_column_letter(ws.max_column) + str(ws.max_row)
                ws.move_range(range, rows=0, cols=-1)

                # move d180 dD after conductance with T_sigma (conductivity) suppresion
                range = "E5:G" + str(ws.max_row)
                ws.move_range(range, rows=0, cols=6)
                range = "H5:" + openpyxl.utils.cell.get_column_letter(ws.max_column) + str(ws.max_row)
                ws.move_range(range, rows=0, cols=-3)

                # add notation row=6
                ws.insert_rows(6)

                ws['A6'] = 'd_1'
                ws['B6'] = 'd_2'
                ws['C6'] = 'd'
                ws['D5'] = 'salinity'
                ws['D6'] = 'S'
                ws['E5'] = 'conductivity'
                ws['E6'] = 'σ'
                ws['E7'] = 'mS/cm'
                ws['F5'] = 'conductivity measurement temperature'
                ws['F6'] = 'T_σ'
                ws['G5'] = 'specific conductance'
                ws['G6'] = 'κ'
                ws['G7'] = 'mS/cm'
                ws['H6'] = 'd180'
                ws['I6'] = 'dD'
                ws['J6'] = ''

            if "T_ice" in wb.sheetnames:
                ws = wb["T_ice"]

                # add reference row=4
                ws.insert_rows(4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws.insert_rows(6)
                ws['A6'] = 'd'
                ws['B6'] = 'T'

            if "oil_content" in wb.sheetnames:
                ws = wb["oil_content"]
                # add reference row=4
                ws.insert_rows(4)
                ws['A4'] = 'vertical reference'
                ws['C4'] = v_ref

                # add notation row=6
                ws.insert_rows(6)
                ws['A6'] = 'd_1'
                ws['B6'] = 'd_2'
                ws['C6'] = 'd'
                ws['D6'] = 'V_i'
                ws['E6'] = 'h_menisc'
                ws['F6'] = 'd_menisc'
                ws['G6'] = 'd_center'

            if "S-figure" in wb.sheetnames:
                ws = wb['S-figure']
                wb.remove_sheet(ws)

            if "T-figure" in wb.sheetnames:
                ws = wb['T-figure']
                wb.remove_sheet(ws)

    wb.save(ic_path)
    wb.close()


def inverse_dict(map):
    """
    return the inverse of a dictionnary with non-unique values
    :param map: dictionnary
    :return inv_map: dictionnary
    """
    revdict = {}
    for k, v in map.items():
        if isinstance(v, list):
            for vi in v:
                revdict[vi] = revdict.get(vi, [])
                revdict[vi].append(k)
        else:
            revdict[v] = revdict.get(v, [])
            revdict[v].append(k)
    return revdict