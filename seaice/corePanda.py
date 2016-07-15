#! /usr/bin/python3
# -*- coding: utf-8 -*-
"""
corePanda.py: ice core data is a toolbox to import ice core data file from xlsx spreadsheet, formatted according to the
 template developped by the Sea Ice Group of the Geophysical Institute of the University of Alaska, Fairbanks.
 corePanda.py integrate the module panda into the module core version 2.1 to simplify the operation and decrease
 computation time. Core profiles are considered as collection of point in depth, time and properties (salinity, temperature or other variable)
"""

## TODO:take in account the timezone

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import pandas as pd
import logging
import openpyxl
import datetime
from seaice.properties import si_prop_list
from seaice.properties import si_prop_unit
from seaice.properties import si_state_variable
import seaice.icdtools

__author__ = "Marc Oggier"
__license__ = "GPL"
__version__ = "3.0.0"
__maintainer__ = "Marc Oggier"
__contact__ = "Marc Oggier"
__email__ = "marc.oggier@gi.alaska.edu"
__status__ = "development"
__date__ = "2014/11/25"
__comment__ = "corePanda integrate the module Panda to simplifiy the operation "
__all__ = ['Core', 'CoreSet', 'Profile']

LOG_FILENAME = 'import.log'
LOG_LEVELS = {'debug': logging.DEBUG,
              'info': logging.INFO,
              'warning': logging.WARNING,
              'error': logging.ERROR,
              'critical': logging.CRITICAL}
nan_value = float('nan')

import time

from .properties import si_prop_unit
import seaice.icdtools

def timing(f):
    def wrap(*args):
        time1 = time.time()
        ret = f(*args)
        time2 = time.time()
        print('%s function took %0.3f ms' % (f.func_name, (time2-time1)*1000.0))
        return ret
    return wrap


class Core:
    """
    Core
    """
    def __init__(self, core_name, coring_date, coring_location, ice_thickness, snow_thickness, comment=None, note=None):
        """
        :param name:
        :param date:
        :param location:
        :param ice_thickness:
        :param snow_thickness:
        :param comment:
        :return:
        """
        self.core_name = core_name
        self.coring_date = coring_date
        self.coring_location = coring_location
        if isinstance(core_name, list):
            self.core_collection = core_name
        else:
            self.core_collection = [core_name]
        self.ice_thickness = ice_thickness
        self.snow_thickness = snow_thickness
        self.profiles = pd.DataFrame()
        self.comment = None
        if comment is not None:
            self.add_comment(comment)
        self.note = None
        if note is not None:
            self.add_comment(note)

    def add_profile(self, profile):
        """
        :param profile:
        :return:
        """
        core_info = ['core_name', 'coring_date', 'location', 'core_collection', 'ice_thickness', 'snow_thickness', 'coring_comment', 'coring_note']
        core_data = pd.DataFrame([[self.core_name, self.coring_date, self.coring_location, self.core_collection, self.ice_thickness, self.snow_thickness, self.comment, self.note]], columns=core_info, index=profile.index)
        profile = profile.join(core_data)
        self.profiles = self.profiles.append(profile)

    def del_profile(self, core, variable=None):
        self.profiles = self.profiles[(self.profiles.core != core) & (self.profiles.variable != variable)]

    def add_corenames(self, core_name):
        """
        :param core_name:
        :return:
        """
        if isinstance(core_name, list):
            for ii_core in core_name:
                if ii_core not in self.core_collection:
                    self.core_collection.append(ii_core)
        elif core_name not in self.core_collection:
            self.core_collection.append(core_name)

    def add_comment(self, comment):
        """
        :param comment:
        :return:
        """
        if self.comment is None:
            self.comment = [comment]
        else:
            self.comment.append(comment)

    # def calc_prop(self, property):
    #     """
    #     :param property:
    #     :return:
    #     """
    #     # check properties variables
    #     if property not in si_prop_list.keys():
    #         logging.warning('property %s not defined in the ice core property module' % property)
    #         return None
    #     elif 'salinity' not in self.profiles:
    #         logging.warning('ice core %s is missing salinity profile for further calculation' % self.name)
    #         return None
    #     elif 'temperature' not in self.profiles:
    #         logging.warning('ice core %s is missing temperature profile for further calculation' % self.name)
    #         return None
    #     else:
    #         import seaice.properties
    #         variable = si_prop_list[property]
    #         function = getattr(seaice.properties, variable.replace(" ", "_"))
    #         profilet = self.profiles['temperature']
    #         profiles = self.profiles['salinity']
    #         y = np.sort(np.unique(np.concatenate((profilet.y, profiles.y))))
    #         y = y[np.where(y <= max(max(profilet.y), max(profiles.y)))]
    #         y = y[np.where(y >= max(min(profilet.y), min(profiles.y)))]
    #         y_mid = y[0:-1] + np.diff(y) / 2
    #         length = max(y) - min(y)
    #
    #         xt = np.interp(y_mid, profilet.y, profilet.x, left=np.nan, right=np.nan)
    #         xs = np.interp(y_mid, profiles.y[:-1] + np.diff(profiles.y) / 2, profiles.x, left=np.nan, right=np.nan)
    #         x = function(xt, xs)
    #
    #         note = 'computed from ' + self.profiles['temperature'].profile_label + ' temperature profile and ' + \
    #                self.profiles['salinity'].profile_label + ' salinity profile'
    #         profile_label = self.name
    #         prop_profile = Profile(x, y, profile_label, variable, comment=None, note=note, length=length)
    #         self.add_profile(prop_profile, variable)
    #         self.add_comment('computed %s' % variable)

    def plot_variable(self, ax, variable, param_dict=None):
        """
        :param ax:
        :param variable:
        :param param_dict:
        :return:
        """
        if variable in si_state_variable.keys():
            profile_label = si_state_variable[variable.lower()]
        elif variable in si_prop_list.keys():
            profile_label = si_prop_list[variable.lower()]
            if profile_label not in self.profiles.keys():
                logging.warning('computing %s' % variable)
                self.calc_prop(variable)
        else:
            logging.warning('no data available to plot')
            return None

        profile = self.profiles[profile_label]
        y = profile.y
        x = profile.x
        if profile.y.__len__() > profile.x.__len__():
            x = np.concatenate((x, np.atleast_1d(x[-1])))
            if param_dict is None:
                out = ax.step(x, y)
            else:
                out = ax.step(x, y, **param_dict)
        else:
            if param_dict is None:
                out = ax.plot(x, y)
            else:
                out = ax.plot(x, y, **param_dict)

        ax.set_xlabel(profile_label + ' ' + si_prop_unit[profile_label])
        ax.set_ylim(max(ax.get_ylim()), 0)
        return out

    def plot_state_variable(self, flag_figure_number=None, param_dict=None):
        """
        :param flag_figure_number:
        :param param_dict:
        :return:
        """
        if flag_figure_number is None:
            fig = plt.figure()
            ax1 = fig.add_subplot(1, 2, 1)
            ax2 = fig.add_subplot(1, 2, 2, sharey=ax1)
        else:
            fig = plt.figure(flag_figure_number)
            ax1 = fig.add_subplot(1, 2, 1)
            ax2 = fig.add_subplot(1, 2, 2, sharey=ax1)

        if 'salinity' in self.profiles.keys():
            self.plot_variable(ax1, 'salinity', param_dict)
            ax1.set_ylabel('depth [m]')
        else:
            logging.warning('salinity profile missing for %s' % self.name)

        if 'temperature' in self.profiles.keys():
            self.plot_variable(ax2, 'temperature', param_dict)
        else:
            logging.warning('temperature profile missing for %s' % self.name)

        ax_fig = [ax1, ax2]
        return fig, ax_fig

    def get_profile_variable(self):
        return sorted(self.profiles.keys())

    def rescale(self, variable=None, section_thickness=0.05):
        return make_section(self, variable, section_thickness)


class CoreStack(pd.DataFrame):
    @property
    def _constructor(self):
        return CoreStack
    pass

    def __init__(self, *args, **kwargs):
        super(CoreStack, self).__init__(*args, **kwargs)

    def add_profiles(self, profile, ignore_index=True, verify_integrity=False):
        temp = self.append(profile, ignore_index=ignore_index, verify_integrity=verify_integrity)
        return CoreStack(temp)

    def remove_profiles(self, core, variable=None):
        if variable is None:
            temp = self[self.core!= core]
        elif isinstance(variable, list):
            for ii_variable in core:
                temp = self[(self.core != core) & (self.variable != ii_variable)]
        else:
            temp = self[(self.core != core) & (self.variable != variable)]
        return CoreStack(temp)

    def select_profile(self, variable_dict):
        str_select = '('
        ii_var = []
        ii = 0
        for ii_key in variable_dict.keys():
            if ii_key in self.columns.values:
                ii_var.append(variable_dict[ii_key])
                str_select = str_select + 'self.' + ii_key + '==ii_var[' + str('%d' % ii) + ']) & ('
                ii += 1
        str_select = str_select[:-4]

        index_select = self[eval(str_select)].index
        index_deselect = [ii for ii in self.index.tolist() if ii not in index_select]
        return self.iloc[index_select], self.iloc[index_deselect]

    def add_variable(self, variable_dict, data):
        for col in data.columns.tolist():
            if col not in self.columns.tolist():
                self[col] = np.nan

        data_select, data_deselect = self.select_profile(variable_dict)
        temp = data_select.drop(data.columns, axis=1)
        if data.__len__() != data_select.__len__():
            index = data_select.index.tolist()
            data = pd.DataFrame([data.iloc[0]], columns=data.columns, index=index)

        temp = pd.concat([temp, data], axis=1)
        ics_data_stack = pd.concat([data_deselect, temp])
        return CoreStack(ics_data_stack)

    def discretize(self, y_bins=None, y_mid=None, variables=None, comment='y', display_figure='n'):
        """
        :param y_bins:
        :param y_mid:
        :param variables:
        :param comment: {0, 1}, Default 0
        :param display_figure: {0, 1}, Default 0
        :return:
        """

        if y_bins is None and y_mid is None:
            y_bins = pd.Series(self.y_low.dropna().tolist()+self.y_sup.dropna().tolist()).sort_values().unique()
            y_mid = self.y_mid.dropna().sort_values().unique()

        elif y_mid is None:
            y_mid = self.y_mid.dropna().sort_values().unique()

        for ii_core in sorted(self.core_name.unique().tolist()):
            if comment == 'y':
                print(ii_core)
            ic_data = self[self.core_name == ii_core]
            ic_data = discretize_profile(ic_data, y_bins=y_bins, y_mid=y_mid, variables=None, comment=comment, display_figure=display_figure)
            self = self[(self.core_name != ii_core)]
            self = self.append(ic_data)

        return CoreStack(self.reset_index())

    def grouped_stat(self, variables, stats, bins_DD, bins_y, comment='n'):
        y_cuts = pd.cut(self.y_mid, bins_y, labels=False)
        t_cuts = pd.cut(self.DD, bins_DD, labels=False)

        if not isinstance(variables, list):
            variables = [variables]
        if not isinstance(stats, list):
            stats = [stats]

        temp_all = pd.DataFrame()
        for ii_variable in variables:
            if comment == 'y':
                print('\ncomputing %s' % ii_variable)
            data = self[self.variable == ii_variable]
            data_grouped = data.groupby([t_cuts, y_cuts])

            for ii_stat in stats:
                if comment == 'y':
                    print('computing %s' % ii_stat)
                func = "groups['" + ii_variable + "']." + ii_stat + "()"
                stat_var = np.nan * np.ones((bins_DD.__len__() - 1, bins_y.__len__()))
                core_var = [[[None] for x in range(bins_y.__len__())] for y in range(bins_DD.__len__() - 1)]
                for k1, groups in data_grouped:
                    stat_var[int(k1[0]), int(k1[1])] = eval(func)
                    core_var[int(k1[0])][int(k1[1])] = [list(groups.dropna(subset=[ii_variable])['core_name'].unique())]
                for ii_bin in range(stat_var.__len__()):
                    temp = pd.DataFrame(stat_var[ii_bin], columns=[ii_variable])
                    temp = temp.join(pd.DataFrame(core_var[ii_bin], columns=['core_collection']))
                    DD_label = 'DD-' + str(bins_DD[ii_bin]) + '_' + str(bins_DD[ii_bin + 1])
                    data = [str(bins_DD[ii_bin]), str(bins_DD[ii_bin + 1]), DD_label, int(ii_bin), ii_stat, ii_variable]
                    columns = ['DD_min', 'DD_max', 'DD_label', 'DD_index', 'stats', 'variable']
                    index = np.array(temp.index.tolist())  #[~np.isnan(temp[ii_variable].tolist())]
                    temp = temp.join(pd.DataFrame([data], columns=columns, index=index))
                    temp = temp.join(pd.DataFrame(index, columns=['y_index'], index=index))
                    columns = ['y_low', 'y_sup', 'y_mid', 'sample_name']
                    t2 = pd.DataFrame(columns=columns)
                    # For step profile, like salinity
                    # if ii_variable in ['salinity']:
                    if not self[self.variable == ii_variable].y_low.isnull().any():
                        for ii_layer in index[:-1]:
                            data = [bins_y[ii_layer], bins_y[ii_layer + 1], (bins_y[ii_layer] + bins_y[ii_layer + 1]) / 2, DD_label + str('-%03d' % ii_layer)]
                            t2 = t2.append(pd.DataFrame([data], columns=columns, index=[ii_layer]))
                    # For linear profile, like temperature
                    # if ii_variable in ['temperature']:
                    elif self[self.variable == ii_variable].y_low.isnull().all():
                        for ii_layer in index[:-1]:
                            data = [np.nan, np.nan, (bins_y[ii_layer] + bins_y[ii_layer + 1]) / 2, DD_label + str('-%03d' % ii_layer)]
                            t2 = t2.append(pd.DataFrame([data], columns=columns, index=[ii_layer]))

                    if temp_all.empty:
                        temp_all = temp.join(t2)
                    else:
                        temp_all = temp_all.append(temp.join(t2), ignore_index=True)

        data_grouped = self.groupby([t_cuts, self['variable']])

        grouped_dict = {}
        for var in variables:
            grouped_dict[var] = [[] for ii_DD in range(bins_DD.__len__()-1)]

        for k1, groups in data_grouped:
            grouped_dict[k1[1]][int(k1[0])] = groups['core'].unique().tolist()

        return CoreStack(temp_all.reset_index()), grouped_dict

    def plot_core(self, core_dict, ax=None, param_dict=None):
        for ii_core in core_dict:
            ic_data = self.select_profile(core_dict)[0]
            ax = plot_profile(ic_data, core_dict['variable'], ax=ax, param_dict=param_dict)
        return ax

    def plot_core_profile(self, core, ax=None, variable=None, param_dict=None):
        """
        :param core:
        :param ax:
        :param variable:
        :param param_dict:
        :return:
        """
        ic_data = self.select_profile({'variable': variable, 'core_name':core})[0].reset_index()
        if ic_data[ic_data.variable == variable].__len__() != 0:
            ax = plot_profile(ic_data, variable, ax = ax, param_dict=param_dict)
        ax.axes.set_xlabel(variable + ' ' + si_prop_unit[variable])
        ax.axes.set_ylim(max(ax.get_ylim()), 0)
        return ax

    def plot_stat_mean(self, ax, variable, bin_index):
        x_mean = self.select_profile({'stats': 'mean', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
        x_max = self.select_profile({'stats': 'max', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
        x_min  = self.select_profile({'stats': 'min', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
        x_std = self.select_profile({'stats': 'std', 'variable': variable, 'DD_index': bin_index})[0].reset_index()

        if x_mean[variable].__len__() !=0:
            plot_profile(x_max, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'r', 'label': 'min'})
            plot_profile(x_min, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'b', 'label': 'max'})
            plot_profile(x_mean, variable, ax=ax,  param_dict={'linewidth': 3, 'color': 'k', 'label': 'mean'})

            if x_std.__len__() < x_mean.__len__():
                index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
                x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))

            if variable in ['salinity']:
                y_low = x_mean['y_low']
                y_sup = x_mean['y_sup']
                x_std_l = x_mean[variable][0] - x_std[variable][0]
                x_std_h = x_mean[variable][0] + x_std[variable][0]
                y_std = y_low[0]
                for ii in range(1, len(x_mean)):
                    x_std_l = np.append(x_std_l, x_mean[variable][ii - 1] - x_std[variable][ii - 1])
                    x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
                    x_std_h = np.append(x_std_h, x_mean[variable][ii - 1] + x_std[variable][ii - 1])
                    x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
                    y_std = np.append(y_std, y_low[ii])
                    y_std = np.append(y_std, y_low[ii])
                if len(x_mean) == 1:
                    ii = 0
                x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
                x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
                y_std = np.append(y_std, y_sup[ii])
            elif variable in ['temperature']:
                y_std = x_mean['y_mid']
                x_std_l = []
                x_std_h = []
                for ii in range(0, len(x_mean)):
                    x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
                    x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
            ax = plt.fill_betweenx(y_std, x_std_l, x_std_h, facecolor='black', alpha=0.3, label=str(r"$\pm$"+"std dev"))
            ax.axes.set_xlabel(variable + ' ' + si_prop_unit[variable])
            ax.axes.set_ylim([max(ax.axes.get_ylim()), min(ax.axes.get_ylim())])
        return ax

    def plot_stat_median(self, ax, variable, bin_index):
        x_mean = self.select_profile({'stats': 'median', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
        x_max = self.select_profile({'stats': 'max', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
        x_min = self.select_profile({'stats': 'min', 'variable': variable, 'DD_index': bin_index})[0].reset_index()
        x_std = self.select_profile({'stats': 'mad', 'variable': variable, 'DD_index': bin_index})[0].reset_index()

        if x_mean[variable].__len__() != 0:
            plot_profile(x_max, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'r', 'label':'min'})
            plot_profile(x_min, variable, ax=ax, param_dict={'linewidth': 3, 'color': 'b', 'label':'max'})
            plot_profile(x_mean, variable, ax=ax,  param_dict={'linewidth': 3, 'color': 'k', 'label':'median'})

            if x_std.__len__() < x_mean.__len__():
                index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
                x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))

            if variable in ['salinity']:
                y_low = x_mean['y_low']
                y_sup = x_mean['y_sup']
                x_std_l = x_mean[variable][0] - x_std[variable][0]
                x_std_h = x_mean[variable][0] + x_std[variable][0]
                y_std = y_low[0]
                for ii in range(1, len(x_mean)):
                    x_std_l = np.append(x_std_l, x_mean[variable][ii - 1] - x_std[variable][ii - 1])
                    x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
                    x_std_h = np.append(x_std_h, x_mean[variable][ii - 1] + x_std[variable][ii - 1])
                    x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
                    y_std = np.append(y_std, y_low[ii])
                    y_std = np.append(y_std, y_low[ii])
                if len(x_mean) == 1:
                    ii = 0
                x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
                x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
                y_std = np.append(y_std, y_sup[ii])
            elif variable in ['temperature']:
                y_std = x_mean['y_mid']
                x_std_l = []
                x_std_h = []
                for ii in range(0, len(x_mean)):
                    x_std_l = np.append(x_std_l, x_mean[variable][ii] - x_std[variable][ii])
                    x_std_h = np.append(x_std_h, x_mean[variable][ii] + x_std[variable][ii])
            ax = plt.fill_betweenx(y_std, x_std_l, x_std_h, facecolor='black', alpha=0.3,
                                            label=str(r"$\pm$"+"mad"))
            ax.axes.set_xlabel(variable + ' ' + si_prop_unit[variable])
            ax.axes.set_ylabel('ice thickness [m]')
            ax.axes.set_ylim([max(ax.axes.get_ylim()), min(ax.axes.get_ylim())])
        return ax

    def core_set(self):
        return list(set(seaice.icdtools.flatten_list(self.core_collection.tolist())))

    def compute_physical_property(self, si_prop, s_profile_shape='linear'):
        ## look for all core belonging to a coring event:
        temp_core_processed = []
        ic_prop = seaice.corePanda.CoreStack()
        for f_core in sorted(self.core_name.unique()):
            # f_core = ics_obs_stack.core_name.unique()[41]
            ic = self[self.core_name == f_core]
            ic_data = seaice.corePanda.CoreStack()
            print('\n')
            if f_core not in temp_core_processed:
                for ff_core in list(set(seaice.icdtools.flatten_list(ic.core_collection.tolist()))):
                    print(ff_core)
                    ic_data = ic_data.add_profiles(self[self.core_name == ff_core])
                ic_prop = ic_prop.append(seaice.corePanda.calc_prop(ic_data, si_prop, s_profile_shape=s_profile_shape))
                temp_core_processed.append(ff_core)

        ics_stack = seaice.corePanda.CoreStack(self)
        ic_prop = seaice.corePanda.CoreStack(ic_prop)

        ics_stack = ics_stack.add_profiles(ic_prop)
        return(ics_stack)

    # def add_property(self, profile_dict, profile_property):
    #     return None
    #
    # def ice_thickness(self):
    #     if 'ice_thickness' in self.columns:
    #         ice_thickness = None
    #     else:
    #         for ii_variable in self.variable.unique():
    #             temp = self.select_profile({ii_variable})
    #         ice_thickness = None
    #     return ice_thickness

#
# class CoreSet:
#     """
#     CoreSet() is a class
#     """
#     def __init__(self, set_name, core, comment=None):
#         """
#         :rtype: CoreSet
#         """
#         self.name = set_name
#         self.core_data = {core.name: core}
#         self.core = [core.name]
#         self.comment = None
#         self.variables = []
#         self.add_variable(core)
#         if comment is not None:
#             self.add_comment(comment)
#
#     def add_variable(self, core):
#         for ii_variable in core.profiles.keys():
#             if ii_variable not in self.variables:
#                 self.variables.append(ii_variable)
#
#     def add_comment(self, comment):
#         """
#         :param comment:
#         :return:
#         """
#         if self.comment is None:
#             self.comment = [comment]
#         else:
#             self.comment.append(comment)
#
#     def add_core(self, core):
#         """
#         :param core:
#         :return:
#         """
#         self.core_data[core.name] = core
#         self.core.append(core.name)
#         self.add_variable(core)
#
#     def ice_thickness(self):
#         """
#         :return:
#         """
#         hi = []
#         for key in self.core_data.keys():
#             if self.core_data[key].ice_thickness is not None:
#                 hi.append(self.core_data[key].ice_thickness)
#         if hi == [] or np.isnan(hi).all():
#             return None
#         else:
#             return hi, np.nanmean(hi), np.nanmax(hi)
#
#     def core_length(self):
#         """
#         :return:
#         """
#         lc = []
#         for key in self.core_data.keys():
#             core = self.core_data[key]
#             for a in dir(core):
#                 try:
#                     temp = core.__getattribute__(a).length
#                 except AttributeError:
#                     logging.warning('core length for %s not defined' % a)
#                 else:
#                     lc.append(temp)
#         return lc, np.nanmean(lc), np.nanmax(lc)
#
#     def merge_bin(self):
#         """
#             :param self: CoreSet
#             :return:
#         """
#         variable = {}
#         for ii_core in self.core:
#             ic_data = self.core_data[ii_core]
#             for ii_variable in ic_data.profiles.keys():
#                 y = ic_data.profiles[ii_variable].y
#                 if ii_variable not in variable.keys():
#                     variable[ii_variable] = np.array(y)
#                 else:
#                     variable[ii_variable] = np.unique(np.append(y, variable[ii_variable]))
#                 variable[ii_variable] = variable[ii_variable][~np.isnan(variable[ii_variable])]
#
#         flag = 0
#         for ii_core in self.core:
#             ic_data = self.core_data[ii_core]
#             ic = Core(ic_data.name, ic_data.date, ic_data.location, ic_data.ice_thickness, ic_data.snow_thickness, ic_data.add_comment('merged bin'))
#
#             for ii_variable in self.variables:
#                 if ii_variable in ic_data.profiles.keys():
#                     x = np.array(ic_data.profiles[ii_variable].x)
#                     y = np.array(ic_data.profiles[ii_variable].y)
#                     if len(x) == len(y):
#                         y_bin = variable[ii_variable]
#                         x_bin = np.interp(y_bin, y[~np.isnan(y)], x[~np.isnan(y)], left=np.nan, right=np.nan)
#                     else:
#                         ii_bin = 1  # cycle from 0 to len(y)
#                         ii = 0  # cycle from 0 to len(x_bin)
#                         x_bin = np.nan*np.ones(len(variable[ii_variable])-1)
#                         y_bin = variable[ii_variable]
#                         while ii < len(x_bin):
#                             if y_bin[ii] < y[ii_bin]:
#                                 x_bin[ii] = x[ii_bin-1]
#                                 ii += 1
#                             elif y_bin[ii] == max(y):
#                                 break
#                             else:
#                                 ii_bin += 1
#
#                     profile = Profile(x_bin, y_bin, ic_data.name, ii_variable, comment='merged bin from core set',
#                                       note=None, length=None)
#                     ic.add_profile(profile, ii_variable)
#
#             if flag == 0:
#                 ics_merged_bin_set = CoreSet(self.name+'-merged_bin', ic)
#                 flag = 1
#             else:
#                 ics_merged_bin_set.add_core(ic)
#             ics_merged_bin_set.add_comment('merge bin cores')
#         return ics_merged_bin_set
#
#     # statistic
#     # def mean(self, var=None):
#     #     """
#     #     :param var:
#     #     :return:
#     #     """
#     #
#     #     variable = {}
#     #     snow_thickness = []
#     #     ice_thickness = []
#     #     date = []
#     #     ics_data = self.merge_bin()
#     #     for ii_core in ics_data.core:
#     #         ic_data = ics_data.core_data[ii_core]
#     #         if var is None:
#     #             for ii_variable in ics_data.variables:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #         else:
#     #             if not isinstance(var, list):
#     #                 var = [var]
#     #             for ii_variable in var:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #
#     #         str_temp = ''
#     #         for ii_core in ics_data.core:
#     #             str_temp += ii_core+', '
#     #         str_temp = str_temp[:-2]
#     #         comment = 'statistic mean computed from ice cores: ' + str_temp
#     #         name = ic_data.name.split('-')[0]+'-'+ic_data.name.split('-')[1][0:8]+'-mean'
#     #
#     #         ic_out = Core(name, date, ic_data.location, np.nanmean(ice_thickness), np.nanmean(snow_thickness), comment='mean value of '+str_temp)
#     #
#     #         for ii_variable in variable.keys():
#     #             y = variable[ii_variable][1]
#     #             x = np.nanmean(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #             count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
#     #
#     #             profile = Profile(x, y, name, 'mean '+ii_variable, comment=comment, note=count, length=None)
#     #             ic_out.add_profile(profile, ii_variable)
#     #     return ic_out
#     #
#     # def std(self, var=None):
#     #     """
#     #     :param var:
#     #     :return:
#     #     """
#     #     variable = {}
#     #     snow_thickness = []
#     #     ice_thickness = []
#     #     date = []
#     #     ics_data = self.merge_bin()
#     #     for ii_core in ics_data.core:
#     #         ic_data = ics_data.core_data[ii_core]
#     #         if var is None:
#     #             for ii_variable in ics_data.variables:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #         else:
#     #             if not isinstance(var, list):
#     #                 var = [var]
#     #             for ii_variable in var:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #
#     #         str_temp = ''
#     #         for ii_core in ics_data.core:
#     #             str_temp += ii_core+', '
#     #         comment = 'statistic mean computed from ice cores: ' + str_temp[:-2]
#     #         name = ic_data.name.split('-')[0]+'-'+ic_data.name.split('-')[1][0:8]+'-mean'
#     #         ic_out = Core(name, date, ic_data.location, np.nanmean(ice_thickness), np.nanmean(snow_thickness), comment='mean value of '+str_temp[:-2])
#     #
#     #         for ii_variable in variable.keys():
#     #             y = variable[ii_variable][1]
#     #             x = np.nanstd(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #             count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
#     #
#     #             profile = Profile(x, y, name, 'mean '+ii_variable, comment=comment, note=count, length=None)
#     #             ic_out.add_profile(profile, ii_variable)
#     #     return ic_out
#     #
#     #
#     # def min(self, var=None):
#     #     """
#     #     :param var:
#     #     :return:
#     #     """
#     #     variable = {}
#     #     snow_thickness = []
#     #     ice_thickness = []
#     #     date = []
#     #     ics_data = self.merge_bin()
#     #     for ii_core in ics_data.core:
#     #         ic_data = ics_data.core_data[ii_core]
#     #         if var is None:
#     #             for ii_variable in ics_data.variables:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #         else:
#     #             if not isinstance(var, list):
#     #                 var = [var]
#     #             for ii_variable in var:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #
#     #         str_temp = ''
#     #         for ii_core in ics_data.core:
#     #             str_temp += ii_core+', '
#     #         comment = 'statistic mean computed from ice cores: ' + str_temp[:-2]
#     #         name = ic_data.name.split('-')[0]+'-'+ic_data.name.split('-')[1][0:8]+'-mean'
#     #         ic_out = Core(name, date, ic_data.location, np.nanmean(ice_thickness), np.nanmean(snow_thickness), comment='mean value of '+str_temp[:-2])
#     #
#     #         for ii_variable in variable.keys():
#     #             y = variable[ii_variable][1]
#     #             x = np.nanmin(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #             count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
#     #
#     #             profile = Profile(x, y, name, 'mean '+ii_variable, comment=comment, note=count, length=None)
#     #             ic_out.add_profile(profile, ii_variable)
#     #     return ic_out
#     #
#     # def max(self, var=None):
#     #     """
#     #     :param var:
#     #     :return:
#     #     """
#     #     variable = {}
#     #     snow_thickness = []
#     #     ice_thickness = []
#     #     date = []
#     #     ics_data = self.merge_bin()
#     #     for ii_core in ics_data.core:
#     #         ic_data = ics_data.core_data[ii_core]
#     #         if var is None:
#     #             for ii_variable in ics_data.variables:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #         else:
#     #             if not isinstance(var, list):
#     #                 var = [var]
#     #             for ii_variable in var:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #
#     #         str_temp = ''
#     #         for ii_core in ics_data.core:
#     #             str_temp += ii_core+', '
#     #         comment = 'statistic mean computed from ice cores: ' + str_temp[:-2]
#     #         name = ic_data.name.split('-')[0]+'-'+ic_data.name.split('-')[1][0:8]+'-mean'
#     #         ic_out = Core(name, date, ic_data.location, np.nanmean(ice_thickness), np.nanmean(snow_thickness), comment='mean value of '+str_temp[:-2])
#     #
#     #         for ii_variable in variable.keys():
#     #             y = variable[ii_variable][1]
#     #             x = np.namax(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #             count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
#     #
#     #             profile = Profile(x, y, name, 'mean '+ii_variable, comment=comment, note=count, length=None)
#     #             ic_out.add_profile(profile, ii_variable)
#     #     return ic_out
#     #
#     # def count(self, var=None):
#     #     """
#     #     :param var:
#     #     :return:
#     #     """
#     #     variable = {}
#     #     snow_thickness = []
#     #     ice_thickness = []
#     #     date = []
#     #     ics_data = self.merge_bin()
#     #     for ii_core in ics_data.core:
#     #         ic_data = ics_data.core_data[ii_core]
#     #         if var is None:
#     #             for ii_variable in ics_data.variables:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #         else:
#     #             if not isinstance(var, list):
#     #                 var = [var]
#     #             for ii_variable in var:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ic_data.snow_thickness is not None:
#     #                         snow_thickness.append(ic_data.snow_thickness)
#     #                     if ic_data.ice_thickness is not None:
#     #                        ice_thickness.append(ic_data.ice_thickness)
#     #                     if ic_data.date not in date:
#     #                         date.append(ic_data.date)
#     #
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #
#     #         str_temp = ''
#     #         for ii_core in ics_data.core:
#     #             str_temp += ii_core+', '
#     #         comment = 'statistic mean computed from ice cores: ' + str_temp[:-2]
#     #         name = ic_data.name.split('-')[0]+'-'+ic_data.name.split('-')[1][0:8]+'-mean'
#     #         ic_out = Core(name, date, ic_data.location, np.nanmean(ice_thickness), np.nanmean(snow_thickness), comment='mean value of '+str_temp[:-2])
#     #
#     #         for ii_variable in variable.keys():
#     #             y = variable[ii_variable][1]
#     #             x = np.nancount(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #             count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
#     #
#     #             profile = Profile(x, y, name, 'mean '+ii_variable, comment=comment, note=count, length=None)
#     #             ic_out.add_profile(profile, ii_variable)
#     #     return ic_out
#     #
#     # def statistic(self, var=None):
#     #     """
#     #     :param var:
#     #     :return:
#     #     """
#     #     ics_merge_bin_set = self.merge_bin()
#     #     variable = {}
#     #     for ii_core in ics_merge_bin_set.core:
#     #         ic_data = ics_merge_bin_set.core_data[ii_core]
#     #         if var is None:
#     #             for ii_variable in ics_merge_bin_set.variables:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #         else:
#     #             if not isinstance(var, list):
#     #                 var = [var]
#     #             for ii_variable in var:
#     #                 if ii_variable in ic_data.profiles.keys():
#     #                     if ii_variable not in variable.keys():
#     #                         variable[ii_variable] = [ic_data.profiles[ii_variable].x, ic_data.profiles[ii_variable].y]
#     #                     elif variable[ii_variable][1] is ic_data.profiles[ii_variable].y:
#     #                         variable[ii_variable][0] = np.vstack((variable[ii_variable][0], ic_data.profiles[ii_variable].x))
#     #                     else:
#     #                         logging.warning('vertical resolution is not the same between the cores')
#     #
#     #     for ii_variable in variable.keys():
#     #         y = variable[ii_variable][1]
#     #         x_mean = np.nanmean(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #         x_std = np.nanstd(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #         x_min = np.nanmin(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #         x_max = np.nanmax(np.atleast_2d(variable[ii_variable][0]), axis=0)
#     #         count = np.sum(~np.isnan(np.atleast_2d(variable[ii_variable][0])), axis=0)
#     #         x = [x_mean, x_std, x_min, x_max, count]
#     #
#     #         profile = Profile(x, y, ic_data.name, ii_variable+'-stat', comment='statistic envelop computed from merged'
#     #                                                                            'bin ice cores', note=None, length=None)
#     #
#     #         if ii_variable in ic_data.profiles.keys():
#     #             ic_data.del_profile(ii_variable)
#     #         ic_data.add_profile(profile, ii_variable)
#     #     return ic_data
#     #
#     # def plot_variable_stat(self, ax, variable, param_dict=None):
#     #     """
#     #     :param ax:
#     #     :param variable:
#     #     :param param_dict:
#     #     :return:
#     #     """
#     #
#     #     if variable not in self.variables:
#     #         logging.warning('variable %s not a statistical variable' % variable)
#     #         return None
#     #
#     #     ic_stat = self.statistic()
#     #
#     #     y = ic_stat.profiles[variable].y
#     #     x = ic_stat.profiles[variable].x
#     #
#     #     if len(y) == len(x[0]):
#     #         out = ax.plot(x[0], y, color='k', label='Mean')
#     #         ax.plot(x[2], y, color='b', label='Min')
#     #         ax.plot(x[3], y, color='r', label='Max')
#     #         ax.fill_betweenx(y, x[0] - x[1], x[0] + x[1], facecolor='black', alpha=0.3, label=str(u"\c2b1") + "std dev")
#     #         ax.set_ylim(max(ax.get_ylim()), 0)
#     #         ax.set_xlabel(variable + ' ' + si_prop_unit[variable])
#     #     else:
#     #         out = ax.step(np.append(x[0], x[0][-1]), y, color='k', label='Mean')
#     #         ax.step(np.append(x[2], x[2][-1]), y, color='b', label='Min')
#     #         ax.step(np.append(x[3], x[3][-1]), y, color='r', label='Max')
#     #
#     #         x_fill_l = [x[0][0]-x[1][0]]
#     #         x_fill_h = [x[0][0]+x[1][0]]
#     #         y_fill = y[0]
#     #         for ii in range(1, len(x[0])):
#     #             x_fill_l = np.append(x_fill_l, x[0][ii-1]-x[1][ii-1])
#     #             x_fill_l = np.append(x_fill_l, x[0][ii]-x[1][ii])
#     #             x_fill_h = np.append(x_fill_h, x[0][ii-1]+x[1][ii-1])
#     #             x_fill_h = np.append(x_fill_h, x[0][ii]+x[1][ii])
#     #             y_fill = np.append(y_fill, y[ii])
#     #             y_fill = np.append(y_fill, y[ii])
#     #
#     #         ax.fill_betweenx(y_fill, x_fill_l, x_fill_h, facecolor='black', alpha=0.3, label=str(u"\c2b1") + "std dev")
#     #         ax.set_ylim(max(ax.get_ylim()), 0)
#     #         ax.set_xlabel(variable + ' ' + si_prop_unit[variable])
#     #     return out

# particular for

def stack_core(ics_data):
    ics_stack = CoreStack()
    for f_core in sorted(ics_data.keys()):
        core_data = ics_data[f_core]
        ics_stack = ics_stack.add_profiles(core_data.profiles)
    return ics_stack


def add_variable(ics_stack, variable_dict, data):
    for col in data.columns.tolist():
        if col not in ics_stack.columns.tolist():
            ics_stack[col] = np.nan

    data_select, data_deselect = ics_stack.select_profile(variable_dict)

    index = data_select.index.tolist()
    data = pd.DataFrame([data.iloc[0]], columns=data.columns, index=index)

    ics_stack = ics_stack.fillna(data)
    return ics_stack


def DD_fillup(ics_stack, DD, freezup_dates):
    for f_day in ics_stack.coring_date.unique():
        #f_day = ics_stack.coring_date.unique()[0]
        data_label = ['FDD', 'TDD', 'DD', 'freezup_day']
        variable_dict = {'coring_date': f_day}

        # look for freezup_day
        if isinstance(f_day, np.datetime64):
            f_day = datetime.datetime.utcfromtimestamp((f_day - np.datetime64('1970-01-01T00:00:00')) / np.timedelta64(1, 's'))
        f_day = datetime.datetime(f_day.year, f_day.month, f_day.day)
        if f_day < datetime.datetime(f_day.year, 9, 1):
            freezup_day = datetime.datetime.fromordinal(freezup_dates[f_day.year])
        else:
            freezup_day = datetime.datetime.fromordinal(freezup_dates[f_day.year+1])

        # look for number of freezing/thawing degree day:

        if DD[f_day][1] < 0:
            data = [[DD[f_day][0], DD[f_day][1], DD[f_day][1], freezup_day]]
        else:
            data = [[DD[f_day][0], DD[f_day][1], DD[f_day][0], freezup_day]]

        data = pd.DataFrame(data, columns=data_label)
        ics_stack = add_variable(ics_stack, variable_dict, data)
    return ics_stack


def stack_DD_fud(ics_data, DD, freezup_dates):
    ics_data_stack = CoreStack()
    for ii_core in ics_data.keys():
        core = ics_data[ii_core]
        ics_data_stack = ics_data_stack.add_profiles(core.profiles)

    for ii_day in ics_data_stack.coring_date.unique():
        variable_dict = {'coring_date': ii_day}
        ii_day = pd.DatetimeIndex([ii_day])[0].to_datetime()

        # freezup day:
        if ii_day < datetime.datetime(ii_day.year, 9, 1):
            freezup_day = datetime.datetime.fromordinal(freezup_dates[ii_day.year - 1])
        else:
            freezup_day = datetime.datetime.fromordinal(freezup_dates[ii_day.year])
        # DD
        if DD[ii_day][1] < 0:
            data = [[DD[ii_day][0], DD[ii_day][1], DD[ii_day][1], np.datetime64(freezup_day)]]
        else:
            data = [[DD[ii_day][0], DD[ii_day][1], DD[ii_day][0], np.datetime64(freezup_day)]]
        data_label = ['coring_date', 'FDD', 'TDD', 'DD', 'freezup_day']
        data = pd.DataFrame(data, columns=data_label)

        ics_data_stack = ics_data_stack.add_variable(variable_dict, data)
    return ics_data_stack


def plot_mean_envelop(ic_data, variable_dict, ax=None, param_dic=None):

    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)


    if 'variable' not in variable_dict.keys():
        print("'a variable should be specified for plotting")
        return 0

    if 'bin_index' not in variable_dict.keys():
        print("DD index should be specified for plotting")
        return 0

    f_variable = variable_dict['variable']
    bin_index = variable_dict['bin_index']

    x_mean = ic_data.select_profile({'stats': 'mean', 'variable': f_variable, 'DD_index': bin_index})[0].reset_index()
    x_std = ic_data.select_profile({'stats': 'std', 'variable': f_variable, 'DD_index': bin_index})[0].reset_index()

    if x_mean[f_variable].__len__() !=0:
        if x_std.__len__() < x_mean.__len__():
            index = [ii for ii in x_mean.index.tolist() if ii not in x_std.index.tolist()]
            x_std = x_std.append(pd.DataFrame(np.nan, columns=x_std.columns.tolist(), index=index))

        if not x_mean[x_mean.variable == f_variable].y_low.isnull().all():
            y_low = x_mean['y_low']
            y_sup = x_mean['y_sup']
            y = np.concatenate((y_low.tolist(), [y_sup.tolist()[-1]]))
            x_std_l = x_mean[f_variable] - x_std[f_variable]
            x_std_h = x_mean[f_variable] + x_std[f_variable]

            x_std_l = seaice.icdtools.plt_step(x_std_l.tolist(), y).transpose()
            x_std_h = seaice.icdtools.plt_step(x_std_h.tolist(), y).transpose()
        elif x_mean[x_mean.variable == f_variable].y_low.isnull().all():
            y_std = x_mean['y_mid']
            x_std_l = np.array([x_mean[f_variable] - x_std[f_variable], y_std])
            x_std_h = np.array([x_mean[f_variable] + x_std[f_variable], y_std])

        ax.fill_betweenx(x_std_l[1,:], x_std_l[0,:], x_std_h[0,:], facecolor='black', alpha=0.3,
                                        label=str(r"$\pm$"+"std dev"))
    return ax


def plot_profile_variable(ic_data, variable_dict, ax=None, param_dict=None):
    """
    :param ic_data:
    :param variable_dict:
    :param ax:
    :param param_dict:
    :return:
    """
    if 'variable' not in variable_dict.keys():
        print("'a variable should be specified for plotting")
        return 0

    profile = ic_data.select_profile(variable_dict)[0]
    ax = plot_profile(profile, variable_dict['variable'], ax=ax, param_dict=param_dict)
    return ax

def plot_profile(ic_data, variable, ax=None, param_dict=None):
    """

    :param ax:
    :param profile:
    :param variable:
    :param param_dict:
    :return:
    """
    if variable is None or isinstance(variable, list):
        raise('one and only one variable should be specified')
        return 0

    if ax is None:
        plt.figure()
        ax = plt.subplot(1, 1, 1)

    if not ic_data[ic_data.variable == variable].y_low.isnull().all():  # salinity like curve
        # step function
        x = []
        y = []
        for ii in ic_data[ic_data.variable == variable].index.tolist():
            y.append(ic_data['y_low'][ii])
            y.append(ic_data['y_sup'][ii])
            x.append(ic_data[variable][ii])
            x.append(ic_data[variable][ii])
        if param_dict is None:
            ax.step(x, y)
        else:
            ax.step(x, y, **param_dict)
    else:
        # linear function
        x = ic_data[variable].tolist()
        y = ic_data.y_mid
        if param_dict is None:
            ax.plot(x, y)
        else:
            ax.plot(x, y, **param_dict)
    return ax


def plot_stat_envelop(ax, profile):
    return None


def plot_state_variable(profile_stack, ax=None, variables='state variables', color_map='core'):
    """
    :param profile_stack:
    :param ax:
    :param variables:
        default 'state variables' which plot salinity and temperature
    :param color:
    :return:
    """
    if variables == None:
        variables = np.unique(profile_stack.variable).tolist()
    elif not isinstance(variables, list):
        variables = [variables]
    elif variables == 'state variables':
        variables = ['salinity, temperature']

    if ax is None:
        fig = plt.figure()
        ax = []
        for ii in range(len(variables)):
            ax.append(plt.subplot(1, len(variables), ii+1))
    elif len(ax) != len(variables):
        logging.warning('ax (len %d) and variables (len %d) should be of same size' %(len(ax), len(variables)))
        return None

    # colors
    color = {}
    if  color_map is 'core':
        n_core = np.unique(profile_stack.core_name).__len__()
        color[ii] = [cm.jet(float(ii)/n_core) for ii in n_core]
    elif color_map is 'year':
        n_year = pd.unique([ii.year for ii in profile_stack.coring_date]).__len__()
        color[ii] = [cm.jet(float(ii)/n_year) for ii in n_year]

    for ii in len(variables):
        var = variables[ii]


def import_core(ic_filepath, variables=None, missing_value=float('nan')):
    """
    :param ic_filepath:
    :param variables:
    :param missing_value:
    """

    column_dict = {'S_ice': ['D', 'AB', 8, 6], 'T_ice': ['B', 'A', 8, 6]}

    wb = openpyxl.load_workbook(filename=ic_filepath, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    if variables == 'state variables':
        variables = ['T_ice', 'S_ice']
    elif variables is None:
        variables = []
        for ii_var in ws_name:
            if (ii_var not in ['summary', 'abreviation', 'locations', 'lists']) and (not ii_var.split('-')[-1]=='figure'):
                variables.append(ii_var)
    elif not isinstance(variables, list):
        variables = [variables]

    ic_name = ws_summary['C21'].value
    print(ic_name)

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value

    # location
    ic_loc = ws_summary['C5'].value

    # snow thickness
    temp_cell = ws_summary['C9']
    if isinstance(temp_cell.value, float):
        ic_snow_depth = temp_cell.value
    elif temp_cell.value in ['n/a', 'unknow', 'unknown', None, 'n/m']:
        ic_snow_depth = np.nan
    else:
        ic_snow_depth = np.nan

    # ice thickness
    temp_cell = ws_summary['C11']
    if isinstance(temp_cell.value, float) :
        ic_ice_thickness = temp_cell.value
    elif temp_cell.value in ['n/a', 'unknow', 'unknown', None, 'n/m']:
        ic_ice_thickness = np.nan
    else:
        ic_ice_thickness = np.nan

    imported_core = Core(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value

    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value

    temp_cell = ws_summary['C17']
    if temp_cell.value is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value

    temp_cell = ws_summary['C18']
    if temp_cell.value is not None or temp_cell.value not in ['n/m', 'n/a', 'unknown']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)] is not None and \
                    ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value is not None:
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value)
        ii_col += 1

    for variable in variables:
        if variable in ws_name:
            profile = read_variable(wb, variable, column_dict[variable])
            imported_core.add_profile(profile)
    return imported_core


def drop_profile(data, core_name, keys):
    data = data[(data.core_name != core_name) | (data.variable != keys)]
    return data


def discretize_profile(ic_data, y_bins=None, y_mid=None, variables=None, comment='n', display_figure='y'):
    """
    :param ic_data:
    :param y_bins:
    :param y_mid:
    :return:
    """
    # VARIABLES CHECK
    if y_bins is None and y_mid is None:
        y_bins = pd.Series(ic_data.y_low.dropna().tolist() + ic_data.y_sup.dropna().tolist()).sort_values().unique()
        y_mid = ic_data.y_mid.dropna().sort_values().unique()
    elif y_mid is None:
        y_mid = ic_data.y_mid.dropna().sort_values().unique()

    if variables is None:
        variables = [ii_variable for ii_variable in ic_data.variable.unique().tolist()]

    if not isinstance(variables, list):
        variables = [variables]

    for ii_variable in variables:
        if comment == 1:
            print(ii_variable)

        # For linear profile, like temperature
        if ic_data[ic_data.variable == ii_variable].y_low.isnull().all():
            yx = ic_data[ic_data.variable == ii_variable].set_index('y_mid').sort_index()[[ii_variable]]
            y2x = yx.reindex(y_mid).astype(float)
            y2x.ix[(y2x.index <= max(yx.index)) & (min(yx.index) <= y2x.index)] = y2x.interpolate(method='index')[(y2x.index <= max(yx.index)) & (min(yx.index) <= y2x.index)]
            temp = pd.DataFrame(columns=ic_data.columns.tolist(), index=range(y_mid.__len__()))
            temp.update(y2x.reset_index('y_mid'))
            yx = yx.reset_index(level='y_mid')

            ic_data_prop = ic_data.head(1)
            ic_data_prop = ic_data_prop.drop(ii_variable, 1)
            ic_data_prop = ic_data_prop.drop('y_low', 1)
            ic_data_prop = ic_data_prop.drop('y_mid', 1)
            ic_data_prop = ic_data_prop.drop('y_sup', 1)
            ic_data_prop = ic_data_prop.drop('sample_name', 1)
            temp.update(pd.DataFrame([ic_data_prop.iloc[0].tolist()], columns=ic_data_prop.columns.tolist(),
                                     index=temp.index.tolist()))
            temp['coring_date'] = temp['coring_date'].astype('datetime64[ns]')

            if display_figure == 'y':
                plt.close()
                plt.plot(yx[ii_variable], yx['y_mid'], 'k')
                plt.plot(temp[ii_variable], temp['y_mid'], 'xr')

        # For step profile, like salinity
        elif not ic_data[ic_data.variable == ii_variable].y_low.isnull().any():
            yx = ic_data[ic_data.variable == ii_variable].set_index('y_mid', drop=False).sort_index().as_matrix(
                ['y_low', 'y_sup', ii_variable])

            x_step = []
            y_step = []
            ii_bin = 0
            if yx[0, 0] < y_bins[0]:
                ii_yx = np.where(yx[:, 0] <= y_bins[0])[0][-1]
            else:
                ii_yx = 0
                while y_bins[ii_bin] < yx[ii_yx, 0]:
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(np.nan)
                    x_step.append(np.nan)
                    ii_bin += 1
                    y_bins[ii_bin]

            while ii_bin < y_bins.__len__() - 1:
                while y_bins[ii_bin + 1] <= yx[ii_yx, 1]:
                    S = yx[ii_yx, 2]
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(S)
                    x_step.append(S)
                    ii_bin += 1

                    if ii_bin == y_bins.__len__() - 1:
                        break

                L = (yx[ii_yx, 1] - y_bins[ii_bin])
                S = (yx[ii_yx, 1] - y_bins[ii_bin]) * yx[ii_yx, 2]
                while ii_yx < len(yx[:, 1]) - 1 and yx[ii_yx + 1, 1] <= y_bins[ii_bin + 1]:
                    L += (yx[ii_yx + 1, 1] - yx[ii_yx + 1, 0])
                    S += (yx[ii_yx + 1, 1] - yx[ii_yx + 1, 0]) * yx[ii_yx + 1, 2]
                    ii_yx += 1
                    if ii_yx == yx[:, 1].__len__() - 1:
                        break

                if ii_bin + 1 == y_bins.__len__():
                    break

                if yx[ii_yx, 1] <= y_bins[ii_bin + 1] and ii_yx + 1 < yx.__len__():
                    L += (y_bins[ii_bin + 1] - yx[ii_yx + 1, 0])
                    S += (y_bins[ii_bin + 1] - yx[ii_yx + 1, 0]) * yx[ii_yx + 1, 2]
                S = S / L
                if S != 0:
                    y_step.append(y_bins[ii_bin])
                    y_step.append(y_bins[ii_bin + 1])
                    x_step.append(S)
                    x_step.append(S)
                    ii_bin += 1
                ii_yx += 1
                if y_bins[ii_bin] >= yx[-1, 1]:
                    while ii_bin + 1 < y_bins.__len__():
                        y_step.append(y_bins[ii_bin])
                        y_step.append(y_bins[ii_bin + 1])
                        x_step.append(np.nan)
                        x_step.append(np.nan)
                        ii_bin += 1

            if display_figure == 'y' or display_figure == 'c':
                if display_figure != 'c':
                    plt.close()
                x = []
                y = []
                for ii in range(yx[:, 0].__len__()):
                    y.append(yx[ii, 0])
                    y.append(yx[ii, 1])
                    x.append(yx[ii, 2])
                    x.append(yx[ii, 2])
                plt.step(x, y)

            temp = pd.DataFrame(columns=ic_data.columns.tolist(), index=range(y_bins[:-1].__len__()))
            temp.update(pd.DataFrame(np.vstack((y_bins[:-1], y_bins[:-1] + np.diff(y_bins) / 2, y_bins[1:],
                                                [x_step[2 * ii] for ii in
                                                 range(int(x_step.__len__() / 2))])).transpose(),
                                     columns=['y_low', 'y_mid', 'y_sup', ii_variable], index=temp.index))

            # properties
            ic_data_prop = ic_data.head(1)
            ic_data_prop = ic_data_prop.drop(ii_variable, 1)
            ic_data_prop = ic_data_prop.drop('y_low', 1)
            ic_data_prop = ic_data_prop.drop('y_mid', 1)
            ic_data_prop = ic_data_prop.drop('y_sup', 1)
            ic_data_prop = ic_data_prop.drop('sample_name', 1)
            temp.update(pd.DataFrame([ic_data_prop.iloc[0].tolist()], columns=ic_data_prop.columns.tolist(),
                                     index=temp.index.tolist()))
            temp['coring_date'] = temp['coring_date'].astype('datetime64[ns]')

        sample_name = [ic_data_prop.core.tolist()[0] + '-' + str('%d' % ii) for ii in range(temp.__len__())]
        temp.update(pd.DataFrame(sample_name, columns=['sample_name'], index=temp.index.tolist()))

        ic_data = ic_data[
            (ic_data.core_name != ic_data.core_name.unique().tolist()[0]) | (ic_data.variable != ii_variable)]
        ic_data = ic_data.append(temp)
    return CoreStack(ic_data)


def make_section(core, variables=None, section_thickness=0.05):
    """
    :param core:
    :param variables:
    :param section_thickness:
    """
    if variables is None:
        variables = sorted(core.profiles.keys())
    if not isinstance(variables, list):
        variables = [variables]

    for ii_profile in variables:
        profile = core.profiles[ii_profile]
        if core.ice_thickness is not None:
            ice_thickness = core.ice_thickness
        else:
            ice_thickness = core.profiles[ii_profile].y[-1]

        y_mid_section = np.arange(section_thickness / 2, ice_thickness, section_thickness)
        delta_y = (ice_thickness + len(y_mid_section) * section_thickness) / 2

        if delta_y < ice_thickness:
            y_mid_section = np.append(y_mid_section, np.atleast_1d(delta_y))
        x = core.profiles[ii_profile].x
        y = np.array(core.profiles[ii_profile].y)

        if len(y) is len(x) + 1:
            y = (y[1:] + y[:-1]) / 2

        x_mid_section = np.interp(y_mid_section, y[~np.isnan(y)], x[~np.isnan(y)], left=np.nan, right=np.nan)

        profile.x = x_mid_section
        profile.y = y_mid_section
        profile.add_comment(
            'artificial section thickness computed with a vertical resolution of ' + str(section_thickness) + 'm')
        core.del_profile(ii_profile)
        core.add_profile(profile, ii_profile)
    return core


def get_filepath(data_dir, data_ext, subdir='no'):
    """
    :param data_dir:
    :param data_ext:
    :param subdir:
    """
    import os
    filepath = []
    if subdir == 'yes':
        for path, subdirs, files in os.walk(data_dir):
            subdirs.sort()
            files.sort()
            for name in files:
                if name.endswith(data_ext):
                    f = os.path.join(path, name)
                    filepath.append(f)
    else:
        files = os.listdir(data_dir)
        for name in files:
            if name.endswith(data_ext):
                f = os.path.join(data_dir, name)
                filepath.append(f)
    return filepath


def generate_source(data_dir, data_ext):
    """
    :param data_dir:
    :param data_ext:
    """
    ic_path_list = sorted(get_filepath(data_dir, data_ext))
    with open(data_dir + '/ic_list.txt', 'w') as f:
        for ii in range(0, len(ic_path_list)):
            f.write(ic_path_list[ii] + "\n")


def read_variable(wb, sheet_name, variable_dict):
    """
    :param wb:
    :param sheet_name:
    :param variable_dict:
    """

    [col_x, col_y, col_c, row_start] = variable_dict
    ice_core_spreadsheet = {'T_ice': ['temperature', 'T'], 'S_ice': ['salinity', 'S']}
    variable_name = {'T_ice':'temperature', 'S_ice':'salinity'}

    if sheet_name in wb.sheetnames:
        columns = [variable_name[sheet_name], 'y_low', 'y_sup', 'y_mid', 'comment', 'variable', 'core', 'note',
                   'ice_core_length', 'sample_name']
        profile = pd.DataFrame(columns=columns)

        # logging.info('%s sheet present, importing profile for %s' % (sheet_name, ice_core_spreadsheet[sheet_name]))

        if not isinstance(col_y, list) and len(col_y) < 2:
            col_y = [col_y]
        col_y_temp = [float('nan')] * len(col_y)
        for ii_col in range(0, len(col_y)):
            if isinstance(col_y[ii_col], str):
                col_y_temp[ii_col] = openpyxl.utils.column_index_from_string(col_y[ii_col])
        col_y = col_y_temp
        del col_y_temp
        if isinstance(col_x, str):
            col_x = openpyxl.utils.column_index_from_string(col_x)
        if isinstance(col_c, str):
            col_c = openpyxl.utils.column_index_from_string(col_c)

        ws = wb.get_sheet_by_name(sheet_name)
        profile_name = ws['C1'].value
        note = ws['C3'].value

        row_jj = row_start
        if len(col_y) == 1:
            col_flag = 1
            y_sup = np.nan
            y_low = np.nan
            while row_jj > 0:
                try:
                    y_mid = ws.cell(column=col_y[0], row=row_jj).value
                except IndexError:
                    break
                if y_mid is not None:
                    x = ws.cell(column=col_x, row=row_jj).value
                    comment = ws.cell(column=col_c, row=row_jj).value
                    if not isinstance(x, (float, int)):
                        if comment is not None:
                            comment = comment + '; value not defined'
                        else:
                            comment = 'value not defined'
                        x = np.nan
                    sample_name = profile_name + '-' + ice_core_spreadsheet[sheet_name][1] + str('-%02d' % (row_jj - row_start + 1))
                    measure = pd.DataFrame(
                        [[x, y_sup, y_low, y_mid, comment, ice_core_spreadsheet[sheet_name][0], profile_name, note, np.nan, sample_name]],
                        columns=columns, index=[sample_name])
                    profile = profile.append(measure)
                else:
                    break
                row_jj += 1
        elif len(col_y) == 2:
            col_flag = 1
            bin_flag = 1
            while bin_flag > 0:
                try:
                    y_sup = ws.cell(column=col_y[0], row=row_jj).value
                    y_low = ws.cell(column=col_y[1], row=row_jj).value
                except IndexError:
                    break
                if y_sup is not None and y_low is not None:
                    y_mid = (y_sup + y_low) / 2
                    x = ws.cell(column=col_x, row=row_jj).value
                    comment = ws.cell(column=col_c, row=row_jj).value
                    if not isinstance(x, (float, int)):
                        if comment is not None:
                            comment = comment + '; value not defined'
                        else:
                            comment = 'value not defined'
                        x = np.nan
                    sample_name = profile_name + '-' + ice_core_spreadsheet[sheet_name][1]+ str('-%02d' % (row_jj - row_start + 1))
                    measure = pd.DataFrame(
                        [[x, y_sup, y_low, y_mid, comment, ice_core_spreadsheet[sheet_name][0], profile_name, note, np.nan, sample_name]],
                        columns=columns, index=[sample_name])
                    profile = profile.append(measure)
                else:
                    break
                row_jj += 1

        # length
        length = ws['C2'].value
        if col_flag is 2:
            y_bin_length = min(profile['y_low'])-max(profile['y_sup'])
            if isinstance(y_bin_length, (float, int)):
                if isinstance(length, (float, int)) and y_bin_length > length:
                    length = y_bin_length
        elif not isinstance(length, (float, int)):
            length = np.nan
        profile.ice_core_length = length

        return profile
    else:
        logging.info('profile %s missing' % ice_core_spreadsheet[sheet_name])
        return None


def import_src(ics_filepath, variables=None):
    """
    import_src import ice core data which path is listed in the text file found in ics_filepath. File formatting: 1 ice
    core by line, entry beginning with # are ignored

    :param ics_filepath:
    :param variables:
    :return ic_dict: dictionnaries of ice core ice_core_name:ice_core_data
    """

    print('Ice core data importation in progress ...')
    a = open(ics_filepath)
    filepath = sorted([line.strip() for line in a])
    ic_dict = {}
    for ii in range(0, len(filepath)):
        if not filepath[ii].startswith('#'):
            ic_filepath = filepath[ii]
            ic_data = import_core(ic_filepath, variables=variables)
            ic_dict[ic_data.core_name] = ic_data
    print('done')
    return ic_dict


def import_list(ics_list, missing_value=float('nan'), log_level='warning', comment='n'):
    """
    :param ics_list:
    :param missing_value:
    :param log_level:
    :param comment:
    """
    print('Ice core data importation in progress ...')

    logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])

    ic_dict = {}
    for ii in range(0, len(ics_list)):
        if comment:
            print(ics_list[ii])
        ic_data = import_core(ics_list[ii], missing_value, comment=comment)
        ic_dict[ic_data.name] = ic_data

    logging.info('Ice core importation complete')
    print('done')

    return ic_dict


def import_variable(ic_path, variable='Salinity', missing_value=float('nan')):
    """
    :param ic_path:
    :param variable: string or array of string
        prop contain variable to import from a spreadsheet wb
    :param missing_value:
    :return:
    """

    variable_sheet_name = {'salinity': 'S_ice', 's': 'S_ice',
                           'temperature': 'T_ice', 't': 'T_ice',
                           'd18O': 'S_ice',
                           'dO2': 'S_ice',
                           'sediment': 'sediment', 'sed': 'sediment',
                           'Chla': 'algal_pigment', 'chlorophyl a': 'algal_pigment',
                           'Phae': 'algal_pigment'}
    variable_coordinate = {'salinity': ['D', 'AB', 8, 6], 's': ['D', 'AB', 8, 6],
                           'temperature': ['B', 'A', 8, 6],
                           'd180': [6, 'AB', 8, 6],
                           'sediment': [6, 'AB', 8, 6], 'sed':[6, 'AB', 8, 6],
                           'dO2': [7, 'AB', 8, 6],
                           'Chla': [5, 'AB', 8, 6], 'chlorophyl a': [5, 'AB', 8, 6],
                           'Phae': [6, 'AB', 8, 6]}

    wb = openpyxl.load_workbook(filename=ic_path, use_iterators=True)  # load the xlsx spreadsheet
    ws_name = wb.get_sheet_names()
    ws_summary = wb.get_sheet_by_name('summary')  # load the data from the summary sheet

    # extract basic information about the ice core
    # name
    ic_name = ws_summary['C21'].value
    print(ic_name)

    # logging.basicConfig(filename=LOG_FILENAME, level=LOG_LEVELS[log_level])
    # logging.info('Processing ' + ic_name + '...')

    # date
    temp_cell = ws_summary['C2']
    if isinstance(temp_cell.value, (float, int)):
        ic_date = datetime.datetime(1899, 12, 30) + datetime.timedelta(temp_cell.value)
    else:
        ic_date = temp_cell.value

    # location
    ic_loc = ws_summary['C5'].value

    # snow thickness
    temp_cell = ws_summary['C9']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_snow_depth = missing_value
    else:
        ic_snow_depth = temp_cell.value

    # ice thickness
    temp_cell = ws_summary['C11']
    if temp_cell is None or temp_cell.value in ['n/m', 'n/a', 'unknow']:
        ic_ice_thickness = missing_value
    else:
        ic_ice_thickness = temp_cell.value

    imported_core = Core(ic_name, ic_date, ic_loc, ic_ice_thickness, ic_snow_depth)

    # surface temperature
    temp_cell = ws_summary['C15']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_air = temp_cell.value
    temp_cell = ws_summary['C16']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_snow = temp_cell.value
    temp_cell = ws_summary['C17']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknow']:
        imported_core.t_ice0 = temp_cell.value
    temp_cell = ws_summary['C18']
    if temp_cell is not None or temp_cell.value not in ['n/m', 'n/a', 'unknown']:
        imported_core.t_water = temp_cell.value

    ii_row = 23
    ii_col = 3
    while ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)] is not None and \
                    ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value is not None:
        imported_core.add_corenames(ws_summary[openpyxl.cell.get_column_letter(ii_col) + str('%.0f' % ii_row)].value)
        ii_col += 1

    # import properties:
    if isinstance(variable, str):
        variable = [variable]
    for ii_variable in variable:
        print(ii_variable)
        if variable_sheet_name[ii_variable.lower()] in ws_name:
            col_x_temp = variable_coordinate[ii_variable.lower()][0]
            col_y_temp = variable_coordinate[ii_variable.lower()][1]
            col_c_temp = variable_coordinate[ii_variable.lower()][2]
            row_start_temp = variable_coordinate[ii_variable.lower()][3]
            profile = read_variable(wb, sheet_name=variable_sheet_name[ii_variable.lower()], col_x=col_x_temp, col_y=col_y_temp,
                                    col_c=col_c_temp, row_start=row_start_temp)
            imported_core.add_profile(profile, ii_variable.lower())
            logging.info('\t%s profile imported' % ii_variable.lower())
        else:
            logging.info('\t%s profile missing' % ii_variable.lower())
    if isinstance(imported_core, Core):
        return imported_core



def bottom_reference(ics_stack, ice_depth=None, comment = 'n'):
    for f_core in ics_stack.core_name.unique():
        ic_data = ics_stack[ics_stack.core_name == f_core]

        if not np.isnan(ic_data.ice_thickness.astype(float)).all():
            hi = ic_data.ice_thickness.astype(float).dropna().unique()
            hi = hi[0]
        elif not np.isnan(ic_data.ice_core_length.astype(float)).all():
            hi = ic_data.ice_core_length.astype(float).dropna().unique()
            hi = hi[0]
        else:
            hi = []
            for ff_core in ic_data.core_collection:
                if not np.isnan(ic_data.ice_thickness.astype(float)).all():
                    hi.append(ic_data.ice_thickness.astype(float).dropna().unique())[0]
                if not np.isnan(ic_data.ice_core_length.astype(float)).all():
                    hi.append(ic_data.ice_thickness.astype(float).dropna().unique())[0]
            hi = np.nanmean(hi)
        if not np.isnan(hi):
            for f_variable in ic_data.variable.unique():
                ics_stack.loc[(ics_stack.core_name == f_core) & (ics_stack.variable == f_variable), 'y_low'] = hi - ics_stack.loc[(ics_stack.core_name == f_core) & (ics_stack.variable == f_variable), 'y_low']
                ics_stack.loc[(ics_stack.core_name == f_core) & (ics_stack.variable == f_variable), 'y_mid'] = hi - ics_stack.loc[(ics_stack.core_name == f_core) & (ics_stack.variable == f_variable), 'y_mid']
                ics_stack.loc[(ics_stack.core_name == f_core) & (ics_stack.variable == f_variable), 'y_sup'] = hi - ics_stack.loc[(ics_stack.core_name == f_core) & (ics_stack.variable == f_variable), 'y_sup']
        else:
            ics_stack = ics_stack.remove_profiles(f_core)
        if comment == 'y':
            print(f_core, hi)
    return ics_stack


def calc_prop(ic_data, si_prop, s_profile_shape = 'linear'):
    """
    :param ic_data:
    :param si_prop:
    :param si_prop: 'linear' or 'step'
    :return:
    """

    if not isinstance(si_prop, list):
        si_prop = [si_prop]

    ## function variable:
    property_stack = seaice.corePanda.CoreStack()

    ## look for variable:
    core_variable = {}
    for f_variable in ic_data.variable.unique():
        core_variable[f_variable] = ic_data[ic_data.variable == f_variable]['core_name'].unique().tolist()

    if 'temperature' not in core_variable or 'salinity' not in core_variable:
        return None

    if core_variable['temperature'].__len__() > 1:
        print('average temperature from cores:', core_variable['temperature'])
        for t_core in core_variable['temperature']:
            ty0 = ic_data[ic_data.core_name == t_core][ic_data.variable == 'temperature']['y_mid'].tolist()
            t_profile0 = ic_data[ic_data.core_name == t_core][ic_data.variable == 'temperature']['temperature'].tolist()

            if t_core == core_variable['temperature'][0]:
                ty = np.array(ty0)
                t_profile = np.array([t_profile0])
            else:
                ty_temp = np.sort(np.unique(np.concatenate((ty, ty0))))
                t_profile_temp = np.interp(ty_temp, ty0, t_profile0)
                for ii in range(t_profile.shape[0]):
                    t_profile_temp = np.vstack((t_profile_temp, np.interp(ty_temp, ty, t_profile[ii])))
                ty = ty_temp
                t_profile = t_profile_temp
        t_profile = np.nanmean(t_profile_temp, axis=0)
    else:
        t_core = core_variable['temperature'][0]
        t_profile = ic_data[ic_data.core_name == t_core][ic_data.variable == 'temperature']['temperature'].tolist()
        ty = np.array(ic_data[ic_data.core_name == t_core][ic_data.variable == 'temperature']['y_mid'].tolist())

    for f_prop in si_prop:
        if f_prop not in seaice.properties.si_prop_list.keys():
            print('property %s not defined in the ice core property module' % property)
            # return None

        property = seaice.properties.si_prop_list[f_prop]

        for s_core in core_variable['salinity']:
            function = getattr(seaice.properties, property.replace(" ", "_"))
            s_profile = ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity']['salinity'].tolist()

            if s_profile_shape == 'linear':
                sy = ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity']['y_low'].tolist()
                xy = np.array([[s_profile[0], sy[0]]])
                for ii in range(s_profile.__len__()):
                    if s_profile[ii] != xy[-1, 0]:
                        if np.isnan(s_profile[ii]):
                            if ~np.isnan(xy[-1, 0]):
                                xy = np.vstack((xy, [s_profile[ii], sy[ii]]))
                        else:
                            xy = np.vstack((xy, [s_profile[ii], sy[ii]]))
                if np.isnan(s_profile[ii]):
                    if np.isnan(xy[-1, 0]):
                        xy = np.vstack((xy, [s_profile[ii], sy[ii]]))

                s_profile = xy[:, 0]
                sy = xy[:, 1]
                sy = np.concatenate((sy, [ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity']['y_sup'].tolist()[-1]]))
                sy = sy[:-1]+np.diff(sy)/2

            else:
                sy = ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity']['y_mid'].tolist()

            if not np.array_equal(sy, ty):
                x = function(np.interp(sy, ty, t_profile), s_profile)
            else:
                x = function(t_profile, s_profile)

            # TODO: check small difference between the curve
            if s_profile_shape == 'linear':
                # plt.figure()
                # plt.plot(x, sy, 'r', linewidth=1)
                x = np.interp(ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity']['y_mid'].tolist(),
                              sy, x, left=np.nan, right=np.nan)
                # plt.plot(x, ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity']['y_mid'].tolist(), 'b')

            index = ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity'].index
            property_frame = pd.DataFrame(x, columns=[property.replace(" ", "_")], index=index)
            variable_frame = pd.DataFrame(property.replace(" ", "_"), columns=['variable'], index=index)
            core_frame = ic_data[ic_data.core_name == s_core][ic_data.variable == 'salinity'].drop('salinity',
                                                                                                   axis=1).drop(
                'variable', axis=1)
            if s_profile_shape == 'linear':
                core_frame = core_frame.drop('y_low', axis=1).drop('y_sup', axis=1)
            core_frame = pd.concat((core_frame, property_frame, variable_frame), axis=1)

            # TODO: add note in the core_frame
            # for ii_index in core_frame.index:
            #     if isinstance(core_frame.iloc[ii_index]['note'], float) and np.isnan(core_frame.iloc[ii_index]['comment']):
            #         core_frame.iloc[ii_index]['note'] = property +' computed from ' + s_core + '(S) and ' + t_core + '(T)'

            property_stack = property_stack.append(core_frame, ignore_index=True, verify_integrity=False)

    return property_stack